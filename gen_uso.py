from lxml import etree
from loguru import logger
import openpyxl
import uuid, os, shutil
from lxml.etree import CDATA

def str_find(str1, arr):
    i = 0
    for el in arr:
        if str(str1).find(el) > -1:
            return True
def translate(str):
    dict = {".":"_",
            "/":"_",
            "\\":"_",
            ",":"_",
            ":":"_",
            ";":"_",
            "А":"A",
            "Б":"_B",
            "В":"B",
            "Г":"G",
            "Д":"D",
            "Е":"E",
            "Ё":"_E",
            "Ж":"J",
            "З":"Z",
            "И":"I",
            "Й":"_I",
            "К":"K",
            "Л":"L",
            "М":"M",
            "Н":"H",
            "О":"O",
            "П":"_P",
            "Р":"P",
            "С":"C",
            "Т":"T",
            "У":"U",
            "Ф":"F",
            "Х":"X",
            "Ц":"C",
            "Ч":"CH",
            "Ш":"SH",
            "Щ":"SCH",
            "Ь":"b",
            "Ы":"_",
            "Ъ":"_",
            "Э":"E",
            "Ю":"U",
            "Я":"YA",
            "а":"a",
            "б":"_b",
            "в":"b",
            "г":"g",
            "д":"d",
            "е":"e",
            "ё":"_e",
            "ж":"j",
            "з":"z",
            "и":"i",
            "й":"_i",
            "к":"k",
            "л":"l",
            "м":"m",
            "н":"h",
            "о":"o",
            "п":"_p",
            "р":"p",
            "с":"c",
            "т":"t",
            "у":"u",
            "ф":"f",
            "х":"x",
            "ц":"c",
            "ч":"ch",
            "ш":"sh",
            "щ":"sch",
            "ь":"b",
            "ы":"_",
            "ъ":"_",
            "э":"e",
            "ю":"u",
            "я":"ya"}

    intab = '.-пПаАфз/еЕсС'
    outtab = '__ppaafz_eEcC'
    trantab = str.maketrans(dict)
    outstr = str.translate(trantab)
    return outstr

# Генерация корзин УСО
@logger.catch
def generate_uso(path_template, exel, flag_ASPT, prefix_system):
    list_module = {'MK-516-008A' : ['AIs', 'type_MK_516_008(AI8)' , '454fa324-27ee-4c5b-852b-10e43769c2fa'],
                   'MK-514-008'  : ['AOs', 'type_MK_514_008(AO)'  , 'e76165af-10c9-4743-b092-8f5dcb3e6e12'],
                   'MK-521-032'  : ['DIs', 'type_MK_521_032(DI)'  , '54337da0-d138-41b0-aefe-20366697201e'],
                   'MK-531-032'  : ['DOs', 'type_MK_531_032(DO)'  , '20cd1522-2d06-49e6-a55d-e0801aeeb4e9'],
                   'MK-545-010'  : ['CNs', 'type_MK_545_010(CN)'  , 'c70fe2c3-a605-4c9d-b471-23e410350ddf'],
                   'MK-550-024'  : ['PSUs','type_MK_550_024(PSU)' , '6d539303-1528-4442-bc2e-1f08a49f1567'],
                   'MK-541-002'  : ['RSs', 'type_MK_541_002(RS)'  , 'dc2b3d53-089e-4f3f-9ecd-4098cdfa823c'],}

    attrib_ss = {'1' : ['X'       , '5', '5' ],
                 '2' : ['Y'       , '5', '34'],
                 '3' : ['Rotation', '5', '0' ],
                 '4' : ['Height'  , '5', '23']}
    attrib_basket = {'X'              : '70',
                     'Y'              : '',
                     'ZValue'         : '0',
                     'Rotation'       : '0',
                     'Scale'          : '1',
                     'Visible'        : 'true',
                     'Opacity'        : '1',
                     'Enabled'        : 'true',
                     'Tooltip'        : '',
                     'Width'          : '730',
                     'Height'         : '160',
                     'RoundingRadius' : '0',
                     'PenColor'       : '4278190080',
                     'PenStyle'       : '0',
                     'PenWidth'       : '1',
                     'BrushColor'     : '4278190080',
                     'BrushStyle'     : '0'}
    attrib_t_input_link_d = {'X'              : '10',
                             'Y'              : '25',
                             'Rotation'       : '0',
                             'Width'          : '220',
                             'Height'         : '20'}
    attrib_t_output_link_d = {'X'              : '10',
                              'Y'              : 'replace',
                              'Rotation'       : '0',
                              'Width'          : '220',
                              'Height'         : '20'}
    attrib_modul = {'X'        : [''   , '5'],
                    'Y'        : ['0'  , '5'],
                    'Rotation' : ['0'  , '5'],
                    'Width'    : ['40' , '5'],
                    'Height'   : ['160', '5']}
    attrib_AIs_AOs = {'RightPopUp'          : 'true',
                      'DownPopUp_faceplate' : '',
                      'UpPopUp_faceplate'   : '',
                      '_init_path'          : ''}
    attrib_DIs_DOs = {'RightPopUp'          : 'true',
                      'DownPopUp_faceplate' : 'true',
                      'UpPopUp_faceplate'   : '',
                      '_init_path'          : ''}
    attrib_PSUs = {'RightPopUp'          : 'true',
                   'DownPopUp_faceplate' : 'true',
                   'UpPopUp_faceplate'   : '',
                   '_init_path'          : ''}

    attrib_RSs = {'RightPopUp'          : 'true',
                   'DownPopUp_faceplate' : 'true',
                   'UpPopUp_faceplate'   : '',
                   '_init_path'          : ''}

    attrib_CNs = {'_init_path'         : '',
                  'eth1_animation'     : 'true',
                  'eth2_animation'     : 'true',
                  'RightPopUp'         : 'true',
                  'port1_device'       : '',
                  'port2_device'       : '',
                  'DownPopUp_faceplate': 'true',
                  'UpPopUp_faceplate'  : '',}

    attrib_link_input_output = {'link_1_is_on'      : 'true',
                                'link_2_is_on'      : 'true',
                                '_init_path_link_1' : '',
                                '_init_path_link_2' : '',
                                'link_1_inv'        : 'true',
                                'link_2_inv'        : 'true'}
    attrib_link_input_output_d = {'X'              : '50',
                                  'Y'              : '60',
                                  'Rotation'       : '0',
                                  'Width'          : '70.5',
                                  'Height'         : '120'}
    attrib_point = {'X' : '0',
                    'Y' : '0'}

    wb = openpyxl.load_workbook(exel, read_only=True)
    sheet_HW  = wb['HW']
    sheet_Net = wb['Net']
    sheet_USO = wb['USO']
    sheet_DI  = wb['DI']
    sheet_AI  = wb['AI']
    # максимальное число рядов и столбцов
    rows_HW     = sheet_HW.max_row
    column_HW   = sheet_HW.max_column
    rows_Net    = sheet_Net.max_row
    rows_USO    = sheet_USO.max_row
    column_USO  = sheet_USO.max_column
    # Пустые переменные
    basket         = []
    signals        = []

    name_uso_      = ''
    in_link        = ''
    out_link       = ''
    name_in        = ''
    name_out       = ''
    number_in      = ''
    number_out     = ''

    flag_service   = False
    # Из табл HW определим корзины и модули из списка
    # Шкаф КЦ исключен из списка
    logger.info(f'Кадры диагностики: сбор данных')
    for i in range(6, rows_HW + 1):
        data_modul     = []
        service_signal = []
        tag_uso    = sheet_HW.cell(row=i, column=3).value
        if tag_uso is None: continue
        tag_cut    = sheet_HW.cell(row=i, column=3).value.partition('_A')[0]
        name_uso   = sheet_HW.cell(row=i, column=4).value
        numer_rack = sheet_HW.cell(row=i, column=5).value

        if name_uso_ != name_uso:
            flag_service = True
            name_uso_ = name_uso
        # Ищем модули и всю необходимую информацию
        for j in range(7, column_HW + 1):
            if j % 2 != 0:
                type_module      = translate(str(sheet_HW.cell(row=i, column=j).value))
                number_modul     = sheet_HW.cell(row=2, column=j).value
                number_modul_cut = sheet_HW.cell(row=2, column=j).value.partition('_')[2]
                if type_module is None: continue
                data_modul.append(dict(type_module      = type_module,
                                       number_modul     = number_modul,
                                       number_modul_cut = number_modul_cut))
        # По тегу УСО ищем концы линка корзин
        for rows in range(4, rows_Net + 1):
            in_          = sheet_Net.cell(row=rows, column=3).value
            out_         = sheet_Net.cell(row=rows, column=4).value
            if in_  == tag_uso: in_link  = out_
            if out_ == tag_uso: out_link = in_
        # Находим соответсвие шкафа для линков
        for rows in range(4, rows_HW + 1):
            name_in_  = sheet_HW.cell(row=rows, column=3).value
            name_out_ = sheet_HW.cell(row=rows, column=3).value
            if name_in_ == in_link  :
                name_in   = sheet_HW.cell(row=rows, column=4).value
                number_in = sheet_HW.cell(row=rows, column=5).value
            if name_out_ == out_link :
                name_out   = sheet_HW.cell(row=rows, column=4).value
                number_out = sheet_HW.cell(row=rows, column=5).value
        # Cлужебные сигналы
        for row in range(4, rows_USO + 1):
            row_name_uso = sheet_USO.cell(row=row, column=4).value
            if (name_uso == row_name_uso) and flag_service is True:
                for column in range(5, column_USO + 1):
                    signal = sheet_USO.cell(row=row, column=column).value
                    if signal is None: continue

                    if str_find(signal, {'DI'}):
                        signal_split = signal.split('.')[0]
                        for row_ in sheet_DI.rows:
                            if signal_split == f'DI[{row_[0].value}]':
                                service_signal.append(dict(type = 'DI',
                                                           word = f'Diskrets.{translate(str(row_[2].value))}'))
                                break
        for row in range(4, rows_USO + 1):
            row_name_uso_ = sheet_USO.cell(row=row, column=4).value
            if (name_uso == row_name_uso_) and flag_service is True:
                # Флаг одного захода для одного шкафа
                flag_service = False
                for column in range(5, column_USO + 1):
                    signal = sheet_USO.cell(row=row, column=column).value
                    if signal is None: continue

                    if str_find(signal, {'AI'}):
                        for row_ in sheet_AI.rows:
                            if str(signal) == f'AI[{row_[0].value}]':
                                service_signal.append(dict(type='AI',
                                                           word=f'Analogs.{translate(str(row_[2].value))}'))
        # Упаковка
        basket.append(dict(data_modul       = data_modul,
                           tag_uso          = tag_uso,
                           tag_cut          = tag_cut,
                           name_uso         = name_uso,
                           numer_rack       = numer_rack,
                           in_link          = in_link,
                           out_link         = out_link,
                           name_in          = name_in,
                           number_in        = number_in,
                           name_out         = name_out,
                           number_out       = number_out,
                           service_signal   = service_signal))
    logger.info(f'Кадры диагностики: данные собраны')
    # Раскладываем на кадры
    uso_save       = ''
    for uso in basket:
        data_modul       = uso['data_modul']
        tag_uso          = uso['tag_uso']
        tag_cut          = uso['tag_cut']
        name_uso         = uso['name_uso']
        numer_rack       = uso['numer_rack']
        in_link          = uso['in_link']
        out_link         = uso['out_link']
        name_in          = uso['name_in']
        name_out         = uso['name_out']
        number_in        = uso['number_in']
        number_out       = uso['number_out']
        service_signal   = uso['service_signal']
        logger.info(f'Кадры диагностики: генерация {name_uso} корзина {numer_rack}')
        # Создадим новую картинку по шаблону, но сначала проверим не существует ли она, а если есть то удалим
        if uso_save != name_uso:
            path_gen_pic = f'{path_template}D_{prefix_system}{tag_cut}.omobj'
            if os.path.isfile(path_gen_pic): os.remove(path_gen_pic)
            shutil.copy2(f'{path_template}D_USO_Template.omobj', path_gen_pic)

            parser = etree.XMLParser(remove_blank_text=True, strip_cdata=False)
            tree   = etree.parse(path_gen_pic, parser)
            root   = tree.getroot()

            uso_save            = name_uso
            t_input_one_0ff_flag = True

            # Редактируем шаблон под нужное УСО
            for lvl_one in root.iter('type'):
                if lvl_one.attrib['name'] == 'name':
                    lvl_one.attrib['name'] = f'D_{prefix_system}{tag_cut}'
                if lvl_one.attrib['display-name'] == 'name':
                    lvl_one.attrib['display-name'] = f'D_{prefix_system}{tag_cut}'
                if lvl_one.attrib['uuid'] == 'uuid':
                    lvl_one.attrib['uuid'] = str(uuid.uuid1())
                if not flag_ASPT:
                    for lvl_two in lvl_one.iter('designed'):
                        # Координата Width
                        if lvl_two.attrib['value'] == '1670':
                            lvl_two.attrib['value'] = '1420'

                for lvl_two in lvl_one.iter('object'):
                    if lvl_two.attrib['name'] == 't_uso_title':
                        for lvl_three in lvl_two.iter('designed'):
                            if lvl_three.attrib['value'] == 'Rename':
                                lvl_three.attrib['value'] = name_uso

                    if not flag_ASPT:
                        if lvl_two.attrib['name'] == 'r_ss':
                            for lvl_three in lvl_two.iter('designed'):
                                if lvl_three.attrib['value'] == '950':
                                    lvl_three.attrib['value'] = '780'

                    if lvl_two.attrib['name'] == 'Rename_link':
                        lvl_two.attrib['name'] = f'_link_D_{prefix_system}{tag_cut}_for_enable'
                    if lvl_two.attrib['display-name'] == 'Rename_link':
                        lvl_two.attrib['display-name'] = f'_link_D_{prefix_system}{tag_cut}_for_enable'

                    for lvl_two in lvl_one.iter('do-on'):
                        if lvl_two.attrib['name'] == 'Handler_1':
                            for lvl_three in lvl_two.iter('body'):
                                lvl_three.text = CDATA(f'_link_D_{prefix_system}{tag_cut}_for_enable.Enabled=false;')

                        if lvl_two.attrib['name'] == 'Handler_2':
                            for lvl_three in lvl_two.iter('body'):
                                lvl_three.text = CDATA(f'_link_D_{prefix_system}{tag_cut}_for_enable.Enabled=true;')
        # Добавляем подписи на входе и выходе УСО
        for lvl_one in root.iter('type'):
            if t_input_one_0ff_flag is True:
                # Вход
                object = etree.Element('object')
                object.attrib['access-modifier'] = 'private'
                object.attrib['name'] = f't_input_link'
                object.attrib['display-name'] = f't_input_link'
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = f'type_text_link'
                object.attrib['base-type-id'] = f'786eccde-a924-4fc9-bf04-e7b38b0e922d'
                object.attrib['ver'] = '5'
                lvl_one.append(object)
                for key, value in attrib_t_input_link_d.items():
                    designed = etree.Element("designed")
                    designed.attrib['target'] = key
                    designed.attrib['value']  = f'{value}'
                    designed.attrib['ver']    = f'5'
                    object.append(designed)
                init = etree.Element("init")
                init.attrib['target'] = f'_init_path'
                init.attrib['ver']    = '5'
                if (str_find(out_link, {'KC'})) and (str_find(out_link, {'A1'}) or str_find(out_link, {'A2'})):
                    init.attrib['value']  = f'Diag.MNs.{out_link}_01'
                else:
                    init.attrib['value']  = f'Diag.CNs.{out_link}_01'
                object.append(init)
                init_ = etree.Element("init")
                init_.attrib['target'] = f'_link_init_ApSource'
                init_.attrib['ver'] = '5'
                init_.attrib['ref'] = f'unit.Global.global_ApSource'
                object.append(init_)

                # Выход
                object = etree.Element('object')
                object.attrib['access-modifier'] = 'private'
                object.attrib['name'] = f't_output_link'
                object.attrib['display-name'] = f't_output_link'
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = f'type_text_link'
                object.attrib['base-type-id'] = f'786eccde-a924-4fc9-bf04-e7b38b0e922d'
                object.attrib['ver'] = '5'
                lvl_one.append(object)

                for key, value in attrib_t_output_link_d.items():
                    designed = etree.Element("designed")
                    designed.attrib['target'] = key
                    designed.attrib['value'] = f'{value}'
                    designed.attrib['ver'] = f'5'
                    object.append(designed)

                init = etree.Element("init")
                init.attrib['target'] = f'_init_path'
                init.attrib['ver'] = '5'
                if (str_find(in_link, {'KC'})) and (str_find(in_link, {'A1'}) or str_find(in_link, {'A2'})):
                    init.attrib['value'] = f'Diag.MNs.{in_link}_01'
                else:
                    init.attrib['value'] = f'Diag.CNs.{in_link}_01'
                in_link_old = init.attrib['value']
                object.append(init)

                in_link_old = init.attrib['value']

                init_ = etree.Element("init")
                init_.attrib['target'] = f'_link_init_ApSource'
                init_.attrib['ver'] = '5'
                init_.attrib['ref'] = f'unit.Global.global_ApSource'
                object.append(init_)

                t_input_one_0ff_flag = False
            # Изменяем координату у выходной подписи УСО
            for lvl_two in lvl_one.iter('object'):
                if lvl_two.attrib['name'] == 't_output_link':
                    for lvl_three in lvl_two.iter('designed'):
                        if numer_rack <= 1:
                            if lvl_three.attrib['value'] == 'replace':
                                lvl_three.attrib['value'] = f'{265 + (177 * (numer_rack - 1))}'
                        else:
                            if lvl_three.attrib['value'] == f'{265 + (177 * (numer_rack - 2))}':
                                lvl_three.attrib['value'] = f'{265 + (177 * (numer_rack - 1))}'

            for lvl_two in lvl_one.iter('object'):
                if lvl_two.attrib['name'] == 't_output_link':
                    for lvl_three in lvl_two.iter('init'):
                        if lvl_three.attrib['target'] == f'_init_path':
                            if (str_find(in_link, {'KC'})) and (str_find(in_link, {'A1'}) or str_find(in_link, {'A2'})):
                                lvl_three.attrib['value'] = f'Diag.MNs.{in_link}_01'
                            else:
                                lvl_three.attrib['value'] = f'Diag.CNs.{in_link}_01'

        # Служебные сигналы
        count_srv = 0
        for srv_sinal in service_signal:
            type = srv_sinal['type']
            if type == 'DI':
                word       = srv_sinal['word']
                count_srv += 1
                for lvl_one in root.iter('type'):
                    for lvl_two in lvl_one.iter('object'):
                        # Находим служебные сигналы
                        if lvl_two.attrib['name'] == 'r_ss':
                            object = etree.Element('object')
                            object.attrib['access-modifier'] = 'private'
                            object.attrib['name'] = f'type_srv_signal_{count_srv}'
                            object.attrib['display-name'] = f'type_srv_signal_{count_srv}'
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = 'type_srv_signal'
                            object.attrib['base-type-id'] = '72176618-ccac-488c-b1d6-d570e5505e1c'
                            object.attrib['ver'] = '5'
                            lvl_two.append(object)
                            for key, value in attrib_ss.items():
                                designed = etree.Element("designed")
                                designed.attrib['target'] = value[0]
                                if key == '2': designed.attrib['value']  = f'{32 + (27 * (count_srv - 1))}'
                                else         : designed.attrib['value'] = value[2]
                                designed.attrib['ver']    = value[1]
                                object.append(designed)
                            init = etree.Element("init")
                            init.attrib['target'] = '_init_path'
                            init.attrib['value']  = word
                            init.attrib['ver']    = '5'
                            object.append(init)
                            init_1 = etree.Element("init")
                            init_1.attrib['target'] = '_link_init_ApSource'
                            init_1.attrib['ver']    = '5'
                            init_1.attrib['ref']    = 'unit.Global.global_ApSource'
                            object.append(init_1)
            if type == 'AI':
                word       = srv_sinal['word']
                count_srv += 1
                for lvl_one in root.iter('type'):
                    for lvl_two in lvl_one.iter('object'):
                        # Находим служебные сигналы
                        if lvl_two.attrib['name'] == 'r_ss':
                            object = etree.Element('object')
                            object.attrib['access-modifier'] = 'private'
                            object.attrib['name'] = 'type_analog_srv'
                            object.attrib['display-name'] = 'type_analog_srv'
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = 'type_analog_srv'
                            object.attrib['base-type-id'] = 'c5d10192-c8ea-4db8-a5ab-15b09b9b2266'
                            object.attrib['ver'] = '5'
                            lvl_two.append(object)
                            for key, value in attrib_ss.items():
                                designed = etree.Element("designed")
                                designed.attrib['target'] = value[0]
                                if key == '2': designed.attrib['value'] = f'{32 + (27 * (count_srv - 1))}'
                                else         : designed.attrib['value'] = value[2]
                                designed.attrib['ver']    = value[1]
                                object.append(designed)
                            init = etree.Element("init")
                            init.attrib['target'] = '_init_path'
                            init.attrib['value']  = word
                            init.attrib['ver']    = '5'
                            object.append(init)
                            init_1 = etree.Element("init")
                            init_1.attrib['target'] = '_link_init_ApSource'
                            init_1.attrib['ver']    = '5'
                            init_1.attrib['ref']    = 'unit.Global.global_ApSource'
                            object.append(init_1)
        # Создадим корзины и заполним модули
        for lvl_one in root.iter('type'):
            object = etree.Element('object')
            object.attrib['access-modifier'] = 'private'
            object.attrib['name'] = f'r_basket_{numer_rack}'
            object.attrib['display-name'] = f'r_basket_{numer_rack}'
            object.attrib['uuid'] = str(uuid.uuid1())
            object.attrib['base-type'] = f'Rectangle'
            object.attrib['base-type-id'] = f'15726dc3-881e-4d8d-b0fa-a8f8237f08ca'
            object.attrib['ver'] = '5'
            lvl_one.append(object)
            for key, value in attrib_basket.items():
                designed = etree.Element("designed")
                designed.attrib['target'] = key
                if key == 'Y': designed.attrib['value'] = f'{70 + (180 * (numer_rack - 1))}'
                else         : designed.attrib['value'] = value
                designed.attrib['ver'] = f'5'
                object.append(designed)
            # Cчетчики модулей в корзине
            count_AIs    = 0
            count_AOs    = 0
            count_DIs    = 0
            count_DOs    = 0
            count_CNs    = 0
            count_CPUs   = 0
            count_PSUs   = 0
            count_RSs    = 0
            count_modul  = 0
            # Заполним модули
            for data in data_modul:
                type_mod   = data['type_module']
                number     = data['number_modul']
                number_cut = data['number_modul_cut']
                for key, value in list_module.items():
                    if key == type_mod:
                        count_modul += 1
                        if value[0] == 'AIs'   :
                            count_AIs += 1
                            current_modul = count_AIs
                            type_attrib   = attrib_AIs_AOs
                        if value[0] == 'AOs'   :
                            count_AOs    += 1
                            current_modul = count_AOs
                            type_attrib   = attrib_AIs_AOs
                        if value[0] == 'DIs'   :
                            count_DIs    += 1
                            current_modul = count_DIs
                            type_attrib   = attrib_DIs_DOs
                        if value[0] == 'DOs'   :
                            count_DOs    += 1
                            current_modul = count_DOs
                            type_attrib   = attrib_DIs_DOs
                        if value[0] == 'CNs'   :
                            count_CNs    += 1
                            current_modul = count_CNs
                            type_attrib   = attrib_CNs
                        if value[0] == 'PSUs'  :
                            count_PSUs   += 1
                            current_modul = count_PSUs
                            type_attrib   = attrib_PSUs
                        if value[0] == 'RSs'   :
                            count_RSs    += 1
                            current_modul = count_RSs
                            type_attrib   = attrib_RSs
                        # Находим корзину и добавляем нужный модуль
                        for lvl_two in lvl_one.iter('object'):
                            if lvl_two.attrib['name'] == f'r_basket_{numer_rack}':
                                object = etree.Element('object')
                                object.attrib['access-modifier'] = 'private'
                                object.attrib['name'] = f'{value[1]}_{current_modul}'
                                object.attrib['display-name'] = ''
                                object.attrib['uuid'] = str(uuid.uuid1())
                                object.attrib['base-type'] = value[1]
                                object.attrib['base-type-id'] = value[2]
                                object.attrib['ver'] = '5'
                                lvl_two.append(object)

                                for key_d, value_d in attrib_modul.items():
                                    designed = etree.Element("designed")
                                    designed.attrib['target'] = key_d
                                    if key_d == 'X': designed.attrib['value'] = f'{40 * (count_modul - 1)}'
                                    else           : designed.attrib['value'] = value_d[0]
                                    designed.attrib['ver'] = value_d[1]
                                    object.append(designed)

                                for key_i, value_i in type_attrib.items():
                                    init = etree.Element("init")
                                    init.attrib['target'] = key_i
                                    init.attrib['ver']    = '5'

                                    if key_i == '_init_path': init.attrib['value']  = f'Diag.{value[0]}.{tag_uso}{number}'
                                    if key_i == 'port1_device': init.attrib['value']  = f'{name_out} корзина А{number_out}'
                                    if key_i == 'port2_device': init.attrib['value']  = f'{name_in} корзина А{number_in}'
                                    if numer_rack >= 2:
                                        if key_i == 'UpPopUp_faceplate'  : init.attrib['value'] = 'true'
                                        if key_i == 'DownPopUp_faceplate': init.attrib['value'] = 'false'
                                    if numer_rack <= 2:
                                        if key_i == 'UpPopUp_faceplate'  : init.attrib['value'] = 'false'
                                        if key_i == 'DownPopUp_faceplate': init.attrib['value'] = 'true'
                                    if key_i == 'RightPopUp'    : init.attrib['value'] = value_i
                                    if key_i == 'eth1_animation': init.attrib['value'] = value_i
                                    if key_i == 'eth2_animation': init.attrib['value'] = value_i
                                    init.attrib['ver']    = '5'
                                    object.append(init)

                                init_ = etree.Element("init")
                                init_.attrib['target'] = f'_link_init_ApSource'
                                init_.attrib['ver']    = '5'
                                init_.attrib['ref']    = f'unit.Global.global_ApSource'
                                object.append(init_)
        # Добавляем линки к корзинам - Port 1
        for lvl_one in root.iter('type'):
            # Входная линия - Port 1
            object = etree.Element('object')
            object.attrib['access-modifier'] = 'private'
            object.attrib['name'] = f'l_input_to_A{numer_rack}'
            object.attrib['display-name'] = f'l_input_to_A{numer_rack}'
            object.attrib['uuid'] = str(uuid.uuid1())
            object.attrib['base-type'] = f'type_line_for_connect'
            object.attrib['base-type-id'] = f'9ce8edc0-9c10-4a3b-9263-da44abb267e1'
            object.attrib['ver'] = '5'
            lvl_one.append(object)

            for key, value in attrib_link_input_output_d.items():
                designed = etree.Element("designed")
                designed.attrib['target'] = key
                designed.attrib['ver'] = f'5'
                if key == 'Rotation': designed.attrib['value'] = f'{value}'
                if key == 'Width': designed.attrib['value'] = f'{value}'
                if key == 'X':
                    if numer_rack <= 1:
                        designed.attrib['value'] = f'{value}'
                    else:
                        designed.attrib['value'] = f'120.5'
                if key == 'Y':
                    if numer_rack == 1:
                        designed.attrib['value'] = f'{value}'
                    elif numer_rack == 2:
                        designed.attrib['value'] = f'207'
                    elif numer_rack == 3:
                        designed.attrib['value'] = f'387'
                    else:
                        designed.attrib['value'] = f'567'
                if key == 'Height':
                    if numer_rack <= 1:
                        designed.attrib['value'] = f'{value}'
                    else:
                        designed.attrib['value'] = f'103'
                object.append(designed)

            for key, value in attrib_link_input_output.items():
                init = etree.Element("init")
                init.attrib['target'] = key
                init.attrib['ver'] = f'5'
                if key == '_init_path_link_1':
                    if (str_find(out_link, {'KC'})) and (str_find(out_link, {'A1'}) or str_find(out_link, {'A2'})):
                        init.attrib['value'] = f'Diag.MNs.{out_link}_01.ch_CN_02.ePNotLink'
                    else:
                        init.attrib['value'] = f'Diag.CNs.{out_link}_01.ch_CN_02.ePNotLink'
                if key == '_init_path_link_2':
                    if (str_find(tag_uso, {'KC'})) and (str_find(tag_uso, {'A1'}) or str_find(tag_uso, {'A2'})):
                        init.attrib['value'] = f'Diag.MNs.{tag_uso}_01.ch_CN_01.ePNotLink'
                    else:
                        init.attrib['value'] = f'Diag.CNs.{tag_uso}_01.ch_CN_01.ePNotLink'
                if key == 'link_1_is_on': init.attrib['value'] = f'{value}'
                if key == 'link_2_is_on': init.attrib['value'] = f'{value}'
                if key == 'link_1_inv'  : init.attrib['value'] = f'{value}'
                if key == 'link_2_inv'  : init.attrib['value'] = f'{value}'
                object.append(init)

            init_ = etree.Element("init")
            init_.attrib['target'] = f'_link_init_ApSource'
            init_.attrib['ver'] = '5'
            init_.attrib['ref'] = f'unit.Global.global_ApSource'
            object.append(init_)

            for i in range(3):
                object_ = etree.Element('object')
                object_.attrib['access-modifier'] = 'private'
                object_.attrib['name'] = f'Point_{i + 1}'
                object_.attrib['display-name'] = f'Point_{i + 1}'
                object_.attrib['uuid'] = str(uuid.uuid1())
                object_.attrib['base-type'] = f'Point'
                object_.attrib['base-type-id'] = f'467f1af0-7bb4-4a61-b6fb-06e7bfd530d6'
                object_.attrib['ver'] = '5'
                object.append(object_)

                for key, value in attrib_point.items():
                    designed_ = etree.Element("designed")
                    designed_.attrib['target'] = key
                    designed_.attrib['ver'] = f'5'
                    if key == 'X':
                        if (numer_rack == 1) and (i + 1) == 1: designed_.attrib['value'] = f'0'
                        if (numer_rack == 1) and (i + 1) == 2: designed_.attrib['value'] = f'0'
                        if (numer_rack == 1) and (i + 1) == 3: designed_.attrib['value'] = f'70.5'

                        if (numer_rack >= 2) and (i + 1) == 1: designed_.attrib['value'] = f'-70.5'
                        if (numer_rack >= 2) and (i + 1) == 2: designed_.attrib['value'] = f'-70.5'
                        if (numer_rack >= 2) and (i + 1) == 3: designed_.attrib['value'] = f'0'

                    if key == 'Y':
                        if (numer_rack == 1) and (i + 1) == 1: designed_.attrib['value'] = f'0'
                        if (numer_rack == 1) and (i + 1) == 2: designed_.attrib['value'] = f'120'
                        if (numer_rack == 1) and (i + 1) == 3: designed_.attrib['value'] = f'120'

                        if (numer_rack >= 2) and (i + 1) == 1: designed_.attrib['value'] = f'50'
                        if (numer_rack >= 2) and (i + 1) == 2: designed_.attrib['value'] = f'153'
                        if (numer_rack >= 2) and (i + 1) == 3: designed_.attrib['value'] = f'153'
                    object_.append(designed_)
        # Добавляем линки к корзинам - Port 2
        for lvl_one in root.iter('type'):
            # Входная линия - Port 2
            object = etree.Element('object')
            object.attrib['access-modifier'] = 'private'
            object.attrib['name'] = f'l_output_to_A{numer_rack}'
            object.attrib['display-name'] = f'l_output_to_A{numer_rack}'
            object.attrib['uuid'] = str(uuid.uuid1())
            object.attrib['base-type'] = f'type_line_for_connect'
            object.attrib['base-type-id'] = f'9ce8edc0-9c10-4a3b-9263-da44abb267e1'
            object.attrib['ver'] = '5'
            lvl_one.append(object)

            for key, value in attrib_link_input_output_d.items():
                designed = etree.Element("designed")
                designed.attrib['target'] = key
                designed.attrib['ver'] = f'5'
                if key == 'Rotation': designed.attrib['value'] = f'{value}'
                if key == 'Width'   : designed.attrib['value'] = f'{value}'
                if key == 'Height'  : designed.attrib['value'] = f'50'

                if key == 'X':
                    if numer_rack <= 2: designed.attrib['value'] = f'120.75'
                    else              : designed.attrib['value'] = f'120.5'
                if key == 'Y':
                    if   numer_rack == 1: designed.attrib['value'] = f'207.5'
                    elif numer_rack == 2: designed.attrib['value'] = f'387.5'
                    elif numer_rack == 3: designed.attrib['value'] = f'567'
                    else                : designed.attrib['value'] = f'747'
                object.append(designed)

            for key, value in attrib_link_input_output.items():
                init = etree.Element("init")
                init.attrib['target'] = key
                init.attrib['ver'] = f'5'
                if key == '_init_path_link_1':
                    if (str_find(tag_uso, {'KC'})) and (str_find(tag_uso, {'A1'}) or str_find(tag_uso, {'A2'})):
                        init.attrib['value'] = f'Diag.MNs.{tag_uso}_01.ch_CN_02.ePNotLink'
                    else:
                        init.attrib['value'] = f'Diag.CNs.{tag_uso}_01.ch_CN_02.ePNotLink'
                if key == '_init_path_link_2':
                    if (str_find(in_link, {'KC'})) and (str_find(in_link, {'A1'}) or str_find(in_link, {'A2'})):
                        init.attrib['value'] = f'Diag.MNs.{in_link}_01.ch_CN_01.ePNotLink'
                    else:
                        init.attrib['value'] = f'Diag.CNs.{in_link}_01.ch_CN_01.ePNotLink'
                if key == 'link_1_is_on': init.attrib['value'] = f'{value}'
                if key == 'link_2_is_on': init.attrib['value'] = f'{value}'
                if key == 'link_1_inv'  : init.attrib['value'] = f'{value}'
                if key == 'link_2_inv'  : init.attrib['value'] = f'{value}'
                object.append(init)

            init_ = etree.Element("init")
            init_.attrib['target'] = f'_link_init_ApSource'
            init_.attrib['ver'] = '5'
            init_.attrib['ref'] = f'unit.Global.global_ApSource'
            object.append(init_)

            for i in range(3):
                object_ = etree.Element('object')
                object_.attrib['access-modifier'] = 'private'
                object_.attrib['name'] = f'Point_{i + 1}'
                object_.attrib['display-name'] = f'Point_{i + 1}'
                object_.attrib['uuid'] = str(uuid.uuid1())
                object_.attrib['base-type'] = f'Point'
                object_.attrib['base-type-id'] = f'467f1af0-7bb4-4a61-b6fb-06e7bfd530d6'
                object_.attrib['ver'] = '5'
                object.append(object_)

                for key, value in attrib_point.items():
                    designed_ = etree.Element("designed")
                    designed_.attrib['target'] = key
                    designed_.attrib['ver'] = f'5'
                    if key == 'X':
                        if (i + 1) == 1: designed_.attrib['value'] = f'0'
                        if (i + 1) == 2: designed_.attrib['value'] = f'-70.5'
                        if (i + 1) == 3: designed_.attrib['value'] = f'-70.5'

                    if key == 'Y':
                        if (i + 1) == 1: designed_.attrib['value'] = f'0'
                        if (i + 1) == 2: designed_.attrib['value'] = f'0'
                        if (i + 1) == 3: designed_.attrib['value'] = f'50'

                    object_.append(designed_)
        tree.write(path_gen_pic, pretty_print=True,encoding='utf-8')

# Отдельный лист со служебными сигналами
@logger.catch
def generate_serv_signal(path_template, exel, flag_ASPT, prefix_system):
    attrib_ss = {'1': ['X', '5', '5'],
                 '2': ['Y', '5', '34'],
                 '3': ['Rotation', '5', '0'],
                 '4': ['Height', '5', '23']}
    wb = openpyxl.load_workbook(exel, read_only=True)
    sheet_USO = wb['USO']
    sheet_DI  = wb['DI']
    sheet_AI  = wb['AI']
    # максимальное число рядов и столбцов
    rows_USO    = sheet_USO.max_row
    column_USO  = sheet_USO.max_column
    # Пустые переменные
    service_signal = []
    count_uso      = []
    logger.info(f'Генерация служебных сигналов на отдельный кадр начата')
    # Cлужебные сигналы
    for row in range(4, rows_USO + 1):
        name_uso = sheet_USO.cell(row=row, column=4).value
        tag_uso  = sheet_USO.cell(row=row, column=3).value
        count_uso.append(dict(uso     = name_uso,
                              tag_uso = tag_uso))
    # DI
    for row in range(4, rows_USO + 1):
        name_uso = sheet_USO.cell(row=row, column=4).value
        for column in range(5, column_USO + 1):
            signal = sheet_USO.cell(row=row, column=column).value
            if signal is None: continue

            if str_find(signal, {'DI'}):
                signal_split = signal.split('.')[0]
                for row_ in sheet_DI.rows:
                    if signal_split == f'DI[{row_[0].value}]':
                        service_signal.append(dict(type         = 'DI',
                                                   word         = f'Diskrets.{translate(str(row_[2].value))}',
                                                   row_name_uso = name_uso))
                        break
    # AI
    for row in range(4, rows_USO + 1):
        name_uso = sheet_USO.cell(row=row, column=4).value
        for column in range(5, column_USO + 1):
            signal = sheet_USO.cell(row=row, column=column).value
            if signal is None: continue

            if str_find(signal, {'AI'}):
                for row_ in sheet_AI.rows:
                    if str(signal) == f'AI[{row_[0].value}]':
                        service_signal.append(dict(type         = 'AI',
                                                   word         = f'Analogs.{translate(str(row_[2].value))}',
                                                   row_name_uso = name_uso))

    # Раскладываем на кадры
    uso_save = ''
    for data in count_uso:
        uso     = data['uso']
        tag_cut = data['tag_uso']

        logger.info(f'Генерация служебных сигналов на отдельный кадр: генерация {uso}')
        # Создадим новую картинку по шаблону, но сначала проверим не существует ли она, а если есть то удалим
        if uso_save != uso:
            path_gen_pic = f'{path_template}D_{prefix_system}{tag_cut}.omobj'
            if os.path.isfile(path_gen_pic): os.remove(path_gen_pic)
            shutil.copy2(f'{path_template}D_USO_Template.omobj', path_gen_pic)

            parser = etree.XMLParser(remove_blank_text=True, strip_cdata=False)
            tree   = etree.parse(path_gen_pic, parser)
            root   = tree.getroot()

            uso_save = uso

            # Редактируем шаблон под нужное УСО
            for lvl_one in root.iter('type'):
                if lvl_one.attrib['name'] == 'name':
                    lvl_one.attrib['name'] = f'D_{prefix_system}{tag_cut}'
                if lvl_one.attrib['display-name'] == 'name':
                    lvl_one.attrib['display-name'] = f'D_{prefix_system}{tag_cut}'
                if lvl_one.attrib['uuid'] == 'uuid':
                    lvl_one.attrib['uuid'] = str(uuid.uuid1())
                if not flag_ASPT:
                    for lvl_two in lvl_one.iter('designed'):
                        # Координата Width
                        if lvl_two.attrib['value'] == '1670':
                            lvl_two.attrib['value'] = '1420'

                for lvl_two in lvl_one.iter('object'):
                    if lvl_two.attrib['name'] == 't_uso_title':
                        for lvl_three in lvl_two.iter('designed'):
                            if lvl_three.attrib['value'] == 'Rename':
                                lvl_three.attrib['value'] = name_uso

                    if not flag_ASPT:
                        if lvl_two.attrib['name'] == 'r_ss':
                            for lvl_three in lvl_two.iter('designed'):
                                if lvl_three.attrib['value'] == '950':
                                    lvl_three.attrib['value'] = '780'

                    if lvl_two.attrib['name'] == 'Rename_link':
                        lvl_two.attrib['name'] = f'_link_D_{prefix_system}{tag_cut}_for_enable'
                    if lvl_two.attrib['display-name'] == 'Rename_link':
                        lvl_two.attrib['display-name'] = f'_link_D_{prefix_system}{tag_cut}_for_enable'

                    for lvl_two in lvl_one.iter('do-on'):
                        if lvl_two.attrib['name'] == 'Handler_1':
                            for lvl_three in lvl_two.iter('body'):
                                lvl_three.text = CDATA(f'_link_D_{prefix_system}{tag_cut}_for_enable.Enabled=false;')

                        if lvl_two.attrib['name'] == 'Handler_2':
                            for lvl_three in lvl_two.iter('body'):
                                lvl_three.text = CDATA(f'_link_D_{prefix_system}{tag_cut}_for_enable.Enabled=true;')

        # Служебные сигналы
        count_srv = 0
        for srv_sinal in service_signal:
            type     = srv_sinal['type']
            name_uso = srv_sinal['row_name_uso']

            if name_uso == uso:
                if type == 'DI':
                    word       = srv_sinal['word']
                    count_srv += 1
                    for lvl_one in root.iter('type'):
                        for lvl_two in lvl_one.iter('object'):
                            # Находим служебные сигналы
                            if lvl_two.attrib['name'] == 'r_ss':
                                object = etree.Element('object')
                                object.attrib['access-modifier'] = 'private'
                                object.attrib['name'] = f'type_srv_signal_{count_srv}'
                                object.attrib['display-name'] = f'type_srv_signal_{count_srv}'
                                object.attrib['uuid'] = str(uuid.uuid1())
                                object.attrib['base-type'] = 'type_srv_signal'
                                object.attrib['base-type-id'] = '72176618-ccac-488c-b1d6-d570e5505e1c'
                                object.attrib['ver'] = '5'
                                lvl_two.append(object)
                                for key, value in attrib_ss.items():
                                    designed = etree.Element("designed")
                                    designed.attrib['target'] = value[0]
                                    if key == '2': designed.attrib['value']  = f'{32 + (27 * (count_srv - 1))}'
                                    else         : designed.attrib['value'] = value[2]
                                    designed.attrib['ver']    = value[1]
                                    object.append(designed)
                                init = etree.Element("init")
                                init.attrib['target'] = '_init_path'
                                init.attrib['value']  = word
                                init.attrib['ver']    = '5'
                                object.append(init)
                                init_1 = etree.Element("init")
                                init_1.attrib['target'] = '_link_init_ApSource'
                                init_1.attrib['ver']    = '5'
                                init_1.attrib['ref']    = 'unit.Global.global_ApSource'
                                object.append(init_1)
                if type == 'AI':
                    word       = srv_sinal['word']
                    count_srv += 1
                    for lvl_one in root.iter('type'):
                        for lvl_two in lvl_one.iter('object'):
                            # Находим служебные сигналы
                            if lvl_two.attrib['name'] == 'r_ss':
                                object = etree.Element('object')
                                object.attrib['access-modifier'] = 'private'
                                object.attrib['name'] = 'type_analog_srv'
                                object.attrib['display-name'] = 'type_analog_srv'
                                object.attrib['uuid'] = str(uuid.uuid1())
                                object.attrib['base-type'] = 'type_analog_srv'
                                object.attrib['base-type-id'] = 'c5d10192-c8ea-4db8-a5ab-15b09b9b2266'
                                object.attrib['ver'] = '5'
                                lvl_two.append(object)
                                for key, value in attrib_ss.items():
                                    designed = etree.Element("designed")
                                    designed.attrib['target'] = value[0]
                                    if key == '2': designed.attrib['value'] = f'{32 + (27 * (count_srv - 1))}'
                                    else         : designed.attrib['value'] = value[2]
                                    designed.attrib['ver']    = value[1]
                                    object.append(designed)
                                init = etree.Element("init")
                                init.attrib['target'] = '_init_path'
                                init.attrib['value']  = word
                                init.attrib['ver']    = '5'
                                object.append(init)
                                init_1 = etree.Element("init")
                                init_1.attrib['target'] = '_link_init_ApSource'
                                init_1.attrib['ver']    = '5'
                                init_1.attrib['ref']    = 'unit.Global.global_ApSource'
                                object.append(init_1)

        tree.write(path_gen_pic, pretty_print=True,encoding='utf-8')