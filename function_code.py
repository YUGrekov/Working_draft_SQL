from lxml import etree
from loguru import logger
import openpyxl
import uuid
import os, re, codecs
import math
import json
#from Connection_DB.connection_db import *


class Initialisation_path():
    def __init__(self, exel, map, map_modbus, map_modbus503, omx, name_prefix, prefix_driver):
        self.exel          = exel
        self.map           = map
        self.map_mb        = map_modbus
        self.map_mb503     = map_modbus503
        self.omx           = omx
        self.name_prefix   = name_prefix
        self.prefix_driver = prefix_driver
        self.data          = self.read_exel()

    actives_table     = ['AI', 'AO', 'DI', 'DO', 'ZD', 'VS', 'VSGRP', 'TrendGRP', 'SPGRP', 'Pic', 'SS', 'VV',
                         'AIGRP', 'КД', 'UMPNA', 'USO', 'UTS', 'KTPRP', 'KTPR', 'KTPRA', 'KTPRS', 'DPS', 'GMPNA',
                         'ModBus', 'PI', 'PZ', 'UPTS', 'MSG', 'NPS','RSreq','RS',
                         'TM_DP', 'TM_TS', 'TM_TU', 'TM_TI2', 'TM_TI4', 'TM_TII', 'TM_TR2', 'TM_TR4']

    zd_hat_table      = ['№', 'Переменная', 'Идентификатор', 'Название', 'КВО', 'КВЗ', 'МПО', 'МПЗ', 'Дист_ф', 'Муфта',
                        'Авар. Привода', 'Открыть', 'Закрыть', 'Остановить', 'Откр. остановить', 'Закр. остановить',
                        'КВО_и', 'КВЗ_и', 'МПО_и', 'МПЗ_и', 'Дист_и', 'Муфта_и', 'Авар. Привода_и', 'Открыть_и',
                        'Закрыть_и', 'Остановить_и', 'Откр. Остановить_и', 'Закр. Остановить_и', 'Наличие ИНТЕРФЕЙСА',
                        'Отсут. связи', 'Закр. с БРУ', 'Стоп с БРУ', 'Напряж.', 'Напряж. ЩСУ',
                        'Напряж. в цепях сигнализации', 'Испр. цепей откр.', 'Испр. цепей закр.', 'ВММО', 'ВММЗ',
                        'Замораж. при подозрит. Изм', 'Это клапан',  '% открытия', 'Pic', 'Тип БУР задвижки',
                        'AlphaHMI', 'AlphaHMI_PIC1', 'AlphaHMI_PIC1_Number_kont', 'AlphaHMI_PIC2',
                        'AlphaHMI_PIC2_Number_kont', 'AlphaHMI_PIC3', 'AlphaHMI_PIC3_Number_kont', 'AlphaHMI_PIC4',
                        'AlphaHMI_PIC4_Number_kont', 'Короткое название']
    vs_hat_table      = ['№', 'Переменная', 'Идентификатор', 'Название', 'Короткое название (для отображения на кадре)',
                        'Группа', 'Номер в группе', 'МП', 'Давл. норма', 'Напр.', 'Напр. СШ', 'Испр. Цепей вкл.',
                        'Внешняя авария', 'Датчик давл. неиспр.', 'Включить', 'Отключить', 'АПВ не требуется', 'Pic',
                        'Таблица сообщений', 'Это клапан / интерфейсная вспомсистема', 'AlphaHMI']
    ai_hat_table      = ['№', 'Идентификатор', 'Название', 'Группа уставок аналогов', 'Единица измерения',
                        'Группа сброса трендов', 'Давление нефти/ нефтепродукта (флаг для пересчета в кгс/см2)',
                        'Группа аналогов', 'Пол. мин.', 'Пол. макс.', 'Инж. Мин.', 'Инж. Макс.', 'Отображаемая точность значения',
                        'Единица измерения физической величины (АЦП)', 'AlphaHMI', 'УСО, модуль, канал',
                        'Исправность канала', 'Сигнализация', 'Сообщение', 'Номер НА или вспом.', 'Предохранитель',
                        'Pic', 'Номер НА или вспом.', 'Вибрация насоса', 'Вибрация ЭД', 'Ток ЭД НА',
                        'Давление на вых. вспом.', '№ уставки мин. авар.', '№ уставки мин. пред.', '№ уставки макс. пред.',
                        '№ уставки макс. авар.']
    ao_hat_table      = ['№', 'Идентификатор', 'Название','УСО, модуль, канал','Исправность канала']
    di_hat_table      = ['№', 'Идентификатор', 'Название', 'Inv','ErrValue','isDI_NC', 'isAI_Warn','isAI_Avar', 'Msg',
                         'isModuleNC','priority[0]', 'priority[1]', 'pNC_AI', 'AlphaHMI','pValue','pHealth',
                         'TS_ID','Pic', 'Подпись на мнемокадре']
    do_hat_table      = ['№','Переменная','Идентификатор','Название','УСО, модуль, канал','Исправность канала']
    trengrp_hat_table = ['ID', 'ParentID', 'Название группы', 'Используется']
    spgrp_hat_table   = ['Название группы']
    pic_hat_table     = ['№', 'Название', 'Кадр IFix(*.grf)','Переменная','Pic']
    ss_hat_table      = ['№', 'Переменная', 'Название', '№ в массиве stateRSreq', 'Pic']
    aigrp_hat_table   = ['Название группы', 'Мин.6', 'Мин.5', 'Мин.4', 'Мин.3', 'Мин.2', 'Мин.',
                                            'Макс.', 'Макс.2', 'Макс.3', 'Макс.4', 'Макс.5', 'Макс.6']
    kd_hat_table      = ['Тип сигнала', 'Шкаф', 'Tэг', 'Наименование', 'КлК', 'Конт', 'Корз', 'Мод', 'Кан']
    uso_hat_table     = ['Идентификатор', 'Название']
    uts_hat_table     = ['№', 'Идентификатор', 'Название', 'Включить', 'Не требует автоотключения', 'Проверка',
                        'Квитирование', 'Сирена', 'Номер листа (для ВУ)', 'Номер порядка (для ВУ)']
    ktpr_hat_table    = ['№', '№ защиты в РД', 'Pic', 'Аварийный параметр', 'Автоматическая деблокировка защиты',
                        'Битовая маска принадлежности защиты группе', 'Временная уставка', 'Группа уставок',
                        'Закрытие воздушных клапанов (жалюзийных решёток) насосного зала',
                        'Закрытие воздушных клапанов (жалюзийных решёток) помещения компрессорной подпора воздуха ЭД',
                        'Закрытие задвижек между ПНС и МНС', 'Закрытие задвижек между РП и ПНС',
                        'Закрытие задвижек на входе и выходе МНА', 'Закрытие задвижек на входе и выходе МНС',
                        'Закрытие задвижек на входе и выходе ПНА', 'Закрытие задвижек на входе и выходе ПНС',
                        'Закрытие задвижек на входе НПС', 'Закрытие задвижек на входе РП',
                        'Закрытие задвижек на входе ССВД', 'Закрытие задвижек на входе узла РД',
                        'Закрытие задвижек на входе ФГУ', 'Закрытие задвижек на выходе НПС',
                        'Закрытие задвижек на выходе узла РД',
                        'Закрытие секущей задвижки узла подключения объекта нефтедобычи/ нефтепереработки',
                        'Запрет маскирования', 'Защита по пожару', 'Идентификатор', 'Название', 'Номер защиты (для ВУ)',
                        'Номер листа (для ВУ)', 'Отключение АВО', 'Отключение антиконденсационных электронагревателей ЭД',
                        'Отключение беспромвальных вентиляторов электрозала',
                        'Отключение вентиляторов водоохлаждения системы оборотного водоснабжения',
                        'Отключение внешнего контура охлаждения ЧРП МНА', 'Отключение внешнего контура охлаждения ЧРП ПНА',
                        'Отключение воздушных охладителей системы запирания торцовых уплотнений всех МНА',
                        'Отключение воздушных охладителей системы запирания торцовых уплотнений отключенных НА',
                        'Отключение вытяжных вентиляторов в помещении РД', 'Отключение вытяжных вентиляторов в помещении ССВД',
                        'Отключение вытяжных вентиляторов в помещении централизованной маслосистемы',
                        'Отключение вытяжных вентиляторов маслоприямка в электрозале',
                        'Отключение вытяжных вентиляторов насосного зала МНС',
                        'Отключение вытяжных вентиляторов насосного зала ПНС',
                        'Отключение компрессоров подпора воздуха ЭД', 'Отключение крышных вентиляторов насосного зала МНС',
                        'Отключение крышных вентиляторов насосного зала ПНС', 'Отключение маслонасосов',
                        'Отключение маслонасосов после сигнала "остановлен" НА', 'Отключение насосов артскважин',
                        'Отключение насосов оборотного водоснабжения',
                        'Отключение насосов откачки из емкостей сбора утечек всех СИКН',
                        'Отключение насосов откачки из емкостей сбора утечек МНС',
                        'Отключение насосов откачки из емкостей сбора утечек ПНС', 'Отключение насосов откачки из емкостей ССВД',
                        'Отключение насосов прокачки нефти/нефтепродукта через БИК',
                        'Отключение насосов прокачки нефти/нефтепродукта через оперативный БИК',
                        'Отключение насосов системы запирания', 'Отключение насосов хозяйственно-питьевого водоснабжения',
                        'Отключение насосов, обеспечивающих подкачку нефти/нефтепродукта от объектов нефтедобычи/нефтепереработки',
                        'Отключение ПНС с выдержкой времени до 5 с после отключения всех МНА',
                        'Отключение подпорных вентиляторов ЭД', 'Отключение подпорных вентиляторов электрозала',
                        'Отключение приточного вентилятора помещения БИК', 'Отключение приточного вентилятора помещения РД',
                        'Отключение приточного вентилятора помещения СИКН',  'Отключение приточного вентилятора помещения ССВД',
                        'Отключение приточных вентиляторов в помещении централизованной маслосистемы и закрытие огнезадерживающих клапанов',
                        'Отключение приточных вентиляторов насосного зала МНС и закрытие огнезадерживающих клапанов',
                        'Отключение приточных вентиляторов насосного зала ПНС и закрытие огнезадерживающих клапанов',
                        'Отключение приточных вентиляторов помещения компрессорной подпора воздуха ЭД и закрытие огнезадерживающих клапанов',
                        'Отключение электронагревателей емкости сбора утечек МНС',
                        'Отключение электронагревателей емкости сбора утечек ПНС',
                        'Отключение электронагревателей емкости сбора утечек СИКН',
                        'Отключение электронагревателей масла', 'Переменная', 'Правило для карты уставок',
                        'Тип остановки НА', 'Тип остановки насосной станции']
    ktprs_hat_table   = ['№', 'Переменная', 'Идентификатор', 'Название', 'Сработка', 'Ссылка на значение',
                         'Приоритет сообщ. при 0', 'Приоритет сообщ. при 1', 'Запрет выдачи сообщений',
                         'Pic', 'Звук сообщения при 0', 'Звук сообщения при 1']
    ktprp_hat_table   = ['№', 'Идентификатор', 'Название', 'Номер листа (для ВУ)', 'Номер защиты (для ВУ)']
    ktpra_hat_table   = ['№', 'Переменная', 'Идентификатор', 'Название', 'НА', 'Аварийный параметр', '№ защиты в РД',
                         'Тип остановки', 'АВР', 'Закрытие задвижек', 'Запрет маскирования', 'Временная уставка', 'Pic',
                         'Группа уставок', 'Правило для карты уставок', 'Номер листа (для ВУ)', 'Номер защиты (для ВУ)',
                         'Номер агрегата (для ВУ)', 'Pic']
    gmpna_hat_table   = ['№', 'Идентификатор', 'Название', 'НА', 'Номер листа (для ВУ)', 'Номер защиты (для ВУ)', 'Номер агрегата (для ВУ)']
    dps_hat_table     = ['№', 'Деблокировка', 'Срабатывание', 'Контроль']
    nps_hat_table     = ['№', 'Переменная', 'Идентификатор', 'Название', 'Значение']
    vsgrp_hat_table   = ['№', 'Переменная', 'Название', 'Пож. или водоорош.', 'Количество вспомсистем в группе',
                         'Требуется выставлять флаг WarnOff, если работает одна вспомсистема в группе']
    umpna_hat_table   = ['№', 'Переменная', 'Идентификатор', 'Название', 'ВВ Включен', 'ВВ Включен дубль',
                         'ВВ отключен', 'ВВ отключен дубль', 'Сила тока >  уставки холостого хода',
                         'Исправность цепей включения ВВ', 'Исправность цепей отключения ВВ',
                         'Исправность цепей отключения ВВ дубль', 'Стоп 1', 'Стоп 2', 'Стоп 3', 'Стоп 4',
                         'Сигнал «Контроль наличия напряжения в цепях оперативного тока»',
                         'Флаг наличия напряжения в двигательной ячейке ЗРУ', 'Тележка ВВ выкачена',
                         'Дистанционный режим управления контроллера РЗиА', 'Наличие связи с контроллером РЗиА',
                         'Состояние возбудителя ЭД', 'Флаг окончания предпусковой продувки двигателя',
                         'Флаг наличия безопасного давления подпора воздуха в корпусе двигателя',
                         'Флаг наличия безопасного давления подпора воздуха в корпусе возбудителя',
                         'Флаг закрытого положения клапана продувки двигателя',
                         'Флаг температуры масла маслосистемы выше 10гр.С на выходе охладителя (для индивидуальной маслосистемы)',
                         'Флаг предельного минимального уровня масла в маслобаке (для индивидуальной маслосистемы)',
                         'Флаг наличия минимального уровня запирающей жидкости в баке системы запирания',
                         'Обобщенный флаг наличия давления запирающей жидкости к торцевому уплотнению',
                         'GMPNA_[49]', 'GMPNA_[50]', 'GMPNA_[51]', 'GMPNA_[52]', 'GMPNA_[53]', 'GMPNA_[54]',
                         'GMPNA_[55]', 'GMPNA_[56]', 'GMPNA_[57]', 'GMPNA_[58]', 'GMPNA_[59]', 'GMPNA_[60]',
                         'GMPNA_[61]', 'GMPNA_[62]', 'GMPNA_[63]', 'GMPNA_[64]',
                         'Команда на включение ВВ (только для UMPNA)', 'Команда на отключение ВВ (выход 1)',
                         'Команда на отключение ВВ (выход 2)', 'НА с ЧРП', 'Тип НА - МНА', 'Насос типа НМ',
                         'Параметр для KTPRAS_1',  'Количество сканов задержки анализа исправности цепей управления ВВ НА',
                         'Номер агрегата вспомсистемы "пуско-резервный маслонасос" (для индивидуальной маслосистемы)',
                         'Номер НПС (1 или 2), к которой относится НА', 'Номер защиты АЧР в массиве станционных защит',
                         'Номер защиты САОН в массиве станционных защит']
    rsreq_hat_table   = ['№', 'Переменная', 'Идентификатор', 'Название', 'Route', 'SlaveId', 'ModbusFunction',
                        'Address', 'Count', 'ResultOffset', 'SingleRequest', 'OnModifyRequest', 'RepeatOverScan',
                        'SkipRepeatsWhenBad', 'Enable']
    rs_hat_table      = ['№', 'Переменная', 'Идентификатор', 'Название', '№ модуля в массиве mRS', 'УСО, модуль, канал']
    vv_hat_table      = ['№', 'Переменная', 'Идентификатор', 'Название', 'Высоковольтный выключатель включен',
                         'Высоковольтный выключатель отключен']
    modbus_hat_table  = ['Переменная Excel', 'Начальный адрес', 'Конечный адрес', 'Число регистров']
    pi_hat_table      = ['№', 'Идентификатор', 'Название', 'Место установки']
    pz_hat_table      = ['№', 'Название', 'Подпись на мнемокадре','Тип', 'Номер готовности', 'Г_1', 'Г_2', 'Г_3',
                         'Г_4', 'Г_5', 'Г_6', 'Г_7', 'Г_8', 'Г_9', 'Г_10', 'Г_11', 'Г_12', 'Г_13', 'Г_14', 'Г_15']
    upts_hat_table    = ['№', 'Идентификатор', 'Название', 'Место установки', 'Короткое название']
    msg_hat_table     = ['№', 'Идентификатор', 'Название', 'Название таблицы БД ВУ', 'Индекс', 'Количество']

    tm_dp_hat_table   = ['№', 'Переменная', 'Название', 'Ссылка на сигнал наличия связи', 'Ссылка на таймаут по умолчанию tmCommon.CSPA_t1']
    tm_ts_hat_table   = ['№', 'Переменная', 'Название', 'Адрес объекта', 'Ссылка на значние']
    tm_tu_hat_table   = ['№', 'Переменная', 'Название', 'Адрес объекта', 'Изменяемая переменная', 'Изменяемый бит', 'descriptionTU (не более 32 символа латиницы)']
    tm_ti2_hat_table  = ['№', 'Переменная', 'Название', 'Адрес объекта', 'Переменная - значение', 'Переменная - статус']
    tm_ti4_hat_table  = ['№', 'Переменная', 'Название', 'Адрес объекта', 'Переменная - значение', 'Переменная - статус', 'Переменная - Aiparam']
    tm_tii_hat_table  = ['№', 'Переменная', 'Название', 'Адрес объекта', 'Переменная - значение', 'Переменная - статус']
    tm_tr2_hat_table  = ['№', 'Переменная', 'Название', 'Адрес объекта', 'Изменяемая переменная', 'descriptionTR4 (не более 16 символов латиницы)']
    tm_tr4_hat_table  = ['№', 'Переменная', 'Название', 'Адрес объекта', 'Изменяемая переменная', 'descriptionTR4 (не более 16 символов латиницы)']

    all_table_hat   = {'ZD'       : zd_hat_table,
                       'VS'       : vs_hat_table,
                       'AI'       : ai_hat_table,
                       'AO'       : ao_hat_table,
                       'DI'       : di_hat_table,
                       'DO'       : do_hat_table,
                       'TrendGRP' : trengrp_hat_table,
                       'SPGRP'    : spgrp_hat_table,
                       'Pic'      : pic_hat_table,
                       'SS'       : ss_hat_table,
                       'AIGRP'    : aigrp_hat_table,
                       'КД'       : kd_hat_table,
                       'UMPNA'    : umpna_hat_table,
                       'USO'      : uso_hat_table,
                       'UTS'      : uts_hat_table,
                       'KTPR'     : ktpr_hat_table,
                       'KTPRP'    : ktprp_hat_table,
                       'KTPRA'    : ktpra_hat_table,
                       'KTPRS'    : ktprs_hat_table,
                       'GMPNA'    : gmpna_hat_table,
                       'ModBus'   : modbus_hat_table,
                       'PI'       : pi_hat_table,
                       'PZ'       : pz_hat_table,
                       'UPTS'     : upts_hat_table,
                       'MSG'      : msg_hat_table,
                       'NPS'      : nps_hat_table,
                       'VSGRP'    : vsgrp_hat_table,
                       'RSreq'    : rsreq_hat_table,
                       'RS'       : rs_hat_table,
                       'VV'       : vv_hat_table,
                       'DPS'      : dps_hat_table,
                       'TM_DP'    : tm_dp_hat_table,
                       'TM_TS'    : tm_ts_hat_table,
                       'TM_TU'    : tm_tu_hat_table,
                       'TM_TI2'   : tm_ti2_hat_table,
                       'TM_TI4'   : tm_ti4_hat_table,
                       'TM_TII'   : tm_tii_hat_table,
                       'TM_TR2'   : tm_tr2_hat_table,
                       'TM_TR4'   : tm_tr4_hat_table,
                      }

    # Метод для поиска в строке - Общий
    def str_find(self, str1, arr):
        i = 0
        for el in arr:
            if str(str1).find(el) > -1:
                return True
    def read_exel(self, actives_table=actives_table, all_table_hat=all_table_hat):
        wb = openpyxl.load_workbook(self.exel, read_only=True, data_only=True)
        out_data = {}
        for table in actives_table:
            hat_tabl = {}
            try:
                sheet = wb[table]
                for data in sheet.rows:
                    for cell in data:
                        for hat in all_table_hat[table]:
                            if cell.value == hat:
                                rows = cell.row + 1
                                hat_tabl[hat] = cell.column - 1
                logger.info(f'Таблица: {table} найдена: {str(hat_tabl)}')
                data = []
                for row in sheet.iter_rows(min_row=rows, min_col=1):
                    keys   = []
                    values = []
                    for name, hat in hat_tabl.items():
                        keys.append(name)
                        values.append(row[hat].value)
                    values.append(table)
                    array = {k: v for k, v in zip(keys, values)}
                    data.append(array)
                out_data[table]=data
            except:
                logger.error(f'Таблица: {table} отсутствует')
        wb.close()
        return out_data
    def translate(self, str):
        dict = {".":"_",
                "/":"_",
                "\\":"_",
                ",":"_",
                ":":"_",
                ";":"_",
                "А":"A",
                "Б":"_B",
                "В":"B",
                "Г":"G",
                "Д":"D",
                "Е":"E",
                "Ё":"_E",
                "Ж":"J",
                "З":"Z",
                "И":"I",
                "Й":"_I",
                "К":"K",
                "Л":"L",
                "М":"M",
                "Н":"H",
                "О":"O",
                "П":"_P",
                "Р":"P",
                "С":"C",
                "Т":"T",
                "У":"U",
                "Ф":"F",
                "Х":"X",
                "Ц":"C",
                "Ч":"CH",
                "Ш":"SH",
                "Щ":"SCH",
                "Ь":"b",
                "Ы":"_",
                "Ъ":"_",
                "Э":"E",
                "Ю":"U",
                "Я":"YA",
                "а":"a",
                "б":"_b",
                "в":"b",
                "г":"g",
                "д":"d",
                "е":"e",
                "ё":"_e",
                "ж":"j",
                "з":"z",
                "и":"i",
                "й":"_i",
                "к":"k",
                "л":"l",
                "м":"m",
                "н":"h",
                "о":"o",
                "п":"_p",
                "р":"p",
                "с":"c",
                "т":"t",
                "у":"u",
                "ф":"f",
                "х":"x",
                "ц":"c",
                "ч":"ch",
                "ш":"sh",
                "щ":"sch",
                "ь":"b",
                "ы":"_",
                "ъ":"_",
                "э":"e",
                "ю":"u",
                "я":"ya"}

        intab = '.-пПаАфз/еЕсС'
        outtab = '__ppaafz_eEcC'
        trantab = str.maketrans(dict)
        outstr = str.translate(trantab)
        return outstr

    def parser_omx(self, directory):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(self.omx, parser)
        root = tree.getroot()
        try:
            for el in root.iter('{automation.deployment}application-object'):
                if el.attrib['name'] == "Application_PLC":
                    for item in el.iter('{automation.control}object'):
                        if item.attrib['name'] == 'Root' + self.name_prefix:
                            for el1 in item.iter('{automation.control}object'):
                                if el1.attrib['name'] == directory:
                                    item.remove(el1)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = directory
                            item.append(object)
                            logger.info(f'{directory}: очистка папки завершена')

                            for el1 in item.iter('{automation.control}object'):
                                if el1.attrib['name'] == directory:
                                    return el1, tree
        except:
            logger.error(f'Корень Root: отсутствует! Работа прекращена')
    def parser_diag_omx(self, directory):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(self.omx, parser)
        root = tree.getroot()
        try:
            for el in root.iter('{automation.deployment}application-object'):
                if el.attrib['name'] == "Application_PLC":
                    for item in el.iter('{automation.control}object'):
                        if item.attrib['name'] == "Diag":
                            for el1 in item.iter('{automation.control}object'):
                                if el1.attrib['name'] == directory:
                                    item.remove(el1)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = directory
                            item.append(object)
                            logger.info(f'{directory}: очистка папки завершена')

                            for el1 in item.iter('{automation.control}object'):
                                if el1.attrib['name'] == directory:
                                    return el1, tree
        except:
            logger.error(f'Корень Root: отсутствует! Работа прекращена')

    def parser_map(self):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(self.map, parser)
        root = tree.getroot()
        return root, tree
    def parser_map_modbus(self):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(self.map_mb, parser)
        root = tree.getroot()
        return root, tree
    def parser_map_modbus503(self):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(self.map_mb503, parser)
        root = tree.getroot()
        return root, tree
    def parser_diag_map(self, path_map):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(path_map, parser)
        root = tree.getroot()
        return root, tree

    def cleaner_map(self, directory, root):
        for item in root.iter('node-path'):
            signal = f'Root{self.name_prefix}{directory}'
            if self.str_find(item.text, {signal}):
                parent = item.getparent()
                root.remove(parent)
    def cleaner_diag_map(self, directory, root):
        for item in root.iter('item'):
            signal = f'Root{self.name_prefix}{directory}'
            if self.str_find(item.attrib['id'], {signal}):
                root.remove(item)

class Equipment(Initialisation_path):
    # Кроме диагностики: чистим объекты и карту адресов
    @logger.catch
    def clear_objects(self, directory, clear_modbus, clear_opcda, clear_opcua):
        # Чистка объектов
        el1, tree  = self.parser_omx(directory)
        tree.write(self.omx, pretty_print=True)
        # Чистка карты адресов OPCDA
        if clear_opcda:
            root_1, tree_1 = self.parser_map()
            self.cleaner_map(f'.{directory}', root_1)
            tree_1.write(self.map, pretty_print=True)
        # Чистка карты адресов ModBus
        if clear_modbus:
            if directory == 'Analogs':
                root_2, tree_2 = self.parser_map_modbus503()
                self.cleaner_map(f'.{directory}', root_2)
                tree_2.write(self.map_mb503, pretty_print=True)
            else:
                root_3, tree_3 = self.parser_map_modbus()
                self.cleaner_map(f'.{directory}', root_3)
                tree_3.write(self.map_mb, pretty_print=True)

        logger.info(f'{directory}: очистка завершена')
        return (f'{directory}: очистка завершена')
    # Диагностика: Очистить объекты, карту атрубутов и карту адресов
    @logger.catch
    def diag_clear(self, directory, clear_attrib, clear_opcda, clear_opcua, clear_modbus, *map_attrib):
        # Чистка объектов
        el, tree = self.parser_diag_omx(directory)
        tree.write(self.omx, pretty_print=True)
        # Чистка карты атрибутов
        if clear_attrib:
            # Цикл по всем xml
            for path in map_attrib:
                root_2, tree_2 = self.parser_diag_map(path)
                self.cleaner_diag_map(directory, root_2) 
                tree_2.write(path, pretty_print=True)
        # Чистка карты адресов OPCDA
        if clear_opcda:
            root_1, tree_1 = self.parser_map()
            self.cleaner_map(directory, root_1)
            tree_1.write(self.map, pretty_print=True)
        # Чистка карты адресов ModBus
        if clear_modbus:
            root_3, tree_3 = self.parser_map_modbus()
            self.cleaner_map(f'.Diag.{directory}', root_3)
            tree_3.write(self.map_mb, pretty_print=True)

        logger.info(f'Diag.{directory}: Очистка завершена')
        return (f'Diag.{directory}: Очистка завершена')
    # Чистка карта атрибутов
    @logger.catch
    def clear_map(self, text_modul, directory, *map_attrib):
        # Чистка карты атрибутов
        # Цикл по всем xml
        for path in map_attrib:
            root, tree = self.parser_diag_map(path)
            self.cleaner_diag_map(directory, root)
            tree.write(path, pretty_print=True)
        logger.info(f'{text_modul}: очистка завершена')
        return (f'{text_modul}: очистка завершена')

    # DevStudio - omx
    # Аналоговые сигналы
    @logger.catch
    def analogs_omx(self):
        dop_analog = {'объем'         : 'V',
                      'объём'         : 'V',
                      'перепад'       : 'dP',
                      'давлени'       : 'P',
                      'загазованность': 'Газ',
                      'вертик'        : 'Xверт',
                      'горизонт'      : 'Xгор',
                      'осевая'        : 'Xос',
                      'попереч'       : 'Xпоп',
                      'осевое'        : 'Xoc',
                      'сила'          : 'I',
                      'температура'   : 'T',
                      'уровень'       : 'L',
                      'утечк'         : 'L',
                      'расход'        : 'Q',
                      'положени'      : 'Q',
                      'затоплен'      : 'L',
                      'частот'        : 'F',
                      'процен'        : 'Q',
                      'заслон'        : 'Q',
                     }
        data      = self.data['AI']
        el1, tree = self.parser_omx('Analogs')
        try:
            for value in data:
                number      = value['№']
                name        = value['Название']
                tag         = value['Идентификатор']
                equ_fiz     = value['Единица измерения физической величины (АЦП)']
                equ         = value['Единица измерения']
                unit_switch = value['Давление нефти/ нефтепродукта (флаг для пересчета в кгс/см2)']
                unit_alt    = 'кгс/см2'
                grp_analog  = value['Группа аналогов']

                if equ_fiz     is None: equ_fiz = ''
                if tag         is None: continue
                if number      is None: continue
                if name        is None: continue
                if equ         is None: continue
                if unit_switch is None: continue

                if grp_analog == 'Уровни' or grp_analog == 'Аналоговые выходы':
                    type = 'unit.Library.PLC_Types.lv_Analog_PLC'
                else:
                    type = 'unit.Library.PLC_Types.Analog_PLC'

                tag_1 = tag
                tag   = self.translate(str(tag))
                unit_switch = True if unit_switch == 1 else False
                # Находим совпадение из словаря с названием сигнала и заполняем подпись на кадре
                sign = ' '
                for key, short in dop_analog.items():
                    if self.str_find(str(name).lower(), {key}):
                        sign = str(short)
                        break

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = tag
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = type
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                atrb1.attrib['value'] = str(number)
                object.append(atrb1)

                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb2.attrib['value'] = sign
                object.append(atrb2)

                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.Library.Attributes.EGU_Desc"
                atrb3.attrib['value'] = equ
                object.append(atrb3)

                atrb4 = etree.Element("attribute")
                atrb4.attrib['type'] = "unit.Library.Attributes.EGU_Desc_phys"
                atrb4.attrib['value'] = equ_fiz
                object.append(atrb4)

                atrb5 = etree.Element("attribute")
                atrb5.attrib['type'] = "unit.System.Attributes.Description"
                atrb5.attrib['value'] = name
                object.append(atrb5)

                atrb6 = etree.Element("attribute")
                atrb6.attrib['type'] = "unit.Library.Attributes.EGU_Desc_Alt"
                atrb6.attrib['value'] = unit_alt
                object.append(atrb6)

                atrb7 = etree.Element("attribute")
                atrb7.attrib['type'] = "unit.Library.Attributes.EGUsChange"
                atrb7.attrib['value'] = str(unit_switch)
                object.append(atrb7)

                atrb8 = etree.Element("attribute")
                atrb8.attrib['type'] = "unit.Library.Attributes.AI_Ref_KZFKP"
                atrb8.attrib['value'] = str(tag_1)
                object.append(atrb8)

                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Analogs: файл omx OK')
            return (f'Analogs: файл omx OK')
        except:
            logger.error(f'Analogs: файл omx FAILED')
            return (f'Analogs: файл omx FAILED')
    # Входные дискретные сигналы
    @logger.catch
    def diskret_in_omx(self):
        logger.info(f'Diskrets: генерация объектов DevStudio')
        dop_discret = {'давлен'        : 'P',
                       'напряж'        : 'U',
                       'уровень'       : 'L',
                       'затоплен'      : 'L',
                       'утечк'         : 'L',
                       'питание'       : 'U',
                       'питание шкафа' : 'U'}
        data_di   = self.data['DI']
        data_ai   = self.data['AI']
        el1, tree = self.parser_omx('Diskrets')
        #try:
        for value in data_di:
            name      = value['Название']
            number_di = value['№']
            pNC_AI    = value['pNC_AI']
            tag       = value['Идентификатор']
            caption   = value['Подпись на мнемокадре']

            if name is None: continue
            if tag  is None: continue
            tag      = self.translate(str(tag))
            tag_ai   = ' '
            tag_ai_1 = ' '
            if not pNC_AI is None:
                isdigit = re.findall('\d+', str(pNC_AI))
                for number in data_ai:
                    number_ai = number['№']
                    tag_ai    = number['Идентификатор']
                    if str(number_ai) == str(isdigit[0]):
                        if tag_ai is None:
                            print('Тэг AI сигнала под номером: ' + number_ai + 'пуст')
                            break
                        else:
                            tag_ai_1 = tag_ai
                            tag_ai   = self.translate(tag_ai)
                            break
            if caption is None:
                for key, value in dop_discret.items():
                    # Находим совпадение из словаря с названием сигнала и заполняем подпись
                    sign = ' '
                    if self.str_find(str(name).lower(), {key}):
                        sign = str(value)
                        break
            else:
                sign = str(caption)
            object = etree.Element("{automation.control}object")
            object.attrib['name'] = str(tag)
            object.attrib['uuid'] = str(uuid.uuid1())
            object.attrib['base-type'] = "unit.Library.PLC_Types.Diskret_PLC"
            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
            atrb1 = etree.Element("attribute")
            atrb1.attrib['type'] = "unit.Library.Attributes.Index"
            atrb1.attrib['value'] = str(number_di)
            object.append(atrb1)
            atrb2 = etree.Element("attribute")
            atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
            atrb2.attrib['value'] = str(sign)
            object.append(atrb2)
            atrb3 = etree.Element("attribute")
            atrb3.attrib['type'] = "unit.System.Attributes.Description"
            atrb3.attrib['value'] = str(name)
            object.append(atrb3)
            atrb4 = etree.Element("attribute")
            atrb4.attrib['type'] = "unit.Library.Attributes.AI_Ref"
            atrb4.attrib['value'] = str(tag_ai)
            object.append(atrb4)
            atrb5 = etree.Element("attribute")
            atrb5.attrib['type'] = "unit.Library.Attributes.AI_Ref_KZFKP"
            atrb5.attrib['value'] = str(tag_ai_1)
            object.append(atrb5)
            el1.append(object)
        tree.write(self.omx, pretty_print=True)
        logger.info(f'Diskrets: файл omx OK')
        return (f'Diskrets: файл omx OK')
        # except:
        #     logger.info(f'Diskrets: файл omx FAILED')
        #     return (f'Diskrets: файл omx FAILED')
    # Индикаторы событий Picture
    @logger.catch
    def picture_omx(self):
        data = self.data['Pic']
        el1, tree = self.parser_omx('Pictures')
        try:
            for value in data:
                number   = value['№']
                name_pic = value['Название']
                screen   = value['Кадр IFix(*.grf)']

                if screen is None: continue
                screen = self.translate(screen)

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = screen
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.Picture_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                atrb1.attrib['value'] = str(number)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb2.attrib['value'] = name_pic
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.System.Attributes.Description"
                atrb3.attrib['value'] = name_pic
                object.append(atrb3)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Pictures: файл omx OK')
            return (f'Pictures: файл omx OK')
        except:
            logger.error(f'Pictures: файл omx FAILED')
            return (f'Pictures: файл omx FAILED')
    # Вспомсистемы
    @logger.catch
    def auxsystem_omx(self):
        data_vs = self.data['VS']
        data_ai = self.data['AI']
        data_di = self.data['DI']
        data_do = self.data['DO']
        el1, tree = self.parser_omx('AuxSystems')
        try:
            for value_vs in data_vs:
                number_vs = value_vs['№']
                name      = value_vs['Название']
                shortdesc = value_vs['Короткое название (для отображения на кадре)']
                sensor    = str(value_vs['Давл. норма'])
                voltage   = str(value_vs['Напр.'])
                close     = str(value_vs['Отключить'])

                if number_vs is None: continue
                if shortdesc is None: shortdesc = ''

                tag = 'VS_' + str(number_vs)
                # Вытаскиваем число из строки
                isdigit = re.findall('\d+', sensor)
                # Ищем давление на выходе из числа AI и DI
                if self.str_find(sensor.lower(), {'di'}):
                    for value_di in data_di:
                        number_di = value_di['№']
                        tag_di    = value_di['Идентификатор']
                        if self.str_find(number_di, isdigit):
                            tag_sensor = self.translate(tag_di)
                            break
                    pc_use = str('1')
                elif self.str_find(sensor.lower(), {'ai'}):
                    for value_ai in data_ai:
                        number_ai = value_ai['№']
                        tag_ai    = value_ai['Идентификатор']
                        if self.str_find(number_ai, isdigit):
                            tag_sensor = self.translate(tag_ai)
                            break
                    pc_use = str('2')
                else:
                    pc_use = str('0')
                    tag_sensor = ' '

                isdigitVoltage = re.findall('\d+', voltage)
                if self.str_find(voltage.lower(), {'di'}):
                    for value_di in data_di:
                        number_di = value_di['№']
                        tag_di_for_diagno = value_di['Идентификатор']
                        if self.str_find(number_di, isdigitVoltage):
                            tag_voltage  = self.translate(tag_di_for_diagno)
                            break

                isdigitCLOSE= re.findall('\d+', close)
                if self.str_find(close.lower(), {'do'}):
                    for value_do in data_do:
                        number_do = value_do['№']
                        tag_do    = value_do['Идентификатор']
                        if self.str_find(number_do, isdigitCLOSE):
                            tag_close  = self.translate(tag_do)
                            break

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = tag
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.AuxSystem_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                atrb1.attrib['value'] = str(number_vs)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb2.attrib['value'] = shortdesc
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.System.Attributes.Description"
                atrb3.attrib['value'] = name
                object.append(atrb3)
                atrb4 = etree.Element("attribute")
                atrb4.attrib['type'] = "unit.Library.Attributes.PC_Use"
                atrb4.attrib['value'] = pc_use
                object.append(atrb4)
                atrb5 = etree.Element("attribute")
                atrb5.attrib['type'] = "unit.Library.Attributes.PC_Ref"
                atrb5.attrib['value'] = str(tag_sensor)
                object.append(atrb5)
                atrb6 = etree.Element("attribute")
                atrb6.attrib['type'] = "unit.Library.Attributes.DI_ref"
                atrb6.attrib['value'] = str(tag_voltage)
                object.append(atrb6)
                atrb7 = etree.Element("attribute")
                atrb7.attrib['type'] = "unit.Library.Attributes.DO_ref"
                atrb7.attrib['value'] = str(tag_close)
                object.append(atrb7)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'AuxSystems: файл omx OK')
            return (f'AuxSystems: файл omx OK')
        except:
            logger.error(f'AuxSystems: файл omx FAILED')
            return (f'AuxSystems: файл omx FAILED')
    # Задвижки
    @logger.catch
    def valves_omx(self):
        data      = self.data['ZD']
        data_di   = self.data['DI']
        data_do   = self.data['DO']
        el1, tree = self.parser_omx('Valves')
        try:
            for value in data:
                number    = value['№']
                name      = value['Название']
                shortdesc = value['Короткое название']
                vmmo      = value['ВММО']
                vmmz      = value['ВММЗ']
                rs        = value['Наличие ИНТЕРФЕЙСА']
                dist_i    = value['Дист_и']
                dist_f    = value['Дист_ф']
                kvo_in_zd = str(value['КВО'])
                open_in_zd= str(value['Открыть'])

                isdigitKVO = re.findall('\d+', kvo_in_zd)
                if self.str_find(kvo_in_zd.lower(), {'di'}):
                    for value_di in data_di:
                        number_di = value_di['№']
                        tag_di    = value_di['Идентификатор']
                        if self.str_find(number_di, isdigitKVO):
                            tag_kvo  = self.translate(tag_di)
                            break

                isdigitOPEN = re.findall('\d+', open_in_zd)
                if self.str_find(open_in_zd.lower(), {'do'}):
                    for value_do in data_do:
                        number_do = value_do['№']
                        tag_do    = value_do['Идентификатор']
                        if self.str_find(number_do, isdigitOPEN):
                            tag_open  = self.translate(tag_do)
                            break

                if number is None: continue
                if name   is None: continue

                tag    = 'ZD_' + str(number)
                # Наличие мутфа, авария
                isBUR = True if (vmmo is None) or (vmmz is None) else False
                # Наличие ключа М/Д смотри по двум полям физика или интерфейс
                isDist = True if (not dist_i is None) or (not dist_f is None) else False
                # Наличие интерфейса
                isRS = True if rs == 1 else False

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = tag
                object.attrib['uuid'] = str(uuid.uuid1())
                if isRS != True:
                    object.attrib['base-type'] = "unit.Library.PLC_Types.Valve_PLC"
                else:
                    object.attrib['base-type'] = "unit.Library.PLC_Types.ex_Valve_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                atrb1.attrib['value'] = str(number)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb2.attrib['value'] = str(shortdesc)
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.System.Attributes.Description"
                atrb3.attrib['value'] = str(name)
                object.append(atrb3)
                atrb4 = etree.Element("attribute")
                atrb4.attrib['type'] = "unit.Library.Attributes.BUR"
                atrb4.attrib['value'] = str(isBUR)
                object.append(atrb4)
                atrb5 = etree.Element("attribute")
                atrb5.attrib['type'] = "unit.Library.Attributes.RS485"
                atrb5.attrib['value'] = str(isRS)
                object.append(atrb5)
                atrb6 = etree.Element("attribute")
                atrb6.attrib['type'] = "unit.Library.Attributes.Dist_key"
                atrb6.attrib['value'] = str(isDist)
                object.append(atrb6)
                atrb7 = etree.Element("attribute")
                atrb7.attrib['type'] = "unit.Library.Attributes.DI_ref"
                atrb7.attrib['value'] = str(tag_kvo)
                object.append(atrb7)
                atrb8 = etree.Element("attribute")
                atrb8.attrib['type'] = "unit.Library.Attributes.DO_ref"
                atrb8.attrib['value'] = str(tag_open)
                object.append(atrb8)
                el1.append(object)
                tree.write(self.omx, pretty_print=True)
            logger.info(f'Valves: файл omx OK')
            return (f'Valves: файл omx OK')
        except:
            logger.error(f'Valves: файл omx FAILED')
            return (f'Valves: файл omx FAILED')
    # Агрегаты
    @logger.catch
    def pumps_omx(self):
        data      = self.data['UMPNA']
        el1, tree = self.parser_omx('NAs')
        try:
            for value in data:
                number    = value['№']
                name      = value['Название']

                if number is None: continue
                if name is None: continue

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = f'NA_{str(number)}'
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.NA_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                atrb1.attrib['value'] = str(number)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb2.attrib['value'] = str(name)
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.System.Attributes.Description"
                atrb3.attrib['value'] = str(name)
                object.append(atrb3)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'NAs: файл omx OK')
            return (f'NAs: файл omx OK')
        except:
            logger.error(f'NAs: файл omx FAILED')
            return (f'NAs: файл omx FAILED')
    # Смежные системы
    @logger.catch
    def relayted_system_omx(self):
        data = self.data['SS']
        el1, tree = self.parser_omx('SSs')
        try:
            for value in data:
                number    = value['№']
                name      = value['Название']

                if number is None: continue
                if name   is None: continue

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = 'SS_' + str(number)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.SS_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                atrb1.attrib['value'] = str(number)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb2.attrib['value'] = str(name)
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.System.Attributes.Description"
                atrb3.attrib['value'] = str(name)
                object.append(atrb3)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'SSs: файл omx OK')
            return (f'SSs: файл omx OK')
        except:
            logger.error(f'SSs: файл omx FAILED')
            return (f'SSs: файл omx FAILED')
    # Табло и сирены(станция) - UTS
    @logger.catch
    def uts_omx(self):
        data = self.data['UTS']
        el1, tree = self.parser_omx('UTSs')
        try:
            for value in data:
                number = value['№']
                name   = value['Название']
                tag    = value['Идентификатор']
                siren  = value['Сирена']

                if tag    is None: continue
                if number is None: continue
                if name   is None: continue

                tag = self.translate(str(tag))

                if int(siren):                                  sign = 'Сирена'
                elif self.str_find(str(name).lower(), {'газ'}): sign = 'Газ'
                else:                                           sign = ''

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = str(tag)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.UTS_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                atrb1.attrib['value'] = str(number)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb2.attrib['value'] = str(sign)
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.System.Attributes.Description"
                atrb3.attrib['value'] = str(name)
                object.append(atrb3)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'UTSs: файл omx OK')
            return (f'UTSs: файл omx OK')
        except:
            logger.error(f'UTSs: файл omx FAILED')
            return (f'UTSs: файл omx FAILED')
    # Табло и сирены(ПТ) - UPTS
    @logger.catch
    def upts_omx(self):
        data = self.data['UPTS']
        el1, tree = self.parser_omx('UPTSs')
        try:
            for value in data:
                number      = value['№']
                name        = value['Идентификатор']
                place       = value['Место установки']
                description = value['Название']
                shortdesc   = value['Короткое название']

                if number      is None: continue
                if name        is None: continue
                if description is None: continue

                name = self.translate(str(name))

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = str(name)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.UPTS_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb1.attrib['value'] = str(shortdesc)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.System.Attributes.Description"
                atrb2.attrib['value'] = str(description)
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.Library.Attributes.Index"
                atrb3.attrib['value'] = str(number)
                object.append(atrb3)
                atrb4 = etree.Element("attribute")
                atrb4.attrib['type'] = "unit.Library.Attributes.Place"
                atrb4.attrib['value'] = str(place)
                object.append(atrb4)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'UPTSs: файл omx OK')
            return     (f'UPTSs: файл omx OK')
        except:
            logger.error(f'UPTSs: файл omx FAILED')
            return      (f'UPTSs: файл omx FAILED')
    # Общестнационные защиты(МНС)
    @logger.catch
    def ktpr_omx(self):
        data = self.data['KTPR']
        el1, tree = self.parser_omx('KTPRs')
        number_group = 0
        try:
            for value in data:
                number_defence = value['№']
                if number_defence is None: continue
                number_group += 1
            count_group = math.ceil(number_group / 4)

            for count in range(count_group):
                object = etree.Element("{automation.control}object")
                object.attrib['name'] = f'Group_{count + 1}'
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.KTPRx_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'KTPRs: файл omx OK')
            return (f'KTPRs: файл omx OK')
        except:
            logger.error(f'KTPRs: файл omx FAILED')
            return (f'KTPRs: файл omx FAILED')
    # Общестнационные защиты(ПТ)
    @logger.catch
    def ktprp_omx(self):
        data = self.data['KTPRP']
        el1, tree = self.parser_omx('KTPRs')
        number_group = 0
        try:
            for value in data:
                number_defence = value['№']
                if number_defence is None: continue
                number_group += 1
            count_group = math.ceil(number_group/4)

            for count in range(count_group):
                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = f'Group_{count + 1}'
                    object.attrib['uuid'] = str(uuid.uuid1())
                    object.attrib['base-type'] = "unit.Library.PLC_Types.KTPRx_PLC"
                    object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                    el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'KTPRPs: файл omx OK')
            return     (f'KTPRPs: файл omx OK')
        except:
            logger.error(f'KTPRPs: файл omx FAILED')
            return      (f'KTPRPs: файл omx FAILED')
    # Агрегатные защиты
    @logger.catch
    def ktpra_omx(self):
        data = self.data['KTPRA']
        el1, tree = self.parser_omx('KTPRAs')
        number_pumps_old = ''
        count_pumps      = 0
        number = 0
        try:
            for value in data:
                number_defence   = value['№']
                number_pumps_int = value['НА']

                if number_defence   is None: continue
                if number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps += 1
                    number_group = 0
                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = f'KTPRAs_{count_pumps}'
                    object.attrib['uuid'] = str(uuid.uuid1())

                if number_defence % 4 == 0:
                    number_group += 1
                    group = etree.Element("{automation.control}object")
                    group.attrib['name'] = f'Group_{number_group}'
                    group.attrib['uuid'] = str(uuid.uuid1())
                    group.attrib['base-type'] = "unit.Library.PLC_Types.KTPRx_PLC"
                    group.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                    object.append(group)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'KTPRAs: файл omx OK')
            return (f'KTPRAs: файл omx OK')
        except:
            logger.error(f'KTPRAs: файл omx FAILED')
            return (f'KTPRAs: файл omx FAILED')
    # Агрегатные готовности
    @logger.catch
    def gmpna_omx(self):
        data = self.data['GMPNA']
        el1, tree = self.parser_omx('GMPNAs')
        number_pumps_old = ''
        count_pumps = 0
        try:
            for value in data:
                number_defence = value['№']
                number_pumps_int = value['НА']

                if number_defence is None: continue
                if number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps += 1
                    number_group = 0
                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = f'GMPNAs_{count_pumps}'
                    object.attrib['uuid'] = str(uuid.uuid1())

                if number_defence % 4 == 0:
                    number_group += 1
                    group = etree.Element("{automation.control}object")
                    group.attrib['name'] = f'Group_{number_group}'
                    group.attrib['uuid'] = str(uuid.uuid1())
                    group.attrib['base-type'] = "unit.Library.PLC_Types.GMPNA_PLC"
                    group.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                    object.append(group)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'GMPNAs: файл omx OK')
            return (f'GMPNAs: файл omx OK')
        except:
            logger.error(f'GMPNAs: файл omx FAILED')
            return (f'GMPNAs: файл omx FAILED')
    # Пожарные извещатели
    @logger.catch
    def pi_omx(self):
        data = self.data['PI']
        el1, tree = self.parser_omx('PIs')
        try:
            for value in data:
                number      = value['№']
                tag         = value['Идентификатор']
                place       = value['Место установки']
                shortdesc   = value['Идентификатор']
                description = value['Название']

                if number      is None: continue
                if tag         is None: continue
                if shortdesc   is None: continue
                if description is None: continue
                if place       is None: place = ''

                tag = self.translate(str(tag))

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = str(tag)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.PI_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb1.attrib['value'] = str(shortdesc)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.System.Attributes.Description"
                atrb2.attrib['value'] = str(description)
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.Library.Attributes.Index"
                atrb3.attrib['value'] = str(number)
                object.append(atrb3)
                atrb4 = etree.Element("attribute")
                atrb4.attrib['type'] = "unit.Library.Attributes.Place"
                atrb4.attrib['value'] = str(place)
                object.append(atrb4)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'PIs: файл omx OK')
            return     (f'PIs: файл omx OK')
        except:
            logger.error(f'PIs: файл omx FAILED')
            return      (f'PIs: файл omx FAILED')
    # Пожарные зоны
    @logger.catch
    def pz_omx(self):
        data = self.data['PZ']
        el1, tree = self.parser_omx('PZs')
        try:
            for value in data:
                number      = value['№']
                shortdesc   = value['Подпись на мнемокадре']
                description = value['Название']

                if number      is None: continue
                if description is None: continue
                if shortdesc   is None: shortdesc = ''

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = f'PZ_{number}'
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.PZ_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                atrb1 = etree.Element("attribute")
                atrb1.attrib['type'] = "unit.Library.Attributes.Sign"
                atrb1.attrib['value'] = str(shortdesc)
                object.append(atrb1)
                atrb2 = etree.Element("attribute")
                atrb2.attrib['type'] = "unit.System.Attributes.Description"
                atrb2.attrib['value'] = str(description)
                object.append(atrb2)
                atrb3 = etree.Element("attribute")
                atrb3.attrib['type'] = "unit.Library.Attributes.Index"
                atrb3.attrib['value'] = str(number)
                object.append(atrb3)
                el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'PZs: файл omx OK')
            return     (f'PZs: файл omx OK')
        except:
            logger.error(f'PZs: файл omx FAILED')
            return      (f'PZs: файл omx FAILED')

    # DevStudio - map
    # Аналоговые сигналы
    @logger.catch
    def analogs_map(self):
        dop_analog = {'AIVisualValue': 'AIVisualValue',
                      'AIElValue'    : 'AIElValue',
                      'AIValue'      : 'AIValue',
                      'AIRealValue'  : 'AIRealValue',
                      'StateAI'      : ''
                      }
        data = self.data['AI']
        root, tree  = self.parser_map()

        # Чистка тэгов
        self.cleaner_map('.Analogs.', root)
        try:
            for value in data:
                number      = value['№']
                name        = value['Название']
                tag         = value['Идентификатор']
                equ_fiz     = value['Единица измерения физической величины (АЦП)']
                equ         = value['Единица измерения']
                unit_switch = value['Давление нефти/ нефтепродукта (флаг для пересчета в кгс/см2)']
                unit_alt    = 'кгс/см2'
                grp_analog  = value['Группа аналогов']

                if tag         is None: continue
                if number      is None: continue
                if name        is None: continue
                if equ         is None: continue
                if unit_switch is None: continue
                tag = self.translate(str(tag))

                for key, value in dop_analog.items():
                    signal = f'Root{self.name_prefix}.Analogs.{tag}.{key}'
                    if key != 'StateAI':
                        arrayposition = str(number - 1)
                        address_name  = self.prefix_driver + str(value)
                    else:
                        if number * 2 <= 512:
                            address_name  = self.prefix_driver + 'stateAI1_HMI'
                            arrayposition = str(2 * (number - 1))
                        else:
                            address_name  = self.prefix_driver + 'stateAI513_HMI'
                            arrayposition = str(2 * (number - 1) - 512)
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(signal)
                    object.append(node_p)

                    address = etree.Element('address')
                    address.text = str(address_name)
                    object.append(address)

                    array_pos = etree.Element('arrayposition')
                    array_pos.text = str(arrayposition)
                    object.append(array_pos)
                    root.append(object)

            tree.write(self.map, pretty_print=True)
            logger.info(f'Analogs: Карта адресов OK')
            return (f'Analogs: Карта адресов OK')
        except:
            logger.error(f'Analogs: Карта адресов FAILED')
            return (f'Analogs: Карта адресов FAILED')
    @logger.catch
    def analogs_map_modbus(self, flag_503):
        dop_analog    = {'AIVisualValue', 'AIElValue', 'AIValue', 'AIRealValue', 'StateAI'}
        dop_analog_lv = {'AIVisualValue', 'AIElValue', 'AIValue', 'AIRealValue', 'StateAI', 'Range_Bottom', 'Range_Top'}

        data        = self.data['AI']
        data_mb     = self.data['ModBus']
        if flag_503 is True: root, tree  = self.parser_map_modbus503()
        else               : root, tree  = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.Analogs.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']

                if variable == 'AIVisualValue': start_AIVisualValue = start_address
                if variable == 'AIElValue'    : start_AIElValue     = start_address
                if variable == 'AIValue'      : start_AIValue       = start_address
                if variable == 'AIRealValue'  : start_AIRealValue   = start_address
                if variable == 'StateAI'      : start_StateAI       = start_address
                if variable == 'AIParam'      : start_AIParam       = start_address

            for value in data:
                number     = value['№']
                tag        = value['Идентификатор']
                name       = value['Название']
                grp_analog = value['Группа аналогов']

                if tag  is None: continue
                if name is None: continue

                # У уровней особый набор параметров
                if grp_analog == 'Уровни': list_analog = dop_analog_lv
                else:                      list_analog = dop_analog

                tag = self.translate(str(tag))

                for key in list_analog:
                    signal = f'Root{self.name_prefix}.Analogs.{tag}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(signal)
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    if key == 'AIVisualValue': address.text = str(start_AIVisualValue + 2 * (number - 1))
                    if key == 'AIElValue'    : address.text = str(start_AIElValue     +     (number - 1))
                    if key == 'AIValue'      : address.text = str(start_AIValue       + 2 * (number - 1))
                    if key == 'AIRealValue'  : address.text = str(start_AIRealValue   + 2 * (number - 1))
                    if key == 'StateAI'      : address.text = str(start_StateAI       + 2 * (number - 1))
                    if key == 'Range_Bottom' : address.text = str(start_AIParam + 4 + 46 * (number - 1))
                    if key == 'Range_Top'    : address.text = str(start_AIParam + 2 + 46 * (number - 1))

                    object.append(address)
                    root.append(object)

                if flag_503 is True: tree.write(self.map_mb503, pretty_print=True)
                else               : tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Analogs: Карта адресов OK')
            return (f'Analogs: Карта адресов OK')
        except:
            logger.error(f'Analogs: Карта адресов FAILED')
            return (f'Analogs: Карта адресов FAILED')
    # Входные дискретные сигналы
    @logger.catch
    def diskret_in_map(self):
        data = self.data['DI']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.Diskrets.', root)
        try:
            for value in data:
                number = value['№']
                tag    = value['Идентификатор']
                name   = value['Название']

                if tag  is None: continue
                if name is None: continue
                tag = self.translate(str(tag))
                signal = 'Root' + self.name_prefix + '.Diskrets.' + tag + '.StateDI'

                if number <= 512:
                    address_name  = self.prefix_driver + 'stateDI1_HMI'
                    arrayposition = str(number - 1)
                elif number <= 1024:
                    address_name  =  self.prefix_driver + 'stateDI513_HMI'
                    arrayposition = str(number - 513)
                elif number <= 1536:
                    address_name  = self.prefix_driver + 'OFS!stateDI1025_HMI'
                    arrayposition = str(number - 1025)
                elif number <= 2048:
                    address_name  = self.prefix_driver + 'OFS!stateDI1537_HMI'
                    arrayposition = str(number - 1537)

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                
                node_p = etree.Element('node-path')
                node_p.text = str(signal)
                object.append(node_p)

                address = etree.Element('address')
                address.text = str(address_name)
                object.append(address)

                array_pos = etree.Element('arrayposition')
                array_pos.text = str(arrayposition)
                object.append(array_pos)

                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'Diskrets: Карта адресов OK')
            return (f'Diskrets: Карта адресов OK')
        except:
            logger.error(f'Diskrets: Карта адресов FAILED')
            return (f'Diskrets: Карта адресов FAILED')
    @logger.catch
    def diskret_in_map_modbus(self):
        data    = self.data['DI']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.Diskrets.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StateDI': start_StateDI = start_address

            for value in data:
                number = value['№']
                tag    = value['Идентификатор']
                name   = value['Название']

                if tag  is None: continue
                if name is None: continue
                tag = self.translate(str(tag))
                signal = f'Root{self.name_prefix}.Diskrets.{tag}.StateDI'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = str(signal)
                object.append(node_p)

                segment = etree.Element('table')
                segment.text = f'Holding Registers'
                object.append(segment)

                address = etree.Element('address')
                address.text = str(start_StateDI + (number - 1))
                object.append(address)
                root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Diskrets: Карта адресов OK')
            return (f'Diskrets: Карта адресов OK')
        except:
            logger.error(f'Diskrets: Карта адресов FAILED')
            return (f'Diskrets: Карта адресов FAILED')
    # Индикаторы событий Picture
    @logger.catch
    def picture_map(self):
        data = self.data['Pic']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.Pictures.', root)
        try:
            for value in data:
                number = value['№']
                name   = value['Название']
                frame  = value['Кадр IFix(*.grf)']

                if frame is None: continue

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.Pictures.{frame}.StatePicture'
                object.append(node_p)
                address = etree.Element('address')
                address.text = f'{self.prefix_driver}statePic_HMI'
                object.append(address)
                array_pos = etree.Element('arrayposition')
                array_pos.text = str(number - 1)
                object.append(array_pos)
                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'Pictures: Карта адресов OK')
            return (f'Pictures: Карта адресов OK')
        except:
            logger.error(f'Pictures: Карта адресов FAILED')
            return (f'Pictures: Карта адресов FAILED')
    @logger.catch
    def picture_map_modbus(self):
        data    = self.data['Pic']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.Pictures.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StatePicture': start_StatePicture = start_address

            for value in data:
                number = value['№']
                name   = value['Название']
                frame  = value['Кадр IFix(*.grf)']

                if frame is None: continue

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.Pictures.{frame}.StatePicture'
                object.append(node_p)

                segment = etree.Element('table')
                segment.text = f'Holding Registers'
                object.append(segment)

                address = etree.Element('address')
                address.text = str(start_StatePicture + (number - 1))
                object.append(address)
                root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Pictures: Карта адресов OK')
            return (f'Pictures: Карта адресов OK')
        except:
            logger.error(f'Pictures: Карта адресов FAILED')
            return (f'Pictures: Карта адресов FAILED')
    # Вспомсистемы
    @logger.catch
    def auxsystem_map(self):
        dop_vs = {'StateAuxSystem'           : 'HMI_VS',
                  'numOfStart'               : 'HMI_Statistic_VS',
                  'operatingTimeCurrentMonth': 'HMI_Statistic_VS',
                  'operatingTimeLastMonth'   : 'HMI_Statistic_VS',
                  'operatingTime'            : 'HMI_Statistic_VS',
                  }
        data = self.data['VS']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.AuxSystems.', root)
        try:
            for value in data:
                number = value['№']

                if number is None: continue

                for key, item in dop_vs.items():
                    tag = 'VS_' + str(number)
                    signal = 'Root' + self.name_prefix + '.AuxSystems.' + tag + '.' + key
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(signal)
                    object.append(node_p)
                    address = etree.Element('address')
                    address.text = self.prefix_driver + item
                    object.append(address)
                    array_pos = etree.Element('arrayposition')
                    if key == 'StateAuxSystem'           : array_pos.text = str(number - 1)
                    if key == 'numOfStart'               : array_pos.text = str((4 * (number - 1)) + 1)
                    if key == 'operatingTimeCurrentMonth': array_pos.text = str((4 * (number - 1)) + 2)
                    if key == 'operatingTimeLastMonth'   : array_pos.text = str((4 * (number - 1)) + 3)
                    if key == 'operatingTime'            : array_pos.text = str(4 * (number - 1))
                    object.append(array_pos)
                    root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'AuxSystems: Карта адресов OK')
            return (f'AuxSystems: Карта адресов OK')
        except:
            logger.error(f'AuxSystems: Карта адресов FAILED')
            return (f'AuxSystems: Карта адресов FAILED')
    @logger.catch
    def auxsystem_map_modbus(self):
        dop_vs = {'StateAuxSystem',
                  'numOfStart',
                  'operatingTimeCurrentMonth',
                  'operatingTimeLastMonth',
                  'operatingTime'}

        data    = self.data['VS']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.AuxSystems.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StateAuxSystem'              : start_StateAuxSystem            = start_address
                if variable == 'vs_numOfStart'               : start_numOfStart                = start_address
                if variable == 'vs_operatingTimeCurrentMonth': start_operatingTimeCurrentMonth = start_address
                if variable == 'vs_operatingTimeLastMonth'   : start_operatingTimeLastMonth    = start_address
                if variable == 'vs_operatingTime'            : start_operatingTime             = start_address

            for value in data:
                number = value['№']

                if number is None: continue

                for key in dop_vs:
                    tag = f'VS_{number}'
                    signal = f'Root{self.name_prefix}.AuxSystems.{tag}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(signal)
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    if key == 'StateAuxSystem'           : address.text = str(start_StateAuxSystem            + 3 * (number - 1))
                    if key == 'numOfStart'               : address.text = str(start_numOfStart                + 2 * (number - 1))
                    if key == 'operatingTimeCurrentMonth': address.text = str(start_operatingTimeCurrentMonth + 2 * (number - 1))
                    if key == 'operatingTimeLastMonth'   : address.text = str(start_operatingTimeLastMonth    + 2 * (number - 1))
                    if key == 'operatingTime'            : address.text = str(start_operatingTime             + 2 * (number - 1))

                    object.append(address)
                    root.append(object)
            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'AuxSystems: Карта адресов OK')
            return     (f'AuxSystems: Карта адресов OK')
        except:
            logger.error(f'AuxSystems: Карта адресов FAILED')
            return      (f'AuxSystems: Карта адресов FAILED')
    # Задвижки
    @logger.catch
    def valves_map(self):
        dop_zd = {'StateValve1'  : 'HMI_ZD',
                  'StateValve2'  : 'HMI_ZD',
                  'StateValve3'  : 'HMI_ZD',
                  'Tm.tmZD'      : 'HMI_ZD',
                  'NumOfOpenings': 'HMI_Statistic_ZD',
                  'NumOfClosings': 'HMI_Statistic_ZD',
                  }
        data = self.data['ZD']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.Valves.', root)
        try:
            for value in data:
                number = value['№']

                if number is None: continue

                for key, item in dop_zd.items():
                    tag = 'ZD_' + str(number)
                    signal = 'Root' + self.name_prefix + '.Valves.' + tag + '.' + key
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(signal)
                    object.append(node_p)
                    address = etree.Element('address')
                    address.text = self.prefix_driver + item
                    object.append(address)
                    array_pos = etree.Element('arrayposition')
                    if key == 'StateValve1'  : array_pos.text = str(5 * (number - 1))
                    if key == 'StateValve2'  : array_pos.text = str((5 * (number - 1)) + 1)
                    if key == 'StateValve3'  : array_pos.text = str((5 * (number - 1)) + 2)
                    if key == 'Percent'      : array_pos.text = str((5 * (number - 1)) + 3)
                    if key == 'Tm.tmZD'      : array_pos.text = str((5 * (number - 1)) + 4)
                    if key == 'NumOfOpenings': array_pos.text = str(2 * (number - 1))
                    if key == 'NumOfClosings': array_pos.text = str(2 * (number - 1) + 1)
                    object.append(array_pos)
                    root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'Valves: Карта адресов OK')
            return (f'Valves: Карта адресов OK')
        except:
            logger.error(f'Valves: Карта адресов FAILED')
            return (f'Valves: Карта адресов FAILED')
    @logger.catch
    def valves_map_modbus(self):
        dop_zd = {'StateValve1',
                  'StateValve2',
                  'StateValve3',
                  'Tm.tmZD',
                  'NumOfOpenings',
                  'NumOfClosings'}

        data    = self.data['ZD']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.Valves.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StateZD'      : start_StateZD       = start_address
                if variable == 'numOfOpenings': start_NumOfOpenings = start_address
                if variable == 'numOfClosings': start_NumOfClosings = start_address

            for value in data:
                number = value['№']

                if number is None: continue

                for key in dop_zd:
                    tag = f'ZD_{number}'
                    signal = f'Root{self.name_prefix}.Valves.{tag}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(signal)
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    if key == 'StateValve1'  : address.text = str(start_StateZD       + 5 * (number - 1))
                    if key == 'StateValve2'  : address.text = str((start_StateZD      + 5 * (number - 1)) + 1)
                    if key == 'StateValve3'  : address.text = str((start_StateZD      + 5 * (number - 1)) + 2)
                    #if key == 'Percent'      : address.text = str(start_operatingTimeLastMonth    + 2 * (number - 1))
                    if key == 'Tm.tmZD'      : address.text = str((start_StateZD      + 5 * (number - 1)) + 4)
                    if key == 'NumOfOpenings': address.text = str(start_NumOfOpenings + 2 * (number - 1))
                    if key == 'NumOfClosings': address.text = str(start_NumOfClosings + 2 * (number - 1))
                    object.append(address)
                    root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Valves: Карта адресов OK')
            return     (f'Valves: Карта адресов OK')
        except:
            logger.error(f'Valves: Карта адресов FAILED')
            return      (f'Valves: Карта адресов FAILED')
    # Агрегаты
    @logger.catch
    def pumps_map(self):
        dop_na = {'StateNA'                         : 'HMI_NA',
                  'StateNAEx'                       : 'HMI_NA',
                  'StateNAStatistic'                : 'HMI_NA',
                  'operatingTimeSinceSwitchingOn'   : 'HMI_statisticNA',
                  'operatingTimeSinceSwitchingOnSet': 'HMI_statisticNA',
                  'operatingTimeBeforeOverhaul'     : 'HMI_statisticNA',
                  'operatingTimeBeforeOverhaulSet'  : 'HMI_statisticNA',
                  'numOfStart'                      : 'HMI_statisticNA',
                  'dateTimeOfStart'                 : 'HMI_statisticNA',
                  'dateTimeOfStop'                  : 'HMI_statisticNA',
                  'operatingTimeCurrentMonth'       : 'HMI_statisticNA',
                  'operatingTimeLastMonth'          : 'HMI_statisticNA',
                  'operatingTimeTO'                 : 'HMI_statisticNA',
                  'operatingTimeTO1'                : 'HMI_statisticNA',
                  'operatingTimeTOSet'              : 'HMI_statisticNA',
                  'operatingTimeMidTO'              : 'HMI_statisticNA',
                  'operatingTimeMidTOSet'           : 'HMI_statisticNA',
                  'operatingTimeThisKvart'          : 'HMI_statisticNA',
                  'operatingTimeLastKvart'          : 'HMI_statisticNA',
                  'operatingTimeFromBegin'          : 'HMI_statisticNA',
                  'operatingTimeED'                 : 'HMI_statisticNA',
                  'operatingTimeEDSet'              : 'HMI_statisticNA',
                  'numOfStartSet'                   : 'HMI_statisticNA',
                  'time24hStart'                    : 'HMI_statisticNA',
                  'timeFromHotStart'                : 'HMI_statisticNA',
                  'numOfStarts24h'                  : 'HMI_statisticNA',
                  'OperatingTimeState'              : '',
                  }
        data = self.data['UMPNA']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.NAs.', root)
        try:
            for value in data:
                number = value['№']

                if number is None: continue

                for key, item in dop_na.items():
                    tag = 'NA_' + str(number)
                    signal = 'Root' + self.name_prefix + '.NAs.' + tag + '.' + key
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(signal)
                    object.append(node_p)
                    address = etree.Element('address')
                    if key == 'OperatingTimeState': address.text = self.prefix_driver + str('statisticNA[' + str(number) + '].state')
                    else                          : address.text = self.prefix_driver + item
                    object.append(address)
                    array_pos = etree.Element('arrayposition')
                    if key == 'StateNA'                         : array_pos.text = str(3 * (number - 1))
                    if key == 'StateNAEx'                       : array_pos.text = str((3 * (number - 1)) + 1)
                    if key == 'StateNAStatistic'                : array_pos.text = str((3 * (number - 1)) + 2)
                    if key == 'operatingTimeSinceSwitchingOn'   : array_pos.text = str(23 * (number - 1))
                    if key == 'operatingTimeSinceSwitchingOnSet': array_pos.text = str((23 * (number - 1)) + 1)
                    if key == 'operatingTimeBeforeOverhaul'     : array_pos.text = str((23 * (number - 1)) + 2)
                    if key == 'operatingTimeBeforeOverhaulSet'  : array_pos.text = str((23 * (number - 1)) + 3)
                    if key == 'numOfStart'                      : array_pos.text = str((23 * (number - 1)) + 4)
                    if key == 'dateTimeOfStart'                 : array_pos.text = str((23 * (number - 1)) + 5)
                    if key == 'dateTimeOfStop'                  : array_pos.text = str((23 * (number - 1)) + 6)
                    if key == 'operatingTimeCurrentMonth'       : array_pos.text = str((23 * (number - 1)) + 7)
                    if key == 'operatingTimeLastMonth'          : array_pos.text = str((23 * (number - 1)) + 8)
                    if key == 'operatingTimeTO'                 : array_pos.text = str((23 * (number - 1)) + 9)
                    if key == 'operatingTimeTO1'                : array_pos.text = str((23 * (number - 1)) + 10)
                    if key == 'operatingTimeTOSet'              : array_pos.text = str((23 * (number - 1)) + 11)
                    if key == 'operatingTimeMidTO'              : array_pos.text = str((23 * (number - 1)) + 12)
                    if key == 'operatingTimeMidTOSet'           : array_pos.text = str((23 * (number - 1)) + 13)
                    if key == 'operatingTimeThisKvart'          : array_pos.text = str((23 * (number - 1)) + 14)
                    if key == 'operatingTimeLastKvart'          : array_pos.text = str((23 * (number - 1)) + 15)
                    if key == 'operatingTimeFromBegin'          : array_pos.text = str((23 * (number - 1)) + 16)
                    if key == 'operatingTimeED'                 : array_pos.text = str((23 * (number - 1)) + 17)
                    if key == 'operatingTimeEDSet'              : array_pos.text = str((23 * (number - 1)) + 18)
                    if key == 'numOfStartSet'                   : array_pos.text = str((23 * (number - 1)) + 19)
                    if key == 'time24hStart'                    : array_pos.text = str((23 * (number - 1)) + 20)
                    if key == 'timeFromHotStart'                : array_pos.text = str((23 * (number - 1)) + 21)
                    if key == 'numOfStarts24h'                  : array_pos.text = str((23 * (number - 1)) + 22)
                    object.append(array_pos)
                    root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'NAs: Карта адресов OK')
            return (f'NAs: Карта адресов OK')
        except:
            logger.error(f'NAs: Карта адресов FAILED')
            return (f'NAs: Карта адресов FAILED')
    @logger.catch
    def pumps_map_modbus(self):
        dop_na = {'StateNA',
                  'StateNAEx',
                  'StateNAStatistic',
                  'operatingTimeSinceSwitchingOn',
                  'operatingTimeSinceSwitchingOnSet',
                  'operatingTimeBeforeOverhaul',
                  'operatingTimeBeforeOverhaulSet',
                  'numOfStart',
                  'dateTimeOfStart',
                  'dateTimeOfStop',
                  'operatingTimeCurrentMonth',
                  'operatingTimeLastMonth',
                  'operatingTimeTO',
                  'operatingTimeTO1',
                  'operatingTimeTOSet',
                  'operatingTimeMidTO',
                  'operatingTimeMidTOSet',
                  'operatingTimeThisKvart',
                  'operatingTimeLastKvart',
                  'operatingTimeFromBegin',
                  'operatingTimeED',
                  'operatingTimeEDSet',
                  'numOfStartSet',
                  'time24hStart',
                  'timeFromHotStart',
                  'numOfStarts24h',
                  'OperatingTimeState',
                  }

        data    = self.data['UMPNA']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.NAs.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StateNA'                         : start_StateNA                          = start_address
                if variable == 'operatingTimeSinceSwitchingOn'   : start_operatingTimeSinceSwitchingOn    = start_address
                if variable == 'operatingTimeSinceSwitchingOnSet': start_operatingTimeSinceSwitchingOnSet = start_address
                if variable == 'operatingTimeBeforeOverhaul'     : start_operatingTimeBeforeOverhaul      = start_address
                if variable == 'operatingTimeBeforeOverhaulSet'  : start_operatingTimeBeforeOverhaulSet   = start_address
                if variable == 'numOfStarts'                     : start_numOfStarts                      = start_address
                if variable == 'numOfStartsSet'                  : start_numOfStartsSet                   = start_address
                if variable == 'dateTimeOfStart'                 : start_dateTimeOfStart                  = start_address
                if variable == 'dateTimeOfStop'                  : start_dateTimeOfStop                   = start_address
                if variable == 'operatingTimeCurrentMonth'       : start_operatingTimeCurrentMonth        = start_address
                if variable == 'operatingTimeLastMonth'          : start_operatingTimeLastMonth           = start_address
                if variable == 'operatingTimeTO'                 : start_operatingTimeTO                  = start_address
                if variable == 'operatingTimeTO1'                : start_operatingTimeTO1                 = start_address
                if variable == 'operatingTimeTOSet'              : start_operatingTimeTOSet               = start_address
                if variable == 'operatingTimeMidTO'              : start_operatingTimeMidTO               = start_address
                if variable == 'operatingTimeMidTOSet'           : start_operatingTimeMidTOSet            = start_address
                if variable == 'operatingTimeThisKvart'          : start_operatingTimeThisKvart           = start_address
                if variable == 'operatingTimeLastKvart'          : start_operatingTimeLastKvart           = start_address
                if variable == 'operatingTimeFromBegin'          : start_operatingTimeFromBegin           = start_address
                if variable == 'operatingTimeED'                 : start_operatingTimeED                  = start_address
                if variable == 'operatingTimeEDSet'              : start_operatingTimeEDSet               = start_address
                if variable == 'operatingTimeState'              : start_operatingTimeState               = start_address

            for value in data:
                number = value['№']

                if number is None: continue

                for key in dop_na:
                    tag    = f'NA_{number}'
                    signal = f'Root{self.name_prefix}.NAs.{tag}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(signal)
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')

                    if key == 'StateNA'                         : address.text = str(start_StateNA  + 11 * (number - 1))
                    if key == 'StateNAEx'                       : address.text = str((start_StateNA + 11 * (number - 1)) + 1)
                    if key == 'StateNAStatistic'                : address.text = str((start_StateNA + 11 * (number - 1)) + 2)
                    if key == 'operatingTimeSinceSwitchingOn'   : address.text = str(start_operatingTimeSinceSwitchingOn    + 42 * (number - 1))
                    if key == 'operatingTimeSinceSwitchingOnSet': address.text = str(start_operatingTimeSinceSwitchingOnSet + 42 * (number - 1))
                    if key == 'operatingTimeBeforeOverhaul'     : address.text = str(start_operatingTimeBeforeOverhaul      + 42 * (number - 1))
                    if key == 'operatingTimeBeforeOverhaulSet'  : address.text = str(start_operatingTimeBeforeOverhaulSet   + 42 * (number - 1))
                    if key == 'numOfStart'                      : address.text = str(start_numOfStarts     + 42 * (number - 1))
                    if key == 'dateTimeOfStart'                 : address.text = str(start_dateTimeOfStart + 42 * (number - 1))
                    if key == 'dateTimeOfStop'                  : address.text = str(start_dateTimeOfStop  + 42 * (number - 1))
                    if key == 'operatingTimeCurrentMonth'       : address.text = str(start_operatingTimeCurrentMonth + 42 * (number - 1))
                    if key == 'operatingTimeLastMonth'          : address.text = str(start_operatingTimeLastMonth    + 42 * (number - 1))
                    if key == 'operatingTimeTO'                 : address.text = str(start_operatingTimeTO    + 42 * (number - 1))
                    if key == 'operatingTimeTO1'                : address.text = str(start_operatingTimeTO1   + 42 * (number - 1))
                    if key == 'operatingTimeTOSet'              : address.text = str(start_operatingTimeTOSet + 42 * (number - 1))
                    if key == 'operatingTimeMidTO'              : address.text = str(start_operatingTimeMidTO    + 42 * (number - 1))
                    if key == 'operatingTimeMidTOSet'           : address.text = str(start_operatingTimeMidTOSet + 42 * (number - 1))
                    if key == 'operatingTimeThisKvart'          : address.text = str(start_operatingTimeThisKvart + 42 * (number - 1))
                    if key == 'operatingTimeLastKvart'          : address.text = str(start_operatingTimeLastKvart + 42 * (number - 1))
                    if key == 'operatingTimeFromBegin'          : address.text = str(start_operatingTimeFromBegin + 42 * (number - 1))
                    if key == 'operatingTimeED'                 : address.text = str(start_operatingTimeED    + 42 * (number - 1))
                    if key == 'operatingTimeEDSet'              : address.text = str(start_operatingTimeEDSet + 42 * (number - 1))
                    if key == 'numOfStartSet'                   : address.text = str(start_numOfStartsSet     + 42 * (number - 1))
                    if key == 'OperatingTimeState'              : address.text = str(start_operatingTimeState + 42 * (number - 1))

                    object.append(address)
                    root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'NAs: Карта адресов OK')
            return     (f'NAs: Карта адресов OK')
        except:
            logger.error(f'NAs: Карта адресов FAILED')
            return      (f'NAs: Карта адресов FAILED')
    # Смежные системы
    @logger.catch
    def relayted_system_map(self):
        data      = self.data['SS']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.SSs.', root)
        try:
            for value in data:
                number = value['№']
                name   = value['Название']

                if number is None: continue
                if name is None: continue

                tag = 'SS_' + str(number)
                # Имя строки используемая в карте
                signal = 'Root' + self.name_prefix + '.SSs.' + tag + '.StateSS'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = str(signal)
                object.append(node_p)
                address = etree.Element('address')
                address.text = str(self.prefix_driver + 'stateDiag.SS[' + str(number) +  '].state')
                object.append(address)
                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'SSs: Карта адресов OK')
            return (f'SSs: Карта адресов OK')
        except:
            logger.error(f'SSs: Карта адресов FAILED')
            return (f'SSs: Карта адресов FAILED')
    @logger.catch
    def relayted_system_map_modbus(self):
        data       = self.data['SS']
        data_mb    = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.SSs.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'SS': start_SS = start_address

            for value in data:
                number = value['№']
                name   = value['Название']

                if number is None: continue
                if name   is None: continue

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.SSs.SS_{number}.StateSS'
                object.append(node_p)

                segment = etree.Element('table')
                segment.text = f'Holding Registers'
                object.append(segment)

                address = etree.Element('address')
                address.text = f'{start_SS + (number - 1)}'
                object.append(address)
                root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'SSs: Карта адресов OK')
            return     (f'SSs: Карта адресов OK')
        except:
            logger.error(f'SSs: Карта адресов FAILED')
            return      (f'SSs: Карта адресов FAILED')
    # Табло и сирены(станция) - UTS
    @logger.catch
    def uts_map(self):
        data = self.data['UTS']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.UTSs.', root)
        try:
            for value in data:
                number = value['№']
                tag    = value['Идентификатор']
                name   = value['Название']

                if number is None: continue
                if name   is None: continue

                tag = self.translate(str(tag))

                # Имя строки используемая в карте
                signal = f'Root{self.name_prefix}.UTSs.{tag}.StateUTS'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = str(signal)
                object.append(node_p)
                address = etree.Element('address')
                address.text = str(f'{self.prefix_driver}stateUTS[{str(number)}].state')
                object.append(address)
                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'UTSs: Карта адресов OK')
            return (f'UTSs: Карта адресов OK')
        except:
            logger.error(f'UTSs: Карта адресов FAILED')
            return (f'UTSs: Карта адресов FAILED')
    @logger.catch
    def uts_map_modbus(self):
        data    = self.data['UTS']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.UTSs.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StateUTS': start_StateUTS = start_address

            for value in data:
                number = value['№']
                tag    = value['Идентификатор']
                name   = value['Название']

                if number is None: continue
                if name   is None: continue
                if tag    is None: continue

                tag = self.translate(str(tag))

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.UTSs.{tag}.StateUTS'
                object.append(node_p)

                segment = etree.Element('table')
                segment.text = f'Holding Registers'
                object.append(segment)

                address = etree.Element('address')
                address.text = str(start_StateUTS + (number - 1))
                object.append(address)
                root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'UTSs: Карта адресов OK')
            return     (f'UTSs: Карта адресов OK')
        except:
            logger.error(f'UTSs: Карта адресов FAILED')
            return      (f'UTSs: Карта адресов FAILED')
    # Табло и сирены(ПТ) - UPTS
    @logger.catch
    def upts_map(self):
        data = self.data['UPTS']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.UPTSs.', root)
        try:
            for value in data:
                number = value['№']
                tag    = value['Идентификатор']
                name   = value['Название']

                if number is None: continue
                if name   is None: continue

                tag = self.translate(str(tag))

                # Имя строки используемая в карте
                signal = f'Root{self.name_prefix}.UPTSs.{tag}.StateUPTS'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = str(signal)
                object.append(node_p)
                address = etree.Element('address')
                address.text = f'{self.prefix_driver}HMI_UPTS'
                object.append(address)
                array_pos = etree.Element('arrayposition')
                array_pos.text = str(number - 1)
                object.append(array_pos)
                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'UPTSs: Карта адресов OK')
            return     (f'UPTSs: Карта адресов OK')
        except:
            logger.error(f'UPTSs: Карта адресов FAILED')
            return      (f'UPTSs: Карта адресов FAILED')
    @logger.catch
    def upts_map_modbus(self):
        data    = self.data['UPTS']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.UPTSs.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StateUPTS': start_StateUPTS = start_address

            for value in data:
                number = value['№']
                tag    = value['Идентификатор']
                name   = value['Название']

                if number is None: continue
                if name   is None: continue
                if tag    is None: continue

                tag = self.translate(str(tag))

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.UPTSs.{tag}.StateUPTS'
                object.append(node_p)

                segment = etree.Element('table')
                segment.text = f'Holding Registers'
                object.append(segment)

                address = etree.Element('address')
                address.text = str(start_StateUPTS + (number - 1))
                object.append(address)
                root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'UPTSs: Карта адресов OK')
            return     (f'UPTSs: Карта адресов OK')
        except:
            logger.error(f'UPTSs: Карта адресов FAILED')
            return      (f'UPTSs: Карта адресов FAILED')
    # Общестнационные защиты(МНС)
    @logger.catch
    def ktpr_map(self):
        data = self.data['KTPR']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.KTPRs.', root)
        number_group = 0
        try:
            for value in data:
                number_defence = value['№']
                if number_defence is None: continue
                number_group += 1
            count_group = math.ceil(number_group/4)

            for count in range(count_group):
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.KTPRs.Group_{count + 1}.StateKTPRx'
                    object.append(node_p)
                    address = etree.Element('address')
                    address.text = str(f'{self.prefix_driver}stateKTPR_HMI')
                    object.append(address)
                    array_pos = etree.Element('arrayposition')
                    array_pos.text = str(count - 1)
                    object.append(array_pos)
                    root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'KTPRs: Карта адресов OK')
            return (f'KTPRs: Карта адресов OK')
        except:
            logger.error(f'KTPRs: Карта адресов FAILED')
            return (f'KTPRs: Карта адресов FAILED')
    @logger.catch
    def ktpr_map_modbus(self):
        data    = self.data['KTPR']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.KTPRs.', root)
        number_group = 0
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'stateKTPRx': start_stateKTPRx = start_address

            for value in data:
                number_defence = value['№']
                if number_defence is None: continue
                number_group += 1
            count_group = math.ceil(number_group/4)

            for count in range(count_group):
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.KTPRs.Group_{count + 1}.StateKTPRx'
                object.append(node_p)

                segment = etree.Element('table')
                segment.text = f'Holding Registers'
                object.append(segment)

                address = etree.Element('address')
                address.text = str(start_stateKTPRx + (count - 1))
                object.append(address)
                root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'KTPRs: Карта адресов OK')
            return     (f'KTPRs: Карта адресов OK')
        except:
            logger.error(f'KTPRs: Карта адресов FAILED')
            return      (f'KTPRs: Карта адресов FAILED')
    # Общестнационные защиты(ПТ)
    @logger.catch
    def ktprp_map(self):
        data = self.data['KTPRP']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.KTPRs.', root)
        number_group = 0
        try:
            for value in data:
                number_defence = value['№']
                if number_defence is None: continue
                number_group += 1
            count_group = math.ceil(number_group / 4)

            for count in range(count_group):
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.KTPRs.Group_{count + 1}.StateKTPRx'
                object.append(node_p)
                address = etree.Element('address')
                address.text = str(f'{self.prefix_driver}stateKTPR_HMI')
                object.append(address)
                array_pos = etree.Element('arrayposition')
                array_pos.text = str(count - 1)
                object.append(array_pos)
                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'KTPRPs: Карта адресов OK')
            return     (f'KTPRPs: Карта адресов OK')
        except:
            logger.error(f'KTPRPs: Карта адресов FAILED')
            return      (f'KTPRPs: Карта адресов FAILED')
    @logger.catch
    def ktprp_map_modbus(self):
        data    = self.data['KTPRP']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.KTPRs.', root)
        number_group = 0
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'stateKTPRx': start_stateKTPRx = start_address

            for value in data:
                number_defence = value['№']
                if number_defence is None: continue
                number_group += 1
            count_group = math.ceil(number_group / 4)

            for count in range(count_group):
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.KTPRs.Group_{count + 1}.StateKTPRx'
                object.append(node_p)

                segment = etree.Element('table')
                segment.text = f'Holding Registers'
                object.append(segment)

                address = etree.Element('address')
                address.text = str(start_stateKTPRx + (count - 1))
                object.append(address)
                root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'KTPRs: Карта адресов OK')
            return     (f'KTPRs: Карта адресов OK')
        except:
            logger.error(f'KTPRs: Карта адресов FAILED')
            return      (f'KTPRs: Карта адресов FAILED')
    # Агрегатные защиты
    @logger.catch
    def ktpra_map(self):
        data = self.data['KTPRA']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.KTPRAs.', root)
        number_pumps_old = ''
        count_pumps = 0
        try:
            for value in data:
                number_defence   = value['№']
                number_pumps_int = value['НА']

                if number_defence   is None: continue
                if number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps += 1
                    number_group = 0

                if number_defence % 4 == 0:
                    number_group += 1
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.KTPRAs.KTPRAs_{count_pumps}.Group_{number_group}.StateKTPRx'
                    object.append(node_p)
                    address = etree.Element('address')
                    address.text = str(f'{self.prefix_driver}stateKTPRA{count_pumps}_HMI')
                    object.append(address)
                    array_pos = etree.Element('arrayposition')
                    array_pos.text = str(number_group - 1)
                    object.append(array_pos)
                    root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'KTPRAs: Карта адресов OK')
            return (f'KTPRAs: Карта адресов OK')
        except:
            logger.error(f'KTPRAs: Карта адресов FAILED')
            return (f'KTPRAs: Карта адресов FAILED')
    @logger.catch
    def ktpra_map_modbus(self):
        data    = self.data['KTPRA']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.KTPRAs.', root)
        number_pumps_old = ''
        count_pumps      = 0
        count            = 0
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'stateKTPRA': start_stateKTPRA = start_address

            for value in data:
                number_defence   = value['№']
                number_pumps_int = value['НА']

                if number_defence   is None: continue
                if number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps  += 1
                    number_group  = 0

                if number_defence % 4 == 0:
                    number_group += 1
                    count += 1

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.KTPRAs.KTPRAs_{count_pumps}.Group_{number_group}.StateKTPRx'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_stateKTPRA + (number_group - 1) + (count_pumps - 1) * 48)
                    object.append(address)
                    root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'KTPRAs: Карта адресов OK')
            return      (f'KTPRAs: Карта адресов OK')
        except:
            logger.error(f'KTPRAs: Карта адресов FAILED')
            return      (f'KTPRAs: Карта адресов FAILED')
    # Агрегатные готовности
    @logger.catch
    def gmpna_map(self):
        data = self.data['GMPNA']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.GMPNAs.', root)
        number_pumps_old = ''
        count_pumps = 0
        try:
            for value in data:
                number_defence   = value['№']
                number_pumps_int = value['НА']

                if number_defence   is None: continue
                if number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps += 1
                    number_group = 0

                if number_defence % 4 == 0:
                    number_group += 1
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.GMPNAs.GMPNAs_{count_pumps}.Group_{number_group}.StateGMPNA'
                    object.append(node_p)
                    address = etree.Element('address')
                    address.text = str(f'{self.prefix_driver}stateGMPNA{count_pumps}_HMI')
                    object.append(address)
                    array_pos = etree.Element('arrayposition')
                    array_pos.text = str(number_group - 1)
                    object.append(array_pos)
                    root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'GMPNAs: Карта адресов OK')
            return (f'GMPNAs: Карта адресов OK')
        except:
            logger.error(f'GMPNAs: Карта адресов FAILED')
            return (f'GMPNAs: Карта адресов FAILED')
    @logger.catch
    def gmpna_map_modbus(self):
        data    = self.data['GMPNA']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.GMPNAs.', root)
        number_pumps_old = ''
        count_pumps = 0
        count       = 0
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'stateGMPNA': start_stateGMPNA = start_address

            for value in data:
                number_defence   = value['№']
                number_pumps_int = value['НА']

                if number_defence   is None: continue
                if number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps += 1
                    number_group = 0

                if number_defence % 4 == 0:
                    number_group += 1
                    count        += 1

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.GMPNAs.GMPNAs_{count_pumps}.Group_{number_group}.StateGMPNA'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_stateGMPNA + (count - 1))
                    object.append(address)
                    root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'GMPNAs: Карта адресов OK')
            return     (f'GMPNAs: Карта адресов OK')
        except:
            logger.error(f'GMPNAs: Карта адресов FAILED')
            return      (f'GMPNAs: Карта адресов FAILED')
    # Пожарные извещатели
    @logger.catch
    def pi_map(self):
        data = self.data['PI']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.PIs.', root)
        try:
            for value in data:
                number = value['№']
                tag    = value['Идентификатор']

                if number is None: continue
                if tag    is None: continue

                tag = self.translate(str(tag))

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.PIs.{tag}.StatePI'
                object.append(node_p)

                address = etree.Element('address')
                address.text = str(self.prefix_driver + 'HMI_PI')
                object.append(address)

                array_pos = etree.Element('arrayposition')
                array_pos.text = str(number - 1)
                object.append(array_pos)
                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'PIs: Карта адресов OK')
            return     (f'PIs: Карта адресов OK')
        except:
            logger.error(f'PIs: Карта адресов FAILED')
            return      (f'PIs: Карта адресов FAILED')
    @logger.catch
    def pi_map_modbus(self):
        data    = self.data['PI']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.PIs.', root)
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StatePI': start_StatePI = start_address

            for value in data:
                number = value['№']
                tag    = value['Идентификатор']

                if number is None: continue
                if tag    is None: continue

                tag = self.translate(str(tag))

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = f'Root{self.name_prefix}.PIs.{tag}.StatePI'
                object.append(node_p)

                segment = etree.Element('table')
                segment.text = f'Holding Registers'
                object.append(segment)

                address = etree.Element('address')
                address.text = str(start_StatePI + (number - 1))
                object.append(address)
                root.append(object)

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'PIs: Карта адресов OK')
            return     (f'PIs: Карта адресов OK')
        except:
            logger.error(f'PIs: Карта адресов FAILED')
            return      (f'PIs: Карта адресов FAILED')
    # Пожарные зоны
    @logger.catch
    def pz_map(self):
        # Зоны с тушением
        dop_pz_ptush = {'StatePZ'         : 'OFS!HMI_PZ',
                        'ReadyFlags'      : 'OFS!HMI_PZ',
                        'TimetoNextAttack': 'OFS!HMI_PZ',
                        'AttackCounter'   : 'OFS!HMI_PZ',
                        'TimetoEvacuation': 'OFS!HMI_PZ',
                        'exStatePZ'       : 'OFS!HMI_PZ',
                        }
        # Зоны без тушения
        dop_pz = {'StatePZ'   : 'OFS!HMI_PZ',
                  'ReadyFlags': 'OFS!HMI_PZ',
                  'exStatePZ' : 'OFS!HMI_PZ',
                  }

        data       = self.data['PZ']
        root, tree = self.parser_map()
        # Чистка тэгов
        self.cleaner_map('.PZs.', root)
        number_array = 0
        try:
            for value in data:
                number    = value['№']
                zone_type = value['Тип']

                if number    is None: continue
                if zone_type is None: continue
                # Выбираем от типа
                set_words = dop_pz if zone_type == 0 else dop_pz_ptush

                for key, value in set_words.items():
                    number_array += 1

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.PZs.PZ_{number}.{key}'
                    object.append(node_p)

                    address = etree.Element('address')
                    address.text = f'{self.prefix_driver}HMI_PZ'
                    object.append(address)

                    array_pos = etree.Element('arrayposition')
                    if key == 'StatePZ'         : array_pos.text = str(6 * (number_array - 1))
                    if key == 'ReadyFlags'      : array_pos.text = str(6 * (number_array - 1) + 1)
                    if key == 'TimetoNextAttack': array_pos.text = str(6 * (number_array - 1) + 2)
                    if key == 'AttackCounter'   : array_pos.text = str(6 * (number_array - 1) + 3)
                    if key == 'TimetoEvacuation': array_pos.text = str(6 * (number_array - 1) + 4)
                    if key == 'exStatePZ'       : array_pos.text = str(6 * (number_array - 1) + 5)
                    object.append(array_pos)
                    root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'PZs: Карта адресов OK')
            return     (f'PZs: Карта адресов OK')
        except:
            logger.error(f'PZs: Карта адресов FAILED')
            return      (f'PZs: Карта адресов FAILED')
    @logger.catch
    def pz_map_modbus(self):
        # Зоны с тушением
        dop_pz_ptush = ['StatePZ', 'exStatePZ', 'ReadyFlags', 'TimetoNextAttack', 'AttackCounter', 'TimetoEvacuation']
        # Зоны без тушения
        #dop_pz = ['StatePZ', 'exStatePZ', 'ReadyFlags']

        data    = self.data['PZ']
        data_mb = self.data['ModBus']
        root, tree = self.parser_map_modbus()
        # Чистка тэгов
        self.cleaner_map('.PZs.', root)
        count = 0
        try:
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'StatePZ': start_StatePZ = start_address

            for value in data:
                number    = value['№']
                zone_type = value['Тип']

                if number    is None: continue
                if zone_type is None: continue
                # Выбираем от типа
                #set_words = dop_pz if zone_type == 0 else dop_pz_ptush

                for key in dop_pz_ptush:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.PZs.PZ_{number}.{key}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_StatePZ + count)
                    object.append(address)
                    root.append(object)
                    count += 1
            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'PZs: Карта адресов OK')
            return     (f'PZs: Карта адресов OK')
        except:
            logger.error(f'PZs: Карта адресов FAILED')
            return      (f'PZs: Карта адресов FAILED')

    # Диагностика SE
    # Аналоговые сигналы входные
    @logger.catch
    def diag_analogs_in(self, MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('AIs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        link_path = MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'ai'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_AI"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.AIs: файл omx OK')
        except:
            logger.error(f'Diag.AIs: файл omx FAILED')
        # 2 этап
        # Заполняем значения атрибутов
        # AttributesMapAI_Ref.xml, AttributesMapKlk.xml,
        # AttributesMapKont.xml, AttributesMapSignalName.xml, AttributesMapTagName.xml
        try:
            # Цикл по всем xml
            for path in link_path:
                root, tree = self.parser_diag_map(path)
                # Чистка тэгов
                self.cleaner_diag_map('.Diag.AIs.', root)
                # Цикл по всем добавленным модулям AI
                for initial_data in signals:
                    name_rus = initial_data['name_uso_rus']
                    name_eng = initial_data['uso_rack_modul']
                    basket   = initial_data['rack']
                    mod      = initial_data['modul_dash']
                    for value in data_kd:
                        klk      = value['КлК']
                        kont     = value['Конт']
                        desc     = value['Наименование']
                        tag      = value['Tэг']
                        uso      = value['Шкаф']
                        basket_v = value['Корз']
                        modul    = value['Мод']
                        channel  = value['Кан']
                        if (name_rus == str(uso)) and (basket == str(basket_v)) and (mod == str(modul)):
                            str_tag = self.translate(str(tag))

                            if klk is None : klk = ' '
                            if kont is None: kont = ' '

                            name_AI = 'Root' + self.name_prefix + '.Diag.AIs.' + name_eng + '.ch_AI_0' + str(channel + 1)
                            object = etree.Element('item')
                            object.attrib['id'] = name_AI
                            if path == MapAI_Ref       :
                                if not str_tag is None : object.attrib['value'] = str(str_tag)
                            if path == MapKlk          :
                                if not klk is None     : object.attrib['value'] = str(klk)
                            if path == MapKont         :
                                if not kont is None    : object.attrib['value'] = str(kont)
                            if path == MapSignalName   :
                                if not desc is None    : object.attrib['value'] = str(desc)
                            if path == MapTagName      :
                                if not str_tag is None : object.attrib['value'] = str(str_tag)
                            root.append(object)
                logger.info(f'Diag.AIs: карта атрибутов: {path} - OK')
                tree.write(path, pretty_print=True)
        except:
            logger.info(f'Diag.AIs: карта атрибутов: {path} - FAILED')
        # 3 этап карта адресов
        try:
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.AIs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            count_HEALT = 0
            # Цикл по всем добавленным модулям AI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству каналов и здоровье в модуле 9 + 1
                count_HEALT += 1
                for i in range(9):
                    if i < 8: count_array += 1
                    num_series = i + 1
                    name_AI       = 'Root' + self.name_prefix + '.Diag.AIs.' + name_eng + '.mAI[' + str(num_series) + ']'
                    name_AI_HEALT = 'Root' + self.name_prefix + '.Diag.AIs.' + name_eng + '.mAI_CH_HEALTH'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    if i > 7: node_p.text = str(name_AI_HEALT)
                    else    : node_p.text = str(name_AI)
                    object.append(node_p)

                    address = etree.Element('address')
                    if i > 7: address.text = self.prefix_driver + 'mAI_CH_HEALTH'
                    else    : address.text = self.prefix_driver + 'mAI'
                    object.append(address)

                    array_pos = etree.Element('arrayposition')
                    if i > 7: array_pos.text = str(count_HEALT - 1)
                    else    : array_pos.text = str(count_array - 1)
                    object.append(array_pos)
                    root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.AIs: файл map OK')
            return (f'Diag.AIs: файл map OK')
        except:
            logger.error(f'Diag.AIs: файл map FAILED')
            return (f'Diag.AIs: файл map FAILED')
    # Аналоговые сигналы выходные
    @logger.catch
    def diag_analogs_out(self, MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('AOs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        link_path = MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'ao'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_AO"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.AOs: файл omx OK')
        except:
            logger.error(f'Diag.AOs: файл omx FAILED')
        # 2 этап
        # Заполняем значения атрибутов
        # AttributesMapAI_Ref.xml, AttributesMapKlk.xml,
        # AttributesMapKont.xml, AttributesMapSignalName.xml, AttributesMapTagName.xml
        try:
            # Цикл по всем xml
            for path in link_path:
                root, tree = self.parser_diag_map(path)
                # Чистка тэгов
                self.cleaner_diag_map('.Diag.AOs.', root)
                # Цикл по всем добавленным модулям AI
                for initial_data in signals:
                    name_rus = initial_data['name_uso_rus']
                    name_eng = initial_data['uso_rack_modul']
                    basket   = initial_data['rack']
                    mod      = initial_data['modul_dash']
                    for value in data_kd:
                        klk      = value['КлК']
                        kont     = value['Конт']
                        desc     = value['Наименование']
                        tag      = value['Tэг']
                        uso      = value['Шкаф']
                        basket_v = value['Корз']
                        modul    = value['Мод']
                        channel  = value['Кан']
                        if (name_rus == str(uso)) and (basket == str(basket_v)) and (mod == str(modul)):
                            str_tag = self.translate(str(tag))

                            if klk is None : klk = ' '
                            if kont is None: kont = ' '

                            name_AO = 'Root' + self.name_prefix + '.Diag.AOs.' + name_eng + '.ch_AI_0' + str(channel + 1)
                            object = etree.Element('item')
                            object.attrib['id'] = name_AO
                            if path == MapAI_Ref       :
                                if not str_tag is None : object.attrib['value'] = str(str_tag)
                            if path == MapKlk          :
                                if not klk is None     : object.attrib['value'] = str(klk)
                            if path == MapKont         :
                                if not kont is None    : object.attrib['value'] = str(kont)
                            if path == MapSignalName   :
                                if not desc is None    : object.attrib['value'] = str(desc)
                            if path == MapTagName      :
                                if not str_tag is None : object.attrib['value'] = str(str_tag)
                            root.append(object)
                logger.info(f'Diag.AOs: карта атрибутов: {path} - OK')
                tree.write(path, pretty_print=True)
        except:
            logger.info(f'Diag.AOs: карта атрибутов: {path} - FAILED')
        # 3 этап карта адресов
        try:
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.AOs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            count_HEALT = 0
            # Цикл по всем добавленным модулям AI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству каналов и здоровье в модуле 9 + 1
                count_HEALT += 1
                for i in range(5):
                    if i < 4: count_array += 1
                    num_series = i + 1
                    name_AO       = 'Root' + self.name_prefix + '.Diag.AOs.' + name_eng + '.mAO[' + str(num_series) + ']'
                    name_AO_HEALT = 'Root' + self.name_prefix + '.Diag.AOs.' + name_eng + '.mAI_CH_HEALTH'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    if i > 3: node_p.text = str(name_AO_HEALT)
                    else    : node_p.text = str(name_AO)
                    object.append(node_p)

                    address = etree.Element('address')
                    if i > 3: address.text = self.prefix_driver + 'mAO_CH_HEALTH'
                    else    : address.text = self.prefix_driver + 'mAO'
                    object.append(address)

                    array_pos = etree.Element('arrayposition')
                    if i > 3: array_pos.text = str(count_HEALT - 1)
                    else    : array_pos.text = str(count_array - 1)
                    object.append(array_pos)
                    root.append(object)
            logger.info(f'Diag.AOs: файл map OK')
            return (f'Diag.AOs: файл map OK')
        except:
            logger.error(f'Diag.AOs: файл map FAILED')
            return (f'Diag.AOs: файл map FAILED')
    # Дискретные сигналы входные
    @logger.catch
    def diag_diskrets_in(self, MapKlk, MapKont, MapSignalName, MapTagName):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('DIs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        link_path = MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'mdi'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_DI"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.DIs: файл omx OK')
        except:
            logger.error(f'Diag.DIs: файл omx FAILED')
        # 2 этап
        # Заполняем значения атрибутов
        try:
        # Цикл по всем xml
            for path in link_path:
                root, tree = self.parser_diag_map(path)
                # Чистка тэгов
                self.cleaner_diag_map('.Diag.DIs.', root)
                # Цикл по всем добавленным модулям DI
                for initial_data in signals:
                    name_rus = initial_data['name_uso_rus']
                    name_eng = initial_data['uso_rack_modul']
                    basket   = initial_data['rack']
                    mod      = initial_data['modul_dash']
                    for value in data_kd:
                        klk      = value['КлК']
                        kont     = value['Конт']
                        desc     = value['Наименование']
                        tag      = value['Tэг']
                        uso      = value['Шкаф']
                        basket_v = value['Корз']
                        modul    = value['Мод']
                        channel  = value['Кан']
                        if (name_rus == str(uso)) and (basket == str(basket_v)) and (mod == str(modul)):
                            str_tag = self.translate(str(tag))

                            if klk is None : klk = ' '
                            if kont is None: kont = ' '

                            if channel <= 8: name_DI = 'Root' + self.name_prefix + '.Diag.DIs.' + name_eng + '.ch_DI_0' + str(channel + 1)
                            else:            name_DI = 'Root' + self.name_prefix + '.Diag.DIs.' + name_eng + '.ch_DI_' + str(channel + 1)

                            object = etree.Element('item')
                            object.attrib['id'] = name_DI
                            if path == MapKlk:
                                if not klk is None :  object.attrib['value'] = str(klk)
                            if path == MapKont:
                                if not kont is None:  object.attrib['value'] = str(kont)
                            if path == MapSignalName :
                                if not desc is None:  object.attrib['value'] = str(desc)
                            if path == MapTagName:
                                if not tag is None :  object.attrib['value'] = str(tag)
                            root.append(object)
                logger.info(f'Diag.DIs: карта атрибутов: {path} - OK')
                tree.write(path, pretty_print=True)
        except:
            logger.info(f'Diag.DIs: карта атрибутов: {path} - FAILED')
        # 3 этап карта адресов
        try:
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.DIs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            # Цикл по всем добавленным модулям DI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству дискретных модулей
                count_array += 1
                name_DI = 'Root' + self.name_prefix + '.Diag.DIs.' + name_eng + '.mDI'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = str(name_DI)
                object.append(node_p)

                address = etree.Element('address')
                address.text = self.prefix_driver + 'mDI'
                object.append(address)

                array_pos = etree.Element('arrayposition')
                array_pos.text = str(count_array - 1)
                object.append(array_pos)
                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.DIs: файл map OK')
            return (f'Diag.DIs: файл map OK')
        except:
            logger.error(f'Diag.DIs: файл map FAILED')
            return (f'Diag.DIs: файл map FAILED')
    # Дискретные сигналы выходные
    @logger.catch
    def diag_diskrets_out(self, MapKlk, MapKont, MapSignalName, MapTagName):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('DOs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        link_path = MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'mdo'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_DI"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.DOs: файл omx OK')
        except:
            logger.error(f'Diag.DOs: файл omx FAILED')
        # 2 этап
        # Заполняем значения атрибутов
        try:
        # Цикл по всем xml
            for path in link_path:
                root, tree = self.parser_diag_map(path)
                # Чистка тэгов
                self.cleaner_diag_map('.Diag.DOs.', root)
                # Цикл по всем добавленным модулям DI
                for initial_data in signals:
                    name_rus = initial_data['name_uso_rus']
                    name_eng = initial_data['uso_rack_modul']
                    basket   = initial_data['rack']
                    mod      = initial_data['modul_dash']
                    for value in data_kd:
                        klk      = value['КлК']
                        kont     = value['Конт']
                        desc     = value['Наименование']
                        tag      = value['Tэг']
                        uso      = value['Шкаф']
                        basket_v = value['Корз']
                        modul    = value['Мод']
                        channel  = value['Кан']
                        if (name_rus == str(uso)) and (basket == str(basket_v)) and (mod == str(modul)):
                            str_tag = self.translate(str(tag))

                            if klk is None : klk = ' '
                            if kont is None: kont = ' '

                            if channel <= 8: name_DI = 'Root' + self.name_prefix + '.Diag.DOs.' + name_eng + '.ch_DI_0' + str(channel + 1)
                            else:            name_DI = 'Root' + self.name_prefix + '.Diag.DOs.' + name_eng + '.ch_DI_' + str(channel + 1)

                            object = etree.Element('item')
                            object.attrib['id'] = name_DI
                            if path == MapKlk:
                                if not klk is None :  object.attrib['value'] = str(klk)
                            if path == MapKont:
                                if not kont is None:  object.attrib['value'] = str(kont)
                            if path == MapSignalName :
                                if not desc is None:  object.attrib['value'] = str(desc)
                            if path == MapTagName:
                                if not tag is None :  object.attrib['value'] = str(tag)
                            root.append(object)
                logger.info(f'Diag.DOs: карта атрибутов: {path} - OK')
                tree.write(path, pretty_print=True)
        except:
            logger.info(f'Diag.DOs: карта атрибутов: {path} - FAILED')
        # 3 этап карта адресов
        try:
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.DOs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            # Цикл по всем добавленным модулям DI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству дискретных модулей
                count_array += 1
                name_DI = 'Root' + self.name_prefix + '.Diag.DOs.' + name_eng + '.mDI'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = str(name_DI)
                object.append(node_p)

                address = etree.Element('address')
                address.text = self.prefix_driver + 'mDO'
                object.append(address)

                array_pos = etree.Element('arrayposition')
                array_pos.text = str(count_array - 1)
                object.append(array_pos)
                root.append(object)
            tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.DOs: файл map OK')
            return (f'Diag.DOs: файл map OK')
        except:
            logger.error(f'Diag.DOs: файл map FAILED')
            return (f'Diag.DOs: файл map FAILED')
    # Модули КЦ
    @logger.catch
    def diag_cpukcs(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('CPUKCs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'cpukc'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_CPUKC"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.CPUKCs: файл omx OK')
        except:
            logger.error(f'Diag.CPUKCs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            dop_CPUKCs = {'.diag'        : 'stateDiag_HMI',
                          '.port1_errcnt': 'stateNetErrCnt_HMI',
                          '.port2_errcnt': 'stateNetErrCnt_HMI',
                          '.port3_errcnt': 'stateNetErrCnt_HMI',
                          '.diag2'       : 'stateDiag_HMI',
                          '.LED'         : 'stateDiag_HMI',
                       }
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.CPUKCs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            # Цикл по всем добавленным модулям DI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству дискретных модулей
                count_array += 1
                for key, value in dop_CPUKCs.items():
                    name_CPU = 'Root' + self.name_prefix + '.Diag.CPUKCs.' + name_eng + key

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(name_CPU)
                    object.append(node_p)

                    address = etree.Element('address')
                    address.text = self.prefix_driver + value
                    object.append(address)

                    array_pos = etree.Element('arrayposition')
                    if key == '.diag'        : array_pos.text = str(146 + (3 * (count_array - 1)))
                    if key == '.diag2'       : array_pos.text = str(147 + (3 * (count_array - 1)))
                    if key == '.LED'         : array_pos.text = str(148 + (3 * (count_array - 1)))
                    if key == '.port1_errcnt': array_pos.text = str(1 + (4 * (count_array - 1)))
                    if key == '.port2_errcnt': array_pos.text = str(2 + (4 * (count_array - 1)))
                    if key == '.port3_errcnt': array_pos.text = str(3 + (4 * (count_array - 1)))
                    object.append(array_pos)
                    root.append(object)
                tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.CPUKCs: файл map OK')
            return (f'Diag.CPUKCs: файл map OK')
        except:
            logger.error(f'Diag.CPUKCs: файл map FAILED')
            return (f'Diag.CPUKCs: файл map FAILED')
    # Модули P58, P34
    @logger.catch
    def diag_cpus(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('CPUs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'mcpu'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_CPU"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.CPUs: файл omx OK')
        except:
            logger.error(f'Diag.CPUs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            dop_CPUs = {'.diag'        : 'stateDiag_HMI',
                        '.port1_errcnt': 'stateDiag_HMI',
                        '.port2_errcnt': 'stateDiag_HMI',
                        '.port3_errcnt': 'stateDiag_HMI',
                       }
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.CPUs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            # Цикл по всем добавленным модулям DI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству дискретных модулей
                count_array += 1
                for key, value in dop_CPUs.items():
                    name_CPU = 'Root' + self.name_prefix + '.Diag.CPUs.' + name_eng + key

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(name_CPU)
                    object.append(node_p)

                    address = etree.Element('address')
                    address.text = self.prefix_driver + value
                    object.append(address)

                    array_pos = etree.Element('arrayposition')
                    if key == '.diag'        : array_pos.text = str(152 + count_array - 1)
                    if key == '.port1_errcnt': array_pos.text = str(152 + count_array - 1)
                    if key == '.port2_errcnt': array_pos.text = str(152 + count_array - 1)
                    if key == '.port3_errcnt': array_pos.text = str(152 + count_array - 1)
                    object.append(array_pos)
                    root.append(object)
                tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.CPUs: файл map OK')
            return (f'Diag.CPUs: файл map OK')
        except:
            logger.error(f'Diag.CPUs: файл map FAILED')
            return (f'Diag.CPUs: файл map FAILED')
    # Модули NOC, NOE
    @logger.catch
    def diag_noc_noe(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('NOC_NOEs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'mnoc_noe'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        isdigit_num  = re.findall('\d+', str(cell_ai))
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash,
                                      isdigit_num = isdigit_num[0])
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_NOC_NOE"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.NOC_NOEs: файл omx OK')
        except:
            logger.error(f'Diag.NOC_NOEs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            dop_NOC_NOEs = {'.diag'        : 'stateDiag_HMI',
                            '.port1_errcnt': 'stateNetErrCnt_HMI',
                            '.port2_errcnt': 'stateNetErrCnt_HMI',
                            '.port3_errcnt': 'stateNetErrCnt_HMI',
                       }
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.NOC_NOEs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            # Цикл по всем добавленным модулям DI
            for initial_data in signals:
                name_eng    = initial_data['uso_rack_modul']
                isdigit_num = int(initial_data['isdigit_num'])
                # Цикл по количеству дискретных модулей
                count_array += 1
                for key, value in dop_NOC_NOEs.items():
                    name_CPU = 'Root' + self.name_prefix + '.Diag.NOC_NOEs.' + name_eng + key

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(name_CPU)
                    object.append(node_p)

                    address = etree.Element('address')
                    address.text = self.prefix_driver + value
                    object.append(address)

                    array_pos = etree.Element('arrayposition')
                    if key == '.diag'        : array_pos.text = str(164 + (isdigit_num - 1))
                    if key == '.port1_errcnt': array_pos.text = str(57 + (4 * (isdigit_num - 1)))
                    if key == '.port2_errcnt': array_pos.text = str(58 + (4 * (isdigit_num - 1)))
                    if key == '.port3_errcnt': array_pos.text = str(59 + (4 * (isdigit_num - 1)))
                    object.append(array_pos)
                    root.append(object)
                tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.NOC_NOEs: файл map OK')
            return (f'Diag.NOC_NOEs: файл map OK')
        except:
            logger.error(f'Diag.NOC_NOEs: файл map FAILED')
            return (f'Diag.NOC_NOEs: файл map FAILED')
    # Модули CRA
    @logger.catch
    def diag_cras(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('CRAs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'mcra'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_CRA"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.CRAs: файл omx OK')
        except:
            logger.error(f'Diag.CRAs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            dop_CRAs = {'.diag'        : 'stateDiag_HMI',
                        '.port1_errcnt': 'stateNetErrCnt_HMI',
                        '.port2_errcnt': 'stateNetErrCnt_HMI',
                        '.port3_errcnt': 'stateNetErrCnt_HMI',
                       }
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.CRAs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            # Цикл по всем добавленным модулям DI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству дискретных модулей
                count_array += 1
                for key, value in dop_CRAs.items():
                    name_CPU = 'Root' + self.name_prefix + '.Diag.CRAs.' + name_eng + key

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = str(name_CPU)
                    object.append(node_p)

                    address = etree.Element('address')
                    address.text = self.prefix_driver + value
                    object.append(address)

                    array_pos = etree.Element('arrayposition')
                    if key == '.diag'        : array_pos.text = str(194 + count_array - 1)
                    if key == '.port1_errcnt': array_pos.text = str(177 + (4 * (count_array - 1)))
                    if key == '.port2_errcnt': array_pos.text = str(178 + (4 * (count_array - 1)))
                    if key == '.port3_errcnt': array_pos.text = str(179 + (4 * (count_array - 1)))
                    object.append(array_pos)
                    root.append(object)
                tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.CRAs: файл map OK')
            return (f'Diag.CRAs: файл map OK')
        except:
            logger.error(f'Diag.CRAs: файл map FAILED')
            return (f'Diag.CRAs: файл map FAILED')
    # Модули NOR
    @logger.catch
    def diag_nors(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('NORs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'mnor'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_NOR"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.NORs: файл omx OK')
        except:
            logger.error(f'Diag.NORs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.NORs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            # Цикл по всем добавленным модулям DI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству дискретных модулей
                count_array += 1

                name_CPU = f'Root{self.name_prefix}.Diag.NORs.{name_eng}.diag'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = str(name_CPU)
                object.append(node_p)

                address = etree.Element('address')
                address.text = f'{self.prefix_driver}stateDiag_HMI'
                object.append(address)

                array_pos = etree.Element('arrayposition')
                array_pos.text = str(194 + count_array - 1)
                object.append(array_pos)

                root.append(object)
                tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.NORs: файл map OK')
            return (f'Diag.NORs: файл map OK')
        except:
            logger.error(f'Diag.NORs: файл map FAILED')
            return (f'Diag.NORs: файл map FAILED')
    # Модули NOM
    @logger.catch
    def diag_noms(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('NOMs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet    = wb['HW']
        sheet_kd = wb['КД']
        # максимальное число рядов и столбцов
        rows = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    count = 0
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'mnom'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack = str(sheet.cell(row=i, column=5).value)
                        type_modul = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash = number_modul
                        modul_point = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Получаем названия портов из таблицы КД
                        description   = 'Резерв'
                        description_1 = 'Резерв'
                        for data in sheet_kd.rows:
                            if (name_uso_rus == str(data[1].value)) and \
                                    (rack == str(data[10].value)) and (modul_dash == str(data[11].value)):
                                count += 1
                                # Если счетчик = 1, тогда 0 порт, иначе 1
                                if count == 1:
                                    channel = str(data[12].value)
                                    description = str(data[3].value)
                                else:
                                    channel_1 = str(data[12].value)
                                    description_1 = str(data[3].value)

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_NOM"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        atrb6 = etree.Element("attribute")
                        atrb6.attrib['type'] = "unit.Library.Attributes.SignalName"
                        atrb6.attrib['value'] = description
                        object.append(atrb6)

                        atrb7 = etree.Element("attribute")
                        atrb7.attrib['type'] = "unit.Library.Attributes.SignalName_2"
                        atrb7.attrib['value'] = description_1
                        object.append(atrb7)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.NOMs: файл omx OK')
        except:
            logger.error(f'Diag.NOMs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            root, tree = self.parser_map()
            # Чистка тэгов
            self.cleaner_map('.Diag.NOMs.', root)
            # Счетчик позиции в массиве
            count_array = 0
            # Цикл по всем добавленным модулям DI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                # Цикл по количеству дискретных модулей
                count_array += 1

                name_CPU = f'Root{self.name_prefix}.Diag.NOMs.{name_eng}.diag'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                node_p = etree.Element('node-path')
                node_p.text = str(name_CPU)
                object.append(node_p)

                address = etree.Element('address')
                address.text = f'{self.prefix_driver}stateDiag_HMI'
                object.append(address)

                array_pos = etree.Element('arrayposition')
                array_pos.text = str(225 + count_array - 1)
                object.append(array_pos)

                root.append(object)
                tree.write(self.map, pretty_print=True)
            logger.info(f'Diag.NOMs: файл map OK')
            return (f'Diag.NOMs: файл map OK')
        except:
            logger.error(f'Diag.NOMs: файл map FAILED')
            return (f'Diag.NOMs: файл map FAILED')
    # Модули CPS
    @logger.catch
    def diag_cps(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('CPSs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if self.str_find(str(cell_ai).lower(), {'mcps'}):
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j - 1).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j - 1).value)
                        modul_dash   = '0'
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.SE.mod_PSU"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.CPSs: файл omx OK')
            return (f'Diag.CPSs: файл omx OK')
        except:
            logger.error(f'Diag.CPSs: файл omx FAILED')
            return (f'Diag.CPSs: файл omx FAILED')

    # Диагностика MK
    # Аналоговые сигналы входные
    @logger.catch
    def diag_mk_analogs_in(self, MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('AIs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        link_path = MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'МК-516-008A':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.MK_Logic.mod_AI"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.AIs: файл omx OK')
        except:
            logger.error(f'Diag.AIs: файл omx FAILED')
        # 2 этап
        # Заполняем значения атрибутов
        # AttributesMapAI_Ref.xml, AttributesMapKlk.xml,
        # AttributesMapKont.xml, AttributesMapSignalName.xml, AttributesMapTagName.xml
        try:
            # Цикл по всем xml
            for path in link_path:
                root, tree = self.parser_diag_map(path)
                # Чистка тэгов
                self.cleaner_diag_map('.Diag.AIs.', root)
                # Цикл по всем добавленным модулям AI
                for initial_data in signals:
                    name_rus = initial_data['name_uso_rus']
                    name_eng = initial_data['uso_rack_modul']
                    basket   = initial_data['rack']
                    mod      = initial_data['modul_dash']
                    for value in data_kd:
                        klk      = value['КлК']
                        kont     = value['Конт']
                        desc     = value['Наименование']
                        tag      = value['Tэг']
                        uso      = value['Шкаф']
                        basket_v = value['Корз']
                        modul    = value['Мод']
                        channel  = value['Кан']
                        if (name_rus == str(uso)) and (basket == str(basket_v)) and (mod == str(modul)):
                            str_tag = self.translate(str(tag))

                            if klk is None : klk = ' '
                            if kont is None: kont = ' '

                            name_AI = 'Root' + self.name_prefix + '.Diag.AIs.' + name_eng + '.ch_AI_0' + str(channel)
                            object = etree.Element('item')
                            object.attrib['id'] = name_AI
                            if path == MapAI_Ref       :
                                if not str_tag is None : object.attrib['value'] = str(str_tag)
                            if path == MapKlk          :
                                if not klk is None     : object.attrib['value'] = str(klk)
                            if path == MapKont         :
                                if not kont is None    : object.attrib['value'] = str(kont)
                            if path == MapSignalName   :
                                if not desc is None    : object.attrib['value'] = str(desc)
                            if path == MapTagName      :
                                if not tag is None     : object.attrib['value'] = str(tag)
                            root.append(object)
                logger.info(f'Diag.AIs: карта атрибутов: {path} - OK')
                tree.write(path, pretty_print=True)
        except:
            logger.info(f'Diag.AIs: карта атрибутов: {path} - FAILED')
        # 3 этап карта адресов
        try:
            list_param = ['mod_State', 'ch_01', 'ch_02', 'ch_03', 'ch_04', 'ch_05', 'ch_06', 'ch_07', 'ch_08' ]
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.AIs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagAI8': start_diagAI8 = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям AI
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.AIs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagAI8 + count_adress)
                    object.append(address)
                    root.append(object)
                    count_adress += 1

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Diag.AIs: файл map OK')
            return     (f'Diag.AIs: файл map OK')
        except:
            logger.error(f'Diag.AIs: файл map FAILED')
            return      (f'Diag.AIs: файл map FAILED')
    # Аналоговые сигналы выходные
    @logger.catch
    def diag_mk_analogs_out(self, MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('AOs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        link_path = MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'MK-514-008':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack      = str(sheet.cell(row=i, column=5).value)
                        type_modul = str(sheet.cell(row=i, column=j).value)
                        modul_dash = number_modul
                        modul_point = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.MK_Logic.mod_AO"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.AOs: файл omx OK')
        except:
            logger.error(f'Diag.AOs: файл omx FAILED')
        # 2 этап
        # Заполняем значения атрибутов
        # AttributesMapAI_Ref.xml, AttributesMapKlk.xml,
        # AttributesMapKont.xml, AttributesMapSignalName.xml, AttributesMapTagName.xml
        try:
            # Цикл по всем xml
            for path in link_path:
                root, tree = self.parser_diag_map(path)
                # Чистка тэгов
                self.cleaner_diag_map('.Diag.AOs.', root)
                # Цикл по всем добавленным модулям AI
                for initial_data in signals:
                    name_rus = initial_data['name_uso_rus']
                    name_eng = initial_data['uso_rack_modul']
                    basket = initial_data['rack']
                    mod = initial_data['modul_dash']
                    for value in data_kd:
                        klk = value['КлК']
                        kont = value['Конт']
                        desc = value['Наименование']
                        tag = value['Tэг']
                        uso = value['Шкаф']
                        basket_v = value['Корз']
                        modul = value['Мод']
                        channel = value['Кан']
                        if (name_rus == str(uso)) and (basket == str(basket_v)) and (mod == str(modul)):
                            str_tag = self.translate(str(tag))

                            if klk is None: klk = ' '
                            if kont is None: kont = ' '

                            name_AO = 'Root' + self.name_prefix + '.Diag.AOs.' + name_eng + '.ch_AO_0' + str(channel)
                            object = etree.Element('item')
                            object.attrib['id'] = name_AO
                            if path == MapAI_Ref:
                                if not str_tag is None: object.attrib['value'] = str(str_tag)
                            if path == MapKlk:
                                if not klk is None: object.attrib['value'] = str(klk)
                            if path == MapKont:
                                if not kont is None: object.attrib['value'] = str(kont)
                            if path == MapSignalName:
                                if not desc is None: object.attrib['value'] = str(desc)
                            if path == MapTagName:
                                if not tag is None: object.attrib['value'] = str(tag)
                            root.append(object)
                logger.info(f'Diag.AOs: карта атрибутов: {path} - OK')
                tree.write(path, pretty_print=True)
        except:
            logger.info(f'Diag.AOs: карта атрибутов: {path} - FAILED')
        # 3 этап карта адресов
        try:
            list_param = ['mod_State', 'chHealth', 'ch_01', 'ch_02', 'ch_03', 'ch_04', 'ch_05', 'ch_06', 'ch_07', 'ch_08']
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.AOs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagAO': start_diagAO = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям AO
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.AOs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagAO + count_adress)
                    object.append(address)
                    root.append(object)
                    count_adress += 1

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Diag.AOs: файл map OK')
            return (f'Diag.AOs: файл map OK')
        except:
            logger.error(f'Diag.AOs: файл map FAILED')
            return (f'Diag.AOs: файл map FAILED')
    # Дискретные сигналы входные
    @logger.catch
    def diag_mk_diskrets_in(self, MapKlk, MapKont, MapSignalName, MapTagName):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('DIs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        link_path = MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'MK-521-032':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack = str(sheet.cell(row=i, column=5).value)
                        type_modul = str(sheet.cell(row=i, column=j).value)
                        modul_dash = number_modul
                        modul_point = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.MK_Logic.mod_DI"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.DIs: файл omx OK')
        except:
            logger.error(f'Diag.DIs: файл omx FAILED')
        # 2 этап
        # Заполняем значения атрибутов
        try:
            # Цикл по всем xml
            for path in link_path:
                root, tree = self.parser_diag_map(path)
                # Чистка тэгов
                self.cleaner_diag_map('.Diag.DIs.', root)
                # Цикл по всем добавленным модулям DI
                for initial_data in signals:
                    name_rus = initial_data['name_uso_rus']
                    name_eng = initial_data['uso_rack_modul']
                    basket = initial_data['rack']
                    mod = initial_data['modul_dash']
                    for value in data_kd:
                        klk = value['КлК']
                        kont = value['Конт']
                        desc = value['Наименование']
                        tag = value['Tэг']
                        uso = value['Шкаф']
                        basket_v = value['Корз']
                        modul = value['Мод']
                        channel = value['Кан']
                        if (name_rus == str(uso)) and (basket == str(basket_v)) and (mod == str(modul)):
                            str_tag = self.translate(str(tag))

                            if klk is None: klk = ' '
                            if kont is None: kont = ' '

                            if channel <= 9:
                                name_DI = 'Root' + self.name_prefix + '.Diag.DIs.' + name_eng + '.ch_DI_0' + str(channel)
                            else:
                                name_DI = 'Root' + self.name_prefix + '.Diag.DIs.' + name_eng + '.ch_DI_' + str(channel)

                            object = etree.Element('item')
                            object.attrib['id'] = name_DI
                            if path == MapKlk:
                                if not klk is None:  object.attrib['value'] = str(klk)
                            if path == MapKont:
                                if not kont is None:  object.attrib['value'] = str(kont)
                            if path == MapSignalName:
                                if not desc is None:  object.attrib['value'] = str(desc)
                            if path == MapTagName:
                                if not tag is None:  object.attrib['value'] = str(tag)
                            root.append(object)
                logger.info(f'Diag.DIs: карта атрибутов: {path} - OK')
                tree.write(path, pretty_print=True)
        except:
            logger.info(f'Diag.DIs: карта атрибутов: {path} - FAILED')
        # 3 этап карта адресов
        try:
            list_param = ['mod_State', 'mDI']
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.DIs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagDI': start_diagDI = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям AO
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.DIs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagDI + count_adress)
                    object.append(address)
                    root.append(object)

                    if i == 'mDI' or i == 'mod_State':
                        count_adress += 2
                    else: count_adress += 1

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Diag.DIs: файл map OK')
            return     (f'Diag.DIs: файл map OK')
        except:
            logger.error(f'Diag.DIs: файл map FAILED')
            return      (f'Diag.DIs: файл map FAILED')
    # Дискретные сигналы выходные
    @logger.catch
    def diag_mk_diskrets_out(self, MapKlk, MapKont, MapSignalName, MapTagName):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('DOs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        link_path = MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'MK-531-032':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack = str(sheet.cell(row=i, column=5).value)
                        type_modul = str(sheet.cell(row=i, column=j).value)
                        modul_dash = number_modul
                        modul_point = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.MK_Logic.mod_DI"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.DOs: файл omx OK')
        except:
            logger.error(f'Diag.DOs: файл omx FAILED')
        # 2 этап
        # Заполняем значения атрибутов
        try:
            # Цикл по всем xml
            for path in link_path:
                root, tree = self.parser_diag_map(path)
                # Чистка тэгов
                self.cleaner_diag_map('.Diag.DOs.', root)
                # Цикл по всем добавленным модулям DI
                for initial_data in signals:
                    name_rus = initial_data['name_uso_rus']
                    name_eng = initial_data['uso_rack_modul']
                    basket = initial_data['rack']
                    mod = initial_data['modul_dash']
                    for value in data_kd:
                        klk = value['КлК']
                        kont = value['Конт']
                        desc = value['Наименование']
                        tag = value['Tэг']
                        uso = value['Шкаф']
                        basket_v = value['Корз']
                        modul = value['Мод']
                        channel = value['Кан']
                        if (name_rus == str(uso)) and (basket == str(basket_v)) and (mod == str(modul)):
                            str_tag = self.translate(str(tag))

                            if klk is None: klk = ' '
                            if kont is None: kont = ' '

                            if channel <= 9:
                                name_DI = 'Root' + self.name_prefix + '.Diag.DOs.' + name_eng + '.ch_DI_0' + str(channel)
                            else:
                                name_DI = 'Root' + self.name_prefix + '.Diag.DOs.' + name_eng + '.ch_DI_' + str(channel)

                            object = etree.Element('item')
                            object.attrib['id'] = name_DI
                            if path == MapKlk:
                                if not klk is None:  object.attrib['value'] = str(klk)
                            if path == MapKont:
                                if not kont is None:  object.attrib['value'] = str(kont)
                            if path == MapSignalName:
                                if not desc is None:  object.attrib['value'] = str(desc)
                            if path == MapTagName:
                                if not tag is None:  object.attrib['value'] = str(tag)
                            root.append(object)
                logger.info(f'Diag.DOs: карта атрибутов: {path} - OK')
                tree.write(path, pretty_print=True)
        except:
            logger.info(f'Diag.DOs: карта атрибутов: {path} - FAILED')
        # 3 этап карта адресов
        try:
            list_param = ['mod_State', 'mDI']
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.DOs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagDO': start_diagDO = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям AO
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.DOs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagDO + count_adress)
                    object.append(address)
                    root.append(object)

                    if i == 'mDI' or i == 'mod_State':
                        count_adress += 2
                    else: count_adress += 1

            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Diag.DOs: файл map OK')
            return     (f'Diag.DOs: файл map OK')
        except:
            logger.error(f'Diag.DOs: файл map FAILED')
            return      (f'Diag.DOs: файл map FAILED')
    # Модуль CPU
    @logger.catch
    def diag_mk_cpus(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('CPUs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'MK-504-120':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.MK_Logic.mod_CPU"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.CPUs: файл omx OK')
        except:
            logger.error(f'Diag.CPUs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            list_param = ['mod_State', 'mod_State_ext', 'mod_State_Err',
                          'CPUMemFree', 'CPULoad', 'ClcCurr', 'ClcMax', 'RsrCRC32']

            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.CPUs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagCPU': start_diagCPU = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям AO
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.CPUs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagCPU + count_adress)
                    object.append(address)
                    root.append(object)

                    if i == 'CPULoad' or i == 'RsrCRC32' or i == 'ClcCurr' or i =='ClcMax':
                        count_adress += 2
                    else: count_adress += 1

            tree.write(self.map_mb, pretty_print=True)

            logger.info(f'Diag.CPUs: файл map OK')
            return (f'Diag.CPUs: файл map OK')
        except:
            logger.error(f'Diag.CPUs: файл map FAILED')
            return (f'Diag.CPUs: файл map FAILED')
    # Модуль CN
    @logger.catch
    def diag_mk_cns(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('CNs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'MK-545-010':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        signals.append(dict(name_uso_rus=name_uso_rus,
                                            uso_rack_modul=uso_rack_modul,
                                            rack=rack,
                                            modul_dash=modul_dash))

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.MK_Logic.mod_CN"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.CNs: файл omx OK')
        except:
            logger.error(f'Diag.CNs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            list_param = ['mod_State', 'mod_State_ext', 'ports_State', 'pwl_ID']
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.CNs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagCN': start_diagCN = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям AO
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.CNs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagCN + count_adress)
                    object.append(address)
                    root.append(object)
                    count_adress += 1

            tree.write(self.map_mb, pretty_print=True)

            logger.info(f'Diag.CNs: файл map OK')
            return     (f'Diag.CNs: файл map OK')
        except:
            logger.error(f'Diag.CNs: файл map FAILED')
            return      (f'Diag.CNs: файл map FAILED')
    # Модуль MN
    @logger.catch
    def diag_mk_mns(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('MNs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'MK-546-010':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        signals.append(dict(name_uso_rus=name_uso_rus,
                                            uso_rack_modul=uso_rack_modul,
                                            rack=rack,
                                            modul_dash=modul_dash))

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.MK_Logic.mod_CN"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.MNs: файл omx OK')
        except:
            logger.error(f'Diag.MNs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            list_param = ['mod_State_ext', 'ports_State']
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.MNs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagMN': start_diagMN = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям AO
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.MNs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagMN + count_adress)
                    object.append(address)
                    root.append(object)
                    count_adress += 1

            tree.write(self.map_mb, pretty_print=True)

            logger.info(f'Diag.MNs: файл map OK')
            return (f'Diag.MNs: файл map OK')
        except:
            logger.error(f'Diag.MNs: файл map FAILED')
            return (f'Diag.MNs: файл map FAILED')
    # Модуль PSU
    @logger.catch
    def diag_mk_psus(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('PSUs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'MK-550-024':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      modul_dash=modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = "unit.Library.PLC_Types.modules.MK_Logic.mod_PSU"
                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.PSUs: файл omx OK')
        except:
            logger.error(f'Diag.PSUs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            list_param = ['mod_State', 'mod_State_ext', 'SupplyVoltage',
                          'CanBusSpeed', 'Can1ErrorCounter', 'Can2ErrorCounter']
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.PSUs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagPSU': start_diagMPSU = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.PSUs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagMPSU + count_adress)
                    object.append(address)
                    root.append(object)
                    if i == 'SupplyVoltage' or i == 'Can2ErrorCounter': count_adress += 2
                    else: count_adress += 1

            tree.write(self.map_mb, pretty_print=True)

            logger.info(f'Diag.PSUs: файл map OK')
            return     (f'Diag.PSUs: файл map OK')
        except:
            logger.error(f'Diag.PSUs: файл map FAILED')
            return      (f'Diag.PSUs: файл map FAILED')
    # Модуль интерфейсный
    @logger.catch
    def diag_mk_rs(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('RSs')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с RS
        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    name_signal_port_1 = ''
                    name_signal_port_2 = ''
                    cell_ai = sheet.cell(row=i, column=j).value
                    if cell_ai == 'MK-541-002':
                        # номер усо, номер модуля для имени, имя усо с корзиной англ,
                        # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                        number_uso   = str(sheet.cell(row=i, column=1).value)
                        number_modul = str(sheet.cell(row=2, column=j).value)
                        name_uso_eng = str(sheet.cell(row=i, column=3).value)
                        name_uso_rus = str(sheet.cell(row=i, column=4).value)
                        rack         = str(sheet.cell(row=i, column=5).value)
                        type_modul   = str(sheet.cell(row=i, column=j).value)
                        modul_dash   = number_modul
                        modul_point  = number_modul

                        if self.str_find(modul_dash, {'_0', '_'}):
                            modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                        if self.str_find(modul_point, {'_'}):
                            modul_point = str(modul_point).replace('_', '.')
                        uso_rack_modul = name_uso_eng + number_modul

                        # Название оборудования на канале
                        for data in data_kd:
                            kd_uso         = str(data['Шкаф'])
                            kd_rack        = str(data['Корз'])
                            kd_modul       = str(data['Мод'])
                            kd_channel     = str(data['Кан'])
                            hw_modul       = str(number_modul).split('_')
                            kd_name_signal = str(data['Наименование'])

                            if kd_uso == name_uso_rus:
                                if kd_rack == rack:
                                    if modul_dash == kd_modul:
                                        if kd_channel == '1': name_signal_port_1 = kd_name_signal
                                        if kd_channel == '2': name_signal_port_2 = kd_name_signal

                        if name_signal_port_1 == '': name_signal_port_1 = 'Резерв'
                        if name_signal_port_2 == '': name_signal_port_2 = 'Резерв'

                        # Заполняем словарь с исходными данными
                        a_dict = dict(name_uso_rus   = name_uso_rus,
                                      uso_rack_modul = uso_rack_modul,
                                      rack           = rack,
                                      modul_dash     = modul_dash)
                        signals.append(a_dict)

                        object = etree.Element("{automation.control}object")
                        object.attrib['name'] = uso_rack_modul
                        object.attrib['uuid'] = str(uuid.uuid1())
                        object.attrib['base-type'] = 'unit.Library.PLC_Types.modules.MK_Logic.mod_RS'
                        object.attrib['aspect'] = 'unit.Library.PLC_Types.PLC'

                        atrb1 = etree.Element("attribute")
                        atrb1.attrib['type'] = "unit.Library.Attributes.ModNumber"
                        atrb1.attrib['value'] = modul_dash
                        object.append(atrb1)

                        atrb2 = etree.Element("attribute")
                        atrb2.attrib['type'] = "unit.Library.Attributes.RackNumber"
                        atrb2.attrib['value'] = number_uso
                        object.append(atrb2)

                        atrb3 = etree.Element("attribute")
                        atrb3.attrib['type'] = "unit.Library.Attributes.ModPosition"
                        atrb3.attrib['value'] = 'A' + rack + modul_point
                        object.append(atrb3)

                        atrb4 = etree.Element("attribute")
                        atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                        atrb4.attrib['value'] = name_uso_rus
                        object.append(atrb4)

                        atrb5 = etree.Element("attribute")
                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                        atrb5.attrib['value'] = type_modul
                        object.append(atrb5)

                        atrb6 = etree.Element("attribute")
                        atrb6.attrib['type'] = "unit.Library.Attributes.SignalName"
                        atrb6.attrib['value'] = name_signal_port_1
                        object.append(atrb6)

                        atrb7 = etree.Element("attribute")
                        atrb7.attrib['type'] = "unit.Library.Attributes.SignalName_2"
                        atrb7.attrib['value'] = name_signal_port_2
                        object.append(atrb7)

                        el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.RSs: файл omx OK')
        except:
            logger.error(f'Diag.RSs: файл omx FAILED')
        # 2 этап карта адресов
        try:
            list_param = ['mod_State', 'mod_State_ext']
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.RSs.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'diagRS': start_diagRS = start_address

            # Счетчик позиции в массиве
            count_adress = 0
            # Цикл по всем добавленным модулям
            for initial_data in signals:
                name_eng = initial_data['uso_rack_modul']
                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.RSs.{name_eng}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    address.text = str(start_diagRS + count_adress)
                    object.append(address)
                    root.append(object)

                    count_adress += 1

            tree.write(self.map_mb, pretty_print=True)

            logger.info(f'Diag.RSs: файл map OK')
            return     (f'Diag.RSs: файл map OK')
        except:
            logger.error(f'Diag.RSs: файл map FAILED')
            return      (f'Diag.RSs: файл map FAILED')

    # RackStates
    @logger.catch
    def diag_rackstates(self):
        data_kd = self.data['КД']
        el1, tree = self.parser_diag_omx('RackStates')
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet = wb['HW']
        # максимальное число рядов и столбцов
        rows   = sheet.max_row
        column = sheet.max_column
        # Создадим пустой массив для дальнейшего использования
        signals = []
        a_dict  = []
        # 1 этап
        # Из табл: HW определим корзины и модули с CPU
        try:
            for i in range(4, rows + 1):
                    cell_rack = sheet.cell(row=i, column=1).value
                    if cell_rack is None: continue

                    # Заполняем словарь с исходными данными
                    a_dict = dict(cell_rack=cell_rack)
                    signals.append(a_dict)

                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = f'rack_{cell_rack}'
                    object.attrib['uuid'] = str(uuid.uuid1())
                    object.attrib['base-type'] = 'unit.Library.PLC_Types.modules.MK_Logic.rack_State'
                    object.attrib['aspect'] = 'unit.Library.PLC_Types.PLC'

                    el1.append(object)
            tree.write(self.omx, pretty_print=True)
            logger.info(f'Diag.RackState: файл omx OK')
        except:
            logger.error(f'Diag.RackState: файл omx FAILED')
        # 2 этап карта адресов
        try:
            list_param = ['mBUS', 'mBUSandCh', 'mBUSblink']
            data_mb    = self.data['ModBus']
            root, tree = self.parser_map_modbus()
            # Чистка тэгов
            self.cleaner_map('.Diag.RackStates.', root)
            for numeric in data_mb:
                variable      = numeric['Переменная Excel']
                start_address = numeric['Начальный адрес']
                end_adress    = numeric['Конечный адрес']

                if variable == 'mBUS'     : start_mBUS      = start_address
                if variable == 'mBUSandCh': start_mBUSandCh = start_address
                if variable == 'mBUSblink': start_mBUSblink = start_address

            # Цикл по всем добавленным модулям
            for initial_data in signals:
                number = initial_data['cell_rack']

                for i in list_param:
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    node_p = etree.Element('node-path')
                    node_p.text = f'Root{self.name_prefix}.Diag.RackStates.rack_{number}.{i}'
                    object.append(node_p)

                    segment = etree.Element('table')
                    segment.text = f'Holding Registers'
                    object.append(segment)

                    address = etree.Element('address')
                    if i == 'mBUS'     : address.text = f'{start_mBUS      + 2 * (number - 1)}'
                    if i == 'mBUSandCh': address.text = f'{start_mBUSandCh + 2 * (number - 1)}'
                    if i == 'mBUSblink': address.text = f'{start_mBUSblink + 2 * (number - 1)}'
                    object.append(address)
                    root.append(object)
            tree.write(self.map_mb, pretty_print=True)
            logger.info(f'Diag.RackState: файл map OK')
            return     (f'Diag.RackState: файл map OK')
        except:
            logger.error(f'Diag.RackState: файл map FAILED')
            return      (f'Diag.RackState: файл map FAILED')

    # Цветовая схема дискретных сигналов
    @logger.catch
    def color_diskrets(self, MapColorScheme):
        dop_color = {"01":"1",
                     "02":"2",
                     "03":"3",
                     "10":"11",
                     "20":"12",
                     "30":"9",
                     "12":"4",
                     "13":"5",
                     "21":"10",
                     "31":"7",
                     "23":"6",
                     "32":"8",
                     "00":"0",
                     "11":"1"}
        data      = self.data['DI']
        root, tree = self.parser_diag_map(MapColorScheme)
        # Чистка тэгов
        self.cleaner_diag_map('.Diskrets.', root)
        try:
            for value in data:
                name       = value['Название']
                tag        = value['Идентификатор']
                priority_0 = value['priority[0]']
                priority_1 = value['priority[1]']

                if priority_0 is None: priority_0 = '0'
                else                 : priority_0 = str(priority_0)

                if priority_1 is None: priority_1 = '0'
                else                 : priority_1 = str(priority_1)

                color_shema=str(dop_color[priority_0 + priority_1])

                if name is None: continue
                if tag  is None: continue
                tag = self.translate(str(tag))

                object = etree.Element('item')
                object.attrib['id'] = 'Root.Diskrets.' + tag + '.s_Config'
                object.attrib['id'] = f'Root{self.name_prefix}.Diskrets.{tag}.s_Config'
                object.attrib['value'] = color_shema
                root.append(object)
                tree.write(MapColorScheme, pretty_print=True)
            logger.info(f'AttributesMapColorScheme.xml: карта атрибутов OK')
            return (f'AttributesMapColorScheme.xml: карта атрибутов OK')
        except:
            logger.error(f'AttributesMapColorScheme.xml: карта атрибутов FAILED')
            return (f'AttributesMapColorScheme.xml: карта атрибутов FAILED')
    # AttributesMapUstName
    @logger.catch
    def analogformat_map(self, MapFormats):
        wb = openpyxl.load_workbook(self.exel, read_only=True, data_only=True)
        data_ai    = self.data['AI']
        data_grp   = wb['AIGRP']

        root, tree = self.parser_diag_map(MapFormats)
        # Чистка тэгов
        self.cleaner_diag_map('.Analogs.', root)
        count_row = 0
        try:
            for value_ai in data_ai:
                tag          = value_ai['Идентификатор']
                group_analog = value_ai['Группа аналогов']
                formatt      = value_ai['Отображаемая точность значения']
                if tag is None: continue
                tag = self.translate(tag)
                for value_grp in data_grp.rows:
                    count_row += 1
                    if count_row > 3:
                        name_group = value_grp[1].value
                        min1       = value_grp[7].value
                        min2       = value_grp[6].value
                        min3       = value_grp[5].value
                        min4       = value_grp[4].value
                        min5       = value_grp[3].value
                        min6       = value_grp[2].value
                        max1       = value_grp[8].value
                        max2       = value_grp[9].value
                        max3       = value_grp[10].value
                        max4       = value_grp[11].value
                        max5       = value_grp[12].value
                        max6       = value_grp[13].value
                        if group_analog == name_group:
                            grp_ai = {'Format': formatt,
                                      'UstName.UstMin1': min1,
                                      'UstName.UstMin2': min2,
                                      'UstName.UstMin3': min3,
                                      'UstName.UstMin4': min4,
                                      'UstName.UstMin5': min5,
                                      'UstName.UstMin6': min6,
                                      'UstName.UstMax1': max1,
                                      'UstName.UstMax2': max2,
                                      'UstName.UstMax3': max3,
                                      'UstName.UstMax4': max4,
                                      'UstName.UstMax5': max5,
                                      'UstName.UstMax6': max6
                                      }
                            for grp, value in grp_ai.items():
                                object = etree.Element('item')
                                object.attrib['id']    = 'Root.' + self.name_prefix + 'Analogs.' + tag + '.' + grp
                                object.attrib['value'] = str(value)
                                root.append(object)
            tree.write(MapFormats, pretty_print=True, encoding='utf-8')
            logger.info(f'AttributesAnalogFormat.xml: карта атрибутов OK')
            return (f'AttributesAnalogFormat.xml: карта атрибутов OK')
        except:
            logger.error(f'AttributesAnalogFormat.xml: карта атрибутов FAILED')
            return (f'AttributesAnalogFormat.xml: карта атрибутов FAILED')
    # AttributesMapEGU
    @logger.catch
    def egu_map(self, MapEgu):
        data       = self.data['AI']
        root, tree = self.parser_diag_map(MapEgu)
        # Чистка тэгов
        self.cleaner_diag_map('.Analogs.', root)
        try:
            for value in data:
                egu = value['Единица измерения']
                tag = value['Идентификатор']

                if egu is None: continue
                if tag is None: continue

                tag = self.translate(str(tag))

                object = etree.Element('item')
                object.attrib['id'] = 'Root' + self.name_prefix +  '.Analogs.' + str(tag) + '.AIValue'
                object.attrib['value'] = str(egu)
                root.append(object)
            tree.write(MapEgu, pretty_print=True, encoding='utf-8')
            logger.info(f'AttributesMapEGU.xml: карта атрибутов OK')
            return (f'AttributesMapEGU.xml: карта атрибутов OK')
        except:
            logger.error(f'AttributesMapEGU.xml: карта атрибутов FAILED')
            return (f'AttributesMapEGU.xml: карта атрибутов FAILED')
    # AttributesAnalogTrends
    @logger.catch
    def analogs_trend(self, file_AnalogTrends):
        deadband_format = { '0' : '0.1',
                            '1' : '0.01',
                            '2' : '0.001',
                            '3' : '0.0001'
        }
        mintime_groups = {'давление нефти' : '100',
                          'вибрация'       : '200',
                          'стоп'           : '100'
        }

        data = self.data['AI']
        root, tree = self.parser_diag_map(file_AnalogTrends)
        # Чистка тэгов
        self.cleaner_diag_map('.Analogs.', root)
        # Цикл по всем сигналам
        try:
            for value_ai in data:
                format = value_ai['Отображаемая точность значения']
                name   = value_ai['Название']
                tag    = value_ai['Идентификатор']

                if tag    is None: continue
                if format is None: continue
                tag = self.translate(str(tag))

                # Deadband зависит от отображаемой точности значения
                for key, value in deadband_format.items():
                    if str(format) == key:
                        deadband = value

                # Вычисляем mintime по группе сигнала
                mintime = '300'
                for key, value in mintime_groups.items():
                    if self.str_find(str(name).lower(), {key}):
                        mintime = value

                name_AI = 'Root' + self.name_prefix + '.Analogs.' + tag + '.AIValue'
                object = etree.Element('item')
                object.attrib['id'] = name_AI
                object.attrib['value'] = str(f'Enable="True" Deadband="{deadband}" MinTime="{mintime}" ServerTime="False"')
                root.append(object)
            tree.write(file_AnalogTrends, pretty_print=True)
            logger.info(f'AttributesAnalogTrends.xml: карта атрибутов OK')
            return (f'AttributesAnalogTrends.xml: карта атрибутов OK')
        except:
            logger.error(f'AttributesAnalogTrends.xml: карта атрибутов FAILED')
            return (f'AttributesAnalogTrends.xml: карта атрибутов FAILED')
    #AttributesMapDescription
    @logger.catch
    def pzs_ready_map(self, MapDescription):
        data = self.data['PZ']
        root, tree = self.parser_diag_map(MapDescription)
        # Чистка тэгов
        self.cleaner_diag_map('.PZs.', root)
        for value in data:
            number  = value['№']
            read_1  = value['Г_1']
            read_2  = value['Г_2']
            read_3  = value['Г_3']
            read_4  = value['Г_4']
            read_5  = value['Г_5']
            read_6  = value['Г_6']
            read_7  = value['Г_7']
            read_8  = value['Г_8']
            read_9  = value['Г_9']
            read_10 = value['Г_10']
            read_11 = value['Г_11']
            read_12 = value['Г_12']
            read_13 = value['Г_13']
            read_14 = value['Г_14']
            read_15 = value['Г_15']

            if number  is None: continue
            if read_1  is None: read_1  = ' '
            if read_2  is None: read_2  = ' '
            if read_3  is None: read_3  = ' '
            if read_4  is None: read_4  = ' '
            if read_5  is None: read_5  = ' '
            if read_6  is None: read_6  = ' '
            if read_7  is None: read_7  = ' '
            if read_8  is None: read_8  = ' '
            if read_9  is None: read_9  = ' '
            if read_10 is None: read_10 = ' '
            if read_11 is None: read_11 = ' '
            if read_12 is None: read_12 = ' '
            if read_13 is None: read_13 = ' '
            if read_14 is None: read_14 = ' '
            if read_15 is None: read_15 = ' '

            readiness = [read_1, read_2, read_3, read_4, read_5, read_6, read_7, read_8,
                         read_9, read_10, read_11, read_12, read_13, read_14, read_15]

            count_read = 0
            for value in readiness:
                count_read += 1
                object = etree.Element('item')
                object.attrib['id'] = f'Root{self.name_prefix}.PZs.PZ_{number}.s_ReadyFlags.Ready{count_read}'
                object.attrib['value'] = f'{value}'
                root.append(object)
        tree.write(MapDescription, pretty_print=True, encoding='utf-8')
        logger.info(f'AttributesMapDescription.xml: карта атрибутов OK')
        return     (f'AttributesMapDescription.xml: карта атрибутов OK')

    # SQL cкрипт для поиска сигналов
    @logger.catch
    def sql_script_search(self, path):
        data_kd  = self.data['КД']
        data_uso = self.data['USO']
        # Скрипт создания таблицы
        script_sql_textfile: str = ('\tCREATE SCHEMA IF NOT EXISTS signals;\n'
                                   '\tCREATE TABLE IF NOT EXISTS signals.allSignals(\n'
                                   '\t\ttag VARCHAR(32),           \n'
                                   '\t\tdescription VARCHAR(1024), \n'
                                   '\t\tklk VARCHAR(32),           \n'
                                   '\t\tkont VARCHAR(32),          \n'
                                   '\t\tinitPath VARCHAR(1024),    \n'
                                   '\t\tposition VARCHAR(32),      \n'
                                   '\t\tcabinet VARCHAR(32),       \n'
                                   '\t\track VARCHAR(32),          \n'
                                   '\t\tmodule VARCHAR(32)         \n'
                                   '\t);\n'
                                   'DELETE FROM signals.allSignals ;\n')

        path_sql = f'{path}\SQLSearch.xml'
        # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
        if not os.path.exists(path_sql):
            file = codecs.open(path_sql, 'w', 'utf-8')
            file.write(script_sql_textfile)
        else:
            os.remove(path_sql)
            file = codecs.open(path_sql, 'w', 'utf-8')
            file.write(script_sql_textfile)

        try:
            for data in data_kd:
                type_str = data['Тип сигнала']
                cabinet  = data['Шкаф']
                tag      = data['Tэг']
                desc     = data['Наименование']
                klk      = data['КлК']
                kont     = data['Конт']
                rack     = data['Корз']
                mod      = data['Мод']
                position = data['Кан']

                if tag is None     : continue
                if type_str is None: continue
                if klk is None     : klk = ''
                if kont is None    : kont = ''
                if self.str_find(str(type_str).lower(), {'rs'}): type_str = 'NOM'

                #tag = self.translate(str(tag))
                for name in data_uso:
                    name_eng = name['Идентификатор']
                    name_rus = name['Название']

                    if name_eng is None: continue
                    if name_rus is None: continue

                    if str(name_rus) == str(cabinet):
                        mod_1 = mod
                        mod_1 = ('0' + str(mod_1)) if (int(mod_1) < 10) else mod_1
                        initPath = 'Diag.' + str(type_str) + 's.' + str(name_eng) + '_A' + str(rack) + '_' + str(mod_1)
                        break
                sql_request = f"INSERT INTO signals.allSignals VALUES('" \
                                f"{str(tag)}','{str(desc)}','{str(klk)}','{str(kont)}','{str(initPath)}','" \
                                f"{str(position)}','{str(cabinet)}','{str(rack)}','{str(mod)}');\n"
                file.write(sql_request)
            file.close()
            logger.info(f'SQL Скрипт готов')
            return (f'SQL Скрипт готов')
        except:
            logger.error(f'SQL Скрипт ошибка')
            return (f'SQL Скрипт ошибка')

    # Тренды
    # Вариант 1 - имеются подгруппы у родителей
    def trends_xml(self, item_alfa, path_file_txt, name_project):
        wb = openpyxl.load_workbook(item_alfa, read_only=True, data_only=True)
        sheet         = wb['items']
        data_ai       = self.data['AI']
        data_trendgrp = self.data['TrendGRP']

        data_parenet = []
        data_child   = []
        a_dict       = []
        snmp_row     = ''
        snmp_parent  = ''
        path_trend   = f'{path_file_txt}\AlphaTrends.xml'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('<?xml version="1.0" encoding="UTF-8"?>\n<Source Type="AlphaPostgres">\n</Source>')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('<?xml version="1.0" encoding="UTF-8"?>\n<Source Type="AlphaPostgres">\n</Source>')
        text_file.close()
        # Определяемся с количеством групп, проверяем на наличие в сигналах AI
        for parent in data_trendgrp:
            array_parent      = []
            array_name_parent = []

            id        = parent['ID']
            parent_id = parent['ParentID']
            name_grp  = parent['Название группы']
            use       = parent['Используется']
            if str(use) == '1':
                if str(parent_id) == '0':
                    array_parent.append(id)
                    array_name_parent.append(name_grp)
                    data_parenet.append({k: v for k, v in zip(array_parent, array_name_parent)})
            for signal_ai in data_ai:
                group_ai = signal_ai['Группа сброса трендов']
                array_child      = []
                array_name_child = []
                if group_ai == id and str(parent_id) != '0':
                    a_dict = dict(array_child      = str(group_ai),
                                  array_name_child = str(name_grp),
                                  parent_id        = str(parent_id))
                    data_child.append(a_dict)
                    break
        # Заполняем файл
        parser = etree.XMLParser(remove_blank_text=True, encoding="utf-8")
        tree = etree.parse(path_trend, parser)
        root = tree.getroot()

        for lvl_one in root.iter('Source'):
            object = etree.Element('Group')
            object.attrib['Name'] = name_project
            for array in data_parenet:
                for key, values in array.items():
                    group = etree.Element('Group')
                    group.attrib['Name'] = str(values)
                    for value in data_ai:
                        name_group = value['Группа сброса трендов']
                        tag        = value['Идентификатор']
                        name       = value['Название']
                        egu        = value['Единица измерения']
                        min        = value['Инж. Мин.']
                        max        = value['Инж. Макс.']
                        if tag is None: continue
                        tag = self.translate(tag)
                        if key == name_group:
                            for data in sheet.rows:
                                for cell in data:
                                    try:
                                        tag_name = str(cell.value).split(',')
                                    except:
                                        logger.info(f'К строке в items невозможно применить split, переход на следующую')
                                    if str(tag_name[2]) == tag:
                                        number_tag = int(tag_name[0]) + 1
                                        group_tag = etree.Element('Tag')
                                        group_tag.attrib['Name'] = str(number_tag)
                                        group_tag.attrib['Alias'] = str(value['Идентификатор'])
                                        group_tag.attrib['EGU'] = str(egu)
                                        group_tag.attrib['Description'] = str(name)
                                        group_tag.attrib['Format'] = '%g'
                                        group_tag.attrib['Min'] = str(min)
                                        group_tag.attrib['Max'] = str(max)
                                        group.append(group_tag)
                                        break
                    # Если есть подгруппы, заполним и их
                    for children in data_child:
                        key_child   = children['array_child']
                        value_child = children['array_name_child']
                        id_child    = children['parent_id']

                        child_name = str(value_child).split(' ')
                        if id_child == str(key):
                            child = etree.Element('Group')
                            try   : child.attrib['Name'] = str(child_name[2])
                            except: child.attrib['Name'] = str(child_name[0])

                            for value in data_ai:
                                name_group = value['Группа сброса трендов']
                                tag        = value['Идентификатор']
                                name       = value['Название']
                                egu        = value['Единица измерения']
                                min        = value['Инж. Мин.']
                                max        = value['Инж. Макс.']
                                if tag is None: continue
                                tag = self.translate(tag)
                                if key_child == str(name_group):
                                    for data in sheet.rows:
                                        for cell in data:
                                            tag_name = str(cell.value).split(',')
                                            if tag_name[2] == tag:
                                                number_tag = int(tag_name[0]) + 1
                                                group_tag = etree.Element('Tag')
                                                group_tag.attrib['Name'] = str(number_tag)
                                                group_tag.attrib['Alias'] = str(value['Идентификатор'])
                                                group_tag.attrib['EGU'] = str(egu)
                                                group_tag.attrib['Description'] = str(name)
                                                group_tag.attrib['Format'] = '%g'
                                                group_tag.attrib['Min'] = str(min)
                                                group_tag.attrib['Max'] = str(max)
                                                child.append(group_tag)
                                                break
                            group.append(child)
                    object.append(group)
        root.append(object)
        # Определяемся с количеством групп
        row = sheet.max_row
        for lvl_one in root.iter('Source'):
            object = etree.Element('Group')
            object.attrib['Name'] = 'SNMP'
            for i in range(2, row + 1):
                snmp_row = str(sheet.cell(row=i, column=1).value).split(',')
                if snmp_row[2].lower() == 'snmp':
                    logger.info(f'Строка SNMP найдена')
                    snmp_parent = snmp_row[0]
                    continue
                if snmp_parent == snmp_row[1]:
                    name_equipment = snmp_row[3]
                    snmp_subject = snmp_row[0]
                    group = etree.Element('Group')
                    group.attrib['Name'] = str(name_equipment)
                    for k in range(i, row + 1):
                        group_snmp = str(sheet.cell(row=k, column=1).value).split(',')
                        if group_snmp[1] == snmp_subject:
                            id_parent = group_snmp[0]
                            pod_group = etree.Element('Group')
                            pod_group.attrib['Name'] = str(group_snmp[2])
                            for s in range(i, row + 1):
                                snmp_id = str(sheet.cell(row=s, column=1).value).split(',')
                                if snmp_id[1] == id_parent:
                                    if self.str_find(group_snmp[2], 'port') or group_snmp[2].lower() == 'system':
                                        group_tag = etree.Element('Tag')
                                        group_tag.attrib['Name'] = str(snmp_id[0])
                                        group_tag.attrib['Alias'] = str(snmp_id[2])
                                        group_tag.attrib['EGU'] = str(snmp_id[4])
                                        group_tag.attrib['Description'] = str(snmp_id[3])
                                        group_tag.attrib['Format'] = str('%g')
                                        group_tag.attrib['Min'] = str(0)
                                        group_tag.attrib['Max'] = str(100)
                                        pod_group.append(group_tag)

                                    elif snmp_id[2].lower() == 'temp':
                                        pod_group = etree.Element('Group')
                                        pod_group.attrib['Name'] = str(snmp_id[2])
                                        for x in range(i, row + 1):
                                            item_temp = str(sheet.cell(row=x, column=1).value).split(',')
                                            if snmp_id[0] == item_temp[1]:
                                                pod_pod_child_group = etree.Element('Group')
                                                pod_pod_child_group.attrib['Name'] = str(item_temp[2])
                                                for v in range(i, row + 1):
                                                    item = str(sheet.cell(row=v, column=1).value).split(',')
                                                    if item_temp[0] == item[1]:
                                                        group_tag = etree.Element('Tag')
                                                        group_tag.attrib['Name'] = str(item[0])
                                                        group_tag.attrib['Alias'] = str(item[2])
                                                        group_tag.attrib['EGU'] = str(item[4])
                                                        group_tag.attrib['Description'] = str(item[3])
                                                        group_tag.attrib['Format'] = str('%g')
                                                        group_tag.attrib['Min'] = str(0)
                                                        group_tag.attrib['Max'] = str(100)
                                                        pod_pod_child_group.append(group_tag)
                                                        break
                                                pod_group.append(pod_pod_child_group)
                                    else:
                                        for w in range(i, row + 1):
                                            item = str(sheet.cell(row=w, column=1).value).split(',')
                                            if snmp_id[0] == item[1]:
                                                group_tag = etree.Element('Tag')
                                                group_tag.attrib['Name'] = str(item[0])
                                                group_tag.attrib['Alias'] = str(item[2])
                                                group_tag.attrib['EGU'] = str(item[4])
                                                group_tag.attrib['Description'] = str(item[3])
                                                group_tag.attrib['Format'] = str('%g')
                                                group_tag.attrib['Min'] = str(0)
                                                group_tag.attrib['Max'] = str(100)
                                                pod_group.append(group_tag)
                                                break
                            group.append(pod_group)
                    object.append(group)
        root.append(object)
        tree.write(path_trend, pretty_print=True, encoding="utf-8")
        logger.info(f'Дерево трендов успешно создано')
        return (f'Выполнено. Генерация файла трендов выполнена')
    # Под Linux
    def trends_linux_xml(self, path_file_txt, name_project):
        data_ai       = self.data['AI']
        data_trendgrp = self.data['TrendGRP']

        path_trend   = f'{path_file_txt}\AlphaTrends.xml'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('<?xml version="1.0" encoding="UTF-8"?>\n<Source Type="NaftaPostgres">\n</Source>')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('<?xml version="1.0" encoding="UTF-8"?>\n<Source Type="NaftaPostgres">\n</Source>')
        text_file.close()
        # Заполняем файл
        parser = etree.XMLParser(remove_blank_text=True, encoding="utf-8")
        tree   = etree.parse(path_trend, parser)
        root   = tree.getroot()
        # Определяемся с количеством групп, проверяем на наличие в сигналах AI
        for lvl_one in root.iter('Source'):
            object = etree.Element('Group')
            object.attrib['Name'] = name_project

            for parent in data_trendgrp:
                parent_id       = parent['ID']
                name_group_id   = parent['Название группы']
                parent_parentID = parent['ParentID']
                count_parent    = 0
                count_child     = 0

                if str(parent_parentID) != '0': continue
                group = etree.Element('Group')
                group.attrib['Name'] = str(name_group_id)

                for ai_data_signal in data_ai:
                    ai_group       = ai_data_signal['Группа сброса трендов']
                    ai_name_signal = ai_data_signal['Название']
                    ai_tag         = ai_data_signal['Идентификатор']
                    ai_egu         = ai_data_signal['Единица измерения']
                    ai_min         = ai_data_signal['Инж. Мин.']
                    ai_max         = ai_data_signal['Инж. Макс.']

                    if parent_id == ai_group:
                        group_tag = etree.Element('Tag')
                        group_tag.attrib['Name']        = str(ai_tag)
                        group_tag.attrib['Format']      = '%g'
                        group_tag.attrib['Description'] = str(ai_name_signal)
                        group_tag.attrib['EGU']         = str(ai_egu)
                        group_tag.attrib['Alias']       = str(ai_tag)
                        group_tag.attrib['Min']         = str(ai_min)
                        group_tag.attrib['Max']         = str(ai_max)
                        group.append(group_tag)
                        count_parent += 1

                for child in data_trendgrp:
                    child_id       = child['ID']
                    child_group_id = child['Название группы']
                    child_parentID = child['ParentID']

                    if parent_id == child_parentID:
                        child_group_tag = etree.Element('Group')
                        child_group_tag.attrib['Name'] = str(child_group_id)
                        for ai_data_signal in data_ai:
                            ai_group       = ai_data_signal['Группа сброса трендов']
                            ai_name_signal = ai_data_signal['Название']
                            ai_tag         = ai_data_signal['Идентификатор']
                            ai_egu         = ai_data_signal['Единица измерения']
                            ai_min         = ai_data_signal['Инж. Мин.']
                            ai_max         = ai_data_signal['Инж. Макс.']

                            if child_id == ai_group:
                                child_signal_tag = etree.Element('Tag')
                                child_signal_tag.attrib['Name']        = str(ai_tag)
                                child_signal_tag.attrib['Format']      = '%g'
                                child_signal_tag.attrib['Description'] = str(ai_name_signal)
                                child_signal_tag.attrib['EGU']         = str(ai_egu)
                                child_signal_tag.attrib['Alias']       = str(ai_tag)
                                child_signal_tag.attrib['Min']         = str(ai_min)
                                child_signal_tag.attrib['Max']         = str(ai_max)
                                child_group_tag.append(child_signal_tag)
                                count_child += 1
                        group.append(child_group_tag)
                # Пропускаем итерацию, если пустые группы
                if count_parent == 0 and count_child == 0: continue
                object.append(group)
        root.append(object)
        tree.write(path_trend, pretty_print=True, encoding="utf-8")

        logger.info(f'Выполнено. Генерация файла трендов выполнена')
        return (f'Выполнено. Генерация файла трендов выполнена')

    # Сообщения MSG
    # PostgreSQL_Messages-Racks.sql
    def msg_racks(self, path_sample, path_request):
        list_modul = ['MK-516-008A', 'MK-514-008', 'MK-521-032', 'MK-541-002', 'MK-504-120',
                      'MK-531-032', 'MK-545-010', 'MK-550-024', 'MK-546-010']
        start_adress = None
        data_msg     = self.data['MSG']
        # Скрипт создания таблицы
        script_sql_textfile =  ('\tCREATE SCHEMA IF NOT EXISTS messages;\n'
                                 '\tCREATE TABLE IF NOT EXISTS messages.OPMessages(\n'
                                 '\t\tCategory INT NOT NULL,\n'
                                 '\t\tMessage VARCHAR(1024),\n'
                                 '\t\tIsAck BOOLEAN NOT NULL,\n'
                                 '\t\tSoundFile VARCHAR(1024),\n'
                                 '\t\tIsCycle BOOLEAN NOT NULL,\n'
                                 '\t\tIsSound BOOLEAN NOT NULL,\n'
                                 '\t\tIsHide BOOLEAN NOT NULL,\n'
                                 '\t\tPriority INT NOT NULL,\n'
                                 '\t\tIsAlert BOOLEAN NOT NULL,\n'
                                 '\t\tCONSTRAINT OPMessages_pkey PRIMARY KEY (Category)\n'
                                 '\t);\n'
                                 'BEGIN TRANSACTION;\n')
        # Проверяем шаблон
        path_sample = f'{path_sample}\TblD_Racks.xml'
        if not os.path.isfile(path_sample):
            logger.error(f'Ошибка. Генерация сообщений для диагностики корзин: Шаблон отсутствует!')
            return (f'Ошибка. Генерация сообщений для диагностики корзин: Шаблон отсутствует!')
        root, tree = self.parser_diag_map(path_sample)
        # Создаём файл запроса
        path_request = f'{path_request}\\PostgreSQL_Messages-Racks.sql'
        if not os.path.exists(path_request):
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(script_sql_textfile)
        else:
            os.remove(path_request)
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(script_sql_textfile)
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        sheet     = wb['HW']
        rows      = sheet.max_row
        column    = sheet.max_column
        save_list = []
        # Определим стартовый адрес
        for msg in data_msg:
            tag      = msg['Название таблицы БД ВУ']
            code_msg = msg['Индекс']

            if tag == 'TblD_Racks': start_adress = code_msg
        # Если стартовый адрес пуст
        if start_adress is None:
            logger.error(f'Ошибка. Генерация сообщений для диагностики корзин: отсутствует стартовый адрес - TblD_Racks!')
            return (f'Ошибка. Генерация сообщений для диагностики корзин: отсутствует стартовый адрес - TblD_Racks!')
        # Из табл HW определим корзины и модули
        for i in range(4, rows + 1):
            name_uso    = str(sheet.cell(row=i, column=4).value)
            number_rack = str(sheet.cell(row=i, column=5).value)
            number      = str(sheet.cell(row=i, column=1).value)
            for j in range(7, column + 1):
                cell = self.translate(str(sheet.cell(row=i, column=j).value))
                if cell is None:continue
                if j % 2 != 0:
                    if cell in list_modul:
                        number_modul = sheet.cell(row=2, column=j).value.partition('_')[2]
                        message = f'Диагностика. {name_uso}. Модуль A{number_rack}.{number_modul} {cell}'
                        save_list.append(dict(message     = message,
                                              number_rack = number))
        # Парсим шаблон и заполняем файл
        rack_default = 0
        count_row    = 0
        # Подсчет строк в шаблоне
        for lvl_one in root.iter('Row'): count_row += 1
        for sign in save_list:
            message = sign['message']
            number  = sign['number_rack']
            if rack_default != int(number):
                # Смещение по корзинам
                adress_offset = (32 * count_row) * (int(number) - 1)
                count_modul   = 0
                rack_default  =  int(number)
            for lvl_one in root.iter('Row'):
                category  = lvl_one.attrib['Category']
                isAck     = lvl_one.attrib['IsAck']
                isCycle   = lvl_one.attrib['IsCycle']
                isSound   = lvl_one.attrib['IsSound']
                isHide    = lvl_one.attrib['IsHide']
                priority  = lvl_one.attrib['Priority']
                isAlert   = lvl_one.attrib['IsAlert']
                mess      = lvl_one.attrib['Message']
                soundFile = lvl_one.attrib['SoundFile']
                nextLink  = lvl_one.attrib['NextLink']
                base      = lvl_one.attrib['Base']
                # Смещение по модулю. На 1 модуль 10 сообщений
                count_modul += 1
                # Запрос
                sql_request = f"DELETE FROM messages.OPMessages WHERE Category ={adress_offset + (start_adress + count_modul)};\n" \
                              f"INSERT INTO messages.OPMessages (Category, Message, IsAck, SoundFile, IsCycle, IsSound, IsHide, Priority, IsAlert)" \
                              f"VALUES({adress_offset + (start_adress + count_modul)}, '{message}. {mess}', " \
                              f"{isAck}, '{soundFile}', {isCycle}, {isSound}, {isHide}, {priority}, {isAlert});\n"
                file.write(sql_request)
        file.write(f'COMMIT;')
        file.close()
        logger.info(f'Выполнено. Генерация сообщений для диагностики корзин')
        return (f'Выполнено. Генерация сообщений для диагностики корзин')
    # PostgreSQL_Messages-Modules.sql
    def msg_modules(self, path_sample, path_request):
        list_modul = ['MK-516-008A', 'MK-514-008', 'MK-521-032', 'MK-541-002', 'MK-504-120',
                      'MK-531-032', 'MK-545-010', 'MK-550-024', 'MK-546-010']
        start_modulesPSU = None
        start_modulesCPU = None
        start_modulesMN  = None
        start_modulesCN  = None

        data_msg   = self.data['MSG']
        # Скрипт создания таблицы
        script_sql_textfile =  ('\tCREATE SCHEMA IF NOT EXISTS messages;\n'
                                 '\tCREATE TABLE IF NOT EXISTS messages.OPMessages(\n'
                                 '\t\tCategory INT NOT NULL,\n'
                                 '\t\tMessage VARCHAR(1024),\n'
                                 '\t\tIsAck BOOLEAN NOT NULL,\n'
                                 '\t\tSoundFile VARCHAR(1024),\n'
                                 '\t\tIsCycle BOOLEAN NOT NULL,\n'
                                 '\t\tIsSound BOOLEAN NOT NULL,\n'
                                 '\t\tIsHide BOOLEAN NOT NULL,\n'
                                 '\t\tPriority INT NOT NULL,\n'
                                 '\t\tIsAlert BOOLEAN NOT NULL,\n'
                                 '\t\tCONSTRAINT OPMessages_pkey PRIMARY KEY (Category)\n'
                                 '\t);\n'
                                 'BEGIN TRANSACTION;\n')
        # Проверяем шаблоны
        path_TblD_ModulesCPU = f'{path_sample}TblD_ModulesCPU.xml'
        path_TblD_ModulesMN  = f'{path_sample}TblD_ModulesMN.xml'
        path_TblD_ModulesCN  = f'{path_sample}TblD_ModulesCN.xml'
        path_TblD_ModulesPSU = f'{path_sample}TblD_ModulesPSU.xml'
        list_path   = [path_TblD_ModulesCPU, path_TblD_ModulesMN, path_TblD_ModulesCN, path_TblD_ModulesPSU]
        name_sample = ['MK-504-120', 'MK-546-010', 'MK-545-010', 'MK-550-024']
        data        = []
        data_sample = []
        count       = 0
        # Упакуем данные с шаблонов
        for sample in list_path:
            if not os.path.isfile(sample):
                logger.error(f'Ошибка. Генерация сообщений для диагностики модулей: Шаблон {sample} отсутствует!')
                return (f'Ошибка. Генерация сообщений для диагностики модулей: Шаблон {sample} отсутствует!')
            root, tree = self.parser_diag_map(sample)
            for lvl_one in root.iter('Row'):
                category  = lvl_one.attrib['Category']
                isAck     = lvl_one.attrib['IsAck']
                isCycle   = lvl_one.attrib['IsCycle']
                isSound   = lvl_one.attrib['IsSound']
                isHide    = lvl_one.attrib['IsHide']
                priority  = lvl_one.attrib['Priority']
                isAlert   = lvl_one.attrib['IsAlert']
                mess      = lvl_one.attrib['Message']
                soundFile = lvl_one.attrib['SoundFile']
                nextLink  = lvl_one.attrib['NextLink']
                base      = lvl_one.attrib['Base']
                data.append(dict(category  = category,
                                 isAck     = isAck,
                                 isCycle   = isCycle,
                                 isSound   = isSound,
                                 isHide    = isHide,
                                 priority  = priority,
                                 isAlert   = isAlert,
                                 mess      = mess,
                                 soundFile = soundFile,
                                 nextLink  = nextLink,
                                 base      = base))
            data_sample.append(dict(name   = name_sample[count],
                                    data   = data))
            data   = []
            count += 1

        # Создаём файл запроса
        path_request = f'{path_request}\\PostgreSQL_Messages-Modules.sql'
        if not os.path.exists(path_request):
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(script_sql_textfile)
        else:
            os.remove(path_request)
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(script_sql_textfile)

        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        sheet     = wb['HW']
        rows      = sheet.max_row
        column    = sheet.max_column
        save_list = []

        # Из табл HW определим корзины и модули
        for i in range(4, rows + 1):
            name_uso    = str(sheet.cell(row=i, column=4).value)
            number_rack = str(sheet.cell(row=i, column=5).value)
            number      = str(sheet.cell(row=i, column=1).value)
            for j in range(7, column + 1):
                cell       = self.translate(str(sheet.cell(row=i, column=j).value))
                if cell is None:continue
                if j % 2 != 0:
                    if cell in list_modul:
                        number_modul = sheet.cell(row=2, column=j).value.partition('_')[2]
                        message = f'Диагностика. {name_uso}. Модуль A{number_rack}.{number_modul} {cell}'
                        save_list.append(dict(message     = message,
                                              number_rack = number,
                                              name_modul  = cell))
        # Определим стартовый адрес
        for msg in data_msg:
            tag       = msg['Название таблицы БД ВУ']
            code_msg  = msg['Индекс']
            count_msg = msg['Количество']
            if tag == 'TblD_ModulesPSU':
                start_modulesPSU = code_msg
                count_modulesPSU = int(count_msg)
            if tag == 'TblD_ModulesCPU':
                start_modulesCPU = code_msg
                count_modulesCPU = int(count_msg)
            if tag == 'TblD_ModulesMN' :
                start_modulesMN  = code_msg
                count_modulesMN  = int(count_msg)
            if tag == 'TblD_ModulesCN' :
                start_modulesCN  = code_msg
                count_modulesCN  = int(count_msg)

        # Если стартовый адрес пуст
        if start_modulesPSU is None or count_modulesPSU is None:
            logger.error(f'Ошибка. Генерация сообщений для диагностики модулей: отсутствует стартовый адрес или количество - TblD_ModulesPSU!')
            return (f'Ошибка. Генерация сообщений для диагностики модулей: отсутствует стартовый адрес или количество - TblD_ModulesPSU!')
        if start_modulesCPU is None or count_modulesCPU is None:
            logger.error(f'Ошибка. Генерация сообщений для диагностики модулей: отсутствует стартовый адрес или количество - TblD_ModulesCPU!')
            return (f'Ошибка. Генерация сообщений для диагностики модулей: отсутствует стартовый адрес или количество - TblD_ModulesCPU!')
        if start_modulesMN is None or count_modulesMN is None:
            logger.error(f'Ошибка. Генерация сообщений для диагностики модулей: отсутствует стартовый адрес или количество - TblD_ModulesMN!')
            return (f'Ошибка. Генерация сообщений для диагностики модулей: отсутствует стартовый адрес или количество - TblD_ModulesMN!')
        if start_modulesCN is None or count_modulesCN is None:
            logger.error(f'Ошибка. Генерация сообщений для диагностики модулей: отсутствует стартовый адрес или количество - TblD_ModulesCN!')
            return (f'Ошибка. Генерация сообщений для диагностики модулей: отсутствует стартовый адрес или количество - TblD_ModulesCN!')

        # Парсим шаблон и заполняем файл
        count_MN  = 0
        count_CN  = 0
        count_PSU = 0
        count_CPU = 0
        for sign in save_list:
            message    = sign['message']
            number     = sign['number_rack']
            name_modul = sign['name_modul']

            if name_modul == 'MK-546-010':
                number_msg = start_modulesMN + (count_MN * count_modulesMN)
                count_MN += 1
            if name_modul == 'MK-545-010':
                number_msg = start_modulesCN + (count_CN * count_modulesCN)
                count_CN += 1
            if name_modul == 'MK-550-024':
                number_msg = start_modulesPSU + (count_PSU * count_modulesPSU)
                count_PSU += 1
            if name_modul == 'MK-504-120':
                number_msg = start_modulesCPU + (count_CPU * count_modulesCPU)
                count_CPU += 1

            for samp in data_sample:
                name  = samp['name']
                files = samp['data']
                if name == name_modul:
                    for i in files:
                        for key, value in i.items():
                            if key == 'category' : category  = value
                            if key == 'isAck'    : isAck     = value
                            if key == 'isCycle'  : isCycle   = value
                            if key == 'isSound'  : isSound   = value
                            if key == 'isHide'   : isHide    = value
                            if key == 'priority' : priority  = value
                            if key == 'isAlert'  : isAlert   = value
                            if key == 'mess'     : mess      = value
                            if key == 'soundFile': soundFile = value
                            if key == 'nextLink' : nextLink  = value
                            if key == 'base'     : base      = value

                        # Запрос
                        sql_request = f"DELETE FROM messages.OPMessages WHERE Category ={number_msg + int(category)};\n" \
                                      f"INSERT INTO messages.OPMessages (Category, Message, IsAck, SoundFile, IsCycle, IsSound, IsHide, Priority, IsAlert)" \
                                      f"VALUES({number_msg + int(category)}, '{message}. {mess}', " \
                                      f"{isAck}, '{soundFile}', {isCycle}, {isSound}, {isHide}, {priority}, {isAlert});\n"
                        file.write(sql_request)
        file.write(f'COMMIT;')
        file.close()
        logger.info(f'Выполнено. Генерация сообщений для диагностики модулей: PSU, CPU, MN, CN')
        return (f'Выполнено. Генерация сообщений для диагностики модулей: PSU, CPU, MN, CN')
    # PostgreSQL_Messages-ModulesRS.sql
    def msg_modules_rs(self, path_sample, path_request):
        start_adress = None
        data_msg     = self.data['MSG']
        # Скрипт создания таблицы
        script_sql_textfile =  ('\tCREATE SCHEMA IF NOT EXISTS messages;\n'
                                 '\tCREATE TABLE IF NOT EXISTS messages.OPMessages(\n'
                                 '\t\tCategory INT NOT NULL,\n'
                                 '\t\tMessage VARCHAR(1024),\n'
                                 '\t\tIsAck BOOLEAN NOT NULL,\n'
                                 '\t\tSoundFile VARCHAR(1024),\n'
                                 '\t\tIsCycle BOOLEAN NOT NULL,\n'
                                 '\t\tIsSound BOOLEAN NOT NULL,\n'
                                 '\t\tIsHide BOOLEAN NOT NULL,\n'
                                 '\t\tPriority INT NOT NULL,\n'
                                 '\t\tIsAlert BOOLEAN NOT NULL,\n'
                                 '\t\tCONSTRAINT OPMessages_pkey PRIMARY KEY (Category)\n'
                                 '\t);\n'
                                 'BEGIN TRANSACTION;\n')
        # Проверяем шаблон
        path_sample = f'{path_sample}\TblD_ModulesRS.xml'
        if not os.path.isfile(path_sample):
            logger.error(f'Ошибка. Генерация сообщений для диагностики модуля RS: Шаблон TblD_ModulesRS отсутствует!')
            return (f'Ошибка. Генерация сообщений для диагностики модуля RS: Шаблон TblD_ModulesRS отсутствует!')
        root, tree = self.parser_diag_map(path_sample)
        # Создаём файл запроса
        path_request = f'{path_request}\\PostgreSQL_Messages-ModulesRS.sql'
        if not os.path.exists(path_request):
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(script_sql_textfile)
        else:
            os.remove(path_request)
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(script_sql_textfile)
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        sheet     = wb['HW']
        rows      = sheet.max_row
        column    = sheet.max_column
        save_list = []
        # Определим стартовый адрес
        for msg in data_msg:
            tag      = msg['Название таблицы БД ВУ']
            code_msg = msg['Индекс']
            count    = msg['Количество']
            if tag == 'TblD_ModulesRS': start_adress = code_msg

        # Если стартовый адрес пуст
        if start_adress is None:
            logger.error(f'Ошибка. Генерация сообщений для диагностики модуля RS: отсутствует стартовый адрес - TblD_ModulesRS!')
            return (f'Ошибка. Генерация сообщений для диагностики модуля RS: отсутствует стартовый адрес - TblD_ModulesRS!')
        # Из табл HW определим корзины и модули
        for i in range(4, rows + 1):
            name_uso    = str(sheet.cell(row=i, column=4).value)
            number_rack = str(sheet.cell(row=i, column=5).value)
            number      = str(sheet.cell(row=i, column=1).value)
            for j in range(7, column + 1):
                cell = self.translate(str(sheet.cell(row=i, column=j).value))
                if cell is None:continue
                if j % 2 != 0:
                    if cell == 'MK-541-002':
                        number_modul = sheet.cell(row=2, column=j).value.partition('_')[2]
                        message = f'Диагностика. {name_uso}. Модуль A{number_rack}.{number_modul} {cell}'
                        save_list.append(dict(message     = message,
                                              number_rack = number))
        # Парсим шаблон и заполняем файл
        count_msg = 0
        for sign in save_list:
            message = sign['message']
            number  = sign['number_rack']

            for lvl_one in root.iter('Row'):
                category  = lvl_one.attrib['Category']
                isAck     = lvl_one.attrib['IsAck']
                isCycle   = lvl_one.attrib['IsCycle']
                isSound   = lvl_one.attrib['IsSound']
                isHide    = lvl_one.attrib['IsHide']
                priority  = lvl_one.attrib['Priority']
                isAlert   = lvl_one.attrib['IsAlert']
                mess      = lvl_one.attrib['Message']
                soundFile = lvl_one.attrib['SoundFile']
                nextLink  = lvl_one.attrib['NextLink']
                base      = lvl_one.attrib['Base']
                # Смещение по модулю. На 1 модуль 4 сообщения
                count_msg += 1
                # Запрос
                sql_request = f"DELETE FROM messages.OPMessages WHERE Category ={start_adress + count_msg};\n" \
                              f"INSERT INTO messages.OPMessages (Category, Message, IsAck, SoundFile, IsCycle, IsSound, IsHide, Priority, IsAlert)" \
                              f"VALUES({start_adress + count_msg}, '{message}. {mess}', " \
                              f"{isAck}, '{soundFile}', {isCycle}, {isSound}, {isHide}, {priority}, {isAlert});\n"
                file.write(sql_request)
        file.write(f'COMMIT;')
        file.close()
        logger.info(f'Выполнено. Генерация сообщений для диагностики модулей: RS')
        return (f'Выполнено. Генерация сообщений для диагностики модулей: RS')
    # PostgreSQL_Messages-DO.sql
    def msg_do(self, path_sample, path_request):
        start_adress = None
        data_msg     = self.data['MSG']
        data_do      = self.data['DO']
        # Скрипт создания таблицы
        script_sql_textfile = ('\tCREATE SCHEMA IF NOT EXISTS messages;\n'
                               '\tCREATE TABLE IF NOT EXISTS messages.OPMessages(\n'
                               '\t\tCategory INT NOT NULL,\n'
                               '\t\tMessage VARCHAR(1024),\n'
                               '\t\tIsAck BOOLEAN NOT NULL,\n'
                               '\t\tSoundFile VARCHAR(1024),\n'
                               '\t\tIsCycle BOOLEAN NOT NULL,\n'
                               '\t\tIsSound BOOLEAN NOT NULL,\n'
                               '\t\tIsHide BOOLEAN NOT NULL,\n'
                               '\t\tPriority INT NOT NULL,\n'
                               '\t\tIsAlert BOOLEAN NOT NULL,\n'
                               '\t\tCONSTRAINT OPMessages_pkey PRIMARY KEY (Category)\n'
                               '\t);\n'
                               'BEGIN TRANSACTION;\n')
        # Проверяем шаблон
        path_sample = f'{path_sample}\TblDO.xml'
        if not os.path.isfile(path_sample):
            logger.error(f'Ошибка. Генерация сообщений DO: Шаблон отсутствует!')
            return (f'Ошибка. Генерация сообщений DO: Шаблон отсутствует!')
        root, tree = self.parser_diag_map(path_sample)

        # Создаём файл запроса
        path_request = f'{path_request}\\PostgreSQL_Messages-DO.sql'
        if not os.path.exists(path_request):
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(script_sql_textfile)
        else:
            os.remove(path_request)
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(script_sql_textfile)

        # Определим стартовый адрес
        for msg in data_msg:
            tag      = msg['Название таблицы БД ВУ']
            code_msg = msg['Индекс']

            if tag == 'TblDO': start_adress = code_msg
        # Если стартовый адрес пуст
        if start_adress is None:
            logger.error(f'Ошибка. Генерация сообщений DO: отсутствует стартовый адрес - TblDO!')
            return (f'Ошибка. Генерация сообщений DO: отсутствует стартовый адрес - TblDO!')

        # Парсим шаблон и заполняем файл
        for data in data_do:
            name   = data['Название']
            number = data['№']

            adress_offset = start_adress + (10 * (int(number) - 1))
            count_signal = 0
            for lvl_one in root.iter('Row'):
                category  = lvl_one.attrib['Category']
                isAck     = lvl_one.attrib['IsAck']
                isCycle   = lvl_one.attrib['IsCycle']
                isSound   = lvl_one.attrib['IsSound']
                isHide    = lvl_one.attrib['IsHide']
                priority  = lvl_one.attrib['Priority']
                isAlert   = lvl_one.attrib['IsAlert']
                mess      = lvl_one.attrib['Message']
                soundFile = lvl_one.attrib['SoundFile']
                nextLink  = lvl_one.attrib['NextLink']
                base      = lvl_one.attrib['Base']
                # Смещение по сигналам. На 1 сигнал 10 сообщений
                count_signal += 1
                # Запрос
                sql_request = f"DELETE FROM messages.OPMessages WHERE Category ={adress_offset + count_signal};\n" \
                              f"INSERT INTO messages.OPMessages (Category, Message, IsAck, SoundFile, IsCycle, IsSound, IsHide, Priority, IsAlert)" \
                              f"VALUES({adress_offset + count_signal}, '{name}. {mess}', " \
                              f"{isAck}, '{soundFile}', {isCycle}, {isSound}, {isHide}, {priority}, {isAlert});\n"
                file.write(sql_request)
        file.write(f'COMMIT;')
        file.close()
        logger.info(f'Выполнено. Генерация сообщений DO')
        return (f'Выполнено. Генерация сообщений DO')

    # СУ
    @logger.catch
    def gen_module(self, path):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(path, parser)
        root = tree.getroot()
        try:
            for el in root.iter():
                if el.attrib == "Name":
                    print(el.text)
                    exit()
        except:
            logger.error(f'Корень Root: отсутствует! Работа прекращена')

    @logger.catch
    def ret_inp_cfg(self, inp):
        stateAI = {'Warn'  : 0,
                   'Avar'  : 1,
                   'LTMin' : 2,
                   'MTMax' : 3,
                   'AlgNdv': 4,
                   'Imit'  : 5,
                   'ExtNdv': 6,
                   'Ndv'   : 7,
                   'Init'  : 8}
        stateAIzone = {'rez_0'                  : 0,
                       'Min6'                   : 1,
                       'Min5'                   : 2,
                       'Min4'                   : 3,
                       'Min3_IsMT10Perc'        : 4,
                       'Min2_IsNdv2ndParam'     : 5,
                       'Min1_IsHighVibStat'     : 6,
                       'Norm'                   : 7,
                       'Max1_IsHighVibStatNMNWR': 8,
                       'Max2_IsHighVibNoStat'   : 9,
                       'Max3_IsAvarVibStat'     : 10,
                       'Max4_IsAvarVibStatNMNWR': 11,
                       'Max5_IsAvarVibNoStat'   : 12,
                       'Max6_IsAvar2Vib'        : 13,
                       'rez_14'                 : 14,
                       'rez_15'                 : 15}
        stateDI = {'Value'    : 0,
                   'ElInput'  : 1,
                   'O'        : 2,
                   'KZ'       : 3,
                   'NC'       : 4,
                   'Imit'     : 5,
                   'ExtNdv'   : 6,
                   'Ndv'      : 7,
                   'priority1': 8,
                   'priority2': 9,
                   'priority3': 10,
                   'rez_11'   : 11,
                   'rez_12'   : 12,
                   'Front_0_1': 13,
                   'Front_1_0': 14,
                   'CfgErr'   : 15}
        stateNPS = {'ModeNPSDst'       : 0,
                    'MNSInWork'        : 1,
                    'IsMNSOff'         : 2,
                    'IsNPSModePsl'     : 3,
                    'IsPressureReady'  : 4,
                    'NeNomFeedInterval': 5,
                    'OIPHighPressure'  : 6,
                    'KTPR_P'           : 7,
                    'KTPR_M'           : 8,
                    'CSPAlinkOK'       : 9,
                    'CSPAWorkDeny'     : 10,
                    'TSstopped'        : 11,
                    'rez_12'           : 12,
                    'stopDisp'         : 13,
                    'stopCSPA'         : 14,
                    'stopARM'          : 15}
        stateFacility = {'longGasPoint1': 0,
                         'longGasPoint2': 1,
                         'longGasPoint3': 2,
                         'longGasPoint4': 3,
                         'longGasPoint5': 4,
                         'longGasPoint6': 5,
                         'longGasPoint7': 6,
                         'longGasPoint8': 7,
                         'rez_8'        : 8,
                         'rez_9'        : 9,
                         'rez_10'       : 10,
                         'rez_11'       : 11,
                         'rez_12'       : 12,
                         'rez_13'       : 13,
                         'rez_14'       : 14,
                         'rez_15'       : 15}
        warnFacility = {'warnGasPoint1': 0,
                        'warnGasPoint2': 1,
                        'warnGasPoint3': 2,
                        'warnGasPoint4': 3,
                        'warnGasPoint5': 4,
                        'warnGasPoint6': 5,
                        'warnGasPoint7': 6,
                        'warnGasPoint8': 7,
                        'rez_8'        : 8,
                        'rez_9'        : 9,
                        'rez_10'       : 10,
                        'rez_11'       : 11,
                        'rez_12'       : 12,
                        'rez_13'       : 13,
                        'rez_14'       : 14,
                        'rez_15'       : 15}
        Facility = {'ndv2Gas'   : 0,
                    'GasLim'    : 1,
                    'GasAv'     : 2,
                    'GasKeep'   : 3,
                    'GasNdvWait': 4,
                    'GasLimWait': 5,
                    'GasNdvProt': 6,
                    'GasAvProt' : 7,
                    'ColdOn'    : 8,
                    'HotOn'     : 9,
                    'rez_10'    : 10,
                    'rez_11'    : 11,
                    'ColdOff'   : 12,
                    'HotOff'    : 13,
                    'rez_14'    : 14,
                    'rez_15'    : 15}
        vsgrpstate = {'REZ_EXIST'           : 0,
                      'REM'                 : 1,
                      'OTKL'                : 2,
                      'OTKL_BY_CMD'         : 3,
                      'VKL_AS_DOP'          : 4,
                      'PUSK_OSN'            : 5,
                      'rez_6'               : 6,
                      'rez_7'               : 7,
                      'rez_8'               : 8,
                      'rez_9'               : 9,
                      'rez_10'              : 10,
                      'rez_11'              : 11,
                      'rez_12'              : 12,
                      'rez_13'              : 13,
                      'LAST_OFF_BY_CMD_ARM ': 14,
                      'ALL_OFF_WITH_BLOCK ' : 15}
        statektpr = {'P': 0,
                     'F': 1,
                     'M': 2}
        state_na = {'MainState_1_VKL'     : 0,
                    'MainState_2_OTKL'    : 1,
                    'MainState_3_PUSK'    : 2,
                    'MainState_4_OSTANOV' : 3,
                    'SubState_1_GP'       : 4,
                    'SubState_2_GORREZ'   : 5,
                    'SubState_3_PP'       : 6,
                    'SubState_4_PO'       : 7,
                    'Mode_1_OSN'          : 8,
                    'Mode_2_TU'           : 9,
                    'Mode_3_REZ'          : 10,
                    'Mode_4_REM'          : 11,
                    'KTPRA_P'             : 12,
                    'SimAgr'              : 13,
                    'Prog_1'              : 14,
                    'Prog_2'              : 15
                    }
        state_na2 = {'HIGHVIB'      : 0,
                     'HIGHVIBNas'   : 1,
                     'QF3A'         : 2,
                     'QF1A'         : 3,
                     'BBon'         : 4,
                     'BBoff'        : 5,
                     'KTPRA_FNM'    : 6,
                     'KTPRA_M'      : 7,
                     'GMPNA_M'      : 8,
                     'BBErrOtkl_All': 9,
                     'BBErrOtkl'    : 10,
                     'BBErrOtkl1'   : 11,
                     'BBErrVkl'     : 12,
                     'GMPNA_P'      : 13,
                     'GMPNA_F'      : 14,
                     'StateAlarm_VV': 15}
        state_na3 = {'KKCAlarm1'   : 0,
                     'KKCAlarm2'   : 1,
                     'KKCAlarm3'   : 2,
                     'KKCAlarm4'   : 3,
                     'InputPath'   : 4,
                     'OutputPath'  : 5,
                     'OIPVib'      : 6,
                     'rez_7'       : 7,
                     'KTPRA_NP'    : 8,
                     'KTPR_ACHR'   : 9,
                     'KTPR_SAON'   : 10,
                     'StopWork'    : 11,
                     'StartWork'   : 12,
                     'SAR_Ramp'    : 13,
                     'needRez'     : 14,
                     'needOverhaul': 15
                     }
        state_na4 = {'StopNoCmd_1'        : 0,
                     'StopNoCmd_2'        : 1,
                     'StartNoCmd'         : 2,
                     'StateAlarm'         : 3,
                     'StateAlarm_ChRP'    : 4,
                     'StateAlarm_All'     : 5,
                     'ChRPRegError'       : 6,
                     'LogicalChRPCrash'   : 7,
                     'ZD_Unprompted_Close': 8,
                     'StopErr'            : 9,
                     'StopErr2'           : 10,
                     'StopErr_All'        : 11,
                     'StartErr'           : 12,
                     'StartErr2'          : 13,
                     'StartErr3'          : 14,
                     'StartErr_All'       : 15,
                     }
        state_na5 = {'ED_IsMT10Perc'         : 0,
                    'ED_IsNdv2ndParam'       : 1,
                    'ED_IsHighVibStat'       : 2,
                    'ED_IsHighVibNoStat'     : 3,
                    'ED_IsAvarVibStat'       : 4,
                    'ED_IsAvarVibNoStat'     : 5,
                    'ED_IsAvar2Vib'          : 6,
                    'Pump_IsMT10Perc'        : 7,
                    'Pump_IsNdv2ndParam'     : 8,
                    'Pump_IsHighVibStat'     : 9,
                    'Pump_IsHighVibStatNMNWR': 10,
                    'Pump_IsHighVibNoStat'   : 11,
                    'Pump_IsAvarVibStat'     : 12,
                    'Pump_IsAvarVibStatNMNWR': 13,
                    'Pump_IsAvarVibNoStat'   : 14,
                    'Pump_IsAvar2Vib'        : 15
                    }
        state_na6 = {'GMPNA_P_2_64': 0,
                     'rez_1': 1,
                     'rez_2': 2,
                     'rez_3': 3,
                     'rez_4': 4,
                     'rez_5': 5,
                     'rez_6': 6,
                     'rez_7': 7,
                     'rez_8': 8,
                     'rez_9': 9,
                     'rez_10': 10,
                     'rez_11': 11,
                     'RptVKL': 12,
                     'UseCT': 13,
                     'ZDin_Unprompted_Close': 14,
                     'ZDout_Unprompted_Close': 15}

        state_zd1  = {'State_1_Opening': 0,
                      'State_2_Opened' : 1,
                      'State_3_Middle' : 2,
                      'State_4_Closing': 3,
                      'State_5_Closed' : 4,
                      'Dist'           : 5,
                      'Imit'           : 6,
                      'NOT_EC'         : 7,
                      'Avar'           : 8,
                      'Diff'           : 9,
                      'WarnClose'      : 10,
                      'Blink'          : 11,
                      'KVO'            : 12,
                      'KVZ'            : 13,
                      'MPO'            : 14,
                      'MPZ'            : 15}
        state_zd2  = {'CorrCO'        : 0,
                      'CorrCZ'        : 1,
                      'VMMO'          : 2,
                      'VMMZ'          : 3,
                      'NOT_ZD_EC_KTP' : 4,
                      'Local'         : 5,
                      'Mufta'         : 6,
                      'Avar_BUR'      : 7,
                      'NeispravVU'    : 8,
                      'ErrMPO'        : 9,
                      'ErrMPZ'        : 10,
                      'EC'            : 11,
                      'RS_OK'         : 12,
                      'Close_Fail'    : 13,
                      'Open_Fail'     : 14,
                      'Stop_Fail'     : 15}
        state_zd3 = {'ECsign'          : 0,
                     'rez_1'           : 1,
                     'rez_2'           : 2,
                     'Unprompted_Open' : 3,
                     'Unprompted_Close': 4,
                     'Neisprav'        : 5,
                     'CorrCOCorrCZ'    : 6,
                     'rez_7'           : 7,
                     'Open'            : 8,
                     'Close'           : 9,
                     'Stop'            : 10,
                     'StopClose'       : 11,
                     'VMMO_save'       : 12,
                     'VMMZ_save'       : 13,
                     'Mufta_save'      : 14,
                     'Avar_BUR_save'   : 15}

        isNum = 0
        isInv = 0
        Inputvar = str(inp).split(".")
        try:
            if self.str_find(Inputvar[0], {'NOT '}):
                isInv = 1
            Inpvr = str(Inputvar[0]).replace('NOT ', '')

            if len(Inputvar) > 2:
                if self.str_find(Inpvr, {'stateSAR'}):
                    pInputnum = str(inp).split('.state.')[1]
                    isNum = 0
                    pInputpInputVar = (str(inp).split('.state.')[0].replace('NOT ', '')) + '.state.reg'
                    if self.str_find(pInputpInputVar,{'stateQ'}):
                        pInputpInputVar = (str(inp).split('.state.')[0].replace('NOT ', '')) + '.reg'
                if self.str_find(Inpvr, {'stateBUF'}):
                    pInputnum = Inputvar[2]
                    isNum = 0
                    pInputpInputVar = Inputvar[0] + '.state.reg'
                if self.str_find(Inpvr, {'mRS'}):
                    pInputnum = Inputvar[2]
                    isNum = 0
                    #pInputpInputVar = Inputvar[1]
                    text = str(inp).replace('NOT ', '').split('.')
                    pInputpInputVar = f'{text[0]}.{text[1]}.reg'
                if self.str_find(Inpvr, {'stateNA'}):
                    pInputnum = Inputvar[2]
                    isNum = 0
                    pInputpInputVar = Inpvr + '.' + Inputvar[1] + '.reg'
            elif len(Inputvar) > 1:
                if (Inputvar[1] in stateAI.keys()) and (self.str_find(Inpvr, {'AI'})):
                    pInputnum = stateAI[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('AI', 'StateAI') + '.state.reg'
                    cfg = '00000000000000' + str(isNum) + str(isInv)
                    return pInputpInputVar, pInputnum, str(hex(int(cfg, 2))).replace("0x", "16#")
                if (Inputvar[1] in stateAIzone.keys()) and (self.str_find(Inpvr, {'AI'})):
                    pInputnum = stateAIzone[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('AI', 'StateAI') + '.stateZone.reg'
                if (Inputvar[1] in vsgrpstate.keys()) and (self.str_find(Inpvr, {'VSGRP'})):
                    pInputnum = vsgrpstate[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('VSGRP', 'stateVSGRP') + '.state.reg'
                if (Inputvar[1] in stateFacility.keys()) and (self.str_find(Inpvr, {'Facility'})):
                    pInputnum = stateFacility[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('Facility', 'stateFacility') + '.longGas.reg'
                if (Inputvar[1] in warnFacility.keys()) and (self.str_find(Inpvr, {'Facility'})):
                    pInputnum = warnFacility[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('Facility', 'stateFacility') + '.warnGas.reg'
                if (Inputvar[1] in Facility.keys()) and (self.str_find(Inpvr, {'Facility'})):
                    pInputnum = Facility[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('Facility', 'stateFacility') + '.state.reg'
                if (Inputvar[1] in stateDI.keys()) and (self.str_find(Inpvr, {'DI'})):
                    pInputnum = stateDI[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('DI', 'StateDI') + '.state.reg'
                if (Inputvar[1] in stateNPS.keys()) and (self.str_find(Inpvr, {'NPS'})):
                    pInputnum = stateNPS[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NPS', 'stateNPS') + '.state.reg'
                if (Inputvar[1] in state_na5.keys()) and (self.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na5[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state5.reg'
                if (Inputvar[1] in state_na.keys()) and (self.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state1.reg'
                if (Inputvar[1] in state_na3.keys()) and (self.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na3[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state3.reg'
                if (Inputvar[1] in state_na6.keys()) and (self.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na6[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state6.reg'
                if (Inputvar[1] in state_na4.keys()) and (self.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na4[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state4.reg'
                if (Inputvar[1] in state_na2.keys()) and (self.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na2[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state2.reg'
                if (Inputvar[1] in statektpr.keys()) and (self.str_find(Inpvr, {'KTPR'})):
                    pInputnum = statektpr[Inputvar[1]]
                    isNum = 0
                    isInv = 0
                    pInputpInputVar = str(Inpvr).replace('KTPR', 'stateKTPR') + '.state.reg'
                if (Inputvar[1] in state_zd1.keys()) and (self.str_find(Inpvr, {'ZD'})):
                    pInputnum = state_zd1[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('ZD', 'stateZD') + '.state1.reg'
                if (Inputvar[1] in state_zd2.keys()) and (self.str_find(Inpvr, {'ZD'})):
                    pInputnum = state_zd2[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('ZD', 'stateZD') + '.state2.reg'
                if (Inputvar[1] in state_zd3.keys()) and (self.str_find(Inpvr, {'ZD'})):
                    pInputnum = state_zd3[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('ZD', 'stateZD') + '.state3.reg'

            if Inputvar[0][:2] == 'AI':
                if Inputvar[1] == 'Value':
                    pInputpInputVar = Inputvar
                    isNum = 1
            cfg = '00000000000000' + str(isNum) + str(isInv)
            return pInputpInputVar, pInputnum, str(hex(int(cfg, 2))).replace("0x", "16#")
        except:
            return 0, 0, 0

    # Cfg_pic
    @logger.catch
    def gen_cfg_pic(self, path, system):
        # соединение с exel
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet_hw   = wb['HW']
        sheet_ss   = wb['SS']
        sheet_tmdp = wb['TM_DP']
        # максимальное число рядов и столбцов
        rows_hw   = sheet_hw.max_row
        rows_ss   = sheet_ss.max_row
        rows_tmdp = sheet_tmdp.max_row

        data       = {}

        data_pic      = self.data['Pic']
        data['AI']    = self.data['AI']
        data['DI']    = self.data['DI']
        data['ZD']    = self.data['ZD']
        data['VS']    = self.data['VS']
        data['SS']    = self.data['SS']
        data['KTPRA'] = self.data['KTPRA']
        data['Pic']   = self.data['Pic']
        
        if system != 'ASPT': 
            data['KTPR'] = self.data['KTPR']


        try:
            path_cfg = f'{path}\cfg_PIC.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_PIC*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_PIC*)\n')
            cfg_txt = ''

            lst = {}

            for info_pic in data_pic:
                pic   = info_pic['№']
                frame = info_pic['Кадр IFix(*.grf)']

                if pic is None: continue

                a = {}
                s_a = []

                lst_ai           = []
                lst_di_err       = []
                lst_di_avar      = []
                lst_zd           = []
                lst_vs           = []
                lst_ktpra        = []
                lst_ktpr         = []
                lst_pic          = []
                lst_hw           = []
                lst_ss_warn      = []
                lst_ss_avar      = []
                lst_net_uso_mn   = []
                lst_net_uso_cn   = []
                ss_warn          = []
                ss_avar          = []
                ss_avar_dp       = []
                lst_net_uso      = {}
                lst_net_ss       = {}
                count            = 0
                count_avar       = 0
                count_net_uso_mn = 0
                count_net_uso_cn = 0

                for el in data['AI']:
                    if el['Pic'] is None: continue
                    s_pic = str(el['Pic']).split(';')
                    for pic_num in s_pic:
                        if str(pic_num) == str(pic):
                            lst_ai.append(el['№'])
                a['AI'] = lst_ai
                count = count + len(lst_ai)
                count_avar = count_avar + len(lst_ai)

                for el in data['DI']:
                    if el['Pic'] is None: continue
                    s_pic = str(el['Pic']).split(';')
                    for pic_num in s_pic:
                        if str(pic_num) == str(pic):
                            if el['priority[0]'] == 3 or el['priority[1]'] == 3:
                                lst_di_avar.append(el['№'])
                                if el['pNC_AI'] is not None:
                                    lst_di_err.append(el['№'])
                            else:
                                lst_di_err.append(el['№'])

                a['DI_err']  = lst_di_err
                a['DI_avar'] = lst_di_avar
                count = count + len(lst_di_err)
                count_avar = count_avar + len(lst_di_avar)

                for el in data['ZD']:
                    s_pic = str(el['Pic']).split(';')
                    for pic_num in s_pic:
                        if str(pic_num) == str(pic):
                            lst_zd.append(el['№'])
                a['ZD'] = lst_zd
                count = count + len(lst_zd)
                count_avar = count_avar + len(lst_zd)

                for el in data['VS']:
                    if el['Pic'] is None: continue
                    s_pic = str(el['Pic']).split(';')
                    for pic_num in s_pic:
                        if str(pic_num) == str(pic):
                            lst_vs.append(el['№'])
                a['VS'] = lst_vs
                count_avar = count_avar + len(lst_vs)

                for el in data['KTPRA']:
                    if el['Pic'] is None: continue
                    s_pic = str(el['Pic']).split(';')
                    for pic_num in s_pic:
                        if str(pic_num) == str(pic):
                            lst_ktpra.append(el['Переменная'])
                a['KTPRA'] = lst_ktpra
                count_avar = count_avar + len(lst_ktpra)

                if system != 'ASPT':
                    for el in data['KTPR']:
                        if el['Pic'] is None: continue
                        s_pic = str(el['Pic']).split(';')
                        for pic_num in s_pic:
                            if str(pic_num) == str(pic):
                                lst_ktpr.append(el['Переменная'])
                    a['KTPR'] = lst_ktpr
                    count_avar = count_avar + len(lst_ktpr)

                for el in data['Pic']:
                    caption_pic = el['№']
                    if el['Pic'] is None: continue
                    if pic == caption_pic:
                        s_pic = str(el['Pic']).split(';')
                        for pic_num in s_pic:
                            #if str(pic_num) == str(pic):
                            lst_pic.append(pic_num)
                a['Pic']   = lst_pic
                count = count + len(lst_pic)
                count_avar = count_avar + len(lst_pic)

                # HW
                for i in range(4, rows_hw + 1):
                    number = sheet_hw.cell(row=i, column=1).value
                    pic_hw = sheet_hw.cell(row=i, column=39).value
                    if number is None or pic_hw is None: continue
                    s_pic = str(pic_hw).split(';')
                    for pic_num in s_pic:
                        if str(pic_num) == str(pic):
                            lst_hw.append(number)
                a['HW']   = lst_hw
                count_avar = count_avar + len(lst_hw)

                #Сеть УСО
                if self.str_find(str(frame).lower(), {'net_uso'}):
                    for i in range(4, rows_hw + 1):
                        cell = sheet_hw.cell(row=i, column=9).value
                        if cell is None: continue
                        if self.str_find(cell, {'MK-546-010'}):
                            count_net_uso_mn += 1
                            lst_net_uso_mn.append(count_net_uso_mn)
                        if self.str_find(cell, {'MK-545-010'}):
                            count_net_uso_cn += 1
                            lst_net_uso_cn.append(count_net_uso_cn)

                    lst_net_uso['MN'] = lst_net_uso_mn
                    lst_net_uso['CN'] = lst_net_uso_cn
                    a['Net_USO'] = lst_net_uso
                count = count + len(lst_net_uso)
                count_avar = count_avar + len(lst_net_uso)

                # Смежные системы
                for i in range(4, rows_ss + 1):
                    cell_port1  = sheet_ss.cell(row=i, column=6).value
                    cell_port2  = sheet_ss.cell(row=i, column=10).value
                    cell_number = sheet_ss.cell(row=i, column=1).value
                    cell_Pic    = sheet_ss.cell(row=i, column=13).value
                    s_pic = str(cell_Pic).split(';')
                    for pic_num in s_pic:
                        if str(pic_num) == str(pic):
                            if not cell_port1 is None and not cell_port2 is None:
                                lst_ss_warn.append(cell_number)
                            lst_ss_avar.append(cell_number)
                a['SS_warn'] = lst_ss_warn
                a['SS_avar'] = lst_ss_avar
                count = count + len(lst_ss_warn)
                count_avar = count_avar + len(lst_ss_avar)

                if self.str_find(str(frame).lower(), {'relatedsystems'}):
                    for i in range(4, rows_ss + 1):
                        cell_port1  = sheet_ss.cell(row=i, column=6).value
                        cell_port2  = sheet_ss.cell(row=i, column=10).value
                        cell_number = sheet_ss.cell(row=i, column=1).value
                        if not cell_port1 is None and not cell_port2 is None:
                            ss_warn.append(cell_number)
                        ss_avar.append(cell_number)

                    for i in range(4, rows_tmdp + 1):
                        name_tdp = sheet_tmdp.cell(row=i, column=4).value
                        link_tdp = sheet_tmdp.cell(row=i, column=5).value
                        if name_tdp is None: continue

                        ss_avar_dp.append(f"TM_DP_linkOk.{str(link_tdp).split('.state.')[1]}")

                    lst_net_ss['WARN'] = ss_warn
                    lst_net_ss['AVAR'] = ss_avar
                    lst_net_ss['DP']   = ss_avar_dp
                    a['SS'] = lst_net_ss
                count = count + len(lst_net_ss)
                count_avar = count_avar + len(lst_net_ss)

                a['count']      = count
                a['count_avar'] = count_avar
                s_a.append(a)
                lst[pic] = s_a

            cfg_txt = cfg_txt + '(* Желтые рамки *)\n'


            for el in data_pic:
                cfg_txt = cfg_txt + f"(*{el['Переменная']} \t{el['Название']}*)\n"
                cfg_txt = cfg_txt + f"ctrlPic[{el['№']}].countWarn :="
                # warning
                for el1 in lst[el['№']]:
                    cnt = 0
                    if el1['count'] == 0:
                        cfg_txt = cfg_txt + f"0;\n"
                        continue
                    else:
                        for el2 in el1['AI']:
                            for i in data['AI']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count'] else ';'
                            cfg_txt = cfg_txt + f"AIcountWarn[{el2}]{znak}                              (* {name_signal} *)\n"

                        for el2 in el1['DI_err']:
                            for i in data['DI']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count'] else ';'
                            cfg_txt = cfg_txt + f"BYTE_TO_UDINT(DIcountWarn[{el2}]){znak}               (* {name_signal} *)\n"

                        for el2 in el1['ZD']:
                            for i in data['ZD']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count'] else ';'
                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(stateZD[{el2}].state2.bits.NeispravVU){znak}            (* {name_signal} *)\n"

                        for el2 in el1['Pic']:
                            for i in data['Pic']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count'] else ';'
                            cfg_txt = cfg_txt + f"ctrlPic[{el2}].countWarn{znak}                        (* {name_signal} *)\n"

                        for el2 in el1['SS_warn']:
                            for i in data['SS']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break

                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(NOT stateDIAG.SS[{el2}].bits.link1Ok)+   (* {name_signal} *)\n"
                            cnt += 1
                            znak = '+' if cnt < len(el1['SS_warn']) else ';'
                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(NOT stateDIAG.SS[{el2}].bits.link2Ok){znak} (* {name_signal} *)\n "

            cfg_txt = cfg_txt + '(* Красные рамки *)\n'
            for el in data_pic:
                cfg_txt = cfg_txt + f"\t(*{el['Переменная']} \t{el['Название']}*)\n"
                cfg_txt = cfg_txt + f"ctrlPic[{el['№']}].countAvar:="
                # cfg_txt = cfg_txt + f"0;\n"
                for el1 in lst[el['№']]:
                    cnt = 0
                    if el1['count_avar'] == 0:
                        cfg_txt = cfg_txt + f"0;\n"
                        continue
                    else:
                        for el2 in el1['AI']:
                            for i in data['AI']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count_avar'] else ';'
                            cfg_txt = cfg_txt + f"AIcountAvar[{el2}]{znak}                                  (* {name_signal} *)\n"
                        for el2 in el1['DI_avar']:
                            for i in data['DI']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count_avar'] else ';'
                            cfg_txt = cfg_txt + f"BYTE_TO_UDINT(DIcountAvar[{el2}]){znak}                   (* {name_signal} *)\n"
                        for el2 in el1['ZD']:
                            for i in data['ZD']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count_avar'] else ';'
                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(stateZD[{el2}].state1.bits.Avar)+\n" \
                                                f"BOOL_TO_UDINT(stateZD[{el2}].state1.bits.NOT_EC){znak}        (* {name_signal} *)\n"
                        for el2 in el1['VS']:
                            for i in data['VS']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count_avar'] else ';'
                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(stateVS[{el2}].state1.bits.NEISPRAV)+\n"
                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(stateVS[{el2}].state1.bits.MPC_CEPI_VKL)+\n"
                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(NOT stateVS[{el2}].state1.bits.EC){znak}        (* {name_signal} *)\n"
                        if system != 'ASPT':
                            for el2 in el1['KTPR']:
                                for i in data['KTPR']:
                                    number = i['Переменная']
                                    name = i['Название']
                                    if number == el2:
                                        name_signal = name
                                        break
                                cnt += 1
                                znak = '+' if cnt < el1['count_avar'] else ';'
                                cfg_txt = cfg_txt + f"BOOL_TO_UDINT((state{el2}.state.bits.F) AND (NOT state{el2}.state.bits.M)){znak}        (* {name_signal} *)\n"
                        for el2 in el1['KTPRA']:
                            for i in data['KTPRA']:
                                number = i['Переменная']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count_avar'] else ';'
                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(state{el2}.state.bits.F AND (NOT state{el2}.state.bits.M)){znak}        (* {name_signal} *)\n"

                        for el2 in el1['Pic']:
                            for i in data['Pic']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count_avar'] else ';'
                            cfg_txt = cfg_txt + f"ctrlPic[{el2}].countAvar{znak}                        (* {name_signal} *)\n"

                        for el2 in el1['HW']:
                            for i in range(4, rows_hw + 1):
                                number = sheet_hw.cell(row=i, column=1).value
                                perem  = sheet_hw.cell(row=i, column=2).value
                                name   = sheet_hw.cell(row=i, column=4).value

                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count_avar'] else ';'
                            cfg_txt = cfg_txt + f"{perem}{znak}                        (* {name_signal} *)\n"

                        for el2 in el1['SS_avar']:
                            for i in data['SS']:
                                number = i['№']
                                name   = i['Название']
                                if number == el2:
                                    name_signal = name
                                    break
                            cnt += 1
                            znak = '+' if cnt < el1['count_avar'] else ';'
                            cfg_txt = cfg_txt + f"BOOL_TO_UDINT(NOT stateDIAG.SS[{el2}].bits.linkOk){znak}  (* {name_signal} *)\n "

                        try:
                            if el1['Net_USO']:
                                list_MN = el1.get('Net_USO').get('MN')
                                list_CN = el1.get('Net_USO').get('CN')

                                for mn in list_MN:
                                    for i in range(2):
                                        cfg_txt = cfg_txt + f"BOOL_TO_UDINT(stateDIAG.diagMN[{mn}].ports_State.bits.eP{i + 1}NotLink)+\n"
                                for cn in list_CN:
                                    cfg_txt = cfg_txt + f"BOOL_TO_UDINT(stateDIAG.diagCN[{cn}].ports_State.bits.eP1NotLink)+\n"

                                    cnt += 1
                                    znak = '+' if cnt < len(list_CN) else ';'
                                    cfg_txt = cfg_txt + f"BOOL_TO_UDINT(stateDIAG.diagCN[{cn}].ports_State.bits.eP2NotLink){znak}\n"
                        except: pass

            text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
           logger.error(f'{self.name_prefix} FAILED')
    # Cfg_ktprs
    @logger.catch
    def gen_cfg_ktprs(self, path):
        data = self.data['KTPRS']

        try:
            path_cfg = f'{path}\Cfg_ktprs.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_ktprs*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_ktprs*)\n')

            for value in data:
                numbers = value['№']
                name    = value['Название']
                srab    = value['Сработка']
                prio0   = value['Приоритет сообщ. при 0']
                prio1   = value['Приоритет сообщ. при 1']
                noMsg   = value['Запрет выдачи сообщений'] if value['Запрет выдачи сообщений'] is not None else '0'
                if numbers is None: continue

                pInputvar, num, cfg = self.ret_inp_cfg(value['Сработка'])
                cfg_txt = (f'(*{numbers} {name}*)\n')
                if pInputvar != 0:
                    cfg_txt = cfg_txt + f'cfgKTPRS[{numbers}].pInputVar.pInputVar               REF={str(pInputvar)};\n' \
                                        f'cfgKTPRS[{numbers}].pInputVar.num                       :={str(num)};\n' \
                                        f'cfgKTPRS[{numbers}].pInputVar.cfg.reg                   :={str(cfg)};\n'

                # cfg_txt = cfg_txt + f'cfgKTPRS[{numbers}].pVal                        REF={"0"};(*не понятно на что ссылаться из экселя не ясно*)\n'
                cfg = '000000000000000' + str(noMsg)
                cfg_txt = cfg_txt + f"cfgKTPRS[{numbers}].cfg.reg                      :={str(hex(int(cfg, 2))).replace('0x', '16#')};\n"
                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_VV
    @logger.catch
    def gen_cfg_VV(self, path):
        data = self.data['VV']
        try:
            path_cfg = f'{path}\cfg_VV.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_VV*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_VV*)\n')

            for value in data:
                numbers = value['№']
                tag   = value['Идентификатор']
                name  = value['Название']
                VVOn  = value['Высоковольтный выключатель включен']
                VVOff = value['Высоковольтный выключатель отключен']
                if numbers is None: continue

                pVVOn, pVVOnnum, pVVOncfg = self.ret_inp_cfg(value['Высоковольтный выключатель включен'])
                pVVOff, pVVOffnum, pVVOffcfg = self.ret_inp_cfg(value['Высоковольтный выключатель отключен'])

                cfg_txt = (f'(*{tag} {name}*)\n')
                if pVVOn != 0:
                    cfg_txt = cfg_txt + f'cfgVV[{numbers}].pBBB.pInputVar               REF={str(pVVOn)};\n' \
                                        f'cfgVV[{numbers}].pBBB.num                       :={str(pVVOnnum)};\n' \
                                        f'cfgVV[{numbers}].pBBB.cfg.reg                       :={str(pVVOncfg)};\n'
                if pVVOff != 0:
                    cfg_txt = cfg_txt + f'cfgVV[{numbers}].pBBO.pInputVar                REF={str(pVVOff)};\n' \
                                        f'cfgVV[{numbers}].pBBO.num                        :={str(pVVOffnum)};\n' \
                                        f'cfgVV[{numbers}].pBBO.cfg.reg                        :={str(pVVOffcfg)};\n'

                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_uts
    @logger.catch
    def gen_cfg_uts(self, path):
        data = self.data['UTS']

        try:
            path_cfg = f'{path}\Cfg_uts.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_uts*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_uts*)\n')

            for value in data:
                numbers = value['№']
                tag          = value['Идентификатор']
                name         = value['Название']
                pVkl         = value['Включить']
                isSiren      = value['Сирена']
                blockAutoOff = value['Не требует автоотключения']
                if numbers is None: continue
                if pVkl    is None: continue

                pCheckInput, cinum, cicfg = self.ret_inp_cfg(value['Проверка'])
                pKvitInput, kvnum, kvcfg  = self.ret_inp_cfg(value['Квитирование'])

                cfg_txt = (f'(*{tag} {name}*)\n')
                if pCheckInput != 0:
                    cfg_txt = cfg_txt + f'cfgUTS[{numbers}].pCheckInput.pInputVar               REF={str(pCheckInput)};\n' \
                                        f'cfgUTS[{numbers}].pCheckInput.num                       :={str(cinum)};\n' \
                                        f'cfgUTS[{numbers}].pCheckInput.cfg.reg                       :={str(cicfg)};\n'
                if pKvitInput != 0:
                    cfg_txt = cfg_txt + f'cfgUTS[{numbers}].pKvitInput.pInputVar                REF={str(pKvitInput)};\n' \
                                        f'cfgUTS[{numbers}].pKvitInput.num                        :={str(kvnum)};\n' \
                                        f'cfgUTS[{numbers}].pKvitInput.cfg.reg                        :={str(kvcfg)};\n'
                cfg_txt = cfg_txt + f'cfgUTS[{numbers}].pVkl                    REF={str(pVkl)};\n'
                cfg_txt = cfg_txt + f'cfgUTS[{numbers}].isSiren                 :={str(isSiren)};\n'
                cfg_txt = cfg_txt + f'cfgUTS[{numbers}].blockAutoOff            :={str(blockAutoOff)};\n'
                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_VSGRP
    @logger.catch
    def gen_cfg_VSGRP(self, path):
        data = self.data['VSGRP']
        try:
            path_cfg = f'{path}\cfg_VSGRP.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_VSGRP*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_VSGRP*)\n')
            for value in data:
                numbers = value['№']
                # tag            = value['Переменная']
                name    = value['Название']
                isPoz   = '1' if value['Пож. или водоорош.'] is not None else '0'
                countVS = value['Количество вспомсистем в группе']
                WarnOff = '1' if value['Требуется выставлять флаг WarnOff, если работает одна вспомсистема в группе'] is not None else '0'
                if name is None: continue
                cfg = str(hex(int("000000" + str(WarnOff) + str(isPoz), 2))).replace('0x', '16#')
                cfg_txt = f'(*{name}*)\n' \
                          f'cfgVSGRP[{numbers}].countVS       :=  {countVS};\n' \
                          f'cfgVSGRP[{numbers}].cfg.reg       :=  {cfg};\n'
                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_nps
    @logger.catch
    def gen_cfg_nps(self, path):
        data = self.data['NPS']
        try:
            path_cfg = f'{path}\cfg_NPS.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_NPS*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_NPS*)\n')

            for value in data:
                s_name = {1: 'pInputVar REF=', 2: 'num:=', 3: 'cfg.reg:='}

                numbers = value['№']
                tag     = value['Идентификатор']
                name    = value['Название']
                perem   = value['Переменная']
                pValue  = value['Значение']
                if numbers is None:
                    if name is not None:
                        cfg_txt = (f'(*{name}*)\n')
                        text_file.write(cfg_txt)
                if pValue is None: continue
                # pInput, pnum, pcfg = self.ret_inp_cfg(value['Аварийный параметр'])
                cfg_txt = (f'(*{name}*)\n')
                if str(pValue).isdigit():
                    cfg_txt = cfg_txt + f'{perem}:={pValue};\n'
                elif self.str_find(pValue, 'AIValue'):
                    cfg_txt = cfg_txt + f'{perem}:={pValue};\n'
                else:
                    a = {}
                    pInput, pnum, pcfg = self.ret_inp_cfg(pValue)
                    a[1] = pInput
                    a[2] = pnum
                    a[3] = pcfg
                    for el in range(1, 4):
                        cfg_txt = cfg_txt + f'{perem}.{s_name[el]}{str(a[el])};\n'
                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_rsreq
    @logger.catch
    def gen_cfg_rsreq(self, path):
        data = self.data['RSreq']

        srsreq = ['Route',
                  'SlaveId',
                  'ModbusFunction',
                  'Address',
                  'Count',
                  'ResultOffset',
                  'SingleRequest',
                  'OnModifyRequest',
                  'RepeatOverScan',
                  'SkipRepeatsWhenBad',
                  'Enable']
        try:
            path_cfg = f'{path}\Cfg_rsreq.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_rsreq*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_rsreq*)\n')

            for value in data:
                name = value['Название']
                if name is None: continue
                numbers = value['№']
                cfg_txt = f'(*{numbers} -- {name}*)\n'
                fname = str(value['Route'][:-3])

                for el in srsreq:
                    num = ""
                    if len(str(value[el])) > 4:
                        a = len(str(value[el])) - 2
                        num = str(value[el][a:-1])
                        break
                for el in srsreq:
                    if el == 'Route': continue
                    if el != 'Enable':
                        cfg_txt = cfg_txt + f'{fname}_Req.Val[{num}].{el}:={value[el]};\n'
                    else:
                        cfg_txt = cfg_txt + f'{fname}_Cmd.Val[{num}].{el}:={value[el]};\n'
                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_na
    @logger.catch
    def gen_cfg_na(self, path):
        data = self.data['UMPNA']
        s_name = {1: 'pInputVar REF=', 2: 'num:=', 3: 'cfg.reg:='}

        try:
            path_cfg = f'{path}\cfg_na.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_na*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_na*)\n')

            for value in data:
                numbers = value['№']
                tag = value['Переменная']
                name = value['Название']
                if name is None: continue
                s_umpna = {'pBBB1': 'ВВ Включен',
                           'pBBO1': 'ВВ отключен',
                           'pBBB2': 'ВВ Включен дубль',
                           'pBBO2': 'ВВ отключен дубль',
                           'pTok': 'Сила тока >  уставки холостого хода',
                           'pECx02': 'Исправность цепей включения ВВ',
                           'pECx03': 'Исправность цепей отключения ВВ',
                           'pECx03_1': 'Исправность цепей отключения ВВ дубль',
                           'pKKC[1]': 'Стоп 1',
                           'pKKC[2]': 'Стоп 2',
                           'pKKC[3]': 'Стоп 3',
                           'pKKC[4]': 'Стоп 4',
                           'pEC1': 'Сигнал «Контроль наличия напряжения в цепях оперативного тока»',
                           'pMotorCellVoltage': 'Флаг наличия напряжения в двигательной ячейке ЗРУ',
                           'pECx04': 'Тележка ВВ выкачена',
                           'pDCx01': 'Дистанционный режим управления контроллера РЗиА',
                           'pRZiALink': 'Наличие связи с контроллером РЗиА',
                           'pExcitReady': 'Состояние возбудителя ЭД',
                           'pOPCx22': 'Флаг окончания предпусковой продувки двигателя',
                           'pSafetyPressureMotor': 'Флаг наличия безопасного давления подпора воздуха в корпусе двигателя',
                           'pSafetyPressureExcit': 'Флаг наличия безопасного давления подпора воздуха в корпусе возбудителя',
                           'pBlowValveClosed': 'Флаг закрытого положения клапана продувки двигателя',
                           'pOIPOilTemperatureFreezer': 'Флаг температуры масла маслосистемы выше 10гр.С на выходе охладителя (для индивидуальной маслосистемы)',
                           'pOIPOilValueMin2': 'Флаг предельного минимального уровня масла в маслобаке (для индивидуальной маслосистемы)',
                           'pOipClosingFluidMinLevel': 'Флаг наличия минимального уровня запирающей жидкости в баке системы запирания',
                           'pOipClosingFluidPressure': 'Обобщенный флаг наличия давления запирающей жидкости к торцевому уплотнению',
                           'pGMPNA[49]': 'GMPNA_[49]',
                           'pGMPNA[50]': 'GMPNA_[50]',
                           'pGMPNA[51]': 'GMPNA_[51]',
                           'pGMPNA[52]': 'GMPNA_[52]',
                           'pGMPNA[53]': 'GMPNA_[53]',
                           'pGMPNA[54]': 'GMPNA_[54]',
                           'pGMPNA[55]': 'GMPNA_[55]',
                           'pGMPNA[56]': 'GMPNA_[56]',
                           'pGMPNA[57]': 'GMPNA_[57]',
                           'pGMPNA[58]': 'GMPNA_[58]',
                           'pGMPNA[59]': 'GMPNA_[59]',
                           'pGMPNA[60]': 'GMPNA_[60]',
                           'pGMPNA[61]': 'GMPNA_[61]',
                           'pGMPNA[62]': 'GMPNA_[62]',
                           'pGMPNA[63]': 'GMPNA_[63]',
                           'pGMPNA[64]': 'GMPNA_[64]',
                           }
                ds_umpna = {'iDelay': 'Количество сканов задержки анализа исправности цепей управления ВВ НА',
                            'nVSprMN': 'Номер агрегата вспомсистемы "пуско-резервный маслонасос" (для индивидуальной маслосистемы)',
                            'GMPNA_49_64_NotMasked': 'Номер агрегата вспомсистемы "пуско-резервный маслонасос" (для индивидуальной маслосистемы)',
                            'nNPS': 'Номер НПС (1 или 2), к которой относится НА',
                            'nProtACHR': 'Номер защиты АЧР в массиве станционных защит',
                            'nProtSAON': 'Номер защиты САОН в массиве станционных защит',
                            'iCounterNdv': 'Параметр для KTPRAS_1',
                            }
                out_umpna = {'pStartWork': 'Команда на включение ВВ (только для UMPNA)',
                             'pStopWork1': 'Команда на отключение ВВ (выход 1)',
                             'pStopWork2': 'Команда на отключение ВВ (выход 2)'}
                cfg_txt   = (f'(*{tag} {name}*)\n')
                IsCMNA    = value['НА с ЧРП']  # (*0 - UMPNA, 1 - CMNA *)
                MPNA_Type = value['Тип НА - МНА']
                IsNM      = value['Насос типа НМ']
                for perem in s_umpna:
                    if value[s_umpna[perem]] is None: continue
                    a = {}
                    pInput, pnum, pcfg = self.ret_inp_cfg(value[s_umpna[perem]])
                    a[1] = pInput
                    a[2] = pnum
                    a[3] = pcfg
                    for el in range(1, 4):
                        cfg_txt = cfg_txt + f'cfg{tag}.{perem}.{s_name[el]}{str(a[el])};\n'
                for perem in ds_umpna:
                    if value[ds_umpna[perem]] is None: continue
                    cfg_txt = cfg_txt + f'cfg{tag}.{perem}:={str(value[ds_umpna[perem]])};\n'
                cfg_txt = cfg_txt + f"cfg{tag}.cfg.reg:={str(hex(int('0000000000000' + str(IsNM) + str(MPNA_Type) + str(IsCMNA), 2))).replace('0x', '16#')};\n"
                for perem in out_umpna:
                    if value[out_umpna[perem]] is None: continue
                    cfg_txt = cfg_txt + f'cfg{tag}.{perem} REF={str(value[out_umpna[perem]])};\n'
                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_ktpra
    @logger.catch
    def gen_cfg_ktpra(self, path):
        data = self.data['KTPRA']
        try:
            path_cfg = f'{path}\cfg_KTPRA.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_KTPRA*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_KTPRA*)\n')   

            for value in data:
                # '№','Переменная','Идентификатор','Название','Короткое название (для отображения на кадре)','Группа','Номер в группе',
                # 'МП','Давл. норма','Напр.','Напр. СШ','Испр. Цепей вкл.','Внешняя авария','Датчик давл. неиспр.','Включить','Отключить','АПВ не требуется',
                # 'Pic','Таблица сообщений','Это клапан / интерфейсная вспомсистема','AlphaHMI'
                numbers = value['№']
                tag     = value['Идентификатор']
                name    = value['Название']
                KTPRA   = value['Переменная']
                pInput  = value['Аварийный параметр']
                NA      = value['НА']

                if name is None: continue
                if tag is None: continue

                avr         = value['АВР'] if value['АВР'] is not None else '0'
                stype       = value['Тип остановки'] if value['Тип остановки'] is not None else '0'
                NotMasked   = value['Запрет маскирования'] if value['Запрет маскирования'] is not None else '0'
                CloseValves = value['Закрытие задвижек'] if value['Закрытие задвижек'] is not None else '0'
                ktpra_cfg   = str(hex(int(str(CloseValves) + str(avr) + str(NotMasked), 2))).replace('0x', '16#')

                pInput, pnum, pcfg = self.ret_inp_cfg(value['Аварийный параметр'])

                cfg_txt = (f'(*{tag} {name}*)\n')
                if pInput != 0:
                    cfg_txt = cfg_txt + f'cfg{KTPRA}.pInput.pInputVar              REF={str(pInput)};\n' \
                                        f'cfg{KTPRA}.pInput.num                      :={str(pnum)};\n' \
                                        f'cfg{KTPRA}.pInput.cfg.reg                  :={str(pcfg)};\n' \
                                        f'cfg{KTPRA}.StopType					     :={str(stype)};\n' \
                                        f'cfg{KTPRA}.cfg.reg		                 :={str(ktpra_cfg)};\n'

                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_VS
    @logger.catch
    def gen_cfg_VS(self, path):
        data = self.data['VS']
        try:
            path_cfg = f'{path}\cfg_VS.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_VS*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_VS*)\n')

            for value in data:
                # '№','Переменная','Идентификатор','Название','Короткое название (для отображения на кадре)','Группа','Номер в группе',
                # 'МП','Давл. норма','Напр.','Напр. СШ','Испр. Цепей вкл.','Внешняя авария','Датчик давл. неиспр.','Включить','Отключить','АПВ не требуется',
                # 'Pic','Таблица сообщений','Это клапан / интерфейсная вспомсистема','AlphaHMI'
                numbers  = value['№']
                tag      = value['Идентификатор']
                name     = value['Название']
                grpNum   = value['Группа']
                numInGrp = value['Номер в группе']
                pVkl     = value['Включить']
                pOtkl    = value['Отключить']
                PC_USE   = '1' if value['Давл. норма'] is not None else '0'
                noAPV    = value['АПВ не требуется'] if value['АПВ не требуется'] is not None else '0'

                if numbers is None: continue

                pMPCpInputVar, pMPCnum, pMPCcfg = self.ret_inp_cfg(value['МП'])
                pPCpInputVar, pPCnum, pPCcfg = self.ret_inp_cfg(value['Давл. норма'])
                pECpInputVar, pECnum, pECcfg = self.ret_inp_cfg(value['Напр.'])
                pSEC_ECpInputVar, pSEC_ECnum, pSEC_ECcfg = self.ret_inp_cfg(value['Напр. СШ'])
                pOPCpInputVar, pOPCnum, pOPCcfg = self.ret_inp_cfg(value['Испр. Цепей вкл.'])
                pDiAVARpInputVar, pDiAVARnum, pDiAVARcfg = self.ret_inp_cfg(value['Внешняя авария'])
                pPC_NEISPRAVpInputVar, pPC_NEISPRAVnum, pPC_NEISPRAVcfg = self.ret_inp_cfg(
                    value['Датчик давл. неиспр.'])

                cfg_txt = (f'(*{tag} {name}*)\n')
                if pMPCpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pMPC.pInputVar               REF={str(pMPCpInputVar)};\n' \
                                        f'cfgVS[{numbers}].pMPC.num                       :={str(pMPCnum)};\n' \
                                        f'cfgVS[{numbers}].pMPC.cfg.reg                       :={str(pMPCcfg)};\n'
                if pPCpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pPC.pInputVar                REF={str(pPCpInputVar)};\n' \
                                        f'cfgVS[{numbers}].pPC.num                        :={str(pPCnum)};\n' \
                                        f'cfgVS[{numbers}].pPC.cfg.reg                        :={str(pPCcfg)};\n'
                if pECpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pEC.pInputVar                REF={str(pECpInputVar)};\n' \
                                        f'cfgVS[{numbers}].pEC.num                        :={str(pECnum)};\n' \
                                        f'cfgVS[{numbers}].pEC.cfg.reg                        :={str(pECcfg)};\n'
                if pSEC_ECpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pSEC_EC.pInputVar            REF={pSEC_ECpInputVar};\n' \
                                        f'cfgVS[{numbers}].pSEC_EC.num                    :={pSEC_ECnum};\n' \
                                        f'cfgVS[{numbers}].pSEC_EC.cfg.reg                    :={pSEC_ECcfg};\n'
                if pOPCpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pOPC.pInputVar               REF={pOPCpInputVar};\n' \
                                        f'cfgVS[{numbers}].pOPC.num                       :={pOPCnum};\n' \
                                        f'cfgVS[{numbers}].pOPC.cfg.reg                       :={pOPCcfg};\n'
                if pDiAVARpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pDiAVAR.pInputVar            REF={pDiAVARpInputVar};\n' \
                                        f'cfgVS[{numbers}].pDiAVAR.num                    :={pDiAVARnum};\n' \
                                        f'cfgVS[{numbers}].pDiAVAR.cfg.reg                    :={pDiAVARcfg};\n'
                if pPC_NEISPRAVpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pPC_NEISPRAV.pInputVar       REF={pPC_NEISPRAVpInputVar};\n' \
                                        f'cfgVS[{numbers}].pPC_NEISPRAV.num               :={pPC_NEISPRAVnum};\n' \
                                        f'cfgVS[{numbers}].pPC_NEISPRAV.cfg.reg               :={pPC_NEISPRAVcfg};\n'
                cfgVS_unioncfgVS = str(hex(int('00000000000000' + str(noAPV) + str(PC_USE), 2))).replace('0x',
                                                                                                         '16#')
                cfg_txt = cfg_txt + f'cfgVS[{numbers}].cfgVS.reg              :={cfgVS_unioncfgVS};\n' \
                                    f'cfgVS[{numbers}].grpNum                         :={grpNum};\n' \
                                    f'cfgVS[{numbers}].numInGrp                       :={numInGrp};\n' \
                                    f'cfgVS[{numbers}].pVkl                         REF={pVkl};\n' \
                                    f'cfgVS[{numbers}].pOtkl                        REF={pOtkl};\n'

                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_ZD не готова!
    @logger.catch
    def gen_cfg_ZD(self, path):
        data  = self.data['ZD']
        data1 = self.data['RS']
        try:
            path_cfg = f'{path}\cfg_ZD.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_ZD*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_ZD*)\n')

            for value in data:
                numbers = value['№']
                tag     = value['Идентификатор']
                name    = value['Название']
                rs_ok   = bool(value['Наличие ИНТЕРФЕЙСА'])

                if tag is None: continue

                IOpKVOpInputVar, IOpKVOnum, IOpKVOcfg_unioncfg_st                = self.ret_inp_cfg(value['КВО'])
                IOpKVZpInputVar, IOpKVZnum, IOpKVZcfg_unioncfg_st                = self.ret_inp_cfg(value['КВЗ'])
                IOpMPOpInputVar, IOpMPOnum, IOpMPOcfg_unioncfg_st                = self.ret_inp_cfg(value['МПО'])
                IOpMPZpInputVar, IOpMPZnum, IOpMPZcfg_unioncfg_st                = self.ret_inp_cfg(value['МПЗ'])
                IOpDIST_KEYpInputVar, IOpDIST_KEYnum, IOpDIST_KEYcfg_unioncfg_st = self.ret_inp_cfg(value['Дист_ф'])
                IOpMuftapInputVar, IOpMuftanum, IOpMuftacfg_unioncfg_st          = self.ret_inp_cfg(value['Муфта'])
                IOpAvar_BURpInputVar, IOpAvar_BURnum, IOpAvar_BURcfg_unioncfg_st = self.ret_inp_cfg(value['Авар. Привода'])
                pNoLinkpInputVar, pNoLinknum, pNoLinkcfg_unioncfg_st             = self.ret_inp_cfg(value['Отсут. связи'])
                pBRUClosepInputVar, pBRUClosenum, pBRUClosecfg_unioncfg_st       = self.ret_inp_cfg(value['Закр. с БРУ'])
                pBRUStoppInputVar, pBRUStopnum, pBRUStopcfg_unioncfg_st          = self.ret_inp_cfg(value['Стоп с БРУ'])
                pECpInputVar, pECnum, pECcfg_unioncfg_st                         = self.ret_inp_cfg(value['Напряж.'])
                pECsignpInputVar, pECsignnum, pECsigncfg_unioncfg_st             = self.ret_inp_cfg(value['Напряж. в цепях сигнализации'])
                pZD_EC_KTPpInputVar, pZD_EC_KTPnum, pZD_EC_KTPcfg_unioncfg_st    = self.ret_inp_cfg(value['Напряж. ЩСУ'])
                pCorrCOpInputVar, pCorrCOnum, pCorrCOcfg_unioncfg_st             = self.ret_inp_cfg(value['Испр. цепей откр.'])
                pCorrCZpInputVar, pCorrCZnum, pCorrCZcfg_unioncfg_st             = self.ret_inp_cfg(value['Испр. цепей закр.'])
                pVMMOpInputVar, pVMMOnum, pVMMOcfg_unioncfg_st                   = self.ret_inp_cfg(value['ВММО'])
                pVMMZpInputVar, pVMMZnum, pVMMZcfg_unioncfg_st                   = self.ret_inp_cfg(value['ВММЗ'])

                IOpOpen        = value['Открыть']
                IOpClose       = value['Закрыть']
                IOpStop        = value['Остановить']
                IOpStopOpen    = value['Откр. остановить']
                IOpStopClose   = value['Закр. остановить']
                typeBURtypeBUR = value['Тип БУР задвижки']
                isClp          = value['Это клапан'] if value['Это клапан'] is not None else '0'
                freeze         = '0'

                if rs_ok is True:
                    RSpKVOpInputVar, RSpKVOnum, RSpKVOcfg_unioncfg_st                = self.ret_inp_cfg(value['КВО_и'])
                    RSpKVZpInputVar, RSpKVZnum, RSpKVZcfg_unioncfg_st                = self.ret_inp_cfg(value['КВЗ_и'])
                    RSpMPOpInputVar, RSpMPOnum, RSpMPOcfg_unioncfg_st                = self.ret_inp_cfg(value['МПО_и'])
                    RSpMPZpInputVar, RSpMPZnum, RSpMPZcfg_unioncfg_st                = self.ret_inp_cfg(value['МПЗ_и'])
                    RSpDIST_KEYpInputVar, RSpDIST_KEYnum, RSpDIST_KEYcfg_unioncfg_st = self.ret_inp_cfg(value['Дист_и'])
                    RSpMuftapInputVar, RSpMuftanum, RSpMuftacfg_unioncfg_st          = self.ret_inp_cfg(value['Муфта_и'])
                    RSpAvar_BURpInputVar, RSpAvar_BURnum, RSpAvar_BURcfg_unioncfg_st = self.ret_inp_cfg(value['Авар. Привода_и'])

                    RSpOpen      = value['Открыть_и']
                    RSpClose     = value['Закрыть_и']
                    RSpStop      = value['Остановить_и']
                    RSpStopOpen  = value['Откр. Остановить_и']
                    RSpStopClose = value['Закр. Остановить_и']

                cfg_txt = (f'(*{numbers}-{tag}-{name}*)\n')
                # RS
                if rs_ok is True:
                    if RSpKVOpInputVar != 0:
                        cfg_txt = cfg_txt + '' \
                                            f"cfgZD[{numbers}].RS.pKVO.pInputVar REF={str(RSpKVOpInputVar)};\n" \
                                            f'cfgZD[{numbers}].RS.pKVO.num:={str(RSpKVOnum)};\n' \
                                            f'cfgZD[{numbers}].RS.pKVO.cfg.reg:={str(RSpKVOcfg_unioncfg_st)};\n'
                    if RSpKVZpInputVar != 0:
                        cfg_txt = cfg_txt + '' \
                                            f'cfgZD[{numbers}].RS.pKVZ.pInputVar REF={str(RSpKVZpInputVar)};\n' \
                                            f'cfgZD[{numbers}].RS.pKVZ.num:={str(RSpKVZnum)};\n' \
                                            f'cfgZD[{numbers}].RS.pKVZ.cfg.reg:={str(RSpKVZcfg_unioncfg_st)};\n'
                    if RSpMPOpInputVar != 0:
                        cfg_txt = cfg_txt + '' \
                                            f'cfgZD[{numbers}].RS.pMPO.pInputVar REF={str(RSpMPOpInputVar)};\n' \
                                            f'cfgZD[{numbers}].RS.pMPO.num:={str(RSpMPOnum)};\n' \
                                            f'cfgZD[{numbers}].RS.pMPO.cfg.reg:={str(RSpMPOcfg_unioncfg_st)};\n'
                    if RSpMPZpInputVar != 0:
                        cfg_txt = cfg_txt + '' \
                                            f'cfgZD[{numbers}].RS.pMPZ.pInputVar REF={str(RSpMPZpInputVar)};\n' \
                                            f'cfgZD[{numbers}].RS.pMPZ.num:={str(RSpMPZnum)};\n' \
                                            f'cfgZD[{numbers}].RS.pMPZ.cfg.reg:={str(RSpMPZcfg_unioncfg_st)};\n'
                    if RSpDIST_KEYpInputVar != 0:
                        cfg_txt = cfg_txt + '' \
                                            f'cfgZD[{numbers}].RS.pDIST_KEY.pInputVar REF={str(RSpDIST_KEYpInputVar)};\n' \
                                            f'cfgZD[{numbers}].RS.pDIST_KEY.num:={str(RSpDIST_KEYnum)};\n' \
                                            f'cfgZD[{numbers}].RS.pDIST_KEY.cfg.reg:={str(RSpDIST_KEYcfg_unioncfg_st)};\n'
                    if RSpMuftapInputVar != 0:
                        cfg_txt = cfg_txt + '' \
                                            f'cfgZD[{numbers}].RS.pMufta.pInputVar REF={str(RSpMuftapInputVar)};\n' \
                                            f'cfgZD[{numbers}].RS.pMufta.num:={str(RSpMuftanum)};\n' \
                                            f'cfgZD[{numbers}].RS.pMufta.cfg.reg:={str(RSpMuftacfg_unioncfg_st)};\n'
                    if RSpAvar_BURpInputVar != 0:
                        cfg_txt = cfg_txt + '' \
                                            f'cfgZD[{numbers}].RS.pAvar_BUR.pInputVar REF={str(RSpAvar_BURpInputVar)};\n' \
                                            f'cfgZD[{numbers}].RS.pAvar_BUR.num:={str(RSpAvar_BURnum)};\n' \
                                            f'cfgZD[{numbers}].RS.pAvar_BUR.cfg.reg:={str(RSpAvar_BURcfg_unioncfg_st)};\n'

                    if RSpOpen != 0            : cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pOpen REF={str(RSpOpen)};\n'
                    if RSpClose is not None    : cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pClose REF={str(RSpClose)};\n'
                    if RSpStop is not None     : cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pStop REF={str(RSpStop)};\n'
                    if RSpStopOpen is not None : cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pStopOpen REF={str(RSpStopOpen)};\n'
                    if RSpStopClose is not None: cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pStopClose REF={str(RSpStopClose)};\n'

                # Not RS
                if IOpKVOpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].IO.pKVO.pInputVar REF={str(IOpKVOpInputVar)};\n' \
                                        f'cfgZD[{numbers}].IO.pKVO.num:={str(IOpKVOnum)};\n' \
                                        f'cfgZD[{numbers}].IO.pKVO.cfg.reg:={str(IOpKVOcfg_unioncfg_st)};\n'
                if IOpKVZpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].IO.pKVZ.pInputVar REF={str(IOpKVZpInputVar)};\n' \
                                        f'cfgZD[{numbers}].IO.pKVZ.num:={str(IOpKVZnum)};\n' \
                                        f'cfgZD[{numbers}].IO.pKVZ.cfg.reg:={str(IOpKVZcfg_unioncfg_st)};\n'
                if IOpMPOpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].IO.pMPO.pInputVar REF={str(IOpMPOpInputVar)};\n' \
                                        f'cfgZD[{numbers}].IO.pMPO.num:={str(IOpMPOnum)};\n' \
                                        f'cfgZD[{numbers}].IO.pMPO.cfg.reg:={str(IOpMPOcfg_unioncfg_st)};\n'
                if IOpMPZpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].IO.pMPZ.pInputVar REF={str(IOpMPZpInputVar)};\n' \
                                        f'cfgZD[{numbers}].IO.pMPZ.num:={str(IOpMPZnum)};\n' \
                                        f'cfgZD[{numbers}].IO.pMPZ.cfg.reg:={str(IOpMPZcfg_unioncfg_st)};\n'
                if IOpDIST_KEYpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].IO.pDIST_KEY.pInputVar REF={str(IOpDIST_KEYpInputVar)};\n' \
                                        f'cfgZD[{numbers}].IO.pDIST_KEY.num:={str(IOpDIST_KEYnum)};\n' \
                                        f'cfgZD[{numbers}].IO.pDIST_KEY.cfg.reg:={str(IOpDIST_KEYcfg_unioncfg_st)};\n'
                if IOpMuftapInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].IO.pMufta.pInputVar REF={str(IOpMuftapInputVar)};\n' \
                                        f'cfgZD[{numbers}].IO.pMufta.num:={str(IOpMuftanum)};\n' \
                                        f'cfgZD[{numbers}].IO.pMufta.cfg.reg:={str(IOpMuftacfg_unioncfg_st)};\n'
                if IOpAvar_BURpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].IO.pAvar_BUR.pInputVar REF={str(IOpAvar_BURpInputVar)};\n' \
                                        f'cfgZD[{numbers}].IO.pAvar_BUR.num:={str(IOpAvar_BURnum)};\n' \
                                        f'cfgZD[{numbers}].IO.pAvar_BUR.cfg.reg:={str(IOpAvar_BURcfg_unioncfg_st)};\n'

                if IOpOpen != 0            : cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pOpen REF={str(IOpOpen)};\n'
                if IOpClose is not None    : cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pClose REF={str(IOpClose)};\n'
                if IOpStop is not None     : cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pStop REF={str(IOpStop)};\n'
                if IOpStopOpen is not None : cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pStopOpen REF={str(IOpStopOpen)};\n'
                if IOpStopClose is not None: cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pStopClose REF={str(IOpStopClose)};\n'

                if pNoLinkpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pNoLink.pInputVar REF={str(pNoLinkpInputVar)};\n' \
                                        f'cfgZD[{numbers}].pNoLink.num:={str(pNoLinknum)};\n' \
                                        f'cfgZD[{numbers}].pNoLink.cfg.reg:={str(pNoLinkcfg_unioncfg_st)};\n'
                if pBRUClosepInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pBRUClose.pInputVar REF={str(pBRUClosepInputVar)};\n' \
                                        f'cfgZD[{numbers}].pBRUClose.num:={str(pBRUClosenum)};\n' \
                                        f'cfgZD[{numbers}].pBRUClose.cfg.reg:={str(pBRUClosecfg_unioncfg_st)};\n'
                if pBRUStoppInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pBRUStop.pInputVar REF={str(pBRUStoppInputVar)};\n' \
                                        f'cfgZD[{numbers}].pBRUStop.num:={str(pBRUStopnum)};\n' \
                                        f'cfgZD[{numbers}].pBRUStop.cfg.reg:={str(pBRUStopcfg_unioncfg_st)};\n'
                if pECpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pEC.pInputVar REF={str(pECpInputVar)};\n' \
                                        f'cfgZD[{numbers}].pEC.num:={str(pECnum)};\n' \
                                        f'cfgZD[{numbers}].pEC.cfg.reg:={str(pECcfg_unioncfg_st)};\n'
                if pECsignpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pECsign.pInputVar REF={str(pECsignpInputVar)};\n' \
                                        f'cfgZD[{numbers}].pECsign.num:={str(pECsignnum)};\n' \
                                        f'cfgZD[{numbers}].pECsign.cfg.reg:={str(pECsigncfg_unioncfg_st)};\n'
                if pZD_EC_KTPpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pZD_EC_KTP.pInputVar REF={str(pZD_EC_KTPpInputVar)};\n' \
                                        f'cfgZD[{numbers}].pZD_EC_KTP.num:={str(pZD_EC_KTPnum)};\n' \
                                        f'cfgZD[{numbers}].pZD_EC_KTP.cfg.reg:={str(pZD_EC_KTPcfg_unioncfg_st)};\n'
                if pCorrCOpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pCorrCO.pInputVar REF={str(pCorrCOpInputVar)};\n' \
                                        f'cfgZD[{numbers}].pCorrCO.num:={str(pCorrCOnum)};\n' \
                                        f'cfgZD[{numbers}].pCorrCO.cfg.reg:={str(pCorrCOcfg_unioncfg_st)};\n'
                if pCorrCZpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pCorrCZ.pInputVar REF={str(pCorrCZpInputVar)};\n' \
                                        f'cfgZD[{numbers}].pCorrCZ.num:={str(pCorrCZnum)};\n' \
                                        f'cfgZD[{numbers}].pCorrCZ.cfg.reg:={str(pCorrCZcfg_unioncfg_st)};\n'
                if pVMMOpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pVMMO.pInputVar REF={str(pVMMOpInputVar)};\n' \
                                        f'cfgZD[{numbers}].pVMMO.num:={str(pVMMOnum)};\n' \
                                        f'cfgZD[{numbers}].pVMMO.cfg.reg:={str(pVMMOcfg_unioncfg_st)};\n'
                if pVMMZpInputVar != 0:
                    cfg_txt = cfg_txt + '' \
                                        f'cfgZD[{numbers}].pVMMZ.pInputVar REF={str(pVMMZpInputVar)};\n' \
                                        f'cfgZD[{numbers}].pVMMZ.num:={str(pVMMZnum)};\n' \
                                        f'cfgZD[{numbers}].pVMMZ.cfg.reg:={str(pVMMZcfg_unioncfg_st)};\n'
                cfg = str(hex(int('00000000000000' + str(isClp) + str(freeze), 2))).replace('0x', '16#')
                cfg_txt = cfg_txt + f'cfgZD[{numbers}].cfg.reg:={str(cfg)};\n'

                if typeBURtypeBUR is not None:
                    cfg_txt = cfg_txt + f'cfgZD[{numbers}].typeBUR.reg:={str(typeBURtypeBUR)};\n'

                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_DO_sim
    @logger.catch
    def gen_cfg_DO_sim(self,path):
        data      = self.data['DO']
        try:
            path_cfg = f'{path}\cfg_DO_sim.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DO_sim*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DO_sim*)\n')

            for value in data:
                numbers         = value['№']
                tag             = value['Идентификатор']
                name            = value['Название']
                pValue          = value['УСО, модуль, канал']
                pHealth         = value['Исправность канала']

                if self.str_find(pHealth,'mDO'):
                    pValue = str(pValue)[str(pValue).index('['):]
                    pHealth = str(pHealth)[str(pHealth).index('['):]
                    cfg_txt = f'(*{tag} {name}*)\n' \
                              f'cfgDO[{numbers}].pValue         REF=  simDO_bool{pHealth}{pValue};\n'
                    text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_DO
    def gen_cfg_DO(self, path):
        data = self.data['DO']
        try:
            path_cfg = f'{path}\cfg_DO.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DO*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DO*)\n')

            for value in data:
                numbers = value['№']
                tag = value['Идентификатор']
                name = value['Название']
                pValue = value['УСО, модуль, канал']
                pHealth = value['Исправность канала']

                if pValue is None: continue

                cfg_txt = f'(*{tag} {name}*)\n' \
                          f'cfgDO[{numbers}].pValue         REF=  {pValue};\n' \
                          f'cfgDO[{numbers}].pHealth        REF=  {pHealth};\n'

                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_DI
    @logger.catch
    def gen_cfg_DI(self, path):
        data = self.data['DI']
        try:
            path_cfg = f'{path}\cfg_DI.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DI*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DI*)\n')

            for value in data:
                s = '000000000'
                for el in ['isModuleNC', 'Msg', 'isAI_Avar', 'isAI_Warn', 'isDI_NC', 'ErrValue', 'Inv']:
                    s = s + str(value[el]) if value[el] is not None else s + '0'
                numbers = value['№']
                tag = value['Идентификатор']
                name = value['Название']
                pValue = value['pValue']
                pHealth = value['pHealth']
                pNC_AI = value['pNC_AI'] if value['pNC_AI'] is not None else '0'
                TS_ID = value['TS_ID'] if value['TS_ID'] is not None else '0'
                priority0 = value['priority[0]'] if value['priority[0]'] is not None else '0'
                priority1 = value['priority[1]'] if value['priority[1]'] is not None else '0'
                cfg = str(hex(int(s, 2))).replace('0x', '16#')
                if pValue is not None:
                    cfg_txt = f'(*{tag} {name}*)\n' \
                              f'cfgDI[{numbers}].pValue         REF=  {pValue};\n' \
                              f'cfgDI[{numbers}].pHealth        REF=  {pHealth};\n' \
                              f'cfgDI[{numbers}].TS_ID            :=  {TS_ID};\n' \
                              f'cfgDI[{numbers}].priority[0]      :=  {priority0};\n' \
                              f'cfgDI[{numbers}].priority[1]      :=  {priority1};\n' \
                              f'cfgDI[{numbers}].cfg.reg          :=  {cfg};\n'
                    # //f'cfgDI[{numbers}].pNC_AI         REF=  {pNC_AI};\n'
                    text_file.write(cfg_txt)
                if value['pNC_AI'] is not None:
                    cfg_txt = f'(*{tag} {name}*)\n' \
                              f'cfgDI[{numbers}].pNC_AI         REF=  {str(pNC_AI).replace("_union.state", ".reg")};\n' \
                              f'cfgDI[{numbers}].TS_ID            :=  {TS_ID};\n' \
                              f'cfgDI[{numbers}].priority[0]      :=  {priority0};\n' \
                              f'cfgDI[{numbers}].priority[1]      :=  {priority1};\n' \
                              f'cfgDI[{numbers}].cfg.reg    :=  {cfg};\n'
                    text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_DI_sim
    @logger.catch
    def gen_cfg_DI_sim(self,path):
        data = self.data['DI']
        try:
            path_cfg = f'{path}\cfg_DI_sim.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DI_sim*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DI_sim*)\n')

            for value in data:
                numbers         = value['№']
                tag             = value['Идентификатор']
                name            = value['Название']
                pValue          = value['pValue']
                pHealth         = value['pHealth']

                if self.str_find(pHealth,'mDI'):
                    pValue = str(pValue)[str(pValue).index('['):]
                    pHealth = str(pHealth)[str(pHealth).index('['):]
                    cfg_txt = f'(*{tag} {name}*)\n' \
                              f'cfgDI[{numbers}].pValue         REF=  simDI_bool{pHealth}{pValue};\n'
                    text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Imitator_Cfg_DI
    @logger.catch
    def gen_cfg_DI_imit(self,path):
        data_DI = self.data['DI']
        data_KD = self.data['КД']
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        sheet = wb['HW']
        try:
            path_cfg = f'{path}\cfg_DI_imitation.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DI_imitation*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DI_imitation*)\n')

            for value in data_DI:
                s = '000000000'
                for el in ['isModuleNC', 'Msg', 'isAI_Avar', 'isAI_Warn', 'isDI_NC', 'ErrValue', 'Inv']:
                    s = s + str(value[el]) if value[el] is not None else s + '0'

                numbers   = value['№']
                tag       = self.translate(str(value['Идентификатор']))
                name      = value['Название']
                TS_ID     = value['TS_ID'] if value['TS_ID'] is not None else '0'
                priority0 = value['priority[0]'] if value['priority[0]'] is not None else '0'
                priority1 = value['priority[1]'] if value['priority[1]'] is not None else '0'
                cfg       = str(hex(int(s, 2))).replace('0x', '16#')

                numb_rack_KD = ''
                numb_modl_KD = ''
                numb_chan    = ''
                name_uso_KD  = ''
                for signal in data_KD:
                    tag_KD  = self.translate(str((signal['Tэг'])))
                    if tag_KD == tag:
                        numb_rack_KD = signal['Корз']
                        numb_modl_KD = signal['Мод']
                        numb_chan    = signal['Кан']
                        name_uso_KD  = signal['Шкаф']
                        break

                exit_True = False
                for i in range(4, sheet.max_row + 1):
                    if exit_True: break
                    name_uso_HW  = sheet.cell(row=i, column=4).value
                    numb_rack_HW = sheet.cell(row=i, column=5).value
                    
                    if (name_uso_HW == name_uso_KD) and (numb_rack_KD == numb_rack_HW):
                        
                        for j in range(11, sheet.max_column + 1):
                            type_modul = sheet.cell(row=i, column=j + 1).value
                            
                            if self.str_find(type_modul, {'mDI'}):
                                 numb_modl_HW = str(sheet.cell(row=2, column=j).value).replace('_0', '').replace('_', '')
                                 
                                 if numb_modl_HW == str(numb_modl_KD):
                                    through_num_mod = int((re.findall(r'\d+', type_modul))[0])
                                    exit_True = True

                                    cfg_txt = f'(*{tag} {name}*)\n' \
                                                f'cfgDI[{numbers}].pValue         REF=  gv_sim.buf_sim_di_bool[{through_num_mod}].Bit_{numb_chan};\n' \
                                                f'cfgDI[{numbers}].pHealth        REF=  1;\n' \
                                                f'cfgDI[{numbers}].TS_ID            :=  {TS_ID};\n' \
                                                f'cfgDI[{numbers}].priority[0]      :=  {priority0};\n' \
                                                f'cfgDI[{numbers}].priority[1]      :=  {priority1};\n' \
                                                f'cfgDI[{numbers}].cfg.reg          :=  {cfg};\n'
                                    text_file.write(cfg_txt)
                                    break
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Imitator_Cfg_AI
    @logger.catch
    def gen_cfg_AI_imit(self,path):
        data_AI = self.data['AI']
        data_KD = self.data['КД']
        wb = openpyxl.load_workbook(self.exel, read_only=True)
        sheet = wb['HW']
        try:
            path_cfg = f'{path}\cfg_AI_imitation.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_AI_imitation*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_AI_imitation*)\n')

            for value in data_AI:
                numbers      = value['№']
                tag          = self.translate(str(value['Идентификатор']))
                name         = value['Название']
                isVibroPump  = value['Вибрация насоса'] if value['Вибрация насоса'] is not None else '0'
                isVibroED    = value['Вибрация ЭД'] if value['Вибрация ЭД'] is not None else '0'
                isCT         = value['Ток ЭД НА'] if value['Ток ЭД НА'] is not None else '0'
                isPressureVS = '1' if value['Номер НА или вспом.'] is not None else '0'
                cfgAI        = str(hex(int('000000000000'+str(isPressureVS)+str(isCT)+str(isVibroED)+str(isVibroPump),2))).replace('0x','16#')
                nNA_VS       = value['Номер НА или вспом.'] if value['Номер НА или вспом.'] is not None else '0'

                numb_rack_KD = ''
                numb_modl_KD = ''
                numb_chan    = ''
                name_uso_KD  = ''
                for signal in data_KD:
                    tag_KD  = self.translate(str((signal['Tэг'])))
                    if tag_KD == tag:
                        numb_rack_KD = signal['Корз']
                        numb_modl_KD = signal['Мод']
                        numb_chan    = signal['Кан']
                        name_uso_KD  = signal['Шкаф']
                        break

                exit_True = False
                for i in range(4, sheet.max_row + 1):
                    if exit_True: break
                    name_uso_HW  = sheet.cell(row=i, column=4).value
                    numb_rack_HW = sheet.cell(row=i, column=5).value
                    
                    if (name_uso_HW == name_uso_KD) and (numb_rack_KD == numb_rack_HW):
                        
                        for j in range(11, sheet.max_column + 1):
                            type_modul = sheet.cell(row=i, column=j + 1).value
                            
                            if self.str_find(type_modul, {'mAI'}):
                                 numb_modl_HW = str(sheet.cell(row=2, column=j).value).replace('_0', '').replace('_', '')
                                 
                                 if numb_modl_HW == str(numb_modl_KD):
                                    through_num_mod = int((re.findall(r'\d+', type_modul))[0])
                                    exit_True = True

                                    cfg_txt = f'(* {tag} {name} *)\n' \
                                            f'cfgAI[{numbers}].pValue                       REF=\tgv_sim.sim_AI[{numbers}].Channel[{numb_chan}];\n'\
                                            f'cfgAI[{numbers}].pHealth                      REF=\t1;\n' \
                                            f'cfgAI[{numbers}].cfgWarnAvar.reg                :=\t16#0000;\n' \
                                            f'cfgAI[{numbers}].cfgAI.reg                      :=\t16#0000;\n' \
                                            f'cfgAI[{numbers}].nNA_VS                         :=\t{nNA_VS};\n' \
                                            f'cfgAI[{numbers}].nFuse                          :=\t0;\n'
                                    text_file.write(cfg_txt)
                                    break
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    
    # Cfg_AO
    @logger.catch
    def gen_cfg_AO(self, path):
        data = self.data['AO']
        try:
            path_cfg = f'{path}\cfg_AO.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_AO*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_AO*)\n')

            for value in data:
                numbers = value['№']
                tag     = value['Идентификатор']
                name    = value['Название']
                pValue  = value['УСО, модуль, канал']
                pHealth = value['Исправность канала']

                if pValue is None: continue

                cfg_txt = f'(*{tag} {name}*)\n' \
                          f'cfgAO[{numbers}].pValue         REF=  {pValue};\n' \
                          f'cfgAO[{numbers}].pHealth        REF=  {pHealth};\n'

                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_AI
    @logger.catch
    def gen_cfg_AI(self,path):
        data      = self.data['AI']
        try:
            path_cfg = f'{path}\cfg_AI.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_AI*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_AI*)\n')
            for value in data:
                numbers         = value['№']
                tag             = value['Идентификатор']
                name            = value['Название']
                pValue          = value['УСО, модуль, канал']
                pHealth         = value['Исправность канала']
                isVibroPump     = value['Вибрация насоса'] if value['Вибрация насоса'] is not None else '0'
                isVibroED       = value['Вибрация ЭД'] if value['Вибрация ЭД'] is not None else '0'
                isCT            = value['Ток ЭД НА'] if value['Ток ЭД НА'] is not None else '0'
                isPressureVS    = '1' if value['Номер НА или вспом.'] is not None else '0'
                isExternal      = value['Давление на вых. вспом.'] if value['Давление на вых. вспом.'] is not None else '0'
                # cfgWarnAvar     = str(hex(int(value['Сигнализация'],2))).replace('0x','16#') if value['Сигнализация'] is not None else '16#0000'
                cfgAI           = str(hex(int('000000000000'+str(isPressureVS)+str(isCT)+str(isVibroED)+str(isVibroPump),2))).replace('0x','16#')
                nNA_VS          = value['Номер НА или вспом.'] if value['Номер НА или вспом.'] is not None else '0'
                nFuse           = value['Предохранитель'] if value['Предохранитель'] is not None else '0'
                minAvar = value['№ уставки мин. авар.']
                minWarn = value['№ уставки мин. пред.']
                maxWarn = value['№ уставки макс. пред.']
                maxAvar = value['№ уставки макс. авар.']
                cfgWarnAvar = '11'
                cfgWarnAvar = 2**(int(minAvar)-1) + 2**(int(minWarn)+4-1)+2**(int(maxWarn)+8-1)+2**(int(maxAvar)+12-1)
                cfgWarnAvar = '16#'+ (str(maxAvar)+str(maxWarn)+str(minWarn)+str(minAvar))

                if pValue is None: continue
                cfg_txt = f'(* {tag} {name} *)\n' \
                          f'cfgAI[{numbers}].pValue                       REF=\t{pValue};\n'
                if pHealth is not None:
                    cfg_txt = cfg_txt + \
                              f'cfgAI[{numbers}].pHealth                      REF=\t{pHealth};\n' \
                              f'cfgAI[{numbers}].pHealthExt                   REF=\t{pHealth};\n'
                cfg_txt = cfg_txt + \
                          f'cfgAI[{numbers}].cfgWarnAvar.reg                :=\t{cfgWarnAvar};\n' \
                          f'cfgAI[{numbers}].cfgAI.reg                      :=\t{str(cfgAI)};\n' \
                          f'cfgAI[{numbers}].nNA_VS                         :=\t{nNA_VS};\n' \
                          f'cfgAI[{numbers}].nFuse                          :=\t{nFuse};\n'

                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_AI_sim
    @logger.catch
    def gen_cfg_AI_sim(self,path):
        data      = self.data['AI']
        try:
            path_cfg = f'{path}\cfg_AI_sim.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_AI_sim*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_AI_sim*)\n')
            for value in data:
                numbers         = value['№']
                tag             = value['Идентификатор']
                name            = value['Название']
                pValue          = value['УСО, модуль, канал']
                pHealth         = value['Исправность канала']
                if pHealth is None: continue
                ch = str(pValue).split('[')
                modul = str(pHealth).split('[')
                num = f'[{str(modul[1]).replace("]","")}][{str(ch[1]).replace("]","")}]'
                cfg_txt = f'(*{tag} {name}*)\n' \
                          f'cfgAI[{numbers}].pValue\tREF=\tsimAI{num};\n' \

                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_DPS
    @logger.catch
    def gen_cfg_DPS(self, path):
        data = self.data['DPS']
        try:
            path_cfg = f'{path}\cfg_DPS.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DPS*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_DPS*)\n')
            dps_hat_table = ['№', 'Деблокировка', 'Контроль']
            for value in data:
                numbers = value['№']
                pDeblock = value['Деблокировка']
                pControl = value['Контроль']
                if pControl is None: continue
                cfg_txt = f'(*{numbers} ДПС*)\n' \
                        f'cfgDPS[{numbers}].pControl        REF=  {pControl};\n'
                if pDeblock is None: continue
                cfg_txt = f'(*{numbers} ДПС*)\n' \
                        f'cfgDPS[{numbers}].pDeblock        REF=  {pDeblock};\n'
                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # Cfg_KTPR
    @logger.catch
    def gen_cfg_ktpr(self,path):
        data      = self.data['KTPR']
        stateAI = {'Warn':0,
                    'Avar':1,
                    'LTMin':2,
                    'MTMax':3,
                    'AlgNdv':4,
                    'Imit':5,
                    'ExtNdv':6,
                    'Ndv':7,
                    'Init':8}
        stateAIzone = {'rez_0':0,
                        'Min6':1,
                        'Min5':2,
                        'Min4':3,
                        'Min3_IsMT10Perc':4,
                        'Min2_IsNdv2ndParam':5,
                        'Min1_IsHighVibStat':6,
                        'Norm':7,
                        'Max1_IsHighVibStatNMNWRBIT;':8,
                        'Max2_IsHighVibNoStat':9,
                        'Max3_IsAvarVibStat':10,
                        'Max4_IsAvarVibStatNMNWRBIT;':11,
                        'Max5_IsAvarVibNoStat':12,
                        'Max6_IsAvar2Vib':13,
                        'rez_14':14,
                        'rez_15':15}
        stateDI = {'Value':0,
                    'ElInput':1,
                    'O':2,
                    'KZ':3,
                    'NC':4,
                    'Imit':5,
                    'ExtNdv':6,
                    'Ndv':7,
                    'priority1':8,
                    'priority2':9,
                    'priority3':10,
                    'rez_11':11,
                    'rez_12':12,
                    'Front_0_1':13,
                    'Front_1_0':14,
                    'CfgErr':15}
        stateNPS = {'ModeNPSDst':0,
                        'MNSInWork':1,
                        'IsMNSOff':2,
                        'IsNPSModePsl':3,
                        'IsPressureReady':4,
                        'NeNomFeedInterval':5,
                        'OIPHighPressure':6,
                        'KTPR_P':7,
                        'KTPR_M':8,
                        'CSPAlinkOK':9,
                        'CSPAWorkDeny':10,
                        'TSstopped':11,
                        'rez_12':12,
                        'stopDisp':13,
                        'stopCSPA':14,
                        'stopARM':15}
        stateFacility = {'longGasPoint1':0,
                        'longGasPoint2':1,
                        'longGasPoint3':2,
                        'longGasPoint4':3,
                        'longGasPoint5':4,
                        'longGasPoint6':5,
                        'longGasPoint7':6,
                        'longGasPoint8':7,
                        'rez_8':8,
                        'rez_9':9,
                        'rez_10':10,
                        'rez_11':11,
                        'rez_12':12,
                        'rez_13':13,
                        'rez_14':14,
                        'rez_15':15}
        warnFacility = {'warnGasPoint1':0,
                        'warnGasPoint2':1,
                        'warnGasPoint3':2,
                        'warnGasPoint4':3,
                        'warnGasPoint5':4,
                        'warnGasPoint6':5,
                        'warnGasPoint7':6,
                        'warnGasPoint8':7,
                        'rez_8':8,
                        'rez_9':9,
                        'rez_10':10,
                        'rez_11':11,
                        'rez_12':12,
                        'rez_13':13,
                        'rez_14':14,
                        'rez_15':15
                        }
        Facility = {'ndv2Gas':0,
                    'GasLim':1,
                    'GasAv':2,
                    'GasKeep':3,
                    'GasNdvWait':4,
                    'GasLimWait':5,
                    'GasNdvProt':6,
                    'GasAvProt':7,
                    'ColdOn':8,
                    'HotOn':9,
                    'rez_10':10,
                    'rez_11':11,
                    'ColdOff':12,
                    'HotOff':13,
                    'rez_14':14,
                    'rez_15':15}
        vsgrpstate = {'REZ_EXIST':0,
                        'REM':1,
                        'OTKL':2,
                        'OTKL_BY_CMD':3,
                        'VKL_AS_DOP':4,
                        'PUSK_OSN':5,
                        'rez_6':6,
                        'rez_7':7,
                        'rez_8':8,
                        'rez_9':9,
                        'rez_10':10,
                        'rez_11':11,
                        'rez_12':12,
                        'rez_13':13,
                        'LAST_OFF_BY_CMD_ARM ':14,
                        'ALL_OFF_WITH_BLOCK ':15}

        ktpr_cfg = ['Отключение ПНС с выдержкой времени до 5 с после отключения всех МНА','Автоматическая деблокировка защиты','Запрет маскирования']
        ktpr_ctrl1 = ['Закрытие задвижек на входе РП',
                      'Закрытие секущей задвижки узла подключения объекта нефтедобычи/ нефтепереработки',
                      'Закрытие задвижек на входе ФГУ',
                      'Закрытие задвижек на входе ССВД',
                      'Закрытие задвижек на выходе узла РД',
                      'Закрытие задвижек на входе узла РД',
                      'Закрытие задвижек на входе и выходе ПНА',
                      'Закрытие задвижек на входе и выходе МНА',
                      'Закрытие задвижек на входе и выходе ПНС',
                      'Закрытие задвижек на входе и выходе МНС',
                      'Закрытие задвижек между РП и ПНС',
                      'Закрытие задвижек между ПНС и МНС',
                      'Закрытие задвижек на выходе НПС',
                      'Закрытие задвижек на входе НПС']

        ktpr_ctrl2 = ['Отключение вентиляторов водоохлаждения системы оборотного водоснабжения',
                      'Отключение АВО',
                      'Отключение насосов артскважин',
                      'Отключение насосов хозяйственно-питьевого водоснабжения',
                      'Отключение насосов прокачки нефти/нефтепродукта через БИК',
                      'Отключение насосов, обеспечивающих подкачку нефти/нефтепродукта от объектов нефтедобычи/нефтепереработки',
                      'Отключение компрессоров подпора воздуха ЭД',
                      'Отключение подпорных вентиляторов электрозала',
                      'Отключение подпорных вентиляторов ЭД',
                      'Отключение беспромвальных вентиляторов электрозала',
                      'Отключение насосов откачки из емкостей ССВД',
                      'Отключение насосов откачки из емкостей сбора утечек ПНС',
                      'Отключение насосов откачки из емкостей сбора утечек МНС',
                      'Отключение насосов оборотного водоснабжения',
                      'Отключение маслонасосов после сигнала "остановлен" НА',
                      'Отключение маслонасосов']
        ktpr_ctrl3 = ['Отключение приточного вентилятора помещения СИКН',
                      'Отключение приточного вентилятора помещения БИК',
                      'Отключение приточных вентиляторов помещения компрессорной подпора воздуха ЭД и закрытие огнезадерживающих клапанов',
                      'Отключение приточного вентилятора помещения ССВД',
                      'Отключение приточного вентилятора помещения РД',
                      'Отключение приточных вентиляторов в помещении централизованной маслосистемы и закрытие огнезадерживающих клапанов',
                      'Отключение приточных вентиляторов насосного зала ПНС и закрытие огнезадерживающих клапанов',
                      'Отключение приточных вентиляторов насосного зала МНС и закрытие огнезадерживающих клапанов',
                      'Отключение крышных вентиляторов насосного зала ПНС',
                      'Отключение крышных вентиляторов насосного зала МНС',
                      'Отключение вытяжных вентиляторов в помещении ССВД',
                      'Отключение вытяжных вентиляторов в помещении РД',
                      'Отключение вытяжных вентиляторов маслоприямка в электрозале',
                      'Отключение вытяжных вентиляторов в помещении централизованной маслосистемы',
                      'Отключение вытяжных вентиляторов насосного зала ПНС',
                      'Отключение вытяжных вентиляторов насосного зала МНС']
        ktpr_ctrl4 = ['Защита по пожару',
                      'Отключение антиконденсационных электронагревателей ЭД',
                      'Отключение насосов откачки из емкостей сбора утечек всех СИКН',
                      'Отключение насосов прокачки нефти/нефтепродукта через оперативный БИК',
                      'Отключение насосов системы запирания',
                      'Отключение внешнего контура охлаждения ЧРП ПНА',
                      'Отключение внешнего контура охлаждения ЧРП МНА',
                      'Отключение воздушных охладителей системы запирания торцовых уплотнений отключенных НА',
                      'Отключение воздушных охладителей системы запирания торцовых уплотнений всех МНА',
                      'Отключение электронагревателей емкости сбора утечек СИКН',
                      'Отключение электронагревателей емкости сбора утечек ПНС',
                      'Отключение электронагревателей емкости сбора утечек МНС',
                      'Отключение электронагревателей масла',
                      'Закрытие воздушных клапанов (жалюзийных решёток) помещения компрессорной подпора воздуха ЭД',
                      'Закрытие воздушных клапанов (жалюзийных решёток) насосного зала']
        try:
            # (*Функция заполнения данными структуры cfgKTPR - генерируется *)
            # (*
            #  cfgKTPR[1].pInput.pInputVar := stateAI[].stateZone;
            #  cfgKTPR[1].pInput.num := 5;
            #  cfgKTPR[1].pInput.cfg_union.cfg := 16  # 00;
            #  cfgKTPR[1].cfg_union.cfg := 16  # 04;
            #  cfgKTPR[1].Group := 1;
            #  cfgKTPR[1].NA_StopType := 1;
            #  cfgKTPR[1].NS_StopType := 1;
            #  cfgKTPR[1].ctrl := 1;
            #  *)
            path_cfg = f'{path}\cfg_KTPR.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_KTPR*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*Cfg_KTPR*)\n')
            for value in data:
                numbers         = value['№']
                tag             = value['Идентификатор']
                name            = value['Название']
                pInputpInputVar= value['Аварийный параметр']
                pInputnum      = value['№']
                cfg_unioncfg = '0000000000000'
                ctrl1 = '00'
                ctrl2=''
                ctrl3=''
                ctrl4=''
                isNum=0
                isInv=0
                Inputvar = str(pInputpInputVar).split(".")

                # if self.str_find(Inputvar[0],'NOT'):
                #     isInv=1
                #     b = str(Inputvar[0]).replace('NOT ','')
                #     if self.str_find(Inputvar[0], 'NPS'):
                #         isInv = 0
                if self.str_find(Inputvar[0], {'NOT '}):
                    isInv = 1
                b = str(Inputvar[0]).replace('NOT ', '')
                if len(Inputvar)>2:
                    if self.str_find(Inputvar[0],'stateBUF'):
                        pInputnum = Inputvar[2]
                        isNum = 0
                        pInputpInputVar = Inputvar[0] + '.state.reg'
                if len(Inputvar)>1:
                    if Inputvar[1] in vsgrpstate.keys():
                        pInputnum = vsgrpstate[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar = str(Inputvar[0]).replace('VSGRP','stateVSGRP') + '.state.reg'
                    if Inputvar[1] in stateFacility.keys():
                        pInputnum = stateFacility[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar = str(Inputvar[0]).replace('Facility','stateFacility') + '.longGas.reg'
                    if Inputvar[1] in warnFacility.keys():
                        pInputnum = warnFacility[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar = str(Inputvar[0]).replace('Facility','stateFacility') + '.warnGas.reg'
                    if Inputvar[1] in stateAI.keys():
                        pInputnum=stateAI[Inputvar[1]]
                        isNum=0
                        pInputpInputVar=str(Inputvar[0]).replace('AI','StateAI')+'.state.reg'
                    if Inputvar[1] in Facility.keys():
                        pInputnum=Facility[Inputvar[1]]
                        isNum=0
                        pInputpInputVar=str(Inputvar[0]).replace('Facility','stateFacility')+'.state.reg'
                    if Inputvar[1] in stateAIzone.keys():
                        pInputnum=stateAIzone[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar=str(Inputvar[0]).replace('AI','StateAI')+'.stateZone.reg'
                    if Inputvar[1] in stateDI.keys():
                        pInputnum=stateDI[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar=str(Inputvar[0]).replace('DI','StateDI')+'.state.reg'
                    if Inputvar[1] in stateNPS.keys():
                        pInputnum=stateNPS[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar=str(b).replace('NPS','stateNPS')+'.state.reg'


                i=0
                for el in ktpr_cfg:
                    i+=1

                    cfg_unioncfg = cfg_unioncfg+str(value[el]) if value[el] is not None else cfg_unioncfg+'0'
                    if i == 2:
                        cfg_unioncfg = cfg_unioncfg + '0'
                        #print(el, ' ', value[el])
                for el in ktpr_ctrl1:
                    ctrl1 = ctrl1+str(value[el]) if value[el] is not None else ctrl1+'0'
                for el in ktpr_ctrl2:
                    ctrl2 = ctrl2+str(value[el]) if value[el] is not None else ctrl2+'0'
                for el in ktpr_ctrl3:
                    ctrl3 = ctrl3+str(value[el]) if value[el] is not None else ctrl3+'0'
                for el in ktpr_ctrl4:
                    ctrl4 = ctrl4+str(value[el]) if value[el] is not None else ctrl4+'0'
                Group           = value['Битовая маска принадлежности защиты группе'] if value['Битовая маска принадлежности защиты группе'] is not None else '0'
                NA_StopType       = value['Тип остановки НА'] if value['Тип остановки НА'] is not None else '0'
                NS_StopType       = value['Тип остановки насосной станции'] if value['Тип остановки насосной станции'] is not None else '0'
                pInputcfg_unioncfg = '00000000000000'+str(isNum)+str(isInv)
                if tag is None: continue
                cfg_txt = f'(*{tag} {name}*)\n' \
                          f'cfgKTPR[{numbers}].pInput.pInputVar REF={pInputpInputVar};\n' \
                          f'cfgKTPR[{numbers}].pInput.num:={pInputnum};\n' \
                          f"cfgKTPR[{numbers}].pInput.cfg.reg:={str(hex(int(pInputcfg_unioncfg,2))).replace('0x','16#')};\n" \
                          f"cfgKTPR[{numbers}].cfg.reg:={str(hex(int(cfg_unioncfg,2))).replace('0x','16#')};\n" \
                          f'cfgKTPR[{numbers}].Group:={Group};\n' \
                          f'cfgKTPR[{numbers}].NA_StopType:={NA_StopType};\n' \
                          f'cfgKTPR[{numbers}].NS_StopType:={NS_StopType};\n' \
                          f"cfgKTPR[{numbers}].ctrl.ctrl1.reg:={str(hex(int(ctrl1,2))).replace('0x','16#')};\n" \
                          f"cfgKTPR[{numbers}].ctrl.ctrl2.reg:={str(hex(int(ctrl2, 2))).replace('0x','16#')};\n" \
                          f"cfgKTPR[{numbers}].ctrl.ctrl3.reg:={str(hex(int(ctrl3, 2))).replace('0x','16#')};\n" \
                          f"cfgKTPR[{numbers}].ctrl.ctrl4.reg:={str(hex(int(ctrl4, 2))).replace('0x','16#')};\n"
                text_file.write(cfg_txt)
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')
    # gv_diag
    @logger.catch
    def gen_gv_diag(self, path):
        type_mk500 = {'MK-550-024' : 'typePSU',
                      'MK-504-120' : 'typeCPU',
                      'MK-545-010' : 'typeCN',
                      'MK-546-010' : 'typeMN',
                      'МК-516-008A': 'typeAI_8ch',
                      'MK-514-008' : 'typeAO',
                      'MK-521-032' : 'typeDI',
                      'MK-531-032' : 'typeDO',
                      'MK-541-002' : 'typeRS485'}

        wb = openpyxl.load_workbook(self.exel, read_only=True)
        # активный лист таблицы
        sheet      = wb['HW']
        sheet_ss   = wb['SS']
        sheet_tmdp = wb['TM_DP']
        # максимальное число рядов и столбцов
        rows      = sheet.max_row
        column    = sheet.max_column
        rows_ss   = sheet_ss.max_row
        rows_tmdp = sheet_tmdp.max_row
        # Создадим пустой массив для дальнейшего использования
        signals      = []
        signals_ss   = []
        signals_tmdp = []
        # Помещаем пути в одну переменную, чтобы на 2 этапе заполнить их в цикле
        # link_path = MapAI_Ref, MapKlk, MapKont, MapSignalName, MapTagName
        # 1 этап
        # Из табл: HW определим корзины и модули с AI

        countPSUmodules     = 0
        countCPUmodules     = 0
        countCNmodules      = 0
        countMNmodules      = 0
        countAI_16CHmodules = 0
        countAI_8CHmodules  = 0
        countAOmodules      = 0
        countDImodules      = 0
        countDOmodules      = 0
        countRS485modules   = 0

        try:
            for i in range(4, rows + 1):
                for j in range(7, column + 1):
                    cell_ai = sheet.cell(row=i, column=j).value

                    # номер усо, номер модуля для имени, имя усо с корзиной англ,
                    # имя усо с корзиной русс, корзина, тип модуля, имя модуля для редактирования
                    number_uso = str(sheet.cell(row=i, column=1).value)
                    abs_num = 0
                    number_modul = str(sheet.cell(row=2, column=j - 1).value)
                    name_uso_eng = str(sheet.cell(row=i, column=3).value)
                    name_uso_rus = str(sheet.cell(row=i, column=4).value)
                    rack = str(sheet.cell(row=i, column=1).value)
                    type_modul = str(sheet.cell(row=i, column=j - 1).value)
                    gvd = str(sheet.cell(row=i, column=j).value)
                    modul_dash = number_modul
                    modul_point = number_modul
                    if type_modul != 'None' and type_modul in type_mk500:
                        if type_mk500[str(type_modul)] == 'typePSU':
                            countPSUmodules += 1
                            abs_num = countPSUmodules
                        if type_mk500[str(type_modul)] == 'typeCPU':
                            countCPUmodules += 1
                            abs_num = countCPUmodules
                        if type_mk500[str(type_modul)] == 'typeCN':
                            countCNmodules += 1
                            abs_num = countCNmodules
                        if type_mk500[str(type_modul)] == 'typeMN':
                            countMNmodules += 1
                            abs_num = countMNmodules
                        if type_mk500[str(type_modul)] == 'typeAI_8ch':
                            countAI_8CHmodules += 1
                            abs_num = countAI_8CHmodules
                        if type_mk500[str(type_modul)] == 'typeAO':
                            countAOmodules += 1
                            abs_num = countAOmodules
                        if type_mk500[str(type_modul)] == 'typeDI':
                            countDImodules += 1
                            abs_num = countDImodules
                        if type_mk500[str(type_modul)] == 'typeDO':
                            countDOmodules += 1
                            abs_num = countDOmodules
                        if type_mk500[str(type_modul)] == 'typeRS485':
                            countRS485modules += 1
                            abs_num = countRS485modules

                    if self.str_find(modul_dash, {'_0', '_'}):
                        modul_dash = str(modul_dash).replace('_0', '').replace('_', '')
                    if self.str_find(modul_point, {'_'}):
                        modul_point = str(modul_point).replace('_', '.')
                    uso_rack_modul = name_uso_eng + number_modul

                    # Заполняем словарь с исходными данными
                    if type_modul != 'None' and type_modul in type_mk500:
                        a_dict = dict(name_uso_rus=name_uso_rus,
                                      uso_rack_modul=uso_rack_modul,
                                      rack=rack,
                                      number_modul=number_modul,
                                      modul_dash=modul_dash,
                                      gvd=gvd,
                                      abs_num=abs_num,
                                      type_modul=type_mk500[str(type_modul)])
                        signals.append(a_dict)
            for i in signals:
                print(i)
            # Смежные системы
            for i in range(4, rows_ss + 1):
                number_ss  = sheet_ss.cell(row=i, column=1).value
                name_ss    = sheet_ss.cell(row=i, column=4).value
                count_ss   = sheet_ss.cell(row=i, column=5).value
                array1_req = sheet_ss.cell(row=i, column=6).value
                num1_nom   = sheet_ss.cell(row=i, column=7).value
                chann1_nom = sheet_ss.cell(row=i, column=8).value
                count_SS   = sheet_ss.cell(row=i, column=9).value
                array2_req = sheet_ss.cell(row=i, column=10).value
                num2_nom   = sheet_ss.cell(row=i, column=11).value
                chann2_nom = sheet_ss.cell(row=i, column=12).value
                time_out   = sheet_ss.cell(row=i, column=14).value

                if number_ss is None: continue
                if name_ss   is None: continue

                signals_ss.append(dict(number_ss  = number_ss,
                                       name_ss    = name_ss,
                                       count_ss   = count_ss,
                                       array1_req = array1_req,
                                       num1_nom   = num1_nom,
                                       chann1_nom = chann1_nom,
                                       count_SS   = count_SS,
                                       array2_req = array2_req,
                                       num2_nom   = num2_nom,
                                       chann2_nom = chann2_nom,
                                       time_out   = time_out))
            # Диспетчерские пункты
            for i in range(4, rows_tmdp + 1):
                number        = sheet_tmdp.cell(row=i, column=1).value
                variable      = sheet_tmdp.cell(row=i, column=2).value
                name          = sheet_tmdp.cell(row=i, column=4).value
                connection    = sheet_tmdp.cell(row=i, column=5).value
                link_time_out = sheet_tmdp.cell(row=i, column=6).value

                if number is None: continue
                if name   is None: continue

                signals_tmdp.append(dict(number       = number,
                                         variable     = variable,
                                         name         = name,
                                         link_time_out= link_time_out,
                                         connection   = connection))

            path_cfg = f'{path}\gv_diag.txt'
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ gv_diag*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ gv_diag*)\n')

            text_file.write('VAR_GLOBAL CONSTANT\n' \
                            f'\tcountPSUmodules\t: UINT := {countPSUmodules};\t(* Количество модулей PSU *)\n' \
                            f'\tcountCPUmodules\t: UINT := {countCPUmodules};\t(* Количество модулей CPU *)\n' \
                            f'\tcountCNmodules\t: UINT := {countCNmodules};\t(* Количество модулей CN *)\n' \
                            f'\tcountMNmodules\t: UINT := {countMNmodules};\t(* Количество модулей MN *)\n' \
                            f'\tcountAI16modules\t: UINT := {countAI_16CHmodules};\t(* Количество 16-ти канальных модулей AI *)\n' \
                            f'\tcountAI8modules\t: UINT := {countAI_8CHmodules};\t(* Количество 8-ми канальных модулей AI *)\n' \
                            f'\tcountAOmodules\t: UINT := {countAOmodules};\t(* Количество модулей AO *)\n' \
                            f'\tcountDImodules\t: UINT := {countDImodules};\t(* Количество модулей DI *)\n' \
                            f'\tcountDOmodules\t: UINT := {countDOmodules};\t(* Количество модулей DO *)\n' \
                            f'\tcountRS485modules\t: UINT := {countRS485modules};\t(* Количество интерфейсных модулей *)\n' \
                            f'\tcontRack:=\t: UINT := {countPSUmodules};\t(* Количество всех корзин *)\n' \
                            f'\tcontSS:=\t: UINT := {countRS485modules};\t(* Количество смежных систем - ПОКА НЕ ГЕНЕРИТСЯ *)\n' \
                            f'END_VAR\n' \
                            'VAR_GLOBAL\n')

            mbusexist = {}
            for i in range(1, 33):
                count = -1
                for value in signals:
                    if value['rack'] == str(i):
                        count = count + 1
                mbusexist[i] = count

            for value in signals:
                uso_rack_modul = value['uso_rack_modul']
                type_modul = value['type_modul']
                if type_modul in ['typeAI_8ch', 'typeAO', 'typeDI', 'typeDO']:
                    if type_modul == 'typeAI_8ch':
                        t = 'AI'
                        count = '8'
                        dData = 'UINT'
                    if type_modul == 'typeDI':
                        t = 'DI'
                        count = '32'
                        dData = 'BOOL'
                    if type_modul == 'typeDO':
                        t = 'DO'
                        count = '32'
                        dData = 'BOOL'
                    if type_modul == 'typeAO':
                        t = 'AO'
                        count = '8'
                        dData = 'UINT'
                    if type_modul == 'typeAO':
                        cfg_txt = f'\t{uso_rack_modul}_Diagnostics: NftIOItfs.IOModuleDiag;\n' \
                                  f'\t{uso_rack_modul}_{t}: ARRAY[1..{count}] OF {dData};\n' \
                                  f'\t{uso_rack_modul}_{t}_Statuses: ARRAY[1..{count}] OF USINT;\n' \
                                  '\t(*-----------------*)\n'
                    else:
                        cfg_txt = f'\t{uso_rack_modul}_Diagnostics: NftIOItfs.IOModuleDiag;\n' \
                                  f'\t{uso_rack_modul}_{t}: ARRAY[1..{count}] OF {dData};\n' \
                                  '\t(*-----------------*)\n'
                    text_file.write(cfg_txt)
                else:

                    cfg_txt = f"\t{uso_rack_modul}:{type_modul};\n"
                    text_file.write(cfg_txt)

            text_file.write('\tCfgDiag:       typeCfgDiag;\t(* Конфигурация диагностики *)\n' \
                            '\tStateDiag:     typeStateDiag;\t(* Данные по диагностике, передаваемые на ВУ *)\n' \
                            '\tMemDiag:       typeMemDiag;\t(* Хранилище данных по диагностике *)\n' \
                            '\tCmdDiag:       WORD;\t(* Команда для диагностики *)\n' \
                            '\tmAI16_HEALTH:  ARRAY[1..countAI16modules] \tOF BOOL;\n' \
                            '\tmAI8_HEALTH:   ARRAY[1..countAI8modules] \tOF BOOL;\n' \
                            '\tmAO_HEALTH:    ARRAY[1..countAOmodules] \tOF BOOL;\n' \
                            '\tmDI_HEALTH:    ARRAY[1..countDImodules] \tOF BOOL;\n' \
                            '\tmDO_HEALTH:    ARRAY[1..countDOmodules] \tOF BOOL;\n' \
                            '\tprjVersion:    typePrjVersion;\n' \
                            '\tprjVersionKvit:typePrjVersion;\n' \
                            'END_VAR\n')
            text_file.close()

            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            path_cfg = f'{path}\cfg_diag.txt'
            if not os.path.exists(path_cfg):
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_diag*)\n')
            else:
                os.remove(path_cfg)
                text_file = open(path_cfg, 'w')
                text_file.write('(*cfg_diag*)\n' \
                                'IF NOT flInit THEN\n')

            for key, value in mbusexist.items():
                if value != -1:
                    st_bin = 0
                    for b in range(0, value + 1):
                        st_bin = st_bin + 2 ** b
                    text_file.write(f"\tCfgDiag.mBUSExists[{key}] := {str(bin(st_bin)).replace('0b', '2#')};\n")

            for mod in signals:
                PortsEnbl = []
                if mod['type_modul'] == 'typePSU':
                    text_file.write(
                        f"\tCfgDiag.PSU[{mod['abs_num']}].mPSU                   REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.PSU[{mod['abs_num']}].nRack                    := {mod['rack']};\n" \
                        f"\tCfgDiag.PSU[{mod['abs_num']}].nMod                     := {mod['modul_dash']};\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeCPU':
                    PortsEnbl = str(mod['gvd']).split(';')
                    pe = PortsEnbl[1] if len(PortsEnbl) > 1 else 7
                    text_file.write(
                        f"\tCfgDiag.CPU[{mod['abs_num']}].mCPU                   REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.CPU[{mod['abs_num']}].nRack                    := {mod['rack']};\n" \
                        f"\tCfgDiag.CPU[{mod['abs_num']}].nMod                     := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.CPU[{mod['abs_num']}].PortsEnbl                := {pe};\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeCN':
                    PortsEnbl = str(mod['gvd']).split(';')
                    pe = PortsEnbl[1] if len(PortsEnbl) > 1 else 3
                    text_file.write(
                        f"\tCfgDiag.CN[{mod['abs_num']}].mCN                     REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.CN[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.CN[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.CN[{mod['abs_num']}].PortsEnbl                 := {pe};\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeMN':
                    PortsEnbl = str(mod['gvd']).split(';')
                    pe = PortsEnbl[1] if len(PortsEnbl) > 1 else 3
                    text_file.write(
                        f"\tCfgDiag.MN[{mod['abs_num']}].mMN                     REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.MN[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.MN[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.MN[{mod['abs_num']}].PortsEnbl                 := {pe};\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeAI_8ch':
                    text_file.write(
                        f"\tCfgDiag.AI8[{mod['abs_num']}].mAI8.AI                REF= {mod['uso_rack_modul']}_AI;\n" \
                        f"\tCfgDiag.AI8[{mod['abs_num']}].nRack                    := {mod['rack']};\n" \
                        f"\tCfgDiag.AI8[{mod['abs_num']}].nMod                     := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.AI8[{mod['abs_num']}].mAI8.Diagnostics       REF= {mod['uso_rack_modul']}_Diagnostics;\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeAO':
                    text_file.write(
                        f"\tCfgDiag.AO[{mod['abs_num']}].mAO.AO                  REF= {mod['uso_rack_modul']}_AO;\n" \
                        f"\tCfgDiag.AO[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.AO[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.AO[{mod['abs_num']}].mAO.Diagnostics         REF= {mod['uso_rack_modul']}_Diagnostics;\n" \
                        f"\tCfgDiag.AO[{mod['abs_num']}].mAO.AOStatuses          REF= {mod['uso_rack_modul']}_AO_Statuses;\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeDI':
                    text_file.write(
                        f"\tCfgDiag.DI[{mod['abs_num']}].mDI.DI                  REF= {mod['uso_rack_modul']}_DI;\n" \
                        f"\tCfgDiag.DI[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.DI[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.DI[{mod['abs_num']}].mDI.Diagnostics         REF= {mod['uso_rack_modul']}_Diagnostics;\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeDO':
                    text_file.write(
                        f"\tCfgDiag.DOs[{mod['abs_num']}].mDO.DOs                 REF= {mod['uso_rack_modul']}_DO;\n" \
                        f"\tCfgDiag.DOs[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.DOs[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.DOs[{mod['abs_num']}].mDO.Diagnostics         REF= {mod['uso_rack_modul']}_Diagnostics;\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeRS485':
                    PortsEnbl = str(mod['gvd']).split(';')
                    pe = PortsEnbl[1] if len(PortsEnbl) > 1 else 3
                    text_file.write(
                        f"\tCfgDiag.RS485[{mod['abs_num']}].mRS485               REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.RS485[{mod['abs_num']}].nRack                  := {mod['rack']};\n" \
                        f"\tCfgDiag.RS485[{mod['abs_num']}].nMod                   := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.RS485[{mod['abs_num']}].PortsEnbl              := {pe};\n" \
                        '(*---------------------*)\n')

            for data in signals_ss:
                number_ss  = data['number_ss']
                name_ss    = data['name_ss']
                count_ss   = data['count_ss']
                array1_req = data['array1_req']
                num1_nom   = data['num1_nom']
                chann1_nom = data['chann1_nom']
                count_SS   = data['count_SS']
                array2_req = data['array2_req']
                num2_nom   = data['num2_nom']
                chann2_nom = data['chann2_nom']
                time_out   = data['time_out']

                if time_out is None: time_out = 'tmCommon.Diag_RStimeout'
                else               : time_out = f'REF= {time_out}'

                if (not num1_nom is None) and (not chann1_nom is None):
                    text_file.write(
                        f'(*------{name_ss}--------*)\n'
                        f"\tcfgDiag.SS[{number_ss}, 1].pTimeOut	         REF= {time_out};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].RSreq                   := {array1_req}; \n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].nRS                     := {num1_nom};   \n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].chRS                    := {chann1_nom}; \n")
                elif not count_ss is None:
                    text_file.write(
                        f'(*------{name_ss}--------*)\n'
                        f"\tcfgDiag.SS[{number_ss}, 1].pTimeOut	         REF= {time_out};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].pCounter	         REF= {count_ss};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].RSreq                   := 0; \n")

                if (not num2_nom is None) and (not chann2_nom is None):
                    text_file.write(
                        f'(*------{name_ss}--------*)\n'
                        f"\tcfgDiag.SS[{number_ss}, 2].pTimeOut	         REF= {time_out};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].RSreq                   := {array2_req}; \n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].nRS                     := {num2_nom};   \n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].chRS                    := {chann2_nom}; \n")
                elif not count_SS is None:
                    text_file.write(
                        f'(*------{name_ss}--------*)\n'
                        f"\tcfgDiag.SS[{number_ss}, 2].pTimeOut	         REF= {time_out};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].pCounter	         REF= {count_SS};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].RSreq                   := 0; \n")

            for data in signals_tmdp:
                number        = data['number']
                variable      = data['variable']
                name          = data['name']
                link_time_out = data['link_time_out']
                connection    = data['connection']

                if link_time_out is None: link_time_out = 'tmCommon.CSPA_t1'

                InputVar, Num, Unioncfg_st = self.ret_inp_cfg(connection)

                text_file.write(
                    f'(*------{name}--------*)\n'
                    f"\tcfgDiag.TM_DP[{number}].pLinkOk.pInputVar    REF= {InputVar};\n" \
                    f"\tcfgDiag.TM_DP[{number}].pLinkOk.num               := {Num}; \n" \
                    f"\tcfgDiag.TM_DP[{number}].pLinkOk.cfg.reg           := {Unioncfg_st};   \n" \
                    f"\tcfgDiag.TM_DP[{number}].pTimeOut                  := {link_time_out}; \n")


            text_file.write("\tflInit := TRUE;\n" \
                            "END_IF;")
            text_file.close()
            logger.info(f'{self.name_prefix} выполнено {path_cfg}')
        except:
            logger.error(f'{self.name_prefix} FAILED')

    # TM
    # TS
    def gen_cfg_TS(self, path):
        data_ts = self.data['TM_TS']

        path_trend   = f'{path}\cfg_TM_TS.txt'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TS*)\n')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TS*)\n')

        for data in data_ts:
            number    = data['№']
            name      = data['Название']
            reference = data['Ссылка на значние']
            adress    = data['Адрес объекта']

            pInput, pnum, pcfg = self.ret_inp_cfg(reference)

            if pInput != 0:
                cfg_txt = f'(* {adress} - {name} *)\n' \
                          f'cfgTM_TS[{number}].pInputVar              REF={str(pInput)};\n' \
                          f'cfgTM_TS[{number}].num                      :={str(pnum)};\n' \
                          f'cfgTM_TS[{number}].cfg.reg                  :={str(pcfg)};\n'
                text_file.write(cfg_txt)
        text_file.close()
        logger.info(f'Выполнено. Генерация файла TM_TS')
        return (f'Выполнено. Генерация файла TM_TS')
    # TU
    def gen_cfg_TU(self, path):
        data_tu = self.data['TM_TU']

        path_trend   = f'{path}\cfg_TM_TU.txt'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TU*)\n')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TU*)\n')

        for data in data_tu:
            number  = data['№']
            name    = data['Название']
            mut_var = data['Изменяемая переменная']
            bits    = data['Изменяемый бит']
            adress  = data['Адрес объекта']

            if mut_var is None: continue
            if bits    is None: continue

            cfg_txt = f'(* {adress} - {name} *)\n' \
                      f'cfgTM_TU[{number}].pVal              REF={str(mut_var)}.reg;\n' \
                      f'cfgTM_TU[{number}].iBit                :={str(bits)};\n'
            text_file.write(cfg_txt)
        text_file.close()
        logger.info(f'Выполнено. Генерация файла TM_TU')
        return (f'Выполнено. Генерация файла TM_TU')
    # TI2
    def gen_cfg_TI2(self, path):
        data_ti2 = self.data['TM_TI2']

        path_trend   = f'{path}\cfg_TM_TI2.txt'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TI2*)\n')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TI2*)\n')

        for data in data_ti2:
            number  = data['№']
            name    = data['Название']
            value   = data['Переменная - значение']
            status  = data['Переменная - статус']
            adress  = data['Адрес объекта']

            if name is None: continue
            if status is None:
                cfg_txt = f'(* {adress} - {name} *)\n' \
                          f'cfgTM_TI2[{number}].pVal              REF={str(value)};\n'
            else:
                cfg_txt = f'(* {adress} - {name} *)\n' \
                          f'cfgTM_TI2[{number}].pVal              REF={str(value)};\n' \
                          f'cfgTM_TI2[{number}].pState            REF={str(status)};\n'
            text_file.write(cfg_txt)
        text_file.close()
        logger.info(f'Выполнено. Генерация файла TM_TI2')
        return (f'Выполнено. Генерация файла TM_TI2')
    # TI4
    def gen_cfg_TI4(self, path):
        data_ti4 = self.data['TM_TI4']

        path_trend   = f'{path}\cfg_TM_TI4.txt'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TI4*)\n')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TI4*)\n')

        for data in data_ti4:
            number   = data['№']
            name     = data['Название']
            value    = data['Переменная - значение']
            status   = data['Переменная - статус']
            ai_param = data['Переменная - Aiparam']
            adress   = data['Адрес объекта']

            if value is None: continue

            if (not status is None) and (not ai_param is None):
                cfg_txt = f'(* {adress} - {name} *)\n' \
                          f'cfgTM_TI4[{number}].pVal              REF={str(value)};\n' \
                          f'cfgTM_TI4[{number}].pState            REF={str(status)}.reg;\n' \
                          f'cfgTM_TI4[{number}].pAIparam          REF={str(ai_param)};\n'
            if (status is None) and (not ai_param is None):
                cfg_txt = f'(* {adress} - {name} *)\n' \
                          f'cfgTM_TI4[{number}].pVal              REF={str(value)};\n' \
                          f'cfgTM_TI4[{number}].pAIparam          REF={str(ai_param)};\n'
            if (not status is None) and (ai_param is None):
                cfg_txt = f'(* {adress} - {name} *)\n' \
                          f'cfgTM_TI4[{number}].pVal              REF={str(value)};\n' \
                          f'cfgTM_TI4[{number}].pState            REF={str(status)}.reg;\n'
            if (status is None) and (ai_param is None):
                cfg_txt = f'(* {adress} - {name} *)\n' \
                          f'cfgTM_TI4[{number}].pVal              REF={str(value)};\n'

            text_file.write(cfg_txt)
        text_file.close()
        logger.info(f'Выполнено. Генерация файла TM_TI4')
        return (f'Выполнено. Генерация файла TM_TI4')
    # TII
    def gen_cfg_TII(self, path):
        data_tii = self.data['TM_TII']

        path_trend   = f'{path}\cfg_TM_TII.txt'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TII*)\n')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TII*)\n')

        for data in data_tii:
            number = data['№']
            name   = data['Название']
            value  = data['Переменная - значение']
            status = data['Переменная - статус']
            adress = data['Адрес объекта']

            if value is None: continue

            if status is None:
                cfg_txt = f'(* {name} - {adress} *)\n' \
                          f'cfgTM_TII[{number}].pVal              REF={str(value)};\n'
            else:
                cfg_txt = f'(* {name} - {adress} *)\n' \
                          f'cfgTM_TII[{number}].pVal              REF={str(value)};\n' \
                          f'cfgTM_TII[{number}].pState            REF={str(status)};\n'

            text_file.write(cfg_txt)
        text_file.close()
        logger.info(f'Выполнено. Генерация файла TM_TII')
        return (f'Выполнено. Генерация файла TM_TII')
    # TR2
    def gen_cfg_TR2(self, path):
        data_tr2 = self.data['TM_TR2']

        path_trend   = f'{path}\cfg_TM_TR2.txt'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TR2*)\n')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TR2*)\n')

        for data in data_tr2:
            number = data['№']
            name   = data['Название']
            value  = data['Изменяемая переменная']
            sign   = data['descriptionTR4 (не более 16 символов латиницы)']
            adress = data['Адрес объекта']

            if value is None: continue

            cfg_txt = f'(* {adress} - {name} - {sign} *)\n' \
                      f'cfgTM_TR2[{number}].pVal              REF={str(value)};\n'

            text_file.write(cfg_txt)
        text_file.close()
        logger.info(f'Выполнено. Генерация файла TM_TR2')
        return (f'Выполнено. Генерация файла TM_TR2')
    # TR4
    def gen_cfg_TR4(self, path):
        data_tr4 = self.data['TM_TR4']

        path_trend   = f'{path}\cfg_TM_TR4.txt'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_trend):
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TR4*)\n')
        else:
            os.remove(path_trend)
            text_file = codecs.open(path_trend, 'w', 'utf-8')
            text_file.write('(*Cfg_TM_TR4*)\n')

        for data in data_tr4:
            number = data['№']
            name   = data['Название']
            value  = data['Изменяемая переменная']
            sign   = data['descriptionTR4 (не более 16 символов латиницы)']
            adress = data['Адрес объекта']

            if value is None: continue

            cfg_txt = f'(* {adress} - {name} - {sign} *)\n' \
                      f'cfgTM_TR4[{number}].pVal              REF={str(value)};\n'

            text_file.write(cfg_txt)
        text_file.close()
        logger.info(f'Выполнено. Генерация файла TM_TR4')
        return (f'Выполнено. Генерация файла TM_TR4')

    # Сборка файла xml для имитатора
    def file_xml_imitator(self, path, mb_AI, mb_DI):
        path_imit= f'{path}\Imitation_file.xml'
        # Проверяем файл на наличие в папке, если есть удаляем, и создаем новый
        if not os.path.exists(path_imit):
            text_file = codecs.open(path_imit, 'w', 'utf-8')
            text_file.write('<?xml version="1.0" encoding="UTF-8"?>\n<Signals>\n</Signals>')
        else:
            os.remove(path_imit)
            text_file = codecs.open(path_imit, 'w', 'utf-8')
            text_file.write('<?xml version="1.0" encoding="UTF-8"?>\n<Signals>\n</Signals>')
        text_file.close()
        # Собираем данные
        try:
            wb = openpyxl.load_workbook(self.exel, read_only=True)
            sheet     = wb['HW']
            signals   = []
            count_reg = 0

            for data_AI in self.data['AI']:
                numb_AI = data_AI['№']
                name_AI = data_AI['Название']
                tag     = self.translate(str(data_AI['Идентификатор']))
                Min_ADC = data_AI['Пол. мин.']
                Min_EGU = data_AI['Инж. Мин.']
                Max_ADC = data_AI['Пол. макс.']
                Max_EGU = data_AI['Инж. Макс.']

                numb_rack_KD = ''
                numb_modl_KD = ''
                name_uso_KD  = ''
                numb_chan_KD = ''
                for signal in self.data['КД']:
                    tag_KD  = self.translate(str((signal['Tэг'])))
                    if tag_KD == tag:
                        name_uso_KD  = signal['Шкаф']
                        numb_rack_KD = signal['Корз']
                        numb_modl_KD = signal['Мод']
                        numb_chan_KD = signal['Кан']
                        break
                
                exit_True = False
                for i in range(4, sheet.max_row + 1):
                    if exit_True: break
                    name_uso_HW  = sheet.cell(row=i, column=4).value
                    numb_rack_HW = sheet.cell(row=i, column=5).value
                    
                    if (name_uso_HW == name_uso_KD) and (numb_rack_KD == numb_rack_HW):
                        
                        for j in range(11, sheet.max_column + 1):
                            type_modul = sheet.cell(row=i, column=j + 1).value
                            
                            if self.str_find(type_modul, {'mAI'}):
                                    numb_modl_HW = str(sheet.cell(row=2, column=j).value).replace('_0', '').replace('_', '')
                                    
                                    if numb_modl_HW == str(numb_modl_KD):
                                        through_num_mod = int((re.findall(r'\d+', type_modul))[0])
                                        exit_True = True

                                        mb_adrr_AI = f'{int(mb_AI) + ((through_num_mod - 1)*8) + (numb_chan_KD-1)}'
                                        signals.append(dict(Index       = numb_AI,
                                                            Type        = 'AI',
                                                            Tag         = tag,
                                                            Description = name_AI,
                                                            Cabinet     = name_uso_KD,
                                                            MBAddr_AI   = mb_adrr_AI,
                                                            Min_ADC     = Min_ADC,
                                                            Min_EGU     = Min_EGU,
                                                            Max_ADC     = Max_ADC,
                                                            Max_EGU     = Max_EGU))
                                        count_reg += 1
                                        break

            for data_DI in self.data['DI']:
                numb_DI = data_DI['№']
                name_DI = data_DI['Название']
                tag     = self.translate(str(data_DI['Идентификатор']))
                
                numb_rack_KD = ''
                numb_modl_KD = ''
                name_uso_KD  = ''
                numb_chan_KD = ''
                for signal in self.data['КД']:
                    tag_KD  = self.translate(str((signal['Tэг'])))
                    if tag_KD == tag:
                        name_uso_KD  = signal['Шкаф']
                        numb_rack_KD = signal['Корз']
                        numb_modl_KD = signal['Мод']
                        numb_chan_KD = signal['Кан']
                        break
                
                exit_True = False
                for i in range(4, sheet.max_row + 1):
                    if exit_True: break
                    name_uso_HW  = sheet.cell(row=i, column=4).value
                    numb_rack_HW = sheet.cell(row=i, column=5).value
                    
                    if (name_uso_HW == name_uso_KD) and (numb_rack_KD == numb_rack_HW):
                        
                        for j in range(11, sheet.max_column + 1):
                            type_modul = sheet.cell(row=i, column=j + 1).value
                            
                            if self.str_find(type_modul, {'mDI'}):
                                    numb_modl_HW = str(sheet.cell(row=2, column=j).value).replace('_0', '').replace('_', '')
                                    
                                    if numb_modl_HW == str(numb_modl_KD):
                                        through_num_mod = int((re.findall(r'\d+', type_modul))[0])

                                        if (numb_chan_KD - 1) > 15: 
                                            num_chan_rasch = (numb_chan_KD - 1) - 16
                                            num_reg_rasch  = 2 * (through_num_mod - 1) + 1
                                        else: 
                                            num_chan_rasch = numb_chan_KD - 1
                                            num_reg_rasch  = (2 * (through_num_mod - 1))
                                        
                                        exit_True = True

                                        mb_adrr_DI = f'{int(mb_DI) + num_reg_rasch}.{num_chan_rasch}'
                                        signals.append(dict(Index       = numb_DI,
                                                            Type        = 'DI',
                                                            Tag         = tag,
                                                            Description = name_DI,
                                                            Cabinet     = name_uso_KD,
                                                            MBAddr_DI   = mb_adrr_DI))
                                        break

            # Формируем файл
            object_one = etree.Element('Signals')
            for data in signals:
                object_two = etree.SubElement(object_one, 'Signal')
                type_signal = data['Type']
                if type_signal == 'DI':
                    Index_DI       = data['Index']
                    Tag_DI         = data['Tag']
                    Description_DI = data['Description']
                    Cabinet_DI     = data['Cabinet']
                    MBAddr_DI      = data['MBAddr_DI']
                    etree.SubElement(object_two, 'Index').text       = f'{Index_DI}'
                    etree.SubElement(object_two, 'Type').text        = f'{type_signal}'
                    etree.SubElement(object_two, 'Tag').text         = f'{Tag_DI}'
                    etree.SubElement(object_two, 'Description').text = f'{Description_DI}'
                    etree.SubElement(object_two, 'Cabinet').text     = f'{Cabinet_DI}'
                    etree.SubElement(object_two, 'MBAddr').text      = f'{MBAddr_DI}'
                elif type_signal == 'AI':
                    Index_AI       = data['Index']
                    Tag_AI         = data['Tag']
                    Description_AI = data['Description']
                    Cabinet_AI     = data['Cabinet']
                    MBAddr_AI      = data['MBAddr_AI']
                    Min_ADC        = data['Min_ADC']
                    Min_EGU        = data['Min_EGU']
                    Max_ADC        = data['Max_ADC']
                    Max_EGU        = data['Max_EGU']
                    etree.SubElement(object_two, 'Index').text       = f'{Index_AI}'
                    etree.SubElement(object_two, 'Type').text        = f'{type_signal}'
                    etree.SubElement(object_two, 'Tag').text         = f'{Tag_AI}'
                    etree.SubElement(object_two, 'Description').text = f'{Description_AI}'
                    etree.SubElement(object_two, 'Cabinet').text     = f'{Cabinet_AI}'
                    etree.SubElement(object_two, 'MBAddr').text      = f'{MBAddr_AI}'
                    object_three = etree.SubElement(object_two, 'Min')
                    etree.SubElement(object_three, 'ADC').text = f'{Min_ADC}'
                    etree.SubElement(object_three, 'EGU').text = f'{Min_EGU}'
                    object_three = etree.SubElement(object_two, 'Max')
                    etree.SubElement(object_three, 'ADC').text = f'{Max_ADC}'
                    etree.SubElement(object_three, 'EGU').text = f'{Max_EGU}'
                etree.tostring(object_one, pretty_print=True).decode('utf-8')
            etree.ElementTree(object_one).write(path_imit)

            logger.info(f'Выполнено. Генерация файла имитатора')
            return (f'Выполнено. Генерация файла имитатора')
        except:
            logger.info(f'Ошибка. Генерация файла имитатора')
            return (f'Ошибка. Генерация файла имитатора')
    # Формирование списка коротких сигналов
    def search_ts_id(self, path):
        path = f'{path}\TS_ID.txt'
        if not os.path.exists(path):
            text_file = codecs.open(path, 'w', 'utf-8')
        else:
            os.remove(path)
            text_file = codecs.open(path, 'w', 'utf-8')

        for data_DI in self.data['DI']:
            ts_id   = data_DI['TS_ID']
            name_DI = data_DI['Название']
            
            if ts_id is None: continue

            for data_KD in self.data['КД']:
                name_KD  = data_KD['Наименование']
                if name_DI == name_KD:
                    name_uso_KD  = data_KD['Шкаф']
                    tag          = data_KD['Tэг']
                    pole_KD      = data_KD['КлК']
                    kl_KD        = data_KD['Конт']

                    if pole_KD is None: pole_KD = f'Клеммник отсутствует'
                    if kl_KD   is None:   kl_KD = f'Контакты отсутствуют'

                    cfg_txt = f'{tag} - {name_KD} - {name_uso_KD} - {pole_KD} - {kl_KD}\n'
                    text_file.write(cfg_txt)
                    break
        text_file.close()
    # Тренды для ДМЗ
    def dmz_trends_tree(self, path):
        path = f'{path}\Tree_PT.json'
        data = {}
        data['UserTree'] = []
        # Проходим по всем наименованиям и тегам, которые получили
        for data_AI in self.data['AI']:
            name        = data_AI['Название']
            tag         = data_AI['Идентификатор']
            equ         = data_AI['Единица измерения']
            group_trend = data_AI['Группа сброса трендов']

            if group_trend is None: continue

            for data_GRP in self.data['TrendGRP']:
                id       = data_GRP['ID']
                name_GRP = data_GRP['Название группы']

                if str(id) == str(group_trend):
                    data["UserTree"].append({"Signal": {"UserTree"   : f'ПТ-1/{name_GRP}/{name}',
                                                        "OpcTag"     : f'Root_PT.Analogs_for_trends.{tag}.AIVisualValue',
                                                        "EUnit"      : equ,
                                                        "Description": name,
                                                        "DataType"   : 'float'}})
        with open(path, 'w', encoding='utf-8') as outfile:
            json.dump(data, outfile, ensure_ascii=False, indent=4)
        print('ДМЗ.Дерево сигналов успешно создано')



# Создание шаблона защит
def defence_gen(path_save):
    sample_defence = (f'<type access-modifier="private" name="name" display-name="name" uuid="uuid" base-type="" base-type-id="ffaf5544-6200-45f4-87ec-9dd24558a9d5" ver="5">\n'
        f'\t<designed target="X" value="0" ver="5"/>\n'
        f'\t<designed target="Y" value="0" ver="5"/>\n'
        f'\t<designed target="Rotation" value="0" ver="5"/>\n'
        f'\t<designed target="Width" value="coordinate_W" ver="5"/>\n'
        f'\t<designed target="Height" value="coordinate_H" ver="5"/>\n'
        f'\t<designed target="ZValue" value="0" ver="5"/>\n'
        f'\t<designed target="Scale" value="1" ver="5"/>\n'
        f'\t<designed target="Visible" value="true" ver="5"/>\n'
        f'\t<designed target="Opacity" value="1" ver="5"/>\n'
        f'\t<designed target="Enabled" value="true" ver="5"/>\n'
        f'\t<designed target="Tooltip" value="" ver="5"/>\n'
        f'\t<designed target="PenColor" value="4278190080" ver="5"/>\n'
        f'\t<designed target="PenStyle" value="0" ver="5"/>\n'
        f'\t<designed target="PenWidth" value="1" ver="5"/>\n'
        f'\t<designed target="BrushColor" value="0xffc4c4c4" ver="5"/>\n'
        f'\t<designed target="BrushStyle" value="1" ver="5"/>\n'
        f'\t<designed target="WindowX" value="0" ver="5"/>\n'
        f'\t<designed target="WindowY" value="0" ver="5"/>\n'
        f'\t<designed target="WindowWidth" value="coordinate_W" ver="5"/>\n'
        f'\t<designed target="WindowHeight" value="coordinate_H" ver="5"/>\n'
        f'\t<designed target="WindowCaption" value="name_list" ver="5"/>\n'
        f'\t<designed target="ShowWindowCaption" value="true" ver="5"/>\n'
        f'\t<designed target="ShowWindowMinimize" value="false" ver="5"/>\n'
        f'\t<designed target="ShowWindowMaximize" value="false" ver="5"/>\n'
        f'\t<designed target="ShowWindowClose" value="true" ver="5"/>\n'
        f'\t<designed target="AlwaysOnTop" value="true" ver="5"/>\n'
        f'\t<designed target="WindowSizeMode" value="0" ver="5"/>\n'
        f'\t<designed target="WindowBorderStyle" value="2" ver="5"/>\n'
        f'\t<designed target="WindowState" value="0" ver="5"/>\n'
        f'\t<designed target="WindowScalingMode" value="0" ver="5"/>\n'
        f'\t<designed target="MonitorNumber" value="0" ver="5"/>\n'
        f'\t<designed target="WindowPosition" value="2" ver="5"/>\n'
        f'\t<designed target="WindowCloseMode" value="0" ver="5"/>\n'
        f'\t<object access-modifier="private" name="defences" display-name="defences" uuid="ea39b4b8-d4c9-4c33-84d8-78d742815125" base-type="ApSource" base-type-id="966603da-f05e-4b4d-8ef0-919efbf8ab2c" ver="5">\n'
        f'\t\t<designed target="Path" value="designed_path" ver="5"/>\n'
        f'\t\t<designed target="Active" value="true" ver="5"/>\n'
        f'\t\t<designed target="ReAdvise" value="0" ver="5"/>\n'
        f'\t\t<init target="ParentSource" ver="5" ref="unit.Global.global_ApSource"/>\n'
        f'\t</object>\n'
        f'\t<object access-modifier="private" name="empty_link" display-name="empty_link" uuid="29b864a8-2dd8-4865-a913-59f8d4b1f783" base-type="type_defence_top" base-type-id="6b175e7c-6060-4e11-a416-88a851f6b4a5" ver="5">\n'
        f'\t\t<designed target="X" value="1" ver="5"/>\n'
        f'\t\t<designed target="Y" value="1" ver="5"/>\n'
        f'\t\t<designed target="Rotation" value="0" ver="5"/>\n'
        f'\t\t<designed target="Width" value="800" ver="5"/>\n'
        f'\t\t<designed target="Height" value="26" ver="5"/>\n'
        f'\t\t<designed target="Visible" value="false" ver="5"/>\n'
        f'\t</object>\n'
        f'\t<designed target="WindowIconPath" value="tn_logo.jpg" ver="5"/>\n'
        f'\t<object access-modifier="private" name="name" display-name="name" uuid="32d80c0a-1695-425e-8bfd-edcb1141713d" base-type="Window" base-type-id="04615219-28bb-4a9a-bba4-50ac66972eb0" ver="5" description="" cardinal="1">\n'
        f'\t\t<do-on access-modifier="private" name="Handler_1" display-name="Handler_1" ver="5" event="MessageReceived" form-action="close"/>\n'
        f'\t\t<init target="Group" ver="5" ref="_Control"/>\n'
        f'\t</object>\n'
        f'</type>\n')
    path_defence = f'{path_save}\Form_Defences_default.omobj'
    # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
    if not os.path.exists(path_defence):
        file = codecs.open(path_defence, 'w', 'utf-8')
        file.write(sample_defence)
        file.close()
    else:
        os.remove(path_defence)
        file = codecs.open(path_defence, 'w', 'utf-8')
        file.write(sample_defence)
        file.close()
# Создание шаблона UTS, UPTS
def uts_upts_gen(path_save):
    uts_upts_sample = (
    f'<type access-modifier="private" name="name" display-name="name" uuid="uuid" base-type="" base-type-id="ffaf5544-6200-45f4-87ec-9dd24558a9d5" ver="5">\n'
	f'\t<designed target="X" value="0" ver="5"/>\n'
	f'\t<designed target="Y" value="0" ver="5"/>\n'
	f'\t<designed target="Rotation" value="0" ver="5"/>\n'
	f'\t<designed target="Width" value="coordinate_W" ver="5"/>\n'
	f'\t<designed target="Height" value="coordinate_H" ver="5"/>\n'
	f'\t<designed target="ZValue" value="0" ver="5"/>\n'
	f'\t<designed target="Scale" value="1" ver="5"/>\n'
	f'\t<designed target="Visible" value="true" ver="5"/>\n'
	f'\t<designed target="Opacity" value="1" ver="5"/>\n'
	f'\t<designed target="Enabled" value="true" ver="5"/>\n'
	f'\t<designed target="Tooltip" value="" ver="5"/>\n'
	f'\t<designed target="PenColor" value="4278190080" ver="5"/>\n'
	f'\t<designed target="PenStyle" value="0" ver="5"/>\n'
	f'\t<designed target="PenWidth" value="1" ver="5"/>\n'
	f'\t<designed target="BrushColor" value="0xfff0f0f0" ver="5"/>\n'
	f'\t<designed target="BrushStyle" value="1" ver="5"/>\n'
	f'\t<designed target="WindowX" value="0" ver="5"/>\n'
	f'\t<designed target="WindowY" value="0" ver="5"/>\n'
	f'\t<designed target="WindowWidth" value="1920" ver="5"/>\n'
	f'\t<designed target="WindowHeight" value="1080" ver="5"/>\n'
	f'\t<designed target="WindowCaption" value="name_list" ver="5"/>\n'
	f'\t<designed target="ShowWindowCaption" value="true" ver="5"/>\n'
	f'\t<designed target="ShowWindowMinimize" value="false" ver="5"/>\n'
	f'\t<designed target="ShowWindowMaximize" value="false" ver="5"/>\n'
	f'\t<designed target="ShowWindowClose" value="true" ver="5"/>\n'
	f'\t<designed target="AlwaysOnTop" value="true" ver="5"/>\n'
	f'\t<designed target="WindowSizeMode" value="2" ver="5"/>\n'
	f'\t<designed target="WindowBorderStyle" value="1" ver="5"/>\n'
	f'\t<designed target="WindowState" value="0" ver="5"/>\n'
	f'\t<designed target="WindowScalingMode" value="0" ver="5"/>\n'
	f'\t<designed target="MonitorNumber" value="0" ver="5"/>\n'
	f'\t<designed target="WindowPosition" value="2" ver="5"/>\n'
	f'\t<designed target="WindowCloseMode" value="0" ver="5"/>\n'
	f'\t<object access-modifier="private" name="t_title" display-name="t_title" uuid="32c58c39-3107-4276-88e5-941449faf1da" base-type="Text" base-type-id="21d59f8d-2ca4-4592-92ca-b4dc48992a0f" ver="5">\n'
		f'\t\t<designed target="X" value="0" ver="5"/>\n'
		f'\t\t<designed target="Y" value="0" ver="5"/>\n'
		f'\t\t<designed target="ZValue" value="0" ver="5"/>\n'
		f'\t\t<designed target="Rotation" value="0" ver="5"/>\n'
		f'\t\t<designed target="Scale" value="1" ver="5"/>\n'
		f'\t\t<designed target="Visible" value="true" ver="5"/>\n'
		f'\t\t<designed target="Opacity" value="1" ver="5"/>\n'
		f'\t\t<designed target="Enabled" value="true" ver="5"/>\n'
		f'\t\t<designed target="Tooltip" value="" ver="5"/>\n'
		f'\t\t<designed target="Width" value="870" ver="5"/>\n'
		f'\t\t<designed target="Height" value="53" ver="5"/>\n'
		f'\t\t<designed target="Text" value="ТАБЛО И СИРЕНЫ" ver="5"/>\n'
		f'\t\t<designed target="Font" value="Arial,16,-1,5,75,0,0,0,0,0,Полужирный" ver="5"/>\n'
		f'\t\t<designed target="FontColor" value="4278190080" ver="5"/>\n'
		f'\t\t<designed target="TextAlignment" value="132" ver="5"/>\n'
	f'\t</object>\n'
	f'\t<object access-modifier="private" name="ApTemplate" display-name="ApTemplate" uuid="3cd004bc-155e-4df0-ae6e-69744e4861fd" base-type="ApSource" base-type-id="966603da-f05e-4b4d-8ef0-919efbf8ab2c" ver="5">\n'
		f'\t\t<designed target="Path" value="designed_path" ver="5"/>\n'
		f'\t\t<designed target="Active" value="true" ver="5"/>\n'
		f'\t\t<designed target="ReAdvise" value="0" ver="5"/>\n'
		f'\t\t<init target="ParentSource" ver="5" ref="unit.Global.global_ApSource"/>\n'
	f'\t</object>\n'
	f'\t<object access-modifier="private" name="empty_link" display-name="empty_link" uuid="c6939328-e0e9-48c6-a45e-ebe2bdc542b4" base-type="Rectangle" base-type-id="15726dc3-881e-4d8d-b0fa-a8f8237f08ca" ver="5">\n'
		f'\t\t<designed target="X" value="8" ver="5"/>\n'
		f'\t\t<designed target="Y" value="53" ver="5"/>\n'
		f'\t\t<designed target="ZValue" value="0" ver="5"/>\n'
		f'\t\t<designed target="Rotation" value="0" ver="5"/>\n'
		f'\t\t<designed target="Scale" value="1" ver="5"/>\n'
		f'\t\t<designed target="Visible" value="false" ver="5"/>\n'
		f'\t\t<designed target="Opacity" value="1" ver="5"/>\n'
		f'\t\t<designed target="Enabled" value="true" ver="5"/>\n'
		f'\t\t<designed target="Tooltip" value="" ver="5"/>\n'
		f'\t\t<designed target="Width" value="854" ver="5"/>\n'
		f'\t\t<designed target="Height" value="26" ver="5"/>\n'
		f'\t\t<designed target="RoundingRadius" value="0" ver="5"/>\n'
		f'\t\t<designed target="PenColor" value="4278190080" ver="5"/>\n'
		f'\t\t<designed target="PenStyle" value="1" ver="5"/>\n'
		f'\t\t<designed target="PenWidth" value="1" ver="5"/>\n'
		f'\t\t<designed target="BrushColor" value="4278190080" ver="5"/>\n'
		f'\t\t<designed target="BrushStyle" value="0" ver="5"/>\n'
	f'\t</object>\n'
	f'\t<designed target="WindowIconPath" value="tn_logo.jpg" ver="5"/>\n'
    f'\t<object access-modifier="private" name="Form_UTS_Control" display-name="Form_UTS_Control" uuid="9193a38e-cafb-44a8-b8f7-e13e85c2de3a" base-type="Window" base-type-id="04615219-28bb-4a9a-bba4-50ac66972eb0" ver="5" description="" cardinal="1">\n'
		f'\t\t<do-on access-modifier="private" name="Handler_1" display-name="Handler_1" ver="5" event="MessageReceived" form-action="close"/>\n'
		f'\t\t<init target="Group" ver="5" ref="unit.WorkspaceControl.Form_UPTS_Control"/>\n'
	f'\t</object>\n'   
	f'\t<do-on access-modifier="private" name="Handler_8" display-name="Handler_8" ver="5" event="Opened">\n'
		f'\t\t<body kind="om">\n'
			f'\t\t\t<![CDATA[if (unit.Variables.MonitorOrientation == 1)\n'
'{\n'
	f'here.SetBoundingRegion(0, 0, 3840, 1080);\n'
'}\n'
f'else\n'
'{\n'
f'\t\there.SetBoundingRegion(0, 0, 1920, 2160);\n'
'}]]>\n'
f'		</body>\n'
f'	</do-on>\n'
f'</type>')
    path_uts_upts = f'{path_save}\Form_UTS_UPTS_default.omobj'
    # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
    if not os.path.exists(path_uts_upts):
        file = codecs.open(path_uts_upts, 'w', 'utf-8')
        file.write(uts_upts_sample)
        file.close()
    else:
        os.remove(path_uts_upts)
        file = codecs.open(path_uts_upts, 'w', 'utf-8')
        file.write(uts_upts_sample)
        file.close()
# Создание шаблона диагностики
def diag_gen(path_save):
    diag_sample = (
        f'<type access-modifier="private" name="name" display-name="name" uuid="uuid" base-type="Form" base-type-id="ffaf5544-6200-45f4-87ec-9dd24558a9d5" ver="5">\n'
        f'\t<designed target="X" value="0" ver="5"/>\n'
        f'\t<designed target="Y" value="0" ver="5"/>\n'
        f'\t<designed target="ZValue" value="0" ver="5"/>\n'
        f'\t<designed target="Rotation" value="0" ver="5"/>\n'
        f'\t<designed target="Scale" value="1" ver="5"/>\n'
        f'\t<designed target="Visible" value="true" ver="5"/>\n'
        f'\t<designed target="Opacity" value="1" ver="5"/>\n'
        f'\t<designed target="Enabled" value="true" ver="5"/>\n'
        f'\t<designed target="Tooltip" value="" ver="5"/>\n'
        f'\t<designed target="Width" value="1670" ver="5"/>\n'
        f'\t<designed target="Height" value="820" ver="5"/>\n'
        f'\t<designed target="PenColor" value="4278190080" ver="5"/>\n'
        f'\t<designed target="PenStyle" value="0" ver="5"/>\n'
        f'\t<designed target="PenWidth" value="1" ver="5"/>\n'
        f'\t<designed target="BrushColor" value="0xfff0f0f0" ver="5"/>\n'
        f'\t<designed target="BrushStyle" value="1" ver="5"/>\n'
        f'\t<designed target="WindowX" value="0" ver="5"/>\n'
        f'\t<designed target="WindowY" value="0" ver="5"/>\n'
        f'\t<designed target="WindowWidth" value="1920" ver="5"/>\n'
        f'\t<designed target="WindowHeight" value="1080" ver="5"/>\n'
        f'\t<designed target="WindowCaption" value="MainForm" ver="5"/>\n'
        f'\t<designed target="ShowWindowCaption" value="true" ver="5"/>\n'
        f'\t<designed target="ShowWindowMinimize" value="true" ver="5"/>\n'
        f'\t<designed target="ShowWindowMaximize" value="true" ver="5"/>\n'
        f'\t<designed target="ShowWindowClose" value="true" ver="5"/>\n'
        f'\t<designed target="AlwaysOnTop" value="false" ver="5"/>\n'
        f'\t<designed target="WindowSizeMode" value="2" ver="5"/>\n'
        f'\t<designed target="WindowBorderStyle" value="1" ver="5"/>\n'
        f'\t<designed target="WindowState" value="0" ver="5"/>\n'
        f'\t<designed target="WindowScalingMode" value="0" ver="5"/>\n'
        f'\t<designed target="MonitorNumber" value="0" ver="5"/>\n'
        f'\t<designed target="WindowPosition" value="2" ver="5"/>\n'
        f'\t<designed target="WindowCloseMode" value="0" ver="5"/>\n'
        f'\t<object access-modifier="private" name="t_uso_title" display-name="t_uso_title" uuid="f08d80b2-d564-4fa4-81d0-c78459ca4c2a" base-type="Text" base-type-id="21d59f8d-2ca4-4592-92ca-b4dc48992a0f" ver="5">\n'
        f'\t\t<designed target="X" value="10" ver="5"/>\n'
        f'\t\t<designed target="Y" value="10" ver="5"/>\n'
        f'\t\t<designed target="ZValue" value="0" ver="5"/>\n'
        f'\t\t<designed target="Rotation" value="0" ver="5"/>\n'
        f'\t\t<designed target="Scale" value="1" ver="5"/>\n'
        f'\t\t<designed target="Visible" value="true" ver="5"/>\n'
        f'\t\t<designed target="Opacity" value="1" ver="5"/>\n'
        f'\t\t<designed target="Enabled" value="true" ver="5"/>\n'
        f'\t\t<designed target="Tooltip" value="" ver="5"/>\n'
        f'\t\t<designed target="Width" value="1650" ver="5"/>\n'
        f'\t\t<designed target="Height" value="30" ver="5"/>\n'
        f'\t\t<designed target="Text" value="Rename" ver="5"/>\n'
        f'\t\t<designed target="Font" value="Arial,18,-1,5,75,0,0,0,0,0,Полужирный" ver="5"/>\n'
        f'\t\t<designed target="FontColor" value="4278190080" ver="5"/>\n'
        f'\t\t<designed target="TextAlignment" value="132" ver="5"/>\n'
        f'\t</object>\n'
        f'\t<object access-modifier="private" name="r_ss" display-name="r_ss" uuid="9f919ed1-143f-492f-aa74-14d14ccae79d" base-type="Rectangle" base-type-id="15726dc3-881e-4d8d-b0fa-a8f8237f08ca" ver="5">\n'
        f'\t\t<designed target="X" value="950" ver="5"/>\n'
        f'\t\t<designed target="Y" value="70" ver="5"/>\n'
        f'\t\t<designed target="ZValue" value="0" ver="5"/>\n'
        f'\t\t<designed target="Rotation" value="0" ver="5"/>\n'
        f'\t\t<designed target="Scale" value="1" ver="5"/>\n'
        f'\t\t<designed target="Visible" value="true" ver="5"/>\n'
        f'\t\t<designed target="Opacity" value="1" ver="5"/>\n'
        f'\t\t<designed target="Enabled" value="true" ver="5"/>\n'
        f'\t\t<designed target="Tooltip" value="" ver="5"/>\n'
        f'\t\t<designed target="Width" value="630" ver="5"/>\n'
        f'\t\t<designed target="Height" value="650" ver="5"/>\n'
        f'\t\t<designed target="RoundingRadius" value="0" ver="5"/>\n'
        f'\t\t<designed target="PenColor" value="4278190080" ver="5"/>\n'
        f'\t\t<designed target="PenStyle" value="1" ver="5"/>\n'
        f'\t\t<designed target="PenWidth" value="2" ver="5"/>\n'
        f'\t\t<designed target="BrushColor" value="0xffdcdcdc" ver="5"/>\n'
        f'\t\t<designed target="BrushStyle" value="1" ver="5"/>\n'
        f'\t\t<object access-modifier="private" name="r_table_top" display-name="r_table_top" uuid="b2d05e1e-9861-4a90-ae60-8259fa404f6c" base-type="Rectangle" base-type-id="15726dc3-881e-4d8d-b0fa-a8f8237f08ca" ver="5">\n'
        f'\t\t\t<designed target="X" value="0" ver="5"/>\n'
        f'\t\t\t<designed target="Y" value="0" ver="5"/>\n'
        f'\t\t\t<designed target="ZValue" value="0" ver="5"/>\n'
        f'\t\t\t<designed target="Rotation" value="0" ver="5"/>\n'
        f'\t\t\t<designed target="Scale" value="1" ver="5"/>\n'
        f'\t\t\t<designed target="Visible" value="true" ver="5"/>\n'
        f'\t\t\t<designed target="Opacity" value="1" ver="5"/>\n'
        f'\t\t\t<designed target="Enabled" value="true" ver="5"/>\n'
        f'\t\t\t<designed target="Tooltip" value="" ver="5"/>\n'
        f'\t\t\t<designed target="Width" value="630" ver="5"/>\n'
        f'\t\t\t<designed target="Height" value="30" ver="5"/>\n'
        f'\t\t\t<designed target="RoundingRadius" value="0" ver="5"/>\n'
        f'\t\t\t<designed target="PenColor" value="4278190080" ver="5"/>\n'
        f'\t\t\t<designed target="PenStyle" value="1" ver="5"/>\n'
        f'\t\t\t<designed target="PenWidth" value="1" ver="5"/>\n'
        f'\t\t\t<designed target="BrushColor" value="0xff001b53" ver="5"/>\n'
        f'\t\t\t<designed target="BrushStyle" value="1" ver="5"/>\n'
        f'\t\t\t<object access-modifier="private" name="t_name_server cabinet" display-name="t_name_server cabinet" uuid="7d992fd2-58fa-4312-b6cd-dd8b4423ebab" base-type="Text" base-type-id="21d59f8d-2ca4-4592-92ca-b4dc48992a0f" ver="5">\n'
        f'\t\t\t\t<designed target="X" value="0" ver="5"/>\n'
        f'\t\t\t\t<designed target="Y" value="0" ver="5"/>\n'
        f'\t\t\t\t<designed target="ZValue" value="0" ver="5"/>\n'
        f'\t\t\t\t<designed target="Rotation" value="0" ver="5"/>\n'
        f'\t\t\t\t<designed target="Scale" value="1" ver="5"/>\n'
        f'\t\t\t\t<designed target="Visible" value="true" ver="5"/>\n'
        f'\t\t\t\t<designed target="Opacity" value="1" ver="5"/>\n'
        f'\t\t\t\t<designed target="Enabled" value="true" ver="5"/>\n'
        f'\t\t\t\t<designed target="Tooltip" value="" ver="5"/>\n'
        f'\t\t\t\t<designed target="Width" value="630" ver="5"/>\n'
        f'\t\t\t\t<designed target="Height" value="30" ver="5"/>\n'
        f'\t\t\t\t<designed target="Text" value="Диагностика шкафа" ver="5"/>\n'
        f'\t\t\t\t<designed target="Font" value="Arial,16,-1,5,75,0,0,0,0,0,Полужирный" ver="5"/>\n'
        f'\t\t\t\t<designed target="FontColor" value="0xffffffff" ver="5"/>\n'
        f'\t\t\t\t<designed target="TextAlignment" value="132" ver="5"/>\n'
        f'\t\t\t</object>\n'
        f'\t\t</object>\n'
        f'\t</object>\n'
        f'\t<designed target="WindowIconPath" value="" ver="5"/>\n'
        f'\t<do-on access-modifier="private" name="Handler_1" display-name="Handler_1" ver="5" event="Opened">\n'
        f'\t\t<body kind="om">\n'
        f'\t\t\t<![CDATA[D_PT_USO_1_1.Enabled = false;]]>\n'
        f'\t\t</body>\n'
        f'\t</do-on>\n'
        f'\t<object access-modifier="private" name="Rename_link" display-name="Rename_link" uuid="0d35b4c2-df7d-4dd6-b1cd-b2b9d891863b" base-type="Action" base-type-id="a9ee9770-1c4a-44c9-b815-157d9fc2ab95" base-const="true" base-ref="true" ver="5" description=""/>\n'
        f'\t<do-on access-modifier="private" name="Handler_2" display-name="Handler_2" ver="5" event="Closed">\n'
        f'\t\t<body kind="om">\n'
        f'\t\t\t<![CDATA[D_PT_USO_1_1.Enabled = true;]]>\n'
        f'\t\t</body>\n'
        f'\t</do-on>\n'
        f'</type>\n')

    path_diag = f'{path_save}\D_USO_Template.omobj'
    # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
    if not os.path.exists(path_diag):
        file = codecs.open(path_diag, 'w', 'utf-8')
        file.write(diag_sample)
        file.close()
    else:
        os.remove(path_diag)
        file = codecs.open(path_diag, 'w', 'utf-8')
        file.write(diag_sample)
        file.close()






