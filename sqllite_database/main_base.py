from models import *
import openpyxl as wb
from lxml import etree
from datetime import datetime
import re, traceback, os, codecs, uuid, math
import psycopg2
today = datetime.now()



# Additional general features
class General_functions():
    def str_find(self, str1, arr):
        i = 0
        for el in arr:
            if str(str1).find(el) > -1:
                return True
    def translate(self, str):
            dict = {".":"_",
                    "/":"_",
                    "\\":"_",
                    ",":"_",
                    ":":"_",
                    ";":"_",
                    "А":"A",
                    "Б":"_B",
                    "В":"B",
                    "Г":"G",
                    "Д":"D",
                    "Е":"E",
                    "Ё":"E",
                    "Ж":"J",
                    "З":"Z",
                    "И":"I",
                    "Й":"I",
                    "К":"K",
                    "Л":"L",
                    "М":"M",
                    "Н":"H",
                    "О":"O",
                    "П":"_P",
                    "Р":"P",
                    "С":"C",
                    "Т":"T",
                    "У":"U",
                    "Ф":"F",
                    "Х":"X",
                    "Ц":"C",
                    "Ч":"CH",
                    "Ш":"SH",
                    "Щ":"SCH",
                    "Ь":"b",
                    "Ы":"_",
                    "Ъ":"_",
                    "Э":"E",
                    "Ю":"U",
                    "Я":"YA",
                    "а":"a",
                    "б":"_b",
                    "в":"b",
                    "г":"g",
                    "д":"d",
                    "е":"e",
                    "ё":"e",
                    "ж":"j",
                    "з":"z",
                    "и":"i",
                    "й":"i",
                    "к":"k",
                    "л":"l",
                    "м":"m",
                    "н":"h",
                    "о":"o",
                    "п":"_p",
                    "р":"p",
                    "с":"c",
                    "т":"t",
                    "у":"u",
                    "ф":"f",
                    "х":"x",
                    "ц":"c",
                    "ч":"ch",
                    "ш":"sh",
                    "щ":"sch",
                    "ь":"b",
                    "ы":"_",
                    "ъ":"_",
                    "э":"e",
                    "ю":"u",
                    "я":"ya"}

            intab = '.-пПаАфз/еЕсС'
            outtab = '__ppaafz_eEcC'
            trantab = str.maketrans(dict)
            outstr = str.translate(trantab)
            return outstr
    def column_check(self, table_used_model, table_used_base, list_column):
        # Logs
        msg = {}
        # Create tabl
        # try:
        #     cursor = db.cursor()
        #     cursor.execute(f'''SELECT * FROM {table_used_base}''')
        #     msg[f'{today} - Таблица: {table_used_base} существует'] = 1
        # except:
        with db.atomic():
            db.create_tables([table_used_model])
            #msg[f'{today} - Таблица: {table_used_base} добавлена в базу данных'] = 3
        # Checking if a column exists
        column_tabl  = []
        new_column   = []
        
        for data_column in db.get_columns(table_used_base):
            if data_column[0] in list_column: column_tabl.append(data_column[0])
        
        for lst in list_column:
            if lst not in column_tabl: 
                msg[f'{today} - Отсутствует обязательный столбец таблицы {table_used_model}: {lst}'] = 2
                new_column.append(lst)
        
        for new_name in new_column:
            msg[f'{today} - Столбец: {new_name} добавлен в таблицу {table_used_model}'] = 3
            migrate(migrator.add_column(table_used_base, new_name, IntegerField(null=True)))
        return msg
    def empty_table(self, table_used):
        cursor = db.cursor()
        cursor.execute(f'''SELECT COUNT (*) FROM "{table_used}"''')
        empty = cursor.fetchall()
        return True if int(empty[0][0]) == 0  else False
    def clear_tabl(self, table_used, table_name, list_tabl):
        msg = {}
        cursor = db.cursor()
        if not table_used in list_tabl:
            msg[f'{today} - Таблица: {table_used} отсутствует!'] = 2
            return msg

        if self.empty_table(table_used): 
            msg[f'{today} - Таблица: {table_used} пустая!'] = 2
            return msg
        
        cursor.execute(f'DELETE FROM "{table_used}"')
        msg[f'{today} - Таблица: {table_used} полностью очищена'] = 1
        return msg
    def search_signal(self, tabl_used_cl, tabl_used_str, tag):
        exists_tag = tabl_used_cl.select().where(tabl_used_cl.tag == tag)
        if bool(exists_tag):
            cursor = db.cursor()
            cursor.execute(f"""SELECT id, tag
                               FROM "{tabl_used_str}"
                               WHERE tag='{tag}'""")
            for id_, tag in cursor.fetchall():
                if tabl_used_str == 'di': return (f'DI[{id_}].Value')
                if tabl_used_str == 'do': return (f'ctrlDO[{id_}]')
                if tabl_used_str == 'ai': return (f'AI[{id_}].Norm')
        else:
            return ''
    def update_signal(self, tabl_used_cl, tabl_used_str, tag, number_NA, column_update_cl, column_update_str):
        msg = {}
        exist_value  = tabl_used_cl.select().where(tabl_used_cl.id == number_NA,
                                                    column_update_cl == tag)
        if not bool(exist_value):
            cursor = db.cursor()
            cursor.execute(f"""UPDATE {tabl_used_str}
                               SET "{column_update_str}"='{tag}' 
                               WHERE "id"='{number_NA}'""")
            msg[f'{today} - Таблица: umpna, NA[{number_NA}] обновлено {column_update_str} = {tag}'] = 3
            return msg
        return msg
    def update_signal_dop(self, tabl_used_cl, tabl_used_str, name, column_update_cl, column_update_str, value):
        msg = {}
        exist_value  = tabl_used_cl.select().where(tabl_used_cl.name == name,
                                                    column_update_cl == value)
        if not bool(exist_value):
            cursor = db.cursor()
            cursor.execute(f"""UPDATE {tabl_used_str}
                               SET "{column_update_str}"='{value}' 
                               WHERE "name"='{name}'""")
            msg[f'{today} - Таблица: {tabl_used_str}, обновлен: {name},  {column_update_str} = {value}'] = 3
            return msg
        return msg
    def parser_sample(self, path, kod_msg, name, flag_write_db, table, *args):
        cursor = db_prj.cursor()
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(path, parser)
        root = tree.getroot()

        list_msg = []

        for lvl_one in root.iter('Row'):
            category  = lvl_one.attrib['Category']
            isAck     = lvl_one.attrib['IsAck']
            isCycle   = lvl_one.attrib['IsCycle']
            isSound   = lvl_one.attrib['IsSound']
            isHide    = lvl_one.attrib['IsHide']
            priority  = lvl_one.attrib['Priority']
            isAlert   = lvl_one.attrib['IsAlert']
            mess      = lvl_one.attrib['Message']
            soundFile = lvl_one.attrib['SoundFile']
            nextLink  = lvl_one.attrib['NextLink']
            base      = lvl_one.attrib['Base']

            if table == 'KTPRAS_1' or table == 'UMPNA':
                if self.str_find(mess, {'%1'}): 
                    mess = str(mess).replace('%1', args[0])
                if self.str_find(mess, {'%2'}): 
                    mess = str(mess).replace('%2', args[1])
            
            if table == 'NPS' or table == 'KRMPN' or table == 'Global': text_mess = mess
            else: text_mess = f'{name}. {mess}'

            del_row_tabl = f"DELETE FROM messages.opmessages WHERE Category ={kod_msg + int(category)};\n"
            ins_row_tabl = f"INSERT INTO messages.opmessages (Category, Message, IsAck, SoundFile, IsCycle, IsSound, IsHide, Priority, IsAlert) VALUES({kod_msg + int(category)}, '{text_mess}', {isAck}, '{soundFile}', {isCycle}, {isSound}, {isHide}, {priority}, {isAlert});\n"
            
            if flag_write_db:
                cursor.execute(del_row_tabl)
                cursor.execute(ins_row_tabl)
            else:
                list_msg.append(dict(delete = del_row_tabl,
                                     insert = ins_row_tabl))
        return list_msg
    def all_tables(self):
        list_tabl = []
        cursor = db.cursor()
        cursor.execute(f"""SELECT table_name 
                           FROM information_schema.tables 
                           WHERE table_schema='public'""")
        for name in cursor.fetchall():
            list_tabl.append(name[0])
        return list_tabl
    # ВУ
    # Подключение к SQL - база проекта
    def connect_by_sql_prj(self, table_used, column):
        try:
            cursor = db_prj.cursor()
            cursor.execute(f'''SELECT {column} FROM {table_used} ORDER BY id''')
        except Exception:
            return 
        return cursor.fetchall()
    # Подключение к SQL - база разработки
    def connect_by_sql(self, table_used, column):
        try:
            cursor = db.cursor()
            cursor.execute(f'''SELECT {column} FROM "{table_used}" ORDER BY id''')
        except Exception:
            return 
        return cursor.fetchall()
    def connect_by_sql_order(self, table_used, column, order):
        try:
            cursor = db.cursor()
            cursor.execute(f'''SELECT {column} FROM "{table_used}" ORDER BY {order}''')
        except Exception:
            return 
        return cursor.fetchall()
    def connect_by_sql_condition(self, table_used, column, condition):
        try:
            cursor = db.cursor()
            cursor.execute(f'''SELECT {column}
                               FROM "{table_used}"
                               WHERE {condition}
                               ORDER BY id''')
        except Exception:
            return 
        return cursor.fetchall()
    def connect_count_row_sql(self, table_used):
        try:
            cursor = db.cursor()
            cursor.execute(f'''SELECT COUNT (*) FROM "{table_used}"''')
        except Exception:
            return True
        return cursor.fetchall()
    def max_value_column(self, table_used, column, condition, *args):
        try:
            cursor = db.cursor()
            if condition is False: cursor.execute(f'''SELECT MAX("{column}") FROM "{table_used}"''')
            else                 : cursor.execute(f'''SELECT MAX("{column}") FROM "{table_used}" WHERE "{args[0]}"={args[1]}''')
        except Exception:
            return 
        return cursor.fetchall()[0][0]
    # Обновление строки
    def update_row(self, tabl_used, tag, column_update, id_):
        cursor = db.cursor()
        cursor.execute(f"""UPDATE {tabl_used}
                           SET "{column_update}"='{tag}' 
                           WHERE "id"={id_}""")
        return f'{column_update}-{tag}, '
        
    # Создание атрибутов
    def new_attr(self, obj, type, value):
         atrb = etree.Element("attribute")
         atrb.attrib['type'] = type
         atrb.attrib['value'] = str(value)
         obj.append(atrb)
    # Создание строки карты адресов
    def new_map_str(self, obj, element, value):
        elem = etree.Element(element)
        elem.text = str(value)
        obj.append(elem)
    # Чистка и парсинг
    def clear_objects_omx(self, directory):
        # Чистка объектов
        msg = {}
        msg_bool, el1, tree = self.parser_omx(directory)
        tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)

        if msg_bool == 1: 
            msg[f'{today} - Файл omx: ошибка при чистке {directory}'] = 2
            return msg
        msg[f'{today} - Файл omx: {directory} удалены'] = 1
        return msg
    def clear_diag_objects_omx(self, directory):
        # Чистка объектов
        msg = {}
        msg_bool, el1, tree = self.parser_diag_omx(directory)
        tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)

        if msg_bool == 1: 
            msg[f'{today} - Файл omx: ошибка при чистке {directory}'] = 2
            return msg
        msg[f'{today} - Файл omx: {directory} удалены'] = 1
        return msg
    def clear_objects_attrib(self, directory, map_attrib):
        msg = {}
        try:
            for path in map_attrib:
                self.parser_diag_map(directory, path)
            msg[f'{today} - Очистка значений {directory}. Успешно'] = 1
        except Exception:
            msg[f'{today} - Ошибка при очистке значений атрибутов: {path}, {traceback.format_exc()}'] = 2
        return msg
    
    def parser_omx(self, directory):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(f'{path_to_devstudio}\\typical_prj.omx', parser)
        root = tree.getroot()
        try:
            for el in root.iter('{automation.deployment}application-object'):
                if el.attrib['name'] == "Application_PLC":
                    for item in el.iter('{automation.control}object'):
                        if item.attrib['name'] == 'Root' + prefix_system:
                            for el1 in item.iter('{automation.control}object'):
                                if el1.attrib['name'] == directory:
                                    item.remove(el1)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = directory
                            item.append(object)

                            for el1 in item.iter('{automation.control}object'):
                                if el1.attrib['name'] == directory:
                                    return 0, el1, tree
        except:
            return 1, 0, 0
    def parser_diag_omx(self, directory):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(f'{path_to_devstudio}\\typical_prj.omx', parser)
        root = tree.getroot()
        try:
            for el in root.iter('{automation.deployment}application-object'):
                if el.attrib['name'] == "Application_PLC":
                    for item in el.iter('{automation.control}object'):
                        if item.attrib['name'] == "Diag":
                            for el1 in item.iter('{automation.control}object'):
                                if el1.attrib['name'] == directory:
                                    item.remove(el1)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = directory
                            item.append(object)

                            for el1 in item.iter('{automation.control}object'):
                                if el1.attrib['name'] == directory:
                                    return 0, el1, tree
        except:
            return 1, 0, 0
    
    def parser_map(self, directory):
        path_map = f'{path_to_devstudio}\\OUA.xml'
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(path_map, parser)
        root = tree.getroot()

        for item in root.iter('node-path'):
            signal = f'Root{prefix_system}{directory}'
            if self.str_find(item.text, {signal}):
                parent = item.getparent()
                root.remove(parent)
        tree.write(path_map, pretty_print=True)
        return root, tree
    def parser_diag_map(self, directory, path_map):
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(path_map, parser)
        root = tree.getroot()
        
        for item in root.iter('item'):
            signal = f'Root{prefix_system}{directory}'
            if self.str_find(item.attrib['id'], {signal}):
                root.remove(item)
        tree.write(path_map, pretty_print=True)
        return root, tree

# Work with filling in the table 'Signals'
class Import_in_SQL():
    def __init__(self, exel):
        self.exel    = exel
        self.connect = wb.load_workbook(self.exel, read_only=True, data_only=True)
    # Read tables from file
    def read_table(self):
        tables = []
        for sheet in self.connect.worksheets:
            tables.append(sheet.title)
        return tables
    # Looking for table hat
    def search_hat_table(self, uso, number_row):
        hat_tabl = []
        for sheet in self.connect.worksheets:
            if sheet.title == uso:
                column = sheet.max_column
                for i in range(int(number_row), int(number_row) + 1):
                    for j in range(1, column + 1):
                        cell = sheet.cell(row=i, column=j).value
                        if cell is None: continue
                        hat_tabl.append(cell)
        return hat_tabl
    # Reading table data
    def import_table(self, uso, number_row, name_column):
        hat_num  = {}
        for sheet in self.connect.worksheets:
            if sheet.title == uso:
                column = sheet.max_column
                for i in range(int(number_row), int(number_row) + 1):
                    for j in range(1, column + 1):
                        cell = sheet.cell(row=i, column=j).value
                        if cell is None: continue
                        for key, value in name_column.items():
                            if value == cell:
                                hat_num[key] = j - 1
                data = []
                for row in sheet.iter_rows(min_row=(int(number_row) + 1)):
                    keys   = []
                    values = []
                    for name, number in hat_num.items():
                        keys.append(name)
                        values.append(row[number].value)
                    values.append(uso)
                    array = {k: v for k, v in zip(keys, values)}
                    data.append(array)
        # Количество строк в таблице
        cursor = db.cursor()
        try:
            cursor.execute(f'''SELECT COUNT(*) FROM signals''')
            count_row = cursor.fetchall()[0][0]
        except:
            count_row = 0

        # Delete basket is None
        data_new = []
        for row in data:
            type_signal = row['type_signal']
            scheme      = row['schema']
            basket      = row['basket']
            module      = row['module']
            channel     = row['channel']

            if basket is None or module is None or channel is None: continue
            count_row += 1

            list_type = ['CPU', 'PSU', 'CN', 'MN', 'AI','AO', 'DI', 'RS','DO']
            for value in list_type:
                if str(scheme).find(value) != -1: 
                    type_signal = value

            dict_column = {'id'          : count_row,
                           'type_signal' : type_signal,
                           'uso'         : uso,
                           'tag'         : row['tag'],
                           'description' : row['description'],
                           'schema'      : row['schema'],
                           'klk'         : row['klk'],
                           'contact'     : row['contact'],
                           'basket'      : basket,
                           'module'      : row['module'],
                           'channel'     : row['channel']}
            if basket is None: continue
            data_new.append(dict_column)
        return data_new
    # Importing into SQL
    def import_for_sql(self, data, uso):
        msg = {}
        # Checking for the existence of a database
        with db.atomic():
            try:
                Signals.insert_many(data).execute()
                msg[f'{today} - Добавлено новое УСО: {uso}'] = 1
            except Exception:
                msg[f'{today} - Таблица: signals, ошибка при заполнении: {traceback.format_exc()}'] = 2
        return(msg)
    # Update Database
    def update_for_sql(self, data, uso):
        msg = {}
        with db:
            try:
                # Filter by uso, basket, modul, channel
                for row_exel in data:
                    exist_row = Signals.select().where(Signals.uso == uso,
                                                    Signals.basket  == str(row_exel['basket']),
                                                    Signals.module  == str(row_exel['module']),
                                                    Signals.channel == str(row_exel['channel']))
                    if not bool(exist_row):
                        # new record
                        Signals.create(
                            type_signal=row_exel['type_signal'],
                            uso        =row_exel['uso'],
                            tag        =row_exel['tag'],
                            description=row_exel['description'],
                            schema     =row_exel['schema'],
                            klk        =row_exel['klk'],
                            contact    =row_exel['contact'],
                            basket     =row_exel['basket'],
                            module     =row_exel['module'],
                            channel    =row_exel['channel'],
                        )
                        msg[f'''{today} - Добавлен новый сигнал: Tag - {row_exel["tag"]}, description - {row_exel["description"]}, 
                                                                basket - {row_exel["basket"]}, module - {row_exel["module"]}, 
                                                                channel - {row_exel["channel"]}'''] = 0

                    for row_sql in Signals.select().dicts():

                        if row_sql['uso']     == uso                     and \
                        row_sql['basket']  == str(row_exel['basket']) and \
                        row_sql['module']  == str(row_exel['module']) and \
                        row_sql['channel'] == str(row_exel['channel']):
                            
                            if str(row_sql['tag'])         == str(row_exel['tag'])         and \
                               str(row_sql['description']) == str(row_exel['description']) and \
                               str(row_sql['scheme'])      == str(row_exel['scheme'])      and \
                               str(row_sql['klk'])         == str(row_exel['klk'])         and \
                               str(row_sql['contact'])     == str(row_exel['contact']):
            
                                continue
                            else:
                                Signals.update(
                                    type_signal=row_exel['type_signal'],
                                    tag        =row_exel['tag'],
                                    description=row_exel['description'],
                                    schema     =row_exel['scheme'],
                                    klk        =row_exel['klk'],
                                    contact    =row_exel['contact'],
                                ).where(Signals.id == row_sql['id']).execute()
                                msg[f'''{today} - Обновление сигнала id = {row_sql["id"]}: Было, 
                                                                                        uso - {row_sql['uso']}, 
                                                                                        type_signal - {row_sql['type_signal']}, 
                                                                                        tag - {row_sql['tag']},                      
                                                                                        description - {row_sql['description']}, 
                                                                                        schema - {row_sql['scheme']}, 
                                                                                        klk - {row_sql['klk']},
                                                                                        contact - {row_sql['contact']} = 
                                                                                        Стало, 
                                                                                        uso - {row_exel['uso']}, 
                                                                                        type_signal - {row_exel['type_signal']}, 
                                                                                        tag - {row_exel['tag']}, 
                                                                                        description - {row_exel['description']}, 
                                                                                        scheme - {row_exel['scheme']}, 
                                                                                        klk - {row_exel['klk']},
                                                                                        contact - {row_exel['contact']}'''] = 3
                                continue
                        else:
                            continue
            except Exception:
                msg[f'{today} - Таблица: signals, ошибка при обновлении: {traceback.format_exc()}'] = 2
        return(msg)
    # Removing all rows
    def clear_tabl(self):
        msg = {}
        for row_sql in Signals.select().dicts():
            Signals.get(Signals.id == row_sql['id']).delete_instance()
        msg[f'{today} - Таблица: signals полностью очищена!'] = 1
        return(msg)
    # Column check
    def column_check(self):
        with db:
            list_default = ['id', 'type_signal', 'uso', 'tag', 'description', 'schema', 'klk', 'contact', 'basket', 'module', 'channel']

            self.dop_func = General_functions()
            msg = self.dop_func.column_check(Signals, 'signals', list_default)
        return msg

# Changing tables SQL
class Editing_table_SQL():
    def __init__(self):
        self.cursor = db.cursor()
        self.dop_function = General_functions()
    def editing_sql(self, table_sql):
        unpacking_  = []
        msg = {}
        try:
            self.cursor.execute(f'SELECT * FROM "{table_sql}" ORDER BY id')
            name_column = next(zip(*self.cursor.description))
            array_name_column = []
            if table_sql in rus_list.keys():

                for tabl, name_c in rus_list.items():

                    if tabl == table_sql:

                        for name in name_column:
                            if name in name_c.keys():

                                for key, value in name_c.items():
                                    if name == key:
                                        array_name_column.append(value)
                                        break
                            else:
                                array_name_column.append(name)
            else:
                array_name_column = name_column

            records = self.cursor.fetchall()
            unpacking_.append(records)

            count_column = len(name_column)
            count_row    = len(records)
            return count_column, count_row, array_name_column, records, msg
        except Exception:
            msg[f'{today} - Ошибка открытия редактора: {traceback.format_exc()}'] = 2
            return 0, 0, array_name_column, records, msg
    def func_chunks_generators(self, lst, n):
        for i in range(0, len(lst), n):
            yield lst[i : i + n]
    # Поиск названия сигнала для подписи
    def search_name(self, tabl, value):
        try:
            isdigit_num  = re.findall('\d+', str(value))
            self.cursor.execute(f"""SELECT name 
                                    FROM "{tabl}"
                                    WHERE id = {int(isdigit_num[0])}""")
            name_row = self.cursor.fetchall()[0][0]
            return name_row
        except:
            return ''
    # Column names
    def column_names(self, table_used):
        self.cursor.execute(f'SELECT * FROM {table_used}')
        return next(zip(*self.cursor.description))
    # Apply request 
    def apply_request_select(self, request, table_used):
        msg = {}
        unpacking = []
        try:
            self.cursor.execute(f'''{request}''')
            name_column = next(zip(*self.cursor.description))

            array_name_column = []
            for tabl, name_c in rus_list.items():

                if tabl == table_used:

                    for name in name_column:
                        if name in name_c.keys():

                            for key, value in name_c.items():
                                if name == key:
                                    array_name_column.append(value)
                                    break
                        else:
                            array_name_column.append(name)

            records = self.cursor.fetchall()
            unpacking.append(records)

            count_column = len(name_column)
            count_row    = len(records)
            return count_column, count_row, array_name_column, records, msg
        except:
            msg[f'{today} - Таблица: {table_used} некорректный запрос!'] = 2
            return 'error', 'error', 'error', 'error', msg
    def other_requests(self, request, table_used):
        msg = {}
        try:
            self.cursor.execute(f'''{request}''')
            msg[f'{today} - Таблица: {table_used} запрос применен!'] = 1
            return msg
        except:
            msg[f'{today} - Таблица: {table_used} некорректный запрос!'] = 2
            return msg
    # Updating cell values
    def update_row_tabl(self, column, text_cell, text_cell_id, table_used, hat_name, flag_NULL):
        msg = {}
        active_column = list(hat_name)[column]
        try:
            if flag_NULL:
                self.cursor.execute(f"""UPDATE {table_used} 
                                        SET "{active_column}"= NULL
                                        WHERE id={text_cell_id}""")
            else:
                self.cursor.execute(f"""UPDATE {table_used} 
                                        SET "{active_column}"='{text_cell}' 
                                        WHERE id={text_cell_id}""")
            return msg
        except Exception:
            msg[f'{today} - Таблица: {table_used}, ошибка при изменении ячейки: {traceback.format_exc()}'] = 2
            return msg
        #table_used.update(**{active_column: text_cell}).where(table_used.id == text_cell_id).execute()
    # Adding new lines
    def add_new_row(self, table_used):
        self.cursor.execute(f'''INSERT INTO {table_used} DEFAULT VALUES''')

        #table_used.insert(**{active_column: ''}).execute()
    # Removing rows
    def delete_row(self, text_cell_id, table_used):
        self.cursor.execute(f'''DELETE FROM {table_used}
                                WHERE id={text_cell_id}''')
        #table_used.get(table_used.id == text_cell_id).delete_instance()
    # Adding new column
    def add_new_column(self, table_used, new_column):
        self.cursor.execute(f'''ALTER TABLE {table_used} 
                                ADD '{new_column}' VARCHAR(255)''')
    # Removing column
    def delete_column(self, column, hat_name, table_used):
        active_column = list(hat_name)[column]
        self.cursor.execute(f'''ALTER TABLE {table_used} 
                                DROP COLUMN {active_column}''')
    # Removing all rows
    def clear_tabl(self, table_used):
        self.cursor.execute(f'''DELETE FROM {table_used}''')
    # Drop table
    def drop_tabl(self, table_used):
        self.cursor.execute(f'''DROP TABLE "{table_used}"''')
    # Table selection window
    def get_tabl(self):
        return db.get_tables()
    def type_column(self, table_used):
        msg       = {}
        type_list = []
        try:
            self.cursor.execute(f"""SELECT column_name, data_type
                                    FROM information_schema.columns
                                    WHERE table_schema = 'public' AND table_name = '{table_used}'""")
            if table_used in rus_list.keys():
                for tabl, name_c in rus_list.items():

                    if tabl == table_used:

                        for i in self.cursor.fetchall():
                            column_name = i[0]
                            data_type   = i[1]

                            if column_name in name_c.keys():

                                for key, value in name_c.items():
                                    if column_name == key:
                                        list_a = [column_name, value, data_type]
                                        type_list.append(list_a)
                                        break
            else:
                for i in self.cursor.fetchall():
                    column_name = i[0]
                    data_type   = i[1]
                    list_a = [column_name, '', data_type]
                    type_list.append(list_a)
        except Exception:
            msg[f'{today} - Окно тип данных: ошибка: {traceback.format_exc()}'] = 2

        return type_list, msg
    def dop_window_signal(self, table_used):
            type_list = []
            try:
                if table_used == 'ktpra':
                    self.cursor.execute(f"""SELECT id, variable, name FROM "{table_used}" ORDER BY id""")
                else:
                    self.cursor.execute(f"""SELECT id, tag, name FROM "{table_used}" ORDER BY id""")
                for i in self.cursor.fetchall():
                    id_  = i[0]
                    tag  = i[1]
                    name = i[2]

                    list_a = [id_, tag, name]
                    type_list.append(list_a)
                msg = 'Таблица открыта'
                color = '#6bdb84'
            except Exception:
                msg = 'Для типа сигнала нет таблицы'
                color = 'yellow'

            return type_list, msg, color
    def filter_text(self, text, list_signal):
        list_request = []
        for i in list_signal:
            id_  = i[0]
            tag  = i[1]
            name = i[2]

            if self.dop_function.str_find(str(name).lower(), {text}):
                list_temp = [id_, tag, name]
                list_request.append(list_temp)
        return list_request

# Generate data SQL
class Generate_database_SQL():
    def __init__(self):
        self.dop_function = General_functions()
    def check_database_connect(self, dbname, user, password, host, port):
        try:
            connect = psycopg2.connect(f"dbname={dbname} user={user} host={host} password={password} port={port} connect_timeout=1 ")
            connect.close()
            return True
        except:
            return False
    def define_number_msg(self, cursor, tag):
        kod_msg     = 0
        addr_offset = 0
        try:
            cursor.execute(f"""SELECT index, count 
                               FROM msg
                               WHERE tag ='{tag}'""")
            for i in cursor.fetchall():
                kod_msg = i[0]
                addr_offset = i[1]
        except Exception:
            return kod_msg, addr_offset
        return kod_msg, addr_offset
    def write_file(self, list_str, tabl, name_file):
        msg = {}
        # Создаём файл запроса
        path_request = f'{path_location_file}\\{name_file}.sql'
        if not os.path.exists(path_request):
            file = codecs.open(path_request, 'w', 'utf-8')
        else:
            os.remove(path_request)
            file = codecs.open(path_request, 'w', 'utf-8')

        if path_location_file == '' or path_location_file is None or len(path_location_file) == 0:
            msg[f'{today} - Сообщения {tabl}: не указана конечная папка'] = 2
            return msg
        begin = ('\tCREATE SCHEMA IF NOT EXISTS messages;\n'
                 '\tCREATE TABLE IF NOT EXISTS messages.OPMessages(\n'
                 '\t\tCategory INT NOT NULL,\n'
                 '\t\tMessage VARCHAR(1024),\n'
                 '\t\tIsAck BOOLEAN NOT NULL,\n'
                 '\t\tSoundFile VARCHAR(1024),\n'
                 '\t\tIsCycle BOOLEAN NOT NULL,\n'
                 '\t\tIsSound BOOLEAN NOT NULL,\n'
                 '\t\tIsHide BOOLEAN NOT NULL,\n'
                 '\t\tPriority INT NOT NULL,\n'
                 '\t\tIsAlert BOOLEAN NOT NULL,\n'
                 '\t\tCONSTRAINT OPMessages_pkey PRIMARY KEY (Category)\n'
                 '\t);\n'
                'BEGIN TRANSACTION;\n')
        file.write(begin)
        if tabl == 'Others':
            for i in list_str:
                delete = i['delete']
                file.write(delete)
                insert = i['insert']
                file.write(insert)
        else:
            for i in list_str:
                for j in i:
                    delete = j['delete']
                    file.write(delete)
                    insert = j['insert']
                    file.write(insert)
        file.write(f'COMMIT;')
        file.close()
        return msg
    def write_in_sql(self, list_tabl, flag_write_db):
        msg = {}
        if len(list_tabl) == 0: return
        for tabl in list_tabl: 
            if tabl == 'AI': 
                msg.update(self.gen_msg_ai(flag_write_db))
                continue
            if tabl == 'DI': 
                msg.update(self.gen_msg_general(flag_write_db, 'di', 'DI', 'PostgreSQL_Messages-DI'))
                continue
            if tabl == 'DO': 
                msg.update(self.gen_msg_general(flag_write_db, 'do', 'DOP', 'PostgreSQL_Messages-DO'))
                continue
            if tabl == 'ZD': 
                msg.update(self.gen_msg_general(flag_write_db, 'zd', 'ZD', 'PostgreSQL_Messages-ZD'))
                continue
            if tabl == 'VS': 
                msg.update(self.gen_msg_general(flag_write_db, 'vs', 'VS', 'PostgreSQL_Messages-VS'))
                continue
            if tabl == 'VV': 
                msg.update(self.gen_msg_defence(flag_write_db, 'vv', 'VV', 'PostgreSQL_Messages-VV', 'TblHighVoltageSwitches'))
                continue
            if tabl == 'UTS': 
                msg.update(self.gen_msg_uts_upts(flag_write_db, 'uts', 'UTS', 'PostgreSQL_Messages-UTS'))
                continue
            if tabl == 'UPTS': 
                msg.update(self.gen_msg_uts_upts(flag_write_db, 'upts', 'UPTS', 'PostgreSQL_Messages-UPTS'))
                continue
            if tabl == 'UMPNA': 
                msg.update(self.gen_msg_umpna(flag_write_db, 'umpna', 'UMPNA', 'PostgreSQL_Messages-Pumps'))
                msg.update(self.gen_msg_umpna(flag_write_db, 'umpna', 'KTPRAS_1', 'PostgreSQL_Messages-KTPRAS_1'))
                continue
            if tabl == 'KTPR': 
                msg.update(self.gen_msg_defence(flag_write_db, 'ktpr', 'KTPR', 'PostgreSQL_Messages-KTPR', 'TblStationDefences'))
                continue
            if tabl == 'KTPRP': 
                msg.update(self.gen_msg_defence(flag_write_db, 'ktprp', 'KTPRP', 'PostgreSQL_Messages-KTPRP', 'TblFireDefences'))
                continue
            if tabl == 'KTPRA': 
                msg.update(self.gen_msg_defence(flag_write_db, 'ktpra', 'KTPRA', 'PostgreSQL_Messages-KTPRA', 'TblPumpDefences'))
                continue
            if tabl == 'GMPNA': 
                msg.update(self.gen_msg_defence(flag_write_db, 'gmpna', 'GMPNA', 'PostgreSQL_Messages-GMPNA', 'TblPumpReadineses'))
                continue
            if tabl == 'KTPRS': 
                msg.update(self.gen_msg_defence(flag_write_db, 'ktprs', 'KTPRS', 'PostgreSQL_Messages-KTPRS', 'TblLimitParameters'))
                continue
            if tabl == 'Diag': 
                msg.update(self.gen_msg_diag(flag_write_db))
                continue
            if tabl == 'SS': 
                msg.update(self.gen_msg_defence(flag_write_db, 'ss', 'DiagSS', 'PostgreSQL_Messages-SS', 'TblD_RelatedSystems'))
                continue
            if tabl == 'DPS': 
                msg.update(self.gen_msg_defence(flag_write_db, 'dps', 'DPS', 'PostgreSQL_Messages-DPS', 'TblPigSignallers'))
                continue
            if tabl == 'Others': 
                msg.update(self.gen_msg_others(flag_write_db, 'msg_others', 'Others', 'PostgreSQL_Messages-Others'))
                continue
            if tabl == 'NPS': 
                msg.update(self.gen_msg_nps(flag_write_db, 'nps', 'NPS','PostgreSQL_Messages-NPS', 'TblNPS'))
                msg.update(self.gen_msg_nps(flag_write_db, 'krmpn', 'KRMPN','PostgreSQL_Messages-KRMPN', 'TblStationCommonKRMPN'))
                continue
            if tabl == 'PZ': 
                msg.update(self.gen_msg_firezone(flag_write_db, 'pz', 'PostgreSQL_Messages-PZ'))
                continue
            if tabl == 'PI': 
                msg.update(self.gen_msg_defence(flag_write_db, 'pi', 'PI', 'PostgreSQL_Messages-PI', 'TblFireDetectors'))
                continue
            if tabl == 'BD': 
                msg.update(self.gen_msg_defence(flag_write_db, 'bd', 'BD', 'PostgreSQL_Messages-BD', 'TblTankDispensers'))
                continue
            if tabl == 'BDGRP': 
                msg.update(self.gen_msg_defence(flag_write_db, 'bdgrp', 'BDGrp', 'PostgreSQL_Messages-BDGrp', 'TblTankDispenserGroups'))
                continue
            if tabl == 'Global': 
                msg.update(self.gen_msg_nps(flag_write_db, 'global', 'Global', 'PostgreSQL_Messages-Global', 'TblGlobal'))
                continue
            if tabl == 'TM_DP': 
                msg.update(self.gen_msg_defence(flag_write_db, 'tm_dp', 'DiagTM_DP', 'PostgreSQL_Messages-TMDP', 'TblD_TM_DP'))
                continue
        return msg
    def write_in_sql_tabl(self, list_tabl, flag_write_db):
            msg = {}
            if len(list_tabl) == 0: return
            for tabl in list_tabl: 
                if tabl == 'AI_tabl': 
                    msg.update(self.gen_table_AI(flag_write_db))
                    continue
                if tabl == 'ZD_tabl': 
                    msg.update(self.gen_table_general(flag_write_db, 'zd_tm', 'TblValveTimeSetpoints'))
                    continue
                if tabl == 'VS_tabl': 
                    msg.update(self.gen_table_general(flag_write_db, 'vs_tm', 'TblAuxSysTimeSetpoints'))
                    continue
                if tabl == 'VSGRP_tabl': 
                    msg.update(self.gen_table_general(flag_write_db, 'vsgrp_tm', 'TblAuxsysgrouptimeSetpoints'))
                    continue
                if tabl == 'Pump_tabl': 
                    msg.update(self.gen_table_general(flag_write_db, 'umpna_tm', 'TblPumptimeSetpoints'))
                    continue
                if tabl == 'UTS_tabl': 
                    msg.update(self.gen_table_general(flag_write_db, 'uts_tm', 'TblSignalingdevicetimeSetpoints'))
                    continue
                if tabl == 'Prj_tabl': 
                    msg.update(self.gen_table_general(flag_write_db, 'prj_tm', 'TblProjecttimeSetpoints'))
                    continue
                if tabl == 'PZ_tabl': 
                    msg.update(self.gen_table_general(flag_write_db, 'pz_tm', 'TblFirezonetimeSetpoints'))
                    continue
                if tabl == 'PumpTime_tabl': 
                    msg.update(self.gen_table_general(flag_write_db, 'umpna_narab_tm', 'TblOpTimeSetpoints'))
                    continue
                if tabl == 'KTPR_tabl': 
                    msg.update(self.gen_table_ktpr(flag_write_db))
                    continue
                if tabl == 'KTPRA_tabl': 
                    msg.update(self.gen_table_pumps(flag_write_db, 'ktpra', 'TblPumpDefencesSetpoints'))
                    continue
                if tabl == 'GMPNA_tabl': 
                    msg.update(self.gen_table_pumps(flag_write_db, 'gmpna', 'TblPumpreadinesesSetpoints'))
                    continue
            return msg
    def synh_in_sql(self, list_tabl):
            msg = {}
            if len(list_tabl) == 0: return
            for tabl in list_tabl: 
                if tabl == 'AI_tabl': 
                    msg.update(self.synh_tabl_ai())
                    continue
                if tabl == 'ZD_tabl': 
                    msg.update(self.synh_tabl('tblvalvetimesetpoints', 'zd_tm', 'TblValveTimeSetPoints'))
                    continue
                if tabl == 'VS_tabl': 
                    msg.update(self.synh_tabl('tblauxsystimesetpoints', 'vs_tm', 'TblAuxsysTimeSetPoints'))
                    continue
                if tabl == 'VSGRP_tabl': 
                    msg.update(self.synh_tabl('tblauxsysgrouptimesetpoints', 'vsgrp_tm', 'TblAuxsysgrouptimeSetpoints'))
                    continue
                if tabl == 'Pump_tabl': 
                    msg.update(self.synh_tabl('tblpumptimesetpoints', 'umpna_tm', 'TblPumptimeSetpoints'))
                    continue
                if tabl == 'UTS_tabl': 
                    msg.update(self.synh_tabl('tblsignalingdevicetimesetpoints', 'uts_tm', 'TblSignalingdevicetimeSetpoints'))
                    continue
                if tabl == 'Prj_tabl': 
                    msg.update(self.synh_tabl('tblprojecttimesetpoints', 'prj_tm', 'TblProjecttimeSetpoints'))
                    continue
                # if tabl == 'PZ_tabl': 
                #     msg.update(self.gen_table_general(flag_write_db, 'pz_tm', 'TblFirezonetimeSetpoints'))
                #     continue
                # if tabl == 'PumpTime_tabl': 
                #     msg.update(self.gen_table_general(flag_write_db, 'umpna_narab_tm', 'TblOpTimeSetpoints'))
                #     continue
                if tabl == 'KTPR_tabl':  
                    msg.update(self.synh_tabl('tblstationdefencessetpoints', 'ktpr', 'TblStationDefencesSetpoints'))
                    continue
                if tabl == 'KTPRA_tabl': 
                    msg.update(self.synh_tabl('tblpumpdefencessetpoints', 'ktpra', 'TblPumpDefencesSetpoints'))
                    continue
                if tabl == 'GMPNA_tabl': 
                    msg.update(self.synh_tabl('tblpumpreadinesessetpoints', 'gmpna', 'TblPumpreadinesesSetpoints'))
                    continue
            return msg
    # msg
    def gen_msg_ai(self, flag_write_db):
        msg = {}
        gen_list = []
        cursor = db.cursor()
        try:
            kod_msg, addr_offset = self.define_number_msg(cursor, 'AI')
            if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                msg[f'{today} - Сообщения ai: ошибка. Адреса из таблицы msg не определены'] = 2
                return msg
            cursor.execute(f"""SELECT id, "name", "AnalogGroupId" FROM ai""")
            list_ai = cursor.fetchall()
            for analog in list_ai:
                id_ai    = analog[0]
                name_ai  = analog[1]
                group_ai = analog[2]

                start_addr = kod_msg + ((id_ai - 1) * int(addr_offset))
                try:
                    cursor.execute(f"""SELECT "table_msg" 
                                       FROM ai_grp
                                       WHERE name_group='{group_ai}'""")
                    list_group = cursor.fetchall()[0][0]
                    path = f'{path_sample}\{list_group}.xml'
                    if not os.path.isfile(path):
                        msg[f'{today} - Сообщения ai: отсутствует шаблон - {list_group}'] = 2
                        continue
                    gen_list.append(self.dop_function.parser_sample(path, start_addr, name_ai, flag_write_db, 'AI'))
                except Exception:
                    msg[f'{today} - Сообщения ai: отсутствует шаблон: {id_ai} - {name_ai}'] = 2
                    continue
            if not flag_write_db:
                msg.update(self.write_file(gen_list, 'AI', 'PostgreSQL_Messages-AI'))
                msg[f'{today} - Сообщения ai: файл скрипта создан'] = 1
                return(msg)
        except Exception:
            msg[f'{today} - Сообщения ai: ошибка генерации: {traceback.format_exc()}'] = 2
        msg[f'{today} - Сообщения ai: генерация завершена'] = 1
        return(msg)
    def gen_msg_umpna(self, flag_write_db, tabl, sign, script_file):
        msg = {}
        gen_list = []
        cursor = db.cursor()
        try:
            kod_msg, addr_offset = self.define_number_msg(cursor, sign)
            if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                msg[f'{today} - Сообщения {tabl}: ошибка. Адреса из таблицы msg не определены'] = 2
                return msg
            
            cursor.execute(f"""SELECT id, name, tabl_msg, replacement_uso_signal_vv_1, replacement_uso_signal_vv_2
                                FROM "{tabl}" ORDER BY id""")
            list_signal = cursor.fetchall()
            for signal in list_signal:
                id_       = signal[0]
                name      = signal[1]
                table_msg = signal[2]
                cabinet_1 = signal[3]
                cabinet_2 = signal[4]

                if sign == 'KTPRAS_1': table_msg = 'TblPumpsKTPRAS'

                start_addr = kod_msg + ((id_ - 1) * int(addr_offset))
                path = f'{path_sample}\{table_msg}.xml'
                if not os.path.isfile(path):
                    msg[f'{today} - Сообщения {tabl}: в папке отсутствует шаблон - {table_msg}'] = 2
                    continue
                gen_list.append(self.dop_function.parser_sample(path, start_addr, name, flag_write_db, sign, cabinet_1, cabinet_2))

            if not flag_write_db:
                msg.update(self.write_file(gen_list, sign, script_file))
                msg[f'{today} - Сообщения {tabl}: файл скрипта создан'] = 1
                return(msg)
        except Exception:
            msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
        msg[f'{today} - Сообщения {tabl}: генерация завершена!'] = 1
        return(msg)
    def gen_msg_uts_upts(self, flag_write_db, tabl, sign, script_file):
        msg = {}
        gen_list = []
        cursor = db.cursor()
        try:
            kod_msg, addr_offset = self.define_number_msg(cursor, sign)
            if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                msg[f'{today} - Сообщения {tabl}: ошибка. Адреса из таблицы msg не определены'] = 2
                return msg

            cursor.execute(f"""SELECT id, name FROM "{tabl}" ORDER BY id""")
            list_signal = cursor.fetchall()
            for signal in list_signal:
                id_       = signal[0]
                name      = signal[1]
                
                if tabl == 'uts':
                    if   self.dop_function.str_find(str(name).lower(), {'звонок'}): table_msg = 'TblSignalingDevicesMale'
                    elif self.dop_function.str_find(str(name).lower(), {'табло'}) : table_msg = 'TblSignalingDevices'
                    elif self.dop_function.str_find(str(name).lower(), {'сирена'}): table_msg = 'TblSignalingDevicesFemale'
                    elif self.dop_function.str_find(str(name).lower(), {'сирены'}): table_msg = 'TblSignalingDevicesMany'
                    elif self.dop_function.str_find(str(name).lower(), {'сигнализация'}): table_msg = 'TblSignalingDevicesFemale'
                    else: table_msg = 'TblSignalingDevices'
                else:
                    table_msg = 'TblFireSignalingDevices'

                start_addr = kod_msg + ((id_ - 1) * int(addr_offset))
                path = f'{path_sample}\{table_msg}.xml'
                if not os.path.isfile(path):
                    msg[f'{today} - Сообщения {tabl}: в папке отсутствует шаблон - {table_msg}'] = 2
                    return msg

                gen_list.append(self.dop_function.parser_sample(path, start_addr, name, flag_write_db, sign))

            if not flag_write_db:
                msg.update(self.write_file(gen_list, sign, script_file))
                msg[f'{today} - Сообщения {tabl}: файл скрипта создан'] = 1
                return(msg)
        except Exception:
            msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
        msg[f'{today} - Сообщения {tabl}: генерация завершена!'] = 1
        return(msg)
    def gen_msg_defence(self, flag_write_db, tabl, sign, script_file, table_msg):
        msg = {}
        gen_list = []
        cursor = db.cursor()
        try:
            kod_msg, addr_offset = self.define_number_msg(cursor, sign)
            if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                msg[f'{today} - Сообщения {tabl}: ошибка. Адреса из таблицы msg не определены'] = 2
                return msg
            
            if sign == 'KTPRA' or sign == 'GMPNA':
                cursor.execute(f"""SELECT id, name, "NA" FROM "{tabl}" ORDER BY id""")
            else:
                cursor.execute(f"""SELECT id, name FROM "{tabl}" ORDER BY id""")
            list_signal = cursor.fetchall()
            for signal in list_signal:
                id_       = signal[0]
                name      = signal[1]
                if sign == 'KTPRA' or sign == 'GMPNA': na = signal[2]

                start_addr = kod_msg + ((id_ - 1) * int(addr_offset))
                path = f'{path_sample}\{table_msg}.xml'
                if not os.path.isfile(path):
                    msg[f'{today} - Сообщения {tabl}: в папке отсутствует шаблон - {table_msg}'] = 2
                    return msg
                if sign == 'KTPRA' or sign == 'GMPNA':
                    gen_list.append(self.dop_function.parser_sample(path, start_addr, f'{na}. {name}', flag_write_db, sign))
                if sign == 'KTPRP':
                    gen_list.append(self.dop_function.parser_sample(path, start_addr, f'Пожарная защита. {name}', flag_write_db, sign))
                else:
                    gen_list.append(self.dop_function.parser_sample(path, start_addr, name, flag_write_db, sign))
            if not flag_write_db:
                msg.update(self.write_file(gen_list, sign, script_file))
                msg[f'{today} - Сообщения {tabl}: файл скрипта создан'] = 1
                return(msg)
        except Exception:
            msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
        msg[f'{today} - Сообщения {tabl}: генерация завершена!'] = 1
        return(msg)
    def gen_msg_general(self, flag_write_db, tabl, sign, script_file):
        msg = {}
        gen_list = []
        cursor = db.cursor()
        try:
            kod_msg, addr_offset = self.define_number_msg(cursor, sign)
            if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                msg[f'{today} - Сообщения {tabl}: ошибка. Адреса из таблицы msg не определены'] = 2
                return msg
            
            cursor.execute(f"""SELECT id, name, tabl_msg FROM "{tabl}" ORDER BY id""")
            list_signal = cursor.fetchall()
            for signal in list_signal:
                id_       = signal[0]
                name      = signal[1]
                table_msg = signal[2]

                start_addr = kod_msg + ((id_ - 1) * int(addr_offset))
                path = f'{path_sample}\{table_msg}.xml'
                if not os.path.isfile(path):
                    msg[f'{today} - Сообщения {tabl}: в папке отсутствует шаблон - {table_msg}'] = 2
                    continue
                gen_list.append(self.dop_function.parser_sample(path, start_addr, name, flag_write_db, sign))
            if not flag_write_db:
                msg.update(self.write_file(gen_list, sign, script_file))
                msg[f'{today} - Сообщения {tabl}: файл скрипта создан'] = 1
                return(msg)
        except Exception:
            msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
        msg[f'{today} - Сообщения {tabl}: генерация завершена!'] = 1
        return(msg)
    def gen_msg_diag(self, flag_write_db):
        msg = {}
        modul_list = []
        cursor = db.cursor()

        tabl = 'hardware' 
        count_CN, count_CPU, count_EthEx = 0, 0, 0
        count_MN, count_PCU, count_RS = 0, 0, 0 
        try:
            for column in HardWare.select().dicts():
                id_basket = column['id']
                uso       = column['uso']
                basket    = column['basket']
                for five_column in range(0, 33, 1):
                    if column[f'type_{five_column}'] != '' and column[f'type_{five_column}'] is not None:
                        type_modul = column[f'type_{five_column}']
                        prefix_number = f'0{five_column}' if five_column < 10 else basket

                        value = f'Диагностика. {uso}. Модуль А{basket}.{prefix_number} {type_modul}'
                        modul_list.append(dict(id          = id_basket,
                                                num_modul  = five_column,
                                                type_modul = type_modul,
                                                value      = value,
                                                basket     = basket))
            for i in range(2):
                gen_list = []
                if i == 0:
                    sign = 'DiagMod' 
                    script_file = 'PostgreSQL_Messages-Racks' 
                    table_msg = 'TblD_Racks'
                    tbl_racks = True

                    kod_msg, addr_offset = self.define_number_msg(cursor, sign)
                    if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                        msg[f'{today} - Сообщения {tabl}: адрес {tabl} из таблицы msg не определен'] = 2
                        return msg
                else:
                    script_file = 'PostgreSQL_Messages-Modul' 
                    tbl_racks = False
                    for j in range(6):
                        if   j == 0: sign = 'DiagCN'
                        elif j == 1: sign = 'DiagCPU'
                        elif j == 2: sign = 'DiagEthEx'
                        elif j == 3: sign = 'DiagMN'
                        elif j == 4: sign = 'DiagCPUKC'
                        elif j == 5: sign = 'DiagRS'

                        kod_msg, addr_offset = self.define_number_msg(cursor, sign)
                        if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                            msg[f'{today} - Сообщения {tabl}: адрес {tabl} из таблицы msg не определен'] = 2
                            return msg
                        
                        if   j == 0: 
                            kod_msg_CN     = kod_msg
                            addr_offset_CN = addr_offset
                        elif j == 1: 
                            kod_msg_CPU     = kod_msg
                            addr_offset_CPU = addr_offset
                        elif j == 2: 
                            kod_msg_EthEx     = kod_msg
                            addr_offset_EthEx = addr_offset
                        elif j == 3: 
                            kod_msg_MN     = kod_msg
                            addr_offset_MN = addr_offset
                        elif j == 4: 
                            kod_msg_PCU     = kod_msg
                            addr_offset_PCU = addr_offset
                        elif j == 5: 
                            kod_msg_RS     = kod_msg
                            addr_offset_RS = addr_offset
                    
                for modul in modul_list:
                    id_basket    = modul['id']
                    number_modul = modul['num_modul']
                    type_modul   = modul['type_modul']
                    value_modul  = modul['value']
                    basket       = modul['basket']

                    if tbl_racks is True:
                        offset_basket = 32 * 14 * (int(id_basket) - 1)
                        start_addr = kod_msg + offset_basket + (number_modul* int(addr_offset))
                    else:
                        if   type_modul == 'MK-545-010': 
                            start_addr = kod_msg_CN + (count_CN * int(addr_offset_CN)) 
                            table_msg = 'TblD_ModulesCN'
                            count_CN += 1
                        elif type_modul == 'MK-504-120': 
                            start_addr = kod_msg_CPU + (count_CPU * int(addr_offset_CPU)) 
                            table_msg = 'TblD_ModulesCPU'
                            count_CPU += 1
                        elif type_modul == 'MK-544-040': 
                            start_addr = kod_msg_EthEx + (count_EthEx * int(addr_offset_EthEx)) 
                            table_msg = 'TblD_ModulesEthEx'
                            count_EthEx += 1
                        elif type_modul == 'MK-546-010': 
                            start_addr = kod_msg_MN + (count_MN * int(addr_offset_MN)) 
                            table_msg = 'TblD_ModulesMN'
                            count_MN += 1
                        elif type_modul == 'MK-550-024': 
                            start_addr = kod_msg_PCU + (count_PCU * int(addr_offset_PCU)) 
                            table_msg = 'TblD_ModulesPSU'
                            count_PCU += 1
                        elif type_modul == 'MK-541-002': 
                            start_addr = kod_msg_RS + (count_RS * int(addr_offset_RS)) 
                            table_msg = 'TblD_ModulesRS'
                            count_RS += 1

                    path = f'{path_sample}\{table_msg}.xml'
                    if not os.path.isfile(path):
                        msg[f'{today} - Сообщения {tabl}: в папке отсутствует шаблон - {table_msg}'] = 2
                        return msg

                    gen_list.append(self.dop_function.parser_sample(path, start_addr, value_modul, flag_write_db, sign))

                if not flag_write_db:
                    msg.update(self.write_file(gen_list, sign, script_file))
                    msg[f'{today} - Сообщения {tabl}: файл скрипта создан'] = 1
        except Exception:
            msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
        msg[f'{today} - Сообщения {tabl}: генерация завершена!'] = 1
        return(msg)
    def gen_msg_others(self, flag_write_db, tabl, sign, script_file):
        msg = {}
        gen_list = []
        cursor = db.cursor()
        cursor_prj = db_prj.cursor()
        
        kod_msg, addr_offset = self.define_number_msg(cursor, sign)
        if addr_offset == 0 or kod_msg is None or addr_offset is None: 
            msg[f'{today} - Сообщения {tabl}: ошибка. Адреса из таблицы msg не определены'] = 2
            return msg
        try:
            cursor.execute(f"""SELECT id, text, priority, "isAck", "IsAlert", "IsCycle", "IsSound", "SoundFile", "IsHide"
                                FROM "{tabl}" ORDER BY id""")
            list_signal = cursor.fetchall()
        except Exception:
            msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
            return msg

        for signal in list_signal:
            id_       = signal[0]
            text      = signal[1]
            priority  = signal[2]
            isAck     = signal[3]
            IsAlert   = signal[4]
            IsCycle   = signal[5]
            IsSound   = signal[6]
            SoundFile = signal[7]
            IsHide    = signal[8]

            if SoundFile is None: SoundFile = ''

            del_row_tabl = f"DELETE FROM messages.opmessages WHERE Category ={kod_msg + int(id_)};\n"
            ins_row_tabl = f"INSERT INTO messages.opmessages (Category, Message, IsAck, SoundFile, IsCycle, IsSound, IsHide, Priority, IsAlert) VALUES({kod_msg + int(id_)}, '{text}', {isAck}, '{SoundFile}', {IsCycle}, {IsSound}, {IsHide}, {priority}, {IsAlert});\n"

            if flag_write_db:
                try:
                    cursor_prj.execute(del_row_tabl)
                    cursor_prj.execute(ins_row_tabl)
                except Exception:
                    msg[f'{today} - Сообщения {tabl}: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                    continue
            else:
                gen_list.append(dict(delete = del_row_tabl,
                                     insert = ins_row_tabl))
    
        if not flag_write_db:
            msg.update(self.write_file(gen_list, sign, script_file))
            msg[f'{today} - Сообщения {tabl}: файл скрипта создан'] = 1
            return(msg)

        msg[f'{today} - Сообщения {tabl}: генерация завершена!'] = 1
        return(msg)
    def gen_msg_nps(self, flag_write_db, tabl, sign, script_file, table_msg):
        msg = {}
        gen_list = []
        cursor = db.cursor()

        try:
            kod_msg, addr_offset = self.define_number_msg(cursor, sign)
            if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                msg[f'{today} - Сообщения {tabl}: ошибка. Адреса из таблицы msg не определены'] = 2
                return msg
            
            path = f'{path_sample}\{table_msg}.xml'
            if not os.path.isfile(path):
                msg[f'{today} - Сообщения {tabl}: в папке отсутствует шаблон - {table_msg}'] = 2
                return msg

            gen_list.append(self.dop_function.parser_sample(path, kod_msg, '', flag_write_db, sign))
        
            if not flag_write_db:
                msg.update(self.write_file(gen_list, sign, script_file))
                msg[f'{today} - Сообщения {tabl}: файл скрипта создан'] = 1
                return(msg)
        except Exception:
            msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
        msg[f'{today} - Сообщения {tabl}: генерация завершена!'] = 1
        return(msg)
    def gen_msg_firezone(self, flag_write_db, tabl, script_file):
        msg = {}
        gen_list = []
        cursor = db.cursor()
        try:  
            for j in range(7):
                if   j == 0: sign = 'SPZ'
                elif j == 1: sign = 'GPZFoam'
                elif j == 2: sign = 'GPZWater'
                elif j == 3: sign = 'SUP'
                elif j == 4: sign = 'ATP'
                elif j == 5: sign = 'GPZWOF'
                elif j == 6: sign = 'GPZGas'

                kod_msg, addr_offset = self.define_number_msg(cursor, sign)
                if addr_offset == 0 or kod_msg is None or addr_offset is None: 
                    msg[f'{today} - Сообщения {tabl}: адрес {sign} из таблицы msg не определен'] = 2
                    msg[f'{today} - Сообщения {tabl}: генерация сообщений без: {sign}'] = 2
                    continue 
                
                if   j == 0: 
                    kod_msg_SPZ    = kod_msg
                    addr_offset_SPZ = addr_offset
                elif j == 1: 
                    kod_msg_GPZFoam     = kod_msg
                    addr_offset_GPZFoam = addr_offset
                elif j == 2: 
                    kod_msg_GPZWater     = kod_msg
                    addr_offset_GPZWater = addr_offset
                elif j == 3: 
                    kod_msg_SUP     = kod_msg
                    addr_offset_SUP = addr_offset
                elif j == 4: 
                    kod_msg_ATP     = kod_msg
                    addr_offset_ATP = addr_offset
                elif j == 5: 
                    kod_msg_GPZWOF     = kod_msg
                    addr_offset_GPZWOF = addr_offset
                elif j == 6: 
                    kod_msg_GPZGas     = kod_msg
                    addr_offset_GPZGas = addr_offset
            
            list_sample = ['TblFireZonesState', 'TblFireZonesGPZFoam', 'TblFireZonesGPZWater', 'TblFireZonesMode',
                            'TblFireZonesAPT', 'TblFireZonesGPZWithout', 'TblFireZonesGPZGas']
            for i in list_sample:
                path = f'{path_sample}\{i}.xml'
                if not os.path.isfile(path):
                    msg[f'{today} - Сообщения {tabl}: в папке отсутствует шаблон - {i}'] = 2

            cursor.execute(f"""SELECT id, name, "type_zone" FROM "{tabl}" ORDER BY id""")
            list_zone = cursor.fetchall()

            for zone in list_zone:
                id_       = zone[0]
                name      = zone[1]
                type_zone = zone[2]

                try:
                    for i in range(7):
                        if i == 0:
                            start_addr = kod_msg_SPZ + ((int(id_) - 1) * int(addr_offset_SPZ)) 
                            table_msg = 'TblFireZonesState'
                            text = f'Пожарные зоны. {name}'
                        elif i == 1 and type_zone == -1:
                            start_addr = kod_msg_GPZFoam + ((int(id_) - 1) * int(addr_offset_GPZFoam)) 
                            table_msg = 'TblFireZonesGPZFoam'
                            text = f'Готовности зон. {name}'
                        elif i == 2 and type_zone >= 1:
                            start_addr = kod_msg_GPZWater + ((int(id_) - 1) * int(addr_offset_GPZWater)) 
                            table_msg = 'TblFireZonesGPZWater'
                            text = f'Готовности зон. {name}'
                        elif i == 3:
                            start_addr = kod_msg_SUP + ((int(id_) - 1) * int(addr_offset_SUP)) 
                            table_msg = 'TblFireZonesMode'
                            text = f'Пожарные зоны. {name}'
                        elif i == 4:
                            start_addr = kod_msg_ATP + ((int(id_) - 1) * int(addr_offset_ATP)) 
                            table_msg = 'TblFireZonesAPT'
                            text = f'Пожарные зоны. {name}'
                        elif i == 5 and type_zone == 0:
                            start_addr = kod_msg_GPZWOF + ((int(id_) - 1) * int(addr_offset_GPZWOF)) 
                            table_msg = 'TblFireZonesGPZWithout'
                            text = f'Готовности зон. {name}'
                        elif i == 6 and type_zone == -2:
                            start_addr = kod_msg_GPZGas + ((int(id_) - 1) * int(addr_offset_GPZGas)) 
                            table_msg = 'TblFireZonesGPZGas'
                            text = f'Готовности зон. {name}'
                        else: continue

                        path = f'{path_sample}\{table_msg}.xml'
                        if not os.path.isfile(path):
                            continue
                        gen_list.append(self.dop_function.parser_sample(path, start_addr, text, flag_write_db, sign))
                except Exception:
                    msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
                    msg[f'{today} - Сообщения {tabl}: генерация продолжится'] = 2
                    continue
        
            if not flag_write_db:
                msg.update(self.write_file(gen_list, sign, script_file))
                msg[f'{today} - Сообщения {tabl}: файл скрипта создан'] = 1
                return(msg)
        except Exception:
            msg[f'{today} - Сообщения {tabl}: ошибка генерации: {traceback.format_exc()}'] = 2
        msg[f'{today} - Сообщения {tabl}: генерация завершена!'] = 1
        return(msg)
    # tabl
    def gen_table_AI(self, flag_write_db):
        cursor = db.cursor()
        cursor_prj = db_prj.cursor()
    
        text_start = ('\tCREATE SCHEMA IF NOT EXISTS objects;\n'
                        '\tCREATE TABLE IF NOT EXISTS objects.TblAnalogs(\n'
                        '\t\tId INT NOT NULL,\n'
                        '\t\tPrefix VARCHAR(1024),\n'
                        '\t\tSystemIndex INT NOT NULL,\n'
                        '\t\tTag VARCHAR(1024),\n'
                        '\t\tName VARCHAR(1024),\n'
                       '\t\tAnalogGroupId INT,\n'
                       '\t\tSetpointGroupId INT,\n'
                        '\t\tEgu VARCHAR(1024),\n'
                        '\t\tPhysicEgu VARCHAR(1024),\n'
                        '\t\tIsOilPressure BOOLEAN NOT NULL,\n'
                        '\t\tIsInterface BOOLEAN NOT NULL,\n'
                        '\t\tIsPhysic BOOLEAN NOT NULL,\n'
                        '\t\tIsPumpVibration BOOLEAN,\n'
                        '\t\tPrecision INT NOT NULL,\n'
                        '\t\tIsTrending BOOLEAN NOT NULL,\n'
                        '\t\tTrendingSettings VARCHAR(1024),\n'
                        '\t\tTrendingGroup INT,\n'
                        '\t\tLoLimField DOUBLE PRECISION,\n'
                        '\t\tHiLimField DOUBLE PRECISION,\n'
                        '\t\tLoLimEng DOUBLE PRECISION,\n'
                        '\t\tHiLimEng DOUBLE PRECISION,\n'
                        '\t\tLoLim DOUBLE PRECISION,\n'
                        '\t\tHiLim DOUBLE PRECISION,\n'
                        '\t\tMin6 DOUBLE PRECISION,\n'
                        '\t\tMin5 DOUBLE PRECISION,\n'
                        '\t\tMin4 DOUBLE PRECISION,\n'
                        '\t\tMin3 DOUBLE PRECISION,\n'
                        '\t\tMin2 DOUBLE PRECISION,\n'
                        '\t\tMin1 DOUBLE PRECISION,\n'
                        '\t\tMax1 DOUBLE PRECISION,\n'
                        '\t\tMax2 DOUBLE PRECISION,\n'
                        '\t\tMax3 DOUBLE PRECISION,\n'
                        '\t\tMax4 DOUBLE PRECISION,\n'
                        '\t\tMax5 DOUBLE PRECISION,\n'
                        '\t\tMax6 DOUBLE PRECISION,\n'
                        '\t\tHisteresis DOUBLE PRECISION,\n'
                        '\t\tDeltaHi DOUBLE PRECISION,\n'
                        '\t\tDeltaLo DOUBLE PRECISION,\n'
                        '\t\tDeltaT DOUBLE PRECISION,\n'
                        '\t\tSmoothFactor DOUBLE PRECISION,\n'
                        '\t\tCtrl SMALLINT,\n'
                        '\t\tMsgMask INT,\n'
                        '\t\tSigMask INT,\n'
                        '\t\tCtrlMask SMALLINT,\n'
                        '\t\tTimeFilter DOUBLE PRECISION,\n'
                        '\t\tIsBackup BOOLEAN NOT NULL,\n'
                        '\t\tRuleName VARCHAR(1024),\n'
                        '\t\tCONSTRAINT TblAnalogs_pkey PRIMARY KEY (Id,SystemIndex)\n'
                    '\t);\n'
                    '\tDELETE FROM objects.TblAnalogs  WHERE SystemIndex = 0;\n')
        
        msg = {}
        gen_list = []
        flag_del_tabl = False
        try:
            cursor.execute(f"""SELECT "id", "tag", "name", "AnalogGroupId", "SetpointGroupId", "Egu", "PhysicEgu", "IsOilPressure", 
                                        "IsPumpVibration", "Precision", "TrendingGroup", "LoLimField", "HiLimField", "LoLimEng", 
                                        "HiLimEng", "LoLim", "HiLim", "Min6", "Min5", "Min4", "Min3", "Min2", "Min1", "Max1", "Max2", "Max3", 
                                        "Max4", "Max5", "Max6", "Histeresis", "DeltaT", "MsgMask", "SigMask", "CtrlMask", "RuleName", "TimeFilter", "module","channel"
                                FROM "ai" ORDER BY Id""")
            list_signal = cursor.fetchall()
        except Exception:
            msg[f'{today} - TblAnalogs: ошибка генерации: {traceback.format_exc()}'] = 2
            return msg

        for signal in list_signal:
            try:
                Id, Tag, Name, AnalogGroupId                          = signal[0], signal[1], signal[2], signal[3]
                SetpointGroupId, Egu, PhysicEgu, IsOilPressure        = signal[4], signal[5], signal[6], signal[7]
                IsPumpVibration, Precision, TrendingGroup, LoLimField = signal[8], signal[9], signal[10], signal[11]
                HiLimField, LoLimEng, HiLimEng, LoLim                 = signal[12], signal[13], signal[14], signal[15]
                HiLim, Min6, Min5, Min4                               = signal[16], signal[17], signal[18], signal[19]
                Min3, Min2, Min1, Max1                                = signal[20], signal[21], signal[22], signal[23]
                Max2, Max3, Max4, Max5                                = signal[24], signal[25], signal[26], signal[27]
                Max6, Histeresis, DeltaT, MsgMask                     = signal[28], signal[29], signal[30], signal[31]
                SigMask, CtrlMask, RuleName, TimeFilter               = signal[32], signal[33], signal[34], signal[35]
                module, channel                                       = signal[36], signal[37]

                # Prefix
                Prefix = 'NULL' if prefix_system == '' or prefix_system is None else str(prefix_system)
                # SystemIndex
                SystemIndex = 0
                # AnalogGroupId
                cursor.execute(f"""SELECT id FROM "ai_grp" WHERE name='{AnalogGroupId}'""")
                try   : AnalogGroupId = cursor.fetchall()[0][0]
                except: AnalogGroupId = 'NULL'
                # SetpointGroupId
                cursor.execute(f"""SELECT id FROM "sp_grp" WHERE name_group='{SetpointGroupId}'""")
                try   : SetpointGroupId = cursor.fetchall()[0][0]
                except: SetpointGroupId = 'NULL'
                # IsOilPressure
                IsOilPressure = False if IsOilPressure is None else IsOilPressure
                # IsPumpVibration
                IsPumpVibration = False if IsPumpVibration is None else IsPumpVibration
                if IsPumpVibration == 1: IsPumpVibration = True
                # IsInterface
                IsInterface = False
                # IsBackup
                IsBackup = True if self.dop_function.str_find(str(Name).lower(), {'резерв'}) else False
                # IsPhysic
                IsPhysic = True if module is not None and channel is not None and IsBackup is False else False
                # IsTrending
                IsTrending = True if IsBackup is False else False

                TrendingGroup = 'NULL' if TrendingGroup is None else TrendingGroup
                LoLimEng = 'NULL' if HiLimEng is None else HiLimEng
                LoLim   = 'NULL' if LoLim is None else LoLim
                HiLim = 'NULL' if HiLim is None else HiLim

                # Ctrl
                Ctrl_list = ['0000', '0', '0','0','0','0','0','0','0','0','0','0','0']
                if Min6 is None: Min6 = 'NULL'
                else: 
                    Min6 = Min6
                    Ctrl_list[12] = '1'
                if Min5 is None: Min5 = 'NULL'
                else: 
                    Min5 = Min5
                    Ctrl_list[11] = '1'
                if Min4 is None: Min4 = 'NULL'
                else: 
                    Min4 = Min4
                    Ctrl_list[10] = '1'
                if Min3 is None: Min3 = 'NULL'
                else: 
                    Min3 = Min3
                    Ctrl_list[9] = '1'
                if Min2 is None: Min2 = 'NULL'
                else: 
                    Min2 = Min2
                    Ctrl_list[8] = '1'
                if Min1 is None: Min1 = 'NULL'
                else: 
                    Min1 = Min1
                    Ctrl_list[7] = '1'            
                if Max1 is None: Max1 = 'NULL'
                else: 
                    Max1 = Max1
                    Ctrl_list[6] = '1'
                if Max2 is None: Max2 = 'NULL'
                else: 
                    Max2 = Max2
                    Ctrl_list[5] = '1'
                if Max3 is None: Max3 = 'NULL'
                else: 
                    Max3 = Max3
                    Ctrl_list[4] = '1'
                if Max4 is None: Max4 = 'NULL'
                else: 
                    Max4 = Max4
                    Ctrl_list[3] = '1'
                if Max5 is None: Max5 = 'NULL'
                else: 
                    Max5 = Max5
                    Ctrl_list[2] = '1'
                if Max6 is None: Max6 = 'NULL'
                else: 
                    Max6 = Max6
                    Ctrl_list[1] = '1'
                Ctrl = int(''.join(Ctrl_list), 2)
                # LoLimField
                LoLimField = 'NULL' if LoLimField is None else LoLimField
                # HiLimField
                HiLimField = 'NULL' if HiLimField is None else HiLimField
                # LoLimEng
                LoLimEng = 'NULL' if LoLimEng is None else LoLimEng
                # HiLimEng
                HiLimEng = 'NULL' if HiLimEng is None else HiLimEng
                # LoLim
                LoLim = 'NULL' if LoLim is None else LoLim
                # HiLim
                HiLim = 'NULL' if HiLim is None else HiLim
                # DeltaHi
                DeltaHi = 'NULL'
                # DeltaLo
                DeltaLo = 'NULL'
                # SmoothFactor
                SmoothFactor = 'NULL'
                # MsgMask
                MsgMask = int(str(MsgMask).replace('_', ''), 2)
                # SigMask
                SigMask = int(str(SigMask).replace('_', ''), 2)
                # CtrlMask
                CtrlMask = int(str(CtrlMask).replace('_', ''), 2)
                # RuleName
                cursor.execute(f"""SELECT rule_name FROM "sp_rules" WHERE name_rules='{RuleName}'""")
                try   : RuleName = cursor.fetchall()[0][0]
                except: RuleName = 'NULL'
            except Exception:
                msg[f'{today} - TblAnalogs: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                continue
            
            ins_row_tabl = f"INSERT INTO objects.TblAnalogs (Id, Prefix, SystemIndex, Tag, Name, AnalogGroupId, SetpointGroupId, Egu, PhysicEgu, IsOilPressure, IsInterface, IsPhysic, IsPumpVibration, Precision, IsTrending, TrendingSettings, TrendingGroup, LoLimField, HiLimField, LoLimEng, HiLimEng, LoLim, HiLim, Min6, Min5, Min4, Min3, Min2, Min1, Max1, Max2, Max3, Max4, Max5, Max6, Histeresis, DeltaHi, DeltaLo, DeltaT, SmoothFactor, Ctrl, MsgMask, SigMask, CtrlMask, TimeFilter, IsBackup, RuleName) VALUES({Id}, {Prefix}, {SystemIndex}, '{Tag}','{Name}', {AnalogGroupId}, {SetpointGroupId}, '{Egu}', '{PhysicEgu}', {IsOilPressure}, {IsInterface}, {IsPhysic}, {IsPumpVibration}, {Precision}, {IsTrending}, 'Historian(Collector = NA_ModbusServer, sourceaddress = %MF{999 + 2 * Id}, InputScaling = 0)', {TrendingGroup}, {LoLimField}, {HiLimField}, {LoLimEng}, {HiLimEng}, {LoLim}, {HiLim}, {Min6}, {Min5}, {Min4}, {Min3}, {Min2}, {Min1}, {Max1}, {Max2}, {Max3}, {Max4}, {Max5}, {Max6}, {Histeresis}, {DeltaHi}, {DeltaLo}, {DeltaT}, {SmoothFactor}, {Ctrl}, {MsgMask}, {SigMask}, {CtrlMask}, {TimeFilter}, {IsBackup}, {RuleName});\n"
            
            if flag_write_db:
                try:
                    if flag_del_tabl is False :
                        cursor_prj.execute(text_start)
                        flag_del_tabl = True
                    cursor_prj.execute(ins_row_tabl)
                except Exception:
                    msg[f'{today} - TblAnalogs: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                    continue
            else:
                gen_list.append(ins_row_tabl)
    
        if not flag_write_db:
            try:
                # Создаём файл запроса
                path_request = f'{path_location_file}\\PostgreSQL-TblAnalogs.sql'
                if not os.path.exists(path_request):
                    file = codecs.open(path_request, 'w', 'utf-8')
                else:
                    os.remove(path_request)
                    file = codecs.open(path_request, 'w', 'utf-8')
                if path_location_file == '' or path_location_file is None or len(path_location_file) == 0:
                    msg[f'{today} - TblAnalogs: не указана конечная папка'] = 2
                    return msg
                file.write(text_start)
                for insert in gen_list:
                    file.write(insert)
                file.write(f'COMMIT;')
                file.close()
                msg[f'{today} - TblAnalogs: файл скрипта создан'] = 1
                return(msg)
            except Exception:
                msg[f'{today} - TblAnalogs: ошибка записи в файл: {traceback.format_exc()}'] = 2

        msg[f'{today} - TblAnalogs: генерация завершена!'] = 1
        return(msg)
    def gen_table_general(self, flag_write_db, tabl_sql, sign):
        cursor = db.cursor()
        cursor_prj = db_prj.cursor()
    
        text_start = ('\tCREATE SCHEMA IF NOT EXISTS objects;\n'
                        f'\tCREATE TABLE IF NOT EXISTS objects.{sign}(\n'
                        '\t\tId INT NOT NULL,\n'
                        '\t\tPrefix VARCHAR(1024),\n'
                        '\t\tSetpointGroupId INT,\n'
                        '\t\tTag VARCHAR(1024),\n'
                        '\t\tName VARCHAR(1024),\n'
                        '\t\tSource VARCHAR(1024),\n'
                        '\t\tValue INT,\n'
                        '\t\tEgu VARCHAR(1024),\n'
                        '\t\tRuleName VARCHAR(1024),\n'
                        f'\t\tCONSTRAINT {sign}_pkey PRIMARY KEY (Id)\n'
                        '\t);\n'
                        f'\t\tDELETE FROM objects.{sign} ;\n')
        
        text_start_pump = ('\tCREATE SCHEMA IF NOT EXISTS objects;\n'
                            f'\tCREATE TABLE IF NOT EXISTS objects.{sign}(\n'
                            '\t\tId INT NOT NULL,\n'
                            '\t\tPrefix VARCHAR(1024),\n'
                            '\t\tSetpointGroupId INT,\n'
                            '\t\tTag VARCHAR(1024),\n'
                            '\t\tName VARCHAR(1024),\n'
                            '\t\tSource VARCHAR(1024),\n'
                            '\t\tValue INT,\n'
                            '\t\tValuereal DOUBLE PRECISION,\n'
                            '\t\tEgu VARCHAR(1024),\n'
                            '\t\tRuleName VARCHAR(1024),\n'
                            f'\t\tCONSTRAINT {sign}_pkey PRIMARY KEY (Id)\n'
                            '\t);\n'
                            f'\t\tDELETE FROM objects.{sign} ;\n')
        msg = {}
        gen_list = []
        flag_del_tabl = False
        try:
            cursor.execute(f"""SELECT id, variable, tag, name, unit, used, value_ust, group_ust, rule_map_ust
                                FROM "{tabl_sql}" ORDER BY Id""")
            list_signal = cursor.fetchall()
        except Exception:
            msg[f'{today} - {sign}: ошибка генерации: {traceback.format_exc()}'] = 2
            return msg

        for signal in list_signal:
            try:
                Id, variable, tag, name, unit            = signal[0], signal[1], signal[2], signal[3], signal[4]
                used, value_ust, group_ust, rule_map_ust = signal[5], signal[6], signal[7], signal[8]

                if used == '0': continue

                # Prefix
                Prefix = 'NULL' if prefix_system == '' or prefix_system is None else str(prefix_system)
                
                # SetpointGroupId
                cursor.execute(f"""SELECT id FROM "sp_grp" WHERE name_group='{group_ust}'""")
                try   : SetpointGroupId = cursor.fetchall()[0][0]
                except: SetpointGroupId = 'NULL'

                # RuleName
                cursor.execute(f"""SELECT rule_name FROM "sp_rules" WHERE name_rules='{rule_map_ust}'""")
                try   : RuleName = cursor.fetchall()[0][0]
                except: RuleName = 'NULL'

            except Exception:
                msg[f'{today} - {sign}: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                continue
            
            ins_row_tabl = f"INSERT INTO objects.{sign} (Id, Prefix, SetpointGroupId, Tag, Name, Source, Value, Egu, RuleName) VALUES({Id},'{Prefix}', {SetpointGroupId}, '{tag}', '{name}', '{variable}', {value_ust}, '{unit}', '{RuleName}');\n"
            
            if flag_write_db:
                try:
                    if flag_del_tabl is False :
                        if sign == 'TblPumptimeSetpoints': cursor_prj.execute(text_start_pump)
                        else                             : cursor_prj.execute(text_start)
                        flag_del_tabl = True

                    cursor_prj.execute(ins_row_tabl)
                except Exception:
                    msg[f'{today} - {sign}: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                    continue
            else:
                gen_list.append(ins_row_tabl)
    
        if not flag_write_db:
            try:
                # Создаём файл запроса
                path_request = f'{path_location_file}\\PostgreSQL-{sign}.sql'
                if not os.path.exists(path_request):
                    file = codecs.open(path_request, 'w', 'utf-8')
                else:
                    os.remove(path_request)
                    file = codecs.open(path_request, 'w', 'utf-8')
                if path_location_file == '' or path_location_file is None or len(path_location_file) == 0:
                    msg[f'{today} - {sign}: не указана конечная папка'] = 2
                    return msg
                file.write(text_start)
                for insert in gen_list:
                    file.write(insert)
                file.write(f'COMMIT;')
                file.close()
                msg[f'{today} - {sign}: файл скрипта создан'] = 1
                return(msg)
            except Exception:
                msg[f'{today} - {sign}: ошибка записи в файл: {traceback.format_exc()}'] = 2
        msg[f'{today} - {sign}: генерация завершена!'] = 1
        return(msg)
    def gen_table_ktpr(self, flag_write_db):
        cursor = db.cursor()
        cursor_prj = db_prj.cursor()
    
        text_start = ('\tCREATE SCHEMA IF NOT EXISTS objects;\n'
                        f'\tCREATE TABLE IF NOT EXISTS objects.TblStationDefencesSetpoints(\n'
                        '\t\tId INT NOT NULL,\n'
                        '\t\tPrefix VARCHAR(1024),\n'
                        '\t\tTag VARCHAR(1024),\n'
                        '\t\tName VARCHAR(1024),\n'
                        '\t\tSource VARCHAR(1024),\n'
                        '\t\tValue INT,\n'
                        '\t\tEgu VARCHAR(1024),\n'
                        '\t\tSetpointGroupId INT,\n'
                        '\t\tRuleName VARCHAR(1024),\n'
                        f'\t\tCONSTRAINT TblStationDefencesSetpoints_pkey PRIMARY KEY (Id)\n'
                        '\t);\n'
                        f'\t\tDELETE FROM objects.TblStationDefencesSetpoints ;\n')
        msg = {}
        gen_list = []
        flag_del_tabl = False
        try:
            cursor.execute(f"""SELECT id, variable, tag, name, "time_ust", "group_ust", "rule_map_ust"
                               FROM "ktpr" ORDER BY Id""")
            list_signal = cursor.fetchall()
        except Exception:
            msg[f'{today} - TblStationDefencesSetpoints: ошибка генерации: {traceback.format_exc()}'] = 2
            return msg

        for signal in list_signal:
            try:
                Id, variable, tag, name,          = signal[0], signal[1], signal[2], signal[3],
                time_ust, group_ust, rule_map_ust = signal[4], signal[5], signal[6]

                # Prefix
                Prefix = 'NULL' if prefix_system == '' or prefix_system is None else str(prefix_system)
                # tag
                tag = '' if tag == '' or tag is None else str(tag)
                # Value
                time_ust = 'NULL' if time_ust == '' or time_ust is None else time_ust
                # SetpointGroupId
                cursor.execute(f"""SELECT id FROM "sp_grp" WHERE name_group='{group_ust}'""")
                try   : SetpointGroupId = cursor.fetchall()[0][0]
                except: SetpointGroupId = 'NULL'
                # RuleName
                cursor.execute(f"""SELECT rule_name FROM "sp_rules" WHERE name_rules='{rule_map_ust}'""")
                try   : RuleName = cursor.fetchall()[0][0]
                except: RuleName = 'NULL'
            except Exception:
                msg[f'{today} - TblStationDefencesSetpoints: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                continue
            
            ins_row_tabl = f"INSERT INTO objects.TblStationDefencesSetpoints (Id, Prefix, Tag, Name, Source, Value, Egu, SetpointGroupId, RuleName) VALUES({Id}, {Prefix}, '{tag}', '{name}', 'tm{variable}', {time_ust}, 'c', {SetpointGroupId}, '{RuleName}');\n"

            if flag_write_db:
                try:
                    if flag_del_tabl is False :
                        cursor_prj.execute(text_start)
                        flag_del_tabl = True
                    cursor_prj.execute(ins_row_tabl)
                except Exception:
                    msg[f'{today} - TblStationDefencesSetpoints: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                    continue
            else:
                gen_list.append(ins_row_tabl)
    
        if not flag_write_db:
            try:
                # Создаём файл запроса
                path_request = f'{path_location_file}\\PostgreSQL-TblStationDefencesSetpoints.sql'
                if not os.path.exists(path_request):
                    file = codecs.open(path_request, 'w', 'utf-8')
                else:
                    os.remove(path_request)
                    file = codecs.open(path_request, 'w', 'utf-8')
                if path_location_file == '' or path_location_file is None or len(path_location_file) == 0:
                    msg[f'{today} - TblStationDefencesSetpoints: не указана конечная папка'] = 2
                    return msg
                file.write(text_start)
                for insert in gen_list:
                    file.write(insert)
                file.write(f'COMMIT;')
                file.close()
                msg[f'{today} - TblStationDefencesSetpoints: файл скрипта создан'] = 1
                return(msg)
            except Exception:
                msg[f'{today} - TblStationDefencesSetpoints: ошибка записи в файл: {traceback.format_exc()}'] = 2
        msg[f'{today} - TblStationDefencesSetpoints: генерация завершена!'] = 1
        return(msg)
    def gen_table_pumps(self, flag_write_db, tabl_sql, sign):
        cursor = db.cursor()
        cursor_prj = db_prj.cursor()
    
        text_start = ('\tCREATE SCHEMA IF NOT EXISTS objects;\n'
                        f'\tCREATE TABLE IF NOT EXISTS objects.{sign}(\n'
                        '\t\tId INT NOT NULL,\n'
                        '\t\tPrefix VARCHAR(1024),\n'
                        '\t\tTag VARCHAR(1024),\n'
                        '\t\tName VARCHAR(1024),\n'
                        '\t\tSource VARCHAR(1024),\n'
                        '\t\tValue INT,\n'
                        '\t\tEgu VARCHAR(1024),\n'
                        '\t\tSetpointGroupId INT,\n'
                        '\t\tRuleName VARCHAR(1024),\n'
                        f'\t\tCONSTRAINT {sign}_pkey PRIMARY KEY (Id)\n'
                        '\t);\n'
                        f'\t\tDELETE FROM objects.{sign} ;\n')
        msg = {}
        gen_list = []
        flag_del_tabl = False
        try:
            cursor.execute(f"""SELECT id, variable, tag, name, "NA", "time_ust", "group_ust", "rule_map_ust", "number_pump_VU"
                               FROM "{tabl_sql}" ORDER BY Id, "NA" """)
            list_signal = cursor.fetchall()
        except Exception:
            msg[f'{today} - {sign}: ошибка генерации: {traceback.format_exc()}'] = 2
            return msg

        for signal in list_signal:
            try:
                Id, variable, tag, name, PumpName = signal[0], signal[1], signal[2], signal[3], signal[4]
                time_ust, group_ust, rule_map_ust = signal[5], signal[6], signal[7]

                if time_ust == '' or time_ust is None: continue

                # Prefix
                Prefix = 'NULL' if prefix_system == '' or prefix_system is None else str(prefix_system)
                # PumpName
                PumpName = 'NULL' if PumpName == '' or PumpName is None else str(PumpName)
                # tag
                tag = 'NULL' if tag == '' or tag is None else str(tag)
                # SetpointGroupId
                cursor.execute(f"""SELECT id FROM "sp_grp" WHERE name_group='{group_ust}'""")
                try   : SetpointGroupId = cursor.fetchall()[0][0]
                except: SetpointGroupId = 'NULL'
                # RuleName
                cursor.execute(f"""SELECT rule_name FROM "sp_rules" WHERE name_rules='{rule_map_ust}'""")
                try   : RuleName = cursor.fetchall()[0][0]
                except: RuleName = 'NULL'

            except Exception:
                msg[f'{today} - {sign}: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                continue
            
            ins_row_tabl = f"INSERT INTO objects.{sign} (Id, Prefix, Tag, Name, Source, Value, Egu, SetpointGroupId, RuleName) VALUES({Id}, {Prefix}, '{tag}', '{name}', 'tm{variable}', {time_ust}, 'c', {SetpointGroupId}, '{RuleName}');\n"

            if flag_write_db:
                try:
                    if flag_del_tabl is False :
                        cursor_prj.execute(text_start)
                        flag_del_tabl = True
                    cursor_prj.execute(ins_row_tabl)
                except Exception:
                    msg[f'{today} - {sign}: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                    continue
            else:
                gen_list.append(ins_row_tabl)
    
        if not flag_write_db:
            try:
                # Создаём файл запроса
                path_request = f'{path_location_file}\\PostgreSQL-{sign}.sql'
                if not os.path.exists(path_request):
                    file = codecs.open(path_request, 'w', 'utf-8')
                else:
                    os.remove(path_request)
                    file = codecs.open(path_request, 'w', 'utf-8')
                if path_location_file == '' or path_location_file is None or len(path_location_file) == 0:
                    msg[f'{today} - {sign}: не указана конечная папка'] = 2
                    return msg
                file.write(text_start)
                for insert in gen_list:
                    file.write(insert)
                file.write(f'COMMIT;')
                file.close()
                msg[f'{today} - {sign}: файл скрипта создан'] = 1
                return(msg)
            except Exception:
                msg[f'{today} - {sign}: ошибка записи в файл: {traceback.format_exc()}'] = 2
        msg[f'{today} - {sign}: генерация завершена!'] = 1
        return(msg)
    def gen_table_gmpna(self, flag_write_db):
        cursor = db.cursor()
        cursor_prj = db_prj.cursor()

        text_start = ('\tCREATE SCHEMA IF NOT EXISTS objects;\n'
                        f'\tCREATE TABLE IF NOT EXISTS objects.TblPumpReadinesesSetpoints(\n'
                        '\t\tId INT NOT NULL,\n'
                        '\t\tPumpId INT NOT NULL,\n'
                        '\t\tPrefix VARCHAR(1024),\n'
                        '\t\tName VARCHAR(1024),\n'
                        '\t\tTag VARCHAR(1024),\n'
                        '\t\tPumpName VARCHAR(1024),\n'
                        '\t\tSource VARCHAR(1024),\n'
                        '\t\tValue INT,\n'
                        '\t\tEgu VARCHAR(1024),\n'
                        '\t\tSetpointGroupId INT,\n'
                        '\t\tRuleName VARCHAR(1024),\n'
                        f'\t\tCONSTRAINT TblPumpReadinesesSetpoints_pkey PRIMARY KEY (Id,PumpId)\n'
                        '\t);\n'
                        f'\t\tDELETE FROM objects.TblPumpReadinesesSetpoints ;\n')
        msg = {}
        gen_list = []
        flag_del_tabl = False
        try:
            cursor.execute(f"""SELECT id, variable, tag, name, "NA", "used_time_ust", "time_ust", "group_ust", "rule_map_ust", "number_pump_VU"
                                FROM "gmpna" ORDER BY Id, "number_pump_VU", "NA" """)
            list_signal = cursor.fetchall()
        except Exception:
            msg[f'{today} - TblPumpReadinesesSetpoints: ошибка генерации: {traceback.format_exc()}'] = 2
            return msg

        for signal in list_signal:
            try:
                Id, variable, tag, name, PumpName, used_ust = signal[0], signal[1], signal[2], signal[3], signal[4], signal[5]
                time_ust, group_ust, rule_map_ust, number_pump_VU = signal[6], signal[7], signal[8], signal[9]
                
                if used_ust is not True: continue
                # Prefix
                Prefix = 'NULL' if prefix_system == '' or prefix_system is None else str(prefix_system)
                # PumpId
                PumpId = 'NULL' if number_pump_VU == '' or number_pump_VU is None else number_pump_VU
                # name
                name = '' if name == '' or name is None else str(name)
                # PumpName
                PumpName = 'NULL' if PumpName == '' or PumpName is None else str(PumpName)
                # tag
                tag = 'NULL' if tag == '' or tag is None else f'{tag}'
                # Value
                time_ust = 'NULL' if time_ust == '' or time_ust is None else time_ust
                # SetpointGroupId
                cursor.execute(f"""SELECT id FROM "sp_grp" WHERE name_group='{group_ust}'""")
                try   : SetpointGroupId = cursor.fetchall()[0][0]
                except: SetpointGroupId = 'NULL'
                # RuleName
                cursor.execute(f"""SELECT rule_name FROM "sp_rules" WHERE name_rules='{rule_map_ust}'""")
                try   : RuleName = cursor.fetchall()[0][0]
                except: RuleName = 'NULL'
            except Exception:
                msg[f'{today} - TblPumpReadinesesSetpoints: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                continue
            
            ins_row_tabl = f"INSERT INTO objects.TblPumpReadinesesSetpoints (Id, PumpId, Prefix, Name, Tag, PumpName, Source, Value, Egu, SetpointGroupId, RuleName) VALUES({Id}, {PumpId}, {Prefix}, '{PumpName}. {name}', '{tag}', '{PumpName}', 'tm{variable}', {time_ust}, 'c', {SetpointGroupId}, '{RuleName}');\n"

            if flag_write_db:
                try:
                    if flag_del_tabl is False :
                        cursor_prj.execute(text_start)
                        flag_del_tabl = True
                    cursor_prj.execute(ins_row_tabl)
                except Exception:
                    msg[f'{today} - TblPumpReadinesesSetpoints: ошибка добавления строки, пропускается: {traceback.format_exc()}'] = 2
                    continue
            else:
                gen_list.append(ins_row_tabl)

        if not flag_write_db:
            try:
                # Создаём файл запроса
                path_request = f'{path_location_file}\\PostgreSQL-TblPumpReadinesesSetpoints.sql'
                if not os.path.exists(path_request):
                    file = codecs.open(path_request, 'w', 'utf-8')
                else:
                    os.remove(path_request)
                    file = codecs.open(path_request, 'w', 'utf-8')
                if path_location_file == '' or path_location_file is None or len(path_location_file) == 0:
                    msg[f'{today} - TblPumpReadinesesSetpoints: не указана конечная папка'] = 2
                    return msg
                file.write(text_start)
                for insert in gen_list:
                    file.write(insert)
                file.write(f'COMMIT;')
                file.close()
                msg[f'{today} - TblPumpReadinesesSetpoints: файл скрипта создан'] = 1
                return(msg)
            except Exception:
                msg[f'{today} - TblPumpReadinesesSetpoints: ошибка записи в файл: {traceback.format_exc()}'] = 2
        msg[f'{today} - TblPumpReadinesesSetpoints: генерация завершена!'] = 1
        return(msg)
    # synh_tabl
    def synh_tabl_ai(self):
        def notation(value):
            s, new_s = '', ''
            while value > 0:
                s = str(value % 2) + s
                value //= 2
            if len(s) < 16:
                for i in range(16 - len(s)):
                    s = '0' + s

            for i in range(1, len(s) + 1):
                new_s += s[i - 1]
                if (i % 4 == 0) and (i != 16): 
                    new_s += '_'
            return new_s
        msg = {}
        try:
            ai_prj = self.dop_function.connect_by_sql_prj("objects.tblanalogs", 
                                                          '''"id", "tag", "name", "egu", "precision", "lolimfield", "hilimfield", "lolimeng", "hilimeng", "lolim", "hilim", 
                                                             "min6", "min5", "min4", "min3", "min2", "min1", "max1", "max2", "max3", "max4", "max5", "max6", 
                                                             "histeresis", "deltat", "msgmask", "sigmask", "ctrlmask", "rulename", "timefilter"''')
            for prj in ai_prj:
                id_prj, tag_prj, name_prj  = prj[0], prj[1], prj[2]
                data = self.dop_function.connect_by_sql_condition("ai", '''"Egu", "Precision", "LoLimField", "HiLimField", "LoLimEng", "HiLimEng", "LoLim", "HiLim", 
                                                                        "Min6", "Min5", "Min4", "Min3", "Min2", "Min1", "Max1", "Max2", "Max3", "Max4", "Max5", "Max6", 
                                                                        "Histeresis", "DeltaT", "MsgMask", "SigMask", "CtrlMask", "RuleName", "TimeFilter"''',
                                                                  f'''"id"={id_prj} AND "tag"='{tag_prj}' AND "name"='{name_prj}' ''')
                if len(data) == 0: continue   
                value = data[0]
                    
                egu, precision, lolimfield, hilimfield, lolimeng, hilimeng, lolim, hilim = prj[3], prj[4], prj[5], prj[6], prj[7], prj[8], prj[9], prj[10]
                min6, min5, min4, min3, min2, min1, max1, max2, max3, max4, max5 = prj[11], prj[12], prj[13], prj[14], prj[15], prj[16], prj[17], prj[18], prj[19], prj[20], prj[21]
                max6, histeresis, deltat, msgmask, sigmask, ctrlmask, rulename, timefilter = prj[22], prj[23], prj[24], prj[25], prj[26], prj[27], prj[28], prj[29]
                
                pred_msg = ''
                if egu        != value[0] : pred_msg += self.dop_function.update_row("ai", egu, "Egu", id_prj)
                if precision  != value[1] : pred_msg += self.dop_function.update_row("ai", precision, "Precision", id_prj)
                if lolimfield != value[2] : pred_msg += self.dop_function.update_row("ai", lolimfield, "LoLimField", id_prj)
                if hilimfield != value[3] : pred_msg += self.dop_function.update_row("ai", hilimfield, "HiLimField", id_prj)
                if lolimeng   != value[4] : pred_msg += self.dop_function.update_row("ai", lolimeng, "LoLimEng", id_prj)
                if hilimeng   != value[5] : pred_msg += self.dop_function.update_row("ai", hilimeng, "HiLimEng", id_prj)
                if lolim      != value[6] : pred_msg += self.dop_function.update_row("ai", lolim, "LoLim", id_prj)
                if hilim      != value[7] : pred_msg += self.dop_function.update_row("ai", hilim, "HiLim", id_prj)
                if min6       != value[8] : pred_msg += self.dop_function.update_row("ai", min6, "Min6", id_prj)
                if min5       != value[9] : pred_msg += self.dop_function.update_row("ai", min5, "Min5", id_prj)
                if min4       != value[10]: pred_msg += self.dop_function.update_row("ai", min4, "Min4", id_prj)
                if min3       != value[11]: pred_msg += self.dop_function.update_row("ai", min3, "Min3", id_prj)
                if min2       != value[12]: pred_msg += self.dop_function.update_row("ai", min2, "Min2", id_prj)
                if min1       != value[13]: pred_msg += self.dop_function.update_row("ai", min1, "Min1", id_prj)
                if max1       != value[14]: pred_msg += self.dop_function.update_row("ai", max1, "Max1", id_prj)
                if max2       != value[15]: pred_msg += self.dop_function.update_row("ai", max2, "Max2", id_prj)
                if max3       != value[16]: pred_msg += self.dop_function.update_row("ai", max3, "Max3", id_prj)
                if max4       != value[17]: pred_msg += self.dop_function.update_row("ai", max4, "Max4", id_prj)
                if max5       != value[18]: pred_msg += self.dop_function.update_row("ai", max5, "Max5", id_prj)
                if max6       != value[19]: pred_msg += self.dop_function.update_row("ai", max6, "Max6", id_prj)
                if histeresis != value[20]: pred_msg += self.dop_function.update_row("ai", histeresis, "Histeresis", id_prj)
                if deltat     != value[21]: pred_msg += self.dop_function.update_row("ai", deltat, "DeltaT", id_prj)
                if timefilter != value[26]: pred_msg += self.dop_function.update_row("ai", timefilter, "TimeFilter", id_prj)

                # rulename
                sp_rules = self.dop_function.connect_by_sql("sp_rules", '''"rule_name", "name_rules"''')
                for rule in sp_rules:
                    if str(rule[0]) == rulename:
                        name_rules = rule[1]
                        break
                if name_rules != value[25]: pred_msg += self.dop_function.update_row("ai", name_rules, "RuleName", id_prj) 
                # msgmask
                msgmask_bin = notation(msgmask)
                if msgmask_bin != value[22]: pred_msg += self.dop_function.update_row("ai", msgmask_bin, "MsgMask", id_prj)
                # sigmask
                sigmask_bin = notation(sigmask)
                if sigmask_bin != value[23]: pred_msg += self.dop_function.update_row("ai", sigmask_bin, "SigMask", id_prj)
                # ctrlmask
                ctrlmask_bin = notation(ctrlmask)
                if ctrlmask_bin != value[24]: pred_msg += self.dop_function.update_row("ai", ctrlmask_bin, "CtrlMask", id_prj)
                if pred_msg != '': msg[f'{today} - Изменения: {id_prj}, {name_prj}: {pred_msg}'] = 3
        except Exception:
            msg[f'{today} - TblAnalogs: ошибка синхронизации: {traceback.format_exc()}'] = 2
            return msg
        
        msg[f'{today} - TblAnalogs: синхронизация завершена!'] = 1
        return(msg)
    def synh_tabl(self, tbl_prj, tbl_dev, sign):
        msg = {}
        msg[f'{today} - {sign}: синхронизация базы проекта с базой разработки'] = 2
        try:
            if tbl_dev != 'umpna_tm': tbl = self.dop_function.connect_by_sql_prj(f"objects.{tbl_prj}", '''"id", "tag", "name", "source", "value", "rulename"''')
            else                    : tbl = self.dop_function.connect_by_sql_prj(f"objects.{tbl_prj}", '''"id", "tag", "name", "source", "value", "valuereal", "rulename"''')
            
            for prj in tbl:
                if tbl_dev != 'umpna_tm': 
                    id_prj, tag_prj, name_prj, source_prj, value_prj, rulename_prj = prj[0], prj[1], prj[2], prj[3], prj[4], prj[5]
                    data = self.dop_function.connect_by_sql_condition(f"{tbl_dev}", '''"variable", "value_ust", "rule_map_ust"''',
                                                                      f'''"id"={id_prj} AND "tag"='{tag_prj}' AND "name"='{name_prj}' ''')
                else:      
                    id_prj, tag_prj, name_prj, source_prj, value_prj, valuereal_prj, rulename_prj = prj[0], prj[1], prj[2], prj[3], prj[4], prj[5], prj[6]
                    data = self.dop_function.connect_by_sql_condition(f"{tbl_dev}", '''"variable", "value_ust", "value_real_ust", "rule_map_ust"''',
                                                                    f'''"id"={id_prj} AND "tag"='{tag_prj}' AND "name"='{name_prj}' ''')
                if len(data) == 0: continue   
                value = data[0]
                
                pred_msg = ''
                if source_prj != value[0]: pred_msg += self.dop_function.update_row(f"{tbl_dev}", source_prj, "variable", id_prj)
                if value_prj  != value[1]: pred_msg += self.dop_function.update_row(f"{tbl_dev}", value_prj, "value_ust", id_prj)
                if tbl_dev == 'umpna_tm': 
                    if valuereal_prj != value[2]: pred_msg += self.dop_function.update_row(f"{tbl_dev}", valuereal_prj, "value_real_ust", id_prj)

                # rulename
                sp_rules = self.dop_function.connect_by_sql("sp_rules", '''"rule_name", "name_rules"''')
                for rule in sp_rules:
                    if str(rule[0]) == rulename_prj:
                        name_rules = rule[1]
                        break

                if tbl_dev == 'umpna_tm': 
                    if name_rules != value[3]: pred_msg += self.dop_function.update_row(f"{tbl_dev}", name_rules, "RuleName", id_prj) 
                else: 
                    if name_rules != value[2]: pred_msg += self.dop_function.update_row(f"{tbl_dev}", name_rules, "RuleName", id_prj) 

                if pred_msg != '': msg[f'{today} - Изменения: {id_prj}, {name_prj}: {pred_msg}'] = 3
        except Exception:
            msg[f'{today} - {sign}: ошибка синхронизации: {traceback.format_exc()}'] = 2
            return msg
        msg[f'{today} - {sign}: синхронизация завершена!'] = 1
        return(msg)
    


# Filling attribute DevStudio
class Filling_attribute_DevStudio():
    def __init__(self):
        self.dop_function = General_functions()
    def write_in_omx(self, list_tabl):
        msg = {}
        if len(list_tabl) == 0:             
            msg[f'{today} - Файл omx: не выбраны атрибуты'] = 2
            return msg
        for tabl in list_tabl: 
            if tabl == 'AI': 
                msg.update(self.analogs_omx())
                continue
            if tabl == 'DI': 
                msg.update(self.diskret_in_omx())
                continue
            if tabl == 'VS': 
                msg.update(self.auxsystem_omx())
                continue
            if tabl == 'ZD': 
                msg.update(self.valves_omx())
                continue
            if tabl == 'NA': 
                msg.update(self.pumps_omx())
                continue
            if tabl == 'PIC': 
                msg.update(self.picture_omx())
                continue
            if tabl == 'SS': 
                msg.update(self.relayted_system_omx())
                continue
            if tabl == 'UTS': 
                msg.update(self.uts_omx())
                continue
            if tabl == 'UPTS': 
                msg.update(self.upts_omx())
                continue
            if tabl == 'KTPR': 
                msg.update(self.ktpr_omx())
                continue
            if tabl == 'KTPRP': 
                msg.update(self.ktprp_omx())
                continue
            if tabl == 'KTPRA': 
                msg.update(self.ktpra_omx())
                continue
            if tabl == 'GMPNA': 
                msg.update(self.gmpna_omx())
                continue
            if tabl == 'PI': 
                msg.update(self.pi_omx())
                continue
            if tabl == 'PZ': 
                msg.update(self.pz_omx())
                continue
            if tabl == 'AI_diag': 
                msg.update(self.mklogic_DIAG_omx('AIs', 'MK-516-008A', 'MK_Logic.mod_AI'))
                msg.update(self.mklogic_AI_AO_atrib('AIs', 'MK-516-008A', 'ch_AI_0'))
                continue
            if tabl == 'AO_diag': 
                msg.update(self.mklogic_DIAG_omx('AOs', 'MK-514-008', 'MK_Logic.mod_AO'))
                msg.update(self.mklogic_AI_AO_atrib('AOs', 'MK-514-008', 'ch_AO_0'))
                continue
            if tabl == 'DI_diag': 
                msg.update(self.mklogic_DIAG_omx('DIs', 'MK-521-032', 'MK_Logic.mod_DI'))
                msg.update(self.mklogic_DI_DO_atrib('DIs', 'MK-521-032',))
                continue
            if tabl == 'DO_diag': 
                msg.update(self.mklogic_DIAG_omx('DOs', 'MK-531-032', 'MK_Logic.mod_DI'))
                msg.update(self.mklogic_DI_DO_atrib('DOs', 'MK-531-032'))
                continue
            if tabl == 'CPU_diag': 
                msg.update(self.mklogic_DIAG_omx('CPUs', 'MK-504-120', 'MK_Logic.mod_CPU'))
                continue
            if tabl == 'CN_diag': 
                msg.update(self.mklogic_DIAG_omx('CNs', 'MK-545-010', 'MK_Logic.mod_CN'))
                continue
            if tabl == 'MN_diag': 
                msg.update(self.mklogic_DIAG_omx('MNs', 'MK-546-010', 'MK_Logic.mod_CN'))
                continue
            if tabl == 'PSU_diag': 
                msg.update(self.mklogic_DIAG_omx('PSUs', 'MK-550-024', 'MK_Logic.mod_PSU'))
                continue
            if tabl == 'RS_diag': 
                msg.update(self.mklogic_DIAG_omx('RSs', 'MK-541-002', 'MK_Logic.mod_RS'))
                continue
            if tabl == 'RackStates_diag': 
                msg.update(self.rackstate_omx())
                continue
            if tabl == 'ColorDI': 
                msg.update(self.colorscheme_di())
                continue
            if tabl == 'formatAI': 
                msg.update(self.analogformat_map())
                continue
            if tabl == 'mapEGU': 
                msg.update(self.egu_map())
                continue
        return msg  
    def write_in_map(self, list_tabl):
            list_param_AI_AO = ['mod_State', 'ch_01', 'ch_02', 'ch_03', 'ch_04', 'ch_05', 'ch_06', 'ch_07', 'ch_08' ]
            list_param_DI_DO = ['mod_State', 'mDI']
            list_param_CPU   = ['mod_State', 'mod_State_ext', 'mod_State_Err', 'CPUMemFree', 'CPULoad', 'ClcCurr', 'ClcMax', 'RsrCRC32', 'DiskFreeSpace']
            list_param_CN    = ['mod_State', 'mod_State_ext', 'ports_State', 'pwl_ID']
            list_param_MN    = ['mod_State_ext', 'ports_State']
            list_param_PSU   = ['mod_State', 'mod_State_ext', 'SupplyVoltage', 'CanBusSpeed', 'Can1ErrorCounter', 'Can2ErrorCounter']
            list_param_RS    = ['mod_State', 'mod_State_ext']
            list_param_RackS = ['mBUS', 'mBUSandCh', 'mBUSblink']
            
            msg = {}
            if len(list_tabl) == 0:             
                msg[f'{today} - Файл omx: не выбраны атрибуты'] = 2
                return msg
            for tabl in list_tabl: 
                if tabl == 'AI': 
                    msg.update(self.analogs_maps())
                    continue
                if tabl == 'DI': 
                    msg.update(self.diskret_maps())
                    continue
                if tabl == 'VS': 
                    msg.update(self.auxsystem_maps())
                    continue
                if tabl == 'ZD': 
                    msg.update(self.valves_maps())
                    continue
                if tabl == 'NA': 
                    msg.update(self.na_maps())
                    continue
                if tabl == 'PIC': 
                    msg.update(self.picturs_maps())
                    continue
                if tabl == 'SS': 
                    msg.update(self.ss_maps())
                    continue
                if tabl == 'UTS': 
                    msg.update(self.uts_maps())
                    continue
                if tabl == 'UPTS': 
                    msg.update(self.upts_maps())
                    continue
                if tabl == 'KTPR': 
                    msg.update(self.ktpr_maps())
                    continue
                if tabl == 'KTPRP': 
                    msg.update(self.ktprp_maps())
                    continue
                if tabl == 'KTPRA': 
                    msg.update(self.ktpra_maps())
                    continue
                if tabl == 'GMPNA': 
                    msg.update(self.gmpna_maps())
                    continue
                if tabl == 'PI': 
                    msg.update(self.pi_maps())
                    continue
                if tabl == 'PZ': 
                    msg.update(self.pz_maps())
                    continue
                if tabl == 'AO_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('AOs', 'MK-514-008', list_param_AI_AO))
                    continue
                if tabl == 'AI_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('AIs', 'MK-516-008A', list_param_AI_AO))
                    continue
                if tabl == 'DO_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('DOs', 'MK-531-032', list_param_DI_DO))
                    continue
                if tabl == 'DI_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('DIs', 'MK-521-032', list_param_DI_DO))
                    continue
                if tabl == 'CPU_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('CPUs', 'MK-504-120', list_param_CPU))
                    continue
                if tabl == 'CN_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('CNs', 'MK-545-010', list_param_CN))
                    continue
                if tabl == 'MN_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('MNs', 'MK-546-010', list_param_MN))
                    continue
                if tabl == 'PSU_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('PSUs', 'MK-550-024', list_param_PSU))
                    continue
                if tabl == 'RS_diag': 
                    msg.update(self.mklogic_DIAG_maps_param('RSs', 'MK-541-002', list_param_RS))
                    continue
                if tabl == 'RackStates_diag': 
                    msg.update(self.mklogic_rackstates_maps(list_param_RackS))
                    continue
            return msg
    def clear_omx(self, list_tabl):
        path_AI_AO = [f'{path_to_devstudio}\\AttributesMapAI_Ref.xml', f'{path_to_devstudio}\\AttributesMapKlk.xml', f'{path_to_devstudio}\\AttributesMapKont.xml', 
                      f'{path_to_devstudio}\\AttributesMapSignalName.xml', f'{path_to_devstudio}\\AttributesMapTagName.xml']
        path_DI_DO = [f'{path_to_devstudio}\\AttributesMapKlk.xml', f'{path_to_devstudio}\\AttributesMapKont.xml', 
                      f'{path_to_devstudio}\\AttributesMapSignalName.xml', f'{path_to_devstudio}\\AttributesMapTagName.xml']
        path_ColorDI = [f'{path_to_devstudio}\\AttributesMapColorScheme.xml']
        path_formatAI = [f'{path_to_devstudio}\\AttributesAnalogsFormats.xml']
        path_egu = [f'{path_to_devstudio}\\AttributesMapEGU.xml']
        msg = {}
        if len(list_tabl) == 0: 
            msg[f'{today} - Файл omx: не выбраны атрибуты'] = 2
            return msg
        for tabl in list_tabl: 
            if tabl == 'AI': 
                msg.update(self.dop_function.clear_objects_omx('Analogs'))
                continue
            if tabl == 'DI': 
                msg.update(self.dop_function.clear_objects_omx('Diskrets'))
                continue
            if tabl == 'VS': 
                msg.update(self.dop_function.clear_objects_omx('AuxSystems'))
                continue
            if tabl == 'ZD': 
                msg.update(self.dop_function.clear_objects_omx('Valves'))
                continue
            if tabl == 'NA': 
                msg.update(self.dop_function.clear_objects_omx('NAs'))
                continue
            if tabl == 'PIC': 
                msg.update(self.dop_function.clear_objects_omx('Pictures'))
                continue
            if tabl == 'SS': 
                msg.update(self.dop_function.clear_objects_omx('SSs'))
                continue
            if tabl == 'UTS': 
                msg.update(self.dop_function.clear_objects_omx('UTSs'))
                continue
            if tabl == 'UPTS': 
                msg.update(self.dop_function.clear_objects_omx('UPTSs'))
                continue
            if tabl == 'KTPR': 
                msg.update(self.dop_function.clear_objects_omx('KTPRs'))
                continue
            if tabl == 'KTPRP': 
                msg.update(self.dop_function.clear_objects_omx('KTPRs'))
                continue
            if tabl == 'KTPRA': 
                msg.update(self.dop_function.clear_objects_omx('KTPRAs'))
                continue
            if tabl == 'GMPNA': 
                msg.update(self.dop_function.clear_objects_omx('GMPNAs'))
                continue
            if tabl == 'PI': 
                msg.update(self.dop_function.clear_objects_omx('PIs'))
                continue
            if tabl == 'PZ': 
                msg.update(self.dop_function.clear_objects_omx('PZs'))
                continue
            if tabl == 'AI_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('AIs'))
                msg.update(self.dop_function.clear_objects_attrib('.Diag.AIs', path_AI_AO))
                continue
            if tabl == 'AO_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('AOs'))
                msg.update(self.dop_function.clear_objects_attrib('.Diag.AOs', path_AI_AO))
                continue
            if tabl == 'DI_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('DIs'))
                msg.update(self.dop_function.clear_objects_attrib('.Diag.DIs', path_DI_DO))
                continue
            if tabl == 'DO_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('DOs'))
                msg.update(self.dop_function.clear_objects_attrib('.Diag.DOs', path_DI_DO))
                continue
            if tabl == 'CPU_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('CPUs'))
                continue
            if tabl == 'CN_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('CNs'))
                continue
            if tabl == 'MN_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('MNs'))
                continue
            if tabl == 'PSU_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('PSUs'))
                continue
            if tabl == 'RS_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('RSs'))
                continue
            if tabl == 'RackStates_diag': 
                msg.update(self.dop_function.clear_diag_objects_omx('RackStates'))
                continue
            if tabl == 'ColorDI': 
                msg.update(self.dop_function.clear_objects_attrib('.Diskrets', path_ColorDI))
                continue
            if tabl == 'formatAI': 
                msg.update(self.dop_function.clear_objects_attrib('.Analogs', path_formatAI))
                continue
            if tabl == 'mapEGU': 
                msg.update(self.dop_function.clear_objects_attrib('.Analogs', path_egu))
                continue
        return msg
    def clear_map(self, list_tabl):
        msg = {}
        if len(list_tabl) == 0: 
            msg[f'{today} - Карта адресов: не выбраны типы для очистки'] = 2
            return msg
        for tabl in list_tabl: 
            if tabl == 'AI': 
                self.dop_function.parser_map('.Analogs.')
                msg[f'{today} - Карта адресов: адреса Analogs очищены'] = 1
                continue
            if tabl == 'DI': 
                self.dop_function.parser_map('.Diskrets.')
                msg[f'{today} - Карта адресов: адреса Diskrets очищены'] = 1
                continue
            if tabl == 'VS': 
                self.dop_function.parser_map('.AuxSystems.')
                msg[f'{today} - Карта адресов: адреса AuxSystems очищены'] = 1
                continue
            if tabl == 'ZD': 
                self.dop_function.parser_map('.Valves.')
                msg[f'{today} - Карта адресов: адреса Valves очищены'] = 1
                continue
            if tabl == 'NA': 
                self.dop_function.parser_map('.NAs.')
                msg[f'{today} - Карта адресов: адреса NAs очищены'] = 1
                continue
            if tabl == 'PIC': 
                self.dop_function.parser_map('.Pictures.')
                msg[f'{today} - Карта адресов: адреса Pictures очищены'] = 1
                continue
            if tabl == 'SS': 
                self.dop_function.parser_map('.SSs.')
                msg[f'{today} - Карта адресов: адреса SSs очищены'] = 1
                continue
            if tabl == 'UTS': 
                self.dop_function.parser_map('.UTSs.')
                msg[f'{today} - Карта адресов: адреса UTSs очищены'] = 1
                continue
            if tabl == 'UPTS': 
                self.dop_function.parser_map('.UPTSs.')
                msg[f'{today} - Карта адресов: адреса UPTSs очищены'] = 1
                continue
            if tabl == 'KTPR': 
                self.dop_function.parser_map('.KTPRs.')
                msg[f'{today} - Карта адресов: адреса KTPRs очищены'] = 1
                continue
            if tabl == 'KTPRP': 
                self.dop_function.parser_map('.KTPRs.')
                msg[f'{today} - Карта адресов: адреса KTPRs очищены'] = 1
                continue
            if tabl == 'KTPRA': 
                self.dop_function.parser_map('.KTPRAs.')
                msg[f'{today} - Карта адресов: адреса KTPRAs очищены'] = 1
                continue
            if tabl == 'GMPNA': 
                self.dop_function.parser_map('.GMPNAs.')
                msg[f'{today} - Карта адресов: адреса GMPNAs очищены'] = 1
                continue
            if tabl == 'PI': 
                self.dop_function.parser_map('.PIs.')
                msg[f'{today} - Карта адресов: адреса PIs очищены'] = 1
                continue
            if tabl == 'PZ': 
                self.dop_function.parser_map('PZs.')
                msg[f'{today} - Карта адресов: адреса PZs очищены'] = 1
                continue
            if tabl == 'AI_diag': 
                self.dop_function.parser_map('.Diag.AIs.')
                msg[f'{today} - Карта адресов: адреса Diag.AIs очищены'] = 1
                continue
            if tabl == 'AO_diag': 
                self.dop_function.parser_map('.Diag.AOs.')
                msg[f'{today} - Карта адресов: адреса Diag.AOs очищены'] = 1
                continue
            if tabl == 'DI_diag': 
                self.dop_function.parser_map('.Diag.DIs.')
                msg[f'{today} - Карта адресов: адреса Diag.DIs очищены'] = 1
                continue
            if tabl == 'DO_diag': 
                self.dop_function.parser_map('.Diag.DOs.')
                msg[f'{today} - Карта адресов: адреса Diag.DOs очищены'] = 1
                continue
            if tabl == 'CPU_diag': 
                self.dop_function.parser_map('.Diag.CPUs.')
                msg[f'{today} - Карта адресов: адреса Diag.CPUs очищены'] = 1
                continue
            if tabl == 'CN_diag': 
                self.dop_function.parser_map('.Diag.CNs.')
                msg[f'{today} - Карта адресов: адреса Diag.CNs очищены'] = 1
                continue
            if tabl == 'MN_diag': 
                self.dop_function.parser_map('.Diag.MNs.')
                msg[f'{today} - Карта адресов: адреса Diag.MNs очищены'] = 1
                continue
            if tabl == 'PSU_diag': 
                self.dop_function.parser_map('.Diag.PSUs.')
                msg[f'{today} - Карта адресов: адреса Diag.PSUs очищены'] = 1
                continue
            if tabl == 'RS_diag': 
                self.dop_function.parser_map('.Diag.RSs.')
                msg[f'{today} - Карта адресов: адреса Diag.RSs очищены'] = 1
                continue
            if tabl == 'RackStates_diag': 
                self.dop_function.parser_map('.Diag.RackStates.')
                msg[f'{today} - Карта адресов: адреса Diag.RackStates очищены'] = 1
                continue
        return msg
    def hardware_data(self, type_modul):
        modul = []
        with db:
            for basket in HardWare.select().dicts():
                id_        = basket['id']
                tag        = basket['tag']
                uso        = basket['uso']
                num_basket = basket['basket']
                for key, value in basket.items():
                    if value == type_modul:
                        number_modul = str(key).split('_')[1]
                        if int(number_modul) < 10: 
                            string_name = f'{tag}_0{number_modul}'
                            modPosition = f'A{num_basket}.0{number_modul}'
                        else:
                            string_name = f'{tag}_{number_modul}'
                            modPosition = f'A{num_basket}.{number_modul}'

                        modul.append(dict(id_=id_,
                                          uso=uso,
                                          string_name=string_name,
                                          num_basket=num_basket,
                                          number_modul=number_modul,
                                          modPosition=modPosition))
        return modul
    # Заполнение omx и атрибутов
    def analogs_omx(self):
            msg = {}
            dop_analog = {'объем'         : 'V',
                          'объём'         : 'V',
                          'перепад'       : 'dP',
                          'давлени'       : 'P',
                          'загазованность': 'Газ',
                          'вертик'        : 'Xверт',
                          'горизонт'      : 'Xгор',
                          'осевая'        : 'Xос',
                          'попереч'       : 'Xпоп',
                          'осевое'        : 'Xoc',
                          'сила'          : 'I',
                          'температура'   : 'T',
                          'уровень'       : 'L',
                          'утечк'         : 'L',
                          'расход'        : 'Q',
                          'положени'      : 'Q',
                          'затоплен'      : 'L',
                          'частот'        : 'F',
                          'процен'        : 'Q',
                          'заслон'        : 'Q',
                        }
            try:
                data_value = self.dop_function.connect_by_sql('ai', f'"id", "tag", "name", "PhysicEgu", "Egu", "IsOilPressure", "AnalogGroupId"')
                msg_bool, el1, tree = self.dop_function.parser_omx('Analogs')
                if msg_bool == 1: 
                    msg[f'{today} - Файл omx: ошибка при очистке Analogs'] = 2
                    return msg
                for value in data_value:
                    number      = value[0]
                    tag         = value[1]
                    name        = value[2]
                    equ_fiz     = value[3]
                    equ         = value[4]
                    unit_switch = value[5]
                    grp_analog  = value[6]
                    unit_alt    = 'кгс/см2'

                    if equ_fiz     == '': equ_fiz = ''
                    if tag         == '' or tag is None        : continue
                    if number      == '' or number is None     : continue
                    if name        == '' or name is None       : continue
                    if equ         == '' or equ is None        : continue
                    if unit_switch == '' or unit_switch is None: continue

                    if grp_analog == 'Уровни' or grp_analog == 'Аналоговые выходы':
                        type = 'unit.Library.PLC_Types.lv_Analog_PLC'
                    else:
                        type = 'unit.Library.PLC_Types.Analog_PLC'

                    tag_translate = self.dop_function.translate(str(tag))
                    unit_switch = True if unit_switch == 1 else False
                    # Находим совпадение из словаря с названием сигнала и заполняем подпись на кадре
                    sign = ' '
                    for key, short in dop_analog.items():
                        if self.dop_function.str_find(str(name).lower(), {key}):
                            sign = str(short)
                            break

                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = tag_translate
                    object.attrib['uuid'] = str(uuid.uuid1())
                    object.attrib['base-type'] = type
                    object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                    self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                    self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", sign)
                    self.dop_function.new_attr(object, "unit.Library.Attributes.EGU_Desc", equ)
                    self.dop_function.new_attr(object, "unit.Library.Attributes.EGU_Desc_phys", equ_fiz)
                    self.dop_function.new_attr(object, "unit.System.Attributes.Description", name)
                    self.dop_function.new_attr(object, "unit.Library.Attributes.EGU_Desc_Alt", unit_alt)
                    self.dop_function.new_attr(object, "unit.Library.Attributes.EGUsChange", unit_switch)
                    self.dop_function.new_attr(object, "unit.Library.Attributes.AI_Ref_KZFKP", tag)

                    el1.append(object)
                tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
                msg[f'{today} - Файл omx: Analogs добавлены'] = 1
                return msg
            except Exception:
                msg[f'{today} - Файл omx: ошибка при добавлении Analogs: {traceback.format_exc()}'] = 2
                return msg
    def diskret_in_omx(self):
        msg = {}
        dop_discret = {'давлен'        : 'P',
                       'напряж'        : 'U',
                       'уровень'       : 'L',
                       'затоплен'      : 'L',
                       'утечк'         : 'L',
                       'питание'       : 'U',
                       'питание шкафа' : 'U'}
        try:
            data_di = self.dop_function.connect_by_sql('di', f'"id", "tag", "name", "pNC_AI"')
            data_ai = self.dop_function.connect_by_sql('ai', f'"id", "tag"')
            msg_bool, el1, tree = self.dop_function.parser_omx('Diskrets')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке Diskrets'] = 2
                return msg
            for value in data_di:
                number_di = value[0]
                tag_di    = value[1]
                name      = value[2]
                pNC_AI    = value[3]

                if name == '' or name is None: continue
                if tag_di  == '' or tag_di is None: continue

                tag_di     = self.dop_function.translate(str(tag_di))
                tag_ai     = ' '
                tag_ai_ref = ' '
                try:
                    if pNC_AI is not None: 
                        isdigit = re.findall('\d+', str(pNC_AI))
                        for number in data_ai:
                            number_ai = number[0]
                            tag_ai    = number[1]
                            if str(number_ai) == str(isdigit[0]):
                                if tag_ai == '' or tag_ai is None:
                                    msg[f'{today} - Файл omx: Diskrets. Тэг AI сигнала {number_ai} пуст. Поля AI_Ref_KZFKP и AI_Ref не заполнены'] = 3
                                    break
                                else: 
                                    tag_ai_ref = tag_ai
                                    tag_ai     = self.dop_function.translate(tag_ai)
                                    break
                except Exception:
                    msg[f'{today} - Файл omx: Diskrets, ошибка пропускается: {traceback.format_exc()}'] = 2
                    continue
                sign = ' '
                for key, value in dop_discret.items():
                    if self.dop_function.str_find(str(name).lower(), {key}):
                        sign = str(value)
                        break

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = str(tag_di)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.Diskret_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                
                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number_di)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", sign)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", name)
                self.dop_function.new_attr(object, "unit.Library.Attributes.AI_Ref", tag_ai)
                self.dop_function.new_attr(object, "unit.Library.Attributes.AI_Ref_KZFKP", tag_ai_ref)

                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: Diskrets добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении Diskrets: {traceback.format_exc()}'] = 2
            return msg
    def picture_omx(self):
            msg = {}
            try:
                data = self.dop_function.connect_by_sql('pic', f'"id", "name", "frame"')
                msg_bool, el1, tree = self.dop_function.parser_omx('Pictures')
                if msg_bool == 1: 
                    msg[f'{today} - Файл omx: ошибка при очистке Pictures'] = 2
                    return msg
                for value in data:
                    number   = value[0]
                    name_pic = value[1]
                    screen   = value[2]

                    if screen == '' or screen is None: continue

                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = screen
                    object.attrib['uuid'] = str(uuid.uuid1())
                    object.attrib['base-type'] = "unit.Library.PLC_Types.Picture_PLC"
                    object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                    self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                    self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", name_pic)
                    self.dop_function.new_attr(object, "unit.System.Attributes.Description", name_pic)
                    
                    el1.append(object)
                tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
                msg[f'{today} - Файл omx: Pictures добавлены'] = 1
                return msg
            except Exception:
                msg[f'{today} - Файл omx: ошибка при добавлении Pictures: {traceback.format_exc()}'] = 2
                return msg
    def auxsystem_omx(self):
        msg = {}
        try:
            data_vs = self.dop_function.connect_by_sql('vs', f'"id", "name", "short_name", "Pressure_is_True", "Voltage", "OTKL"')
            data_ai = self.dop_function.connect_by_sql('ai', f'"id", "tag"')
            data_di = self.dop_function.connect_by_sql('di', f'"id", "tag"')
            data_do = self.dop_function.connect_by_sql('do', f'"id", "tag"')
            msg_bool, el1, tree = self.dop_function.parser_omx('AuxSystems')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке AuxSystems'] = 2
                return msg

            for value_vs in data_vs:
                number_vs = value_vs[0]
                name      = value_vs[1]
                shortdesc = value_vs[2]
                sensor    = value_vs[3]
                voltage   = value_vs[4]
                close     = value_vs[5]

                if number_vs == '' or number_vs is None: continue
                if shortdesc == '' or shortdesc is None: shortdesc = ''

                tag = 'VS_' + str(number_vs)
                isdigit = re.findall('\d+', sensor)
                # Ищем давление на выходе из таблицы AI и DI
                tag_sensor = ' '
                if self.dop_function.str_find(sensor.lower(), {'di'}):
                    for value_di in data_di:
                        number_di = value_di[0]
                        tag_di    = value_di[1]
                        if self.dop_function.str_find(number_di, isdigit):
                            tag_sensor = tag_di
                            break
                    pc_use = '1'
                elif self.dop_function.str_find(sensor.lower(), {'ai'}):
                    for value_ai in data_ai:
                        number_ai = value_ai[0]
                        tag_ai    = value_ai[1]
                        if self.dop_function.str_find(number_ai, isdigit):
                            tag_sensor = tag_ai
                            break
                    pc_use = str('2')
                else:
                    pc_use = str('0')
                # Ищем напряжение из таблицы DI
                tag_voltage = ' '
                isdigitVoltage = re.findall('\d+', voltage)
                if self.dop_function.str_find(voltage.lower(), {'di'}):
                    for value_di in data_di:
                        number_di = value_di[0]
                        tag_di_for_diagno = value_di[1]
                        if self.dop_function.str_find(number_di, isdigitVoltage):
                            tag_voltage  = tag_di_for_diagno
                            break
                # Ищем команду закрыть из таблицы DO
                tag_close = ' '
                isdigitCLOSE= re.findall('\d+', close)
                if self.dop_function.str_find(close.lower(), {'do'}):
                    for value_do in data_do:
                        number_do = value_do[0]
                        tag_do    = value_do[1]
                        if self.dop_function.str_find(number_do, isdigitCLOSE):
                            tag_close = tag_do
                            break

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = tag
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.AuxSystem_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number_vs)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", shortdesc)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", name)
                self.dop_function.new_attr(object, "unit.Library.Attributes.PC_Use", pc_use)
                self.dop_function.new_attr(object, "unit.Library.Attributes.PC_Ref", tag_sensor)
                self.dop_function.new_attr(object, "unit.Library.Attributes.DI_ref", tag_voltage)
                self.dop_function.new_attr(object, "unit.Library.Attributes.DO_ref", tag_close)

                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: AuxSystems добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении AuxSystems: {traceback.format_exc()}'] = 2
            return msg
    def valves_omx(self):
        msg = {}
        try:
            data_zd = self.dop_function.connect_by_sql('zd', f'"id", "name", "short_name", "VMMO", "VMMZ", "exists_interface", "Dist", "Dist_i", "KVO", "Open"')
            data_di = self.dop_function.connect_by_sql('di', f'"id", "tag"')
            data_do = self.dop_function.connect_by_sql('do', f'"id", "tag"')
            msg_bool, el1, tree = self.dop_function.parser_omx('Valves')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке Valves'] = 2
                return msg

            for value in data_zd:
                number    = value[0]
                name      = value[1]
                shortdesc = value[2]
                vmmo      = value[3]
                vmmz      = value[4]
                rs        = value[5]
                dist_f    = value[6]
                dist_i    = value[7]
                kvo_in_zd = value[8]
                open_in_zd= value[9]

                if number == '' or number is None: continue
                if name == '' or name   is None: continue

                isdigitKVO = re.findall('\d+', kvo_in_zd)
                if self.dop_function.str_find(kvo_in_zd.lower(), {'di'}):
                    for value_di in data_di:
                        tag_di = value_di[1]
                        if self.dop_function.str_find(value_di[0], isdigitKVO): break

                isdigitOPEN = re.findall('\d+', open_in_zd)
                if self.dop_function.str_find(open_in_zd.lower(), {'do'}):
                    for value_do in data_do:
                        tag_do = value_do[1]
                        if self.dop_function.str_find(value_do[0], isdigitOPEN): break

                tag    = f'ZD_{number}'
                # Наличие мутфа, авария
                isBUR  = True if (vmmo is None or vmmo == '') or (vmmz is None or vmmz == '') else False
                # Наличие ключа М/Д смотри по двум полям физика или интерфейс
                isDist = True if (not dist_i is None or dist_i == '') or (not dist_f is None or dist_f == '') else False
                # Наличие интерфейса
                isRS   = True if rs == 1 else False

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = tag
                object.attrib['uuid'] = str(uuid.uuid1())
                if isRS != True: object.attrib['base-type'] = "unit.Library.PLC_Types.Valve_PLC"
                else           : object.attrib['base-type'] = "unit.Library.PLC_Types.ex_Valve_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", shortdesc)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", name)
                self.dop_function.new_attr(object, "unit.Library.Attributes.BUR", isBUR)
                self.dop_function.new_attr(object, "unit.Library.Attributes.RS485", isRS)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Dist_key", isDist)
                self.dop_function.new_attr(object, "unit.Library.Attributes.DI_ref", tag_di)
                self.dop_function.new_attr(object, "unit.Library.Attributes.DO_ref", tag_do)

                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: Valves добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении Valves: {traceback.format_exc()}'] = 2
            return msg
    def pumps_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('umpna', f'"id", "name"')
            msg_bool, el1, tree = self.dop_function.parser_omx('NAs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке  NAs'] = 2
                return msg
            
            for value in data:
                number    = value[0]
                name      = value[1]

                if number is None: continue
                if name is None: continue

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = f'NA_{str(number)}'
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.NA_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                
                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", name)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", name)
                
                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: NAs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении NAs: {traceback.format_exc()}'] = 2
            return msg
    def relayted_system_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('ss', f'"id", "name"')
            msg_bool, el1, tree = self.dop_function.parser_omx('SSs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке SSs'] = 2
                return msg

            for value in data:
                number = value[0]
                name   = value[1]

                if number is None or number == '': continue
                if name   is None or name == '': continue

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = 'SS_' + str(number)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.SS_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", name)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", name)

                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: SSs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении SSs: {traceback.format_exc()}'] = 2
            return msg
    def uts_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('uts', f'"id", "tag", "name", "siren"')
            msg_bool, el1, tree = self.dop_function.parser_omx('UTSs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке UTSs'] = 2
                return msg

            for value in data:
                number = value[0]
                tag    = value[1]
                name   = value[2]
                siren  = value[3]

                if tag    is None or tag == '': continue
                if number is None or number == '': continue
                if name   is None or name == '': continue

                if int(siren)                                              : sign = 'Сирена'
                elif self.dop_function.str_find(str(name).lower(), {'газ'}): sign = 'Газ'
                else                                                       : sign = ''

                tag = self.dop_function.translate(str(tag))

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = str(tag)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.UTS_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", sign)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", name)
                
                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: UTSs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении UTSs: {traceback.format_exc()}'] = 2
            return msg
    def upts_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('upts', f'"id", "tag", "name", "location", "short_name"')
            msg_bool, el1, tree = self.dop_function.parser_omx('UPTSs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке UPTSs'] = 2
                return msg

            for value in data:
                number      = value[0]
                tag         = value[1]
                description = value[2]
                place       = value[3]
                shortdesc   = value[4]

                if number == '' or number is None: continue
                if tag == '' or tag is None: continue
                if description == '' or description is None: continue

                tag = self.dop_function.translate(str(tag))

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = str(tag)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.UPTS_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", shortdesc)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", description)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Place", place)
                
                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: UPTSs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении UPTSs: {traceback.format_exc()}'] = 2
            return msg
    def ktpr_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('ktpr', f'"id"')
            msg_bool, el1, tree = self.dop_function.parser_omx('KTPRs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке KTPRs'] = 2
                return msg
            number_group = 0

            for value in data:
                number_defence = value[0]
                if number_defence == '' or number_defence is None: continue
                number_group += 1
            count_group = math.ceil(number_group / 4)

            for count in range(count_group):
                object = etree.Element("{automation.control}object")
                object.attrib['name'] = f'Group_{count + 1}'
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.KTPRx_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: KTPRs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении KTPRs: {traceback.format_exc()}'] = 2
            return msg
    def ktprp_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('ktprp', f'"id"')
            msg_bool, el1, tree = self.dop_function.parser_omx('KTPRs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке KTPRs'] = 2
                return msg
            number_group = 0

            for value in data:
                number_defence = value[0]
                if number_defence == '' or number_defence is None: continue
                number_group += 1
            count_group = math.ceil(number_group/4)

            for count in range(count_group):
                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = f'Group_{count + 1}'
                    object.attrib['uuid'] = str(uuid.uuid1())
                    object.attrib['base-type'] = "unit.Library.PLC_Types.KTPRx_PLC"
                    object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                    el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: KTPRPs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении KTPRPs: {traceback.format_exc()}'] = 2
            return msg
    def ktpra_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('ktpra', f'"id", "NA"')
            msg_bool, el1, tree = self.dop_function.parser_omx('KTPRAs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке KTPRAs'] = 2
                return msg

            number_pumps_old = ''
            count_pumps      = 0

            for value in data:
                number_defence   = value[0]
                number_pumps_int = value[1]

                if number_defence == '' or number_defence   is None: continue
                if number_pumps_int == '' or number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps += 1
                    number_group = 0
                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = f'KTPRAs_{count_pumps}'
                    object.attrib['uuid'] = str(uuid.uuid1())

                if number_defence % 4 == 0:
                    number_group += 1
                    group = etree.Element("{automation.control}object")
                    group.attrib['name'] = f'Group_{number_group}'
                    group.attrib['uuid'] = str(uuid.uuid1())
                    group.attrib['base-type'] = "unit.Library.PLC_Types.KTPRx_PLC"
                    group.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                    object.append(group)
                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: KTPRAs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении KTPRAs: {traceback.format_exc()}'] = 2
            return msg
    def gmpna_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('gmpna', f'"id", "NA"')
            msg_bool, el1, tree = self.dop_function.parser_omx('GMPNAs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке GMPNAs'] = 2
                return msg

            number_pumps_old = ''
            count_pumps = 0

            for value in data:
                number_defence = value[0]
                number_pumps_int = value[1]

                if number_defence == '' or number_defence is None: continue
                if number_pumps_int == '' or number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps += 1
                    number_group = 0
                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = f'GMPNAs_{count_pumps}'
                    object.attrib['uuid'] = str(uuid.uuid1())

                if number_defence % 4 == 0:
                    number_group += 1
                    group = etree.Element("{automation.control}object")
                    group.attrib['name'] = f'Group_{number_group}'
                    group.attrib['uuid'] = str(uuid.uuid1())
                    group.attrib['base-type'] = "unit.Library.PLC_Types.GMPNA_PLC"
                    group.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                    object.append(group)
                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: GMPNAs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении GMPNAs: {traceback.format_exc()}'] = 2
            return msg
    def pi_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('pi', f'"id", "tag", "location", "name"')
            msg_bool, el1, tree = self.dop_function.parser_omx('PIs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке PIs'] = 2
                return msg

            for value in data:
                number      = value[0]
                tag         = value[1]
                shortdesc   = value[1]
                place       = value[2]
                description = value[3]

                if number == '' or number is None: continue
                if tag == '' or tag is None: continue
                if shortdesc == '' or shortdesc is None: continue
                if description == '' or description is None: continue
                if place == '' or place is None: place = ''

                tag = self.dop_function.translate(str(tag))

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = str(tag)
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.PI_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                
                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", shortdesc)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", description)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Place", place)

                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: PIs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении PIs: {traceback.format_exc()}'] = 2
            return msg
    def pz_omx(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('pz', f'"id", "name", "short_name"')
            msg_bool, el1, tree = self.dop_function.parser_omx('PZs')
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке PZs'] = 2
                return msg

            for value in data:
                number      = value[0]
                description = value[1]
                shortdesc   = value[2]

                if number == '' or number is None: continue
                if description == '' or description is None: continue
                if shortdesc == '' or shortdesc is None: shortdesc = ''

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = f'PZ_{number}'
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = "unit.Library.PLC_Types.PZ_PLC"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                
                self.dop_function.new_attr(object, "unit.Library.Attributes.Sign", shortdesc)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", description)
                self.dop_function.new_attr(object, "unit.Library.Attributes.Index", number)
                
                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: PZs добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении PZs: {traceback.format_exc()}'] = 2
            return msg
    
    def mklogic_AI_AO_atrib(self, variable_mod, type_mod, prefix):
        link_path = [f'{path_to_devstudio}\\AttributesMapAI_Ref.xml', 
                     f'{path_to_devstudio}\\AttributesMapKlk.xml', 
                     f'{path_to_devstudio}\\AttributesMapKont.xml', 
                     f'{path_to_devstudio}\\AttributesMapSignalName.xml',
                     f'{path_to_devstudio}\\AttributesMapTagName.xml']
        msg = {}
        try:
            data_hw = self.hardware_data(type_mod)

            for path in link_path:
                root, tree = self.dop_function.parser_diag_map(f'.Diag.{variable_mod}.', path)

                for data in data_hw:
                    uso          = data['uso']
                    string_name  = data['string_name']
                    num_basket   = data['num_basket']
                    number_modul = data['number_modul']

                    data_kd = self.dop_function.connect_by_sql_condition('signals', '*', f'''"uso"='{uso}' AND "basket"={int(num_basket)} AND "module"={int(number_modul)}''')
                    for data in data_kd:
                        uso     = data[2]
                        tag     = data[3]
                        name    = data[4]
                        klk     = data[6]
                        contact = data[7]
                        channel = data[10]

                        str_tag = self.dop_function.translate(str(tag))
                        if klk == '' or klk is None: klk = ' '
                        if contact == '' or contact is None: contact = ' '
                        if tag == '' or tag is None: str_tag = ' '

                        name_AO = f'Root{prefix_system}.Diag.{variable_mod}.{string_name}.{prefix}{str(channel)}'

                        object = etree.Element('item')
                        object.attrib['id'] = name_AO

                        if path == f'{path_to_devstudio}\\AttributesMapAI_Ref.xml': 
                            if not str_tag is None or str_tag == '': object.attrib['value'] = str(str_tag)
                        if path == f'{path_to_devstudio}\\AttributesMapKlk.xml': 
                            if not klk is None or klk == '': object.attrib['value'] = str(klk)
                        if path == f'{path_to_devstudio}\\AttributesMapKont.xml': 
                            if not contact is None or contact == '': object.attrib['value'] = str(contact)
                        if path == f'{path_to_devstudio}\\AttributesMapSignalName.xml': 
                            if not name is None or name == '': object.attrib['value'] = str(name)
                        if path == f'{path_to_devstudio}\\AttributesMapTagName.xml': 
                            if not tag is None or tag == '': object.attrib['value'] = str(tag)

                        msg[f'{today} - Значения атрибутов Diag.{variable_mod} файл заполнен: {path}'] = 1
                        root.append(object)
                tree.write(path, pretty_print=True)
            msg[f'{today} - Значения атрибутов Diag.{variable_mod} добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Ошибка при добавлении значений атрибутов Diag.{variable_mod}: {traceback.format_exc()}'] = 2
            return msg
    def mklogic_DI_DO_atrib(self, variable_mod, type_mod):
        link_path = [f'{path_to_devstudio}\\AttributesMapKlk.xml', 
                     f'{path_to_devstudio}\\AttributesMapKont.xml', 
                     f'{path_to_devstudio}\\AttributesMapSignalName.xml',
                     f'{path_to_devstudio}\\AttributesMapTagName.xml']
        msg = {}
        try:
            data_hw = self.hardware_data(type_mod)

            for path in link_path:
                root, tree = self.dop_function.parser_diag_map('.Diag.{variable_mod}.', path)

                for data in data_hw:
                    uso          = data['uso']
                    string_name  = data['string_name']
                    num_basket   = data['num_basket']
                    number_modul = data['number_modul']

                    data_kd = self.dop_function.connect_by_sql_condition('signals', '*', f'''"uso"='{uso}' AND "basket"={int(num_basket)} AND "module"={int(number_modul)}''')
                    for data in data_kd:
                        uso     = data[2]
                        tag     = data[3]
                        name    = data[4]
                        klk     = data[6]
                        contact = data[7]
                        channel = data[10]

                        if klk == '' or klk is None: klk = ' '
                        if contact == '' or contact is None: contact = ' '
                        if tag == '' or tag is None: tag = ' '

                        if channel < 10: name_DI = f'Root{prefix_system}.Diag.{variable_mod}.{string_name}.ch_DI_0{str(channel)}'
                        else           : name_DI = f'Root{prefix_system}.Diag.{variable_mod}.{string_name}.ch_DI_{str(channel)}'

                        object = etree.Element('item')
                        object.attrib['id'] = name_DI

                        if path == f'{path_to_devstudio}\\AttributesMapKlk.xml': 
                            if not klk is None or klk == '': object.attrib['value'] = str(klk)
                        if path == f'{path_to_devstudio}\\AttributesMapKont.xml': 
                            if not contact is None or contact == '': object.attrib['value'] = str(contact)
                        if path == f'{path_to_devstudio}\\AttributesMapSignalName.xml': 
                            if not name is None or name == '': object.attrib['value'] = str(name)
                        if path == f'{path_to_devstudio}\\AttributesMapTagName.xml': 
                            if not tag is None or tag == '': object.attrib['value'] = str(tag)

                        msg[f'{today} - Значения атрибутов Diag.{variable_mod} файл заполнен: {path}'] = 1
                        root.append(object)
                tree.write(path, pretty_print=True)
            msg[f'{today} - Значения атрибутов Diag.{variable_mod} добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Ошибка при добавлении значений атрибутов Diag.{variable_mod}: {traceback.format_exc()}'] = 2
            return msg
    def mklogic_DIAG_omx(self, variable_mod, type_mod, base_type):
        msg = {}
        try:
            msg_bool, el1, tree = self.dop_function.parser_diag_omx(variable_mod)
            if msg_bool == 1: 
                msg[f'{today} - Файл omx: ошибка при очистке Diag.{variable_mod}'] = 2
                return msg
            
            data_hw = self.hardware_data(type_mod)
            for data in data_hw:
                id_          = data['id_']
                uso          = data['uso']
                string_name  = data['string_name']
                number_modul = data['number_modul']
                num_basket   = data['num_basket']
                modPosition  = data['modPosition']

                object = etree.Element("{automation.control}object")
                object.attrib['name'] = string_name
                object.attrib['uuid'] = str(uuid.uuid1())
                object.attrib['base-type'] = f"unit.Library.PLC_Types.modules.{base_type}"
                object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"

                self.dop_function.new_attr(object, "unit.Library.Attributes.ModNumber", number_modul)
                self.dop_function.new_attr(object, "unit.Library.Attributes.RackNumber", id_)
                self.dop_function.new_attr(object, "unit.Library.Attributes.ModPosition", modPosition)
                self.dop_function.new_attr(object, "unit.Library.Attributes.ModUSO", uso)
                self.dop_function.new_attr(object, "unit.System.Attributes.Description", type_mod)
                # Только для RS
                if variable_mod == 'RSs': 
                    rs_port1 = 'Резерв'
                    rs_port2 = 'Резерв'
                    data_kd = self.dop_function.connect_by_sql_condition('signals', '*', f'''"uso"='{uso}' AND "basket"={int(num_basket)} AND "module"={int(number_modul)}''')
                    for data in data_kd:
                        channel = data[10]
                        name_kd = data[4]
                        if channel == 1: rs_port1 = name_kd
                        if channel == 2: rs_port2 = name_kd
                    self.dop_function.new_attr(object, "unit.Library.Attributes.SignalName", rs_port1)
                    self.dop_function.new_attr(object, "unit.Library.Attributes.SignalName_2", rs_port2)

                el1.append(object)
            tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
            msg[f'{today} - Файл omx: Diag.{variable_mod} добавлены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл omx: ошибка при добавлении Diag.{variable_mod}: {traceback.format_exc()}'] = 2   
            return msg
    def rackstate_omx(self):
            msg = {}
            try:
                msg_bool, el1, tree = self.dop_function.parser_diag_omx('RackStates')
                if msg_bool == 1: 
                    msg[f'{today} - Файл omx: ошибка при очистке Diag.RackStates'] = 2
                    return msg
                
                cursor = db.cursor()
                cursor.execute(f'''SELECT COUNT (*) FROM hardware''')
                for i in cursor.fetchall():
                    count = i[0]

                for i in range(count):
                    object = etree.Element("{automation.control}object")
                    object.attrib['name'] = f'rack_{i + 1}'
                    object.attrib['uuid'] = str(uuid.uuid1())
                    object.attrib['base-type'] = f"unit.Library.PLC_Types.modules.MK_Logic.rack_State"
                    object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                    
                    el1.append(object)
                tree.write(f'{path_to_devstudio}\\typical_prj.omx', pretty_print=True)
                msg[f'{today} - Файл omx: Diag.RackStates добавлены'] = 1
                return msg
            except Exception:
                msg[f'{today} - Файл omx: ошибка при добавлении Diag.RackStates: {traceback.format_exc()}'] = 2   
                return msg     
    def colorscheme_di(self):
        link_path = f'{path_to_devstudio}\\AttributesMapColorScheme.xml'
        dop_color = {"01":"1",
                     "02":"2",
                     "03":"3",
                     "10":"11",
                     "20":"12",
                     "30":"9",
                     "12":"4",
                     "13":"5",
                     "21":"10",
                     "31":"7",
                     "23":"6",
                     "32":"8",
                     "00":"0",
                     "11":"1"}
        msg = {}
        try:
            data_di = self.dop_function.connect_by_sql('di', f'"tag", "name", "priority_0", "priority_1"')
            root, tree = self.dop_function.parser_diag_map(f'.Diskrets.', link_path)

            for data in data_di:
                tag  = data[0]
                name = data[1]
                priority_0 = data[2]
                priority_1 = data[3]

                if priority_0 is None: priority_0 = '0'
                else                 : priority_0 = str(priority_0)

                if priority_1 is None: priority_1 = '0'
                else                 : priority_1 = str(priority_1)

                color_shema = str(dop_color[priority_0 + priority_1])

                if name is None: continue
                if tag  is None: continue
                tag = self.dop_function.translate(str(tag))

                object = etree.Element('item')
                object.attrib['id'] = f'Root.Diskrets.{tag}.s_Config'
                object.attrib['id'] = f'Root{prefix_system}.Diskrets.{tag}.s_Config'
                object.attrib['value'] = color_shema
                root.append(object)

                tree.write(link_path, pretty_print=True)
            msg[f'{today} - Карта атрибутов AttributesMapColorScheme заполнена'] = 1
            return msg
        except Exception:
            msg[f'{today} - Ошибка при добавлении значений в карту атрибутов AttributesMapColorScheme: {traceback.format_exc()}'] = 2
            return msg
    def analogformat_map(self):
        link_path = f'{path_to_devstudio}\\AttributesAnalogsFormats.xml'
        msg = {}
        try:
            data_ai = self.dop_function.connect_by_sql('ai', f'"tag", "AnalogGroupId", "Precision"')
            data_aigrp = self.dop_function.connect_by_sql('ai_grp', f'"name", "min6", "min5", "min4", "min3", "min2", "min1", "max1", "max2", "max3", "max4", "max5", "max6"')
            root, tree = self.dop_function.parser_diag_map(f'.Analogs.', link_path)

            for data in data_ai:
                tag          = data[0]
                group_analog = data[1]
                formate      = data[2]

                if tag == '' or tag is None: continue
                tag = self.dop_function.translate(str(tag))

                for grp in data_aigrp:
                     if group_analog == grp[0]:
                        grp_ai = {'Format': formate,
                                  'UstName.UstMin1': grp[6], 'UstName.UstMin2': grp[5], 'UstName.UstMin3': grp[4],
                                  'UstName.UstMin4': grp[3], 'UstName.UstMin5': grp[2], 'UstName.UstMin6': grp[1],
                                  'UstName.UstMax1': grp[7], 'UstName.UstMax2': grp[8], 'UstName.UstMax3': grp[9],
                                  'UstName.UstMax4': grp[10],'UstName.UstMax5': grp[11],'UstName.UstMax6': grp[12]}
                        for grp, value in grp_ai.items():
                            object = etree.Element('item')
                            object.attrib['id']    = f'Root.{prefix_system}Analogs.{tag}.{grp}'
                            object.attrib['value'] = str(value)
                            root.append(object)
                tree.write(link_path, pretty_print=True)
            msg[f'{today} - Карта атрибутов AttributesAnalogsFormats заполнена'] = 1
            return msg
        except Exception:
            msg[f'{today} - Ошибка при добавлении значений в карту атрибутов AttributesAnalogsFormats: {traceback.format_exc()}'] = 2
            return msg
    def egu_map(self):
        link_path = f'{path_to_devstudio}\\AttributesMapEGU.xml'
        msg = {}
        try:
            data_ai = self.dop_function.connect_by_sql('ai', f'"tag", "Egu"')
            root, tree = self.dop_function.parser_diag_map(f'.Analogs.', link_path)

            for data in data_ai:
                tag = data[0]
                egu = data[1]

                if tag == '' or tag is None: continue
                tag = self.dop_function.translate(str(tag))

                object = etree.Element('item')
                object.attrib['id'] = f'Root{prefix_system}.Analogs.{tag}.AIValue'
                object.attrib['value'] = str(egu)
                root.append(object)

                tree.write(link_path, pretty_print=True)
            msg[f'{today} - Карта атрибутов AttributesMapEGU заполнена'] = 1
            return msg
        except Exception:
            msg[f'{today} - Ошибка при добавлении значений в карту атрибутов AttributesMapEGU: {traceback.format_exc()}'] = 2
            return msg
    # Заполнение карты адресов
    def analogs_maps(self):
        dop_analog    = {'AIVisualValue', 'AIElValue', 'AIValue', 'AIRealValue', 'StateAI'}
        dop_analog_lv = {'AIVisualValue', 'AIElValue', 'AIValue', 'AIRealValue', 'StateAI', 'Range_Bottom', 'Range_Top'}
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('ai', f'"id", "tag", "name", "AnalogGroupId"')
            root, tree  = self.dop_function.parser_map('.Analogs.')

            for value in data:
                number     = value[0]
                tag        = value[1]
                name       = value[2]
                grp_analog = value[3]

                if tag == '' or tag is None: continue
                if name == '' or name is None: continue

                # У уровней особый набор параметров
                if grp_analog == 'Уровни': list_analog = dop_analog_lv
                else:                      list_analog = dop_analog

                tag = self.dop_function.translate(str(tag))

                for key in list_analog:
                    signal = f'Root{prefix_system}.Analogs.{tag}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'

                    self.dop_function.new_map_str(object, 'node-path', signal)
                    self.dop_function.new_map_str(object, 'table-path', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', 2 * (number - 1))
                    
                    root.append(object)
                tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса Analogs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов Analogs: {traceback.format_exc()}'] = 2
            return msg
    def diskret_maps(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('di', f'"id", "tag", "name"')
            root, tree  = self.dop_function.parser_map('.Diskrets.')

            for value in data:
                number = value[0]
                tag    = value[1]
                name   = value[2]

                if tag == '' or tag  is None: continue
                if name == '' or name is None: continue

                tag = self.dop_function.translate(str(tag))
                signal = f'Root{prefix_system}.Diskrets.{tag}.StateDI'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'

                self.dop_function.new_map_str(object, 'node-path', signal)
                self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                self.dop_function.new_map_str(object, 'address', 2 * (number - 1))
                    
                root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса Diskrets заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов Diskrets: {traceback.format_exc()}'] = 2
            return msg
    def picturs_maps(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('pic', f'"id", "frame"')
            root, tree  = self.dop_function.parser_map('.Pictures.')

            for value in data:
                number = value[0]
                frame  = value[1]

                if frame == '' or frame is None: continue
                signal = f'Root{prefix_system}.Pictures.{frame}.StatePicture'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
        
                self.dop_function.new_map_str(object, 'node-path', signal)
                self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                self.dop_function.new_map_str(object, 'address', 2 * (number - 1))
                    
                root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса Pictures заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов Pictures: {traceback.format_exc()}'] = 2
            return msg
    def auxsystem_maps(self):
        dop_vs = {'StateAuxSystem', 'numOfStart', 'operatingTimeCurrentMonth', 'operatingTimeLastMonth', 'operatingTime'}
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('vs', f'"id"')
            root, tree  = self.dop_function.parser_map('.AuxSystems.')

            for value in data:
                number = value[0]

                if number == '' or number is None: continue
                for key in dop_vs:
                    tag = f'VS_{number}'
                    signal = f'Root{prefix_system}.AuxSystems.{tag}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'

                    self.dop_function.new_map_str(object, 'node-path', signal)
                    self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', 2 * (number - 1))

                    root.append(object)
                tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса AuxSystems заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов AuxSystems: {traceback.format_exc()}'] = 2
            return msg
    def valves_maps(self):
        dop_zd = {'StateValve1', 'StateValve2', 'StateValve3', 'Tm.tmZD', 'NumOfOpenings', 'NumOfClosings'}
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('zd', f'"id"')
            root, tree  = self.dop_function.parser_map('.Valves.')

            for value in data:
                number = value[0]

                if number == '' or number is None: continue

                for key in dop_zd:
                    tag = f'ZD_{number}'
                    signal = f'Root{prefix_system}.Valves.{tag}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'

                    self.dop_function.new_map_str(object, 'node-path', signal)
                    self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', 2 * (number - 1))

                    root.append(object)
                tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса Valves заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов Valves: {traceback.format_exc()}'] = 2
            return msg
    def na_maps(self):
        dop_na = {'StateNA','StateNAEx','StateNAStatistic','operatingTimeSinceSwitchingOn','operatingTimeSinceSwitchingOnSet','operatingTimeBeforeOverhaul',
                  'operatingTimeBeforeOverhaulSet','numOfStart','dateTimeOfStart','dateTimeOfStop','operatingTimeCurrentMonth','operatingTimeLastMonth',
                  'operatingTimeTO','operatingTimeTO1','operatingTimeTOSet','operatingTimeMidTO','operatingTimeMidTOSet','operatingTimeThisKvart',
                  'operatingTimeLastKvart','operatingTimeFromBegin','operatingTimeED','operatingTimeEDSet','numOfStartSet','time24hStart',
                  'timeFromHotStart','numOfStarts24h','OperatingTimeState',}
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('umpna', f'"id"')
            root, tree  = self.dop_function.parser_map('.NAs.')

            for value in data:
                number = value[0]

                if number == '' or number is None: continue

                for key in dop_na:
                    tag    = f'NA_{number}'
                    signal = f'Root{prefix_system}.NAs.{tag}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    
                    self.dop_function.new_map_str(object, 'node-path', signal)
                    self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', 2 * (number - 1))

                    root.append(object)
                tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса NAs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов NAs: {traceback.format_exc()}'] = 2
            return msg
    def ss_maps(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('ss', f'"id", "name"')
            root, tree  = self.dop_function.parser_map('.SSs.')

            for value in data:
                number = value[0]
                name   = value[1]

                if number == '' or number is None: continue
                if name == '' or name   is None: continue

                signal = f'Root{prefix_system}.SSs.SS_{number}.StateSS'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'

                self.dop_function.new_map_str(object, 'node-path', signal)
                self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                self.dop_function.new_map_str(object, 'address', 2 * (number - 1))

                root.append(object)
                tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса SSs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов SSs: {traceback.format_exc()}'] = 2
            return msg
    def uts_maps(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('uts', f'"id", "tag", "name"')
            root, tree  = self.dop_function.parser_map('.UTSs.')

            for value in data:
                number = value[0]
                tag    = value[1]
                name   = value[2]

                if number == '' or number is None: continue
                if name == '' or name is None: continue
                if tag == '' or tag is None: continue

                tag = self.dop_function.translate(str(tag))
                signal = f'Root{prefix_system}.UTSs.{tag}.StateUTS'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                
                self.dop_function.new_map_str(object, 'node-path', signal)
                self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                self.dop_function.new_map_str(object, 'address', 2 * (number - 1))

                root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса UTSs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов UTSs: {traceback.format_exc()}'] = 2
            return msg
    def upts_maps(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('upts', f'"id", "tag", "name"')
            root, tree  = self.dop_function.parser_map('.UPTSs.')

            for value in data:
                number = value[0]
                tag    = value[1]
                name   = value[2]

                if number == '' or number is None: continue
                if name == '' or name is None: continue
                if tag == '' or tag is None: continue

                tag = self.dop_function.translate(str(tag))
                signal = f'Root{prefix_system}.UPTSs.{tag}.StateUPTS'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'

                self.dop_function.new_map_str(object, 'node-path', signal)
                self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                self.dop_function.new_map_str(object, 'address', 2 * (number - 1))

                root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса UPTSs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов UPTSs: {traceback.format_exc()}'] = 2
            return msg
    def ktpr_maps(self):
        msg = {}
        number_group = 0
        try:
            data = self.dop_function.connect_by_sql('ktpr', f'"id"')
            root, tree  = self.dop_function.parser_map('.KTPRs.')

            for value in data:
                number_defence = value[0]
                if number_defence == '' or number_defence is None: continue
                number_group += 1

            count_group = math.ceil(number_group/4)

            for count in range(count_group):
                signal = f'Root{prefix_system}.KTPRs.Group_{count + 1}.StateKTPRx'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'

                self.dop_function.new_map_str(object, 'node-path', signal)
                self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                self.dop_function.new_map_str(object, 'address', 2 * ((count - 1)))

                root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса KTPRs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов KTPRs: {traceback.format_exc()}'] = 2
            return msg
    def ktprp_maps(self):
        msg = {}
        number_group = 0
        try:
            data = self.dop_function.connect_by_sql('ktprp', f'"id"')
            root, tree  = self.dop_function.parser_map('.KTPRs.')

            for value in data:
                number_defence = value[0]
                if number_defence == '' or number_defence is None: continue
                number_group += 1

            count_group = math.ceil(number_group / 4)

            for count in range(count_group):
                signal = f'Root{prefix_system}.KTPRs.Group_{count + 1}.StateKTPRx'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                
                self.dop_function.new_map_str(object, 'node-path', signal)
                self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                self.dop_function.new_map_str(object, 'address', 2 * ((count - 1)))

                root.append(object)

            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса KTPRPs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов KTPRPs: {traceback.format_exc()}'] = 2
            return msg 
    def ktpra_maps(self):
        msg = {}
        number_pumps_old = ''
        count_pumps      = 0
        count            = 0
        try:
            data = self.dop_function.connect_by_sql('ktpra', f'"id", "NA"')
            root, tree  = self.dop_function.parser_map('.KTPRAs.')

            for value in data:
                number_defence   = value[0]
                number_pumps_int = value[1]

                if number_defence == '' or  number_defence   is None: continue
                if number_pumps_int == '' or number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps  += 1
                    number_group  = 0

                if number_defence % 4 == 0:
                    number_group += 1
                    count += 1
                    signal = f'Root{prefix_system}.KTPRAs.KTPRAs_{count_pumps}.Group_{number_group}.StateKTPRx'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'

                    self.dop_function.new_map_str(object, 'node-path', signal)
                    self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', 2 * ((count - 1)))

                    root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса KTPRAs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов KTPRAs: {traceback.format_exc()}'] = 2
            return msg  
    def gmpna_maps(self):
        msg = {}
        number_pumps_old = ''
        count_pumps      = 0
        count            = 0
        try:
            data = self.dop_function.connect_by_sql('gmpna', f'"id", "NA"')
            root, tree  = self.dop_function.parser_map('.GMPNAs.')

            for value in data:
                number_defence   = value[0]
                number_pumps_int = value[1]

                if number_defence == '' or  number_defence   is None: continue
                if number_pumps_int == '' or number_pumps_int is None: continue

                if number_pumps_int != number_pumps_old:
                    number_pumps_old = number_pumps_int
                    count_pumps += 1
                    number_group = 0

                if number_defence % 4 == 0:
                    number_group += 1
                    count        += 1
                    signal = f'Root{prefix_system}.GMPNAs.GMPNAs_{count_pumps}.Group_{number_group}.StateGMPNA'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'

                    self.dop_function.new_map_str(object, 'node-path', signal)
                    self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', 2 * ((count - 1)))

                    root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса GMPNAs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов GMPNAs: {traceback.format_exc()}'] = 2
            return msg  
    def pi_maps(self):
        msg = {}
        try:
            data = self.dop_function.connect_by_sql('pi', f'"id", "tag"')
            root, tree  = self.dop_function.parser_map('.PIs.')

            for value in data:
                number = value[0]
                tag    = value[1]

                if number == '' or number is None: continue
                if tag == '' or tag    is None: continue

                tag = self.translate(str(tag))
                signal = f'Root{prefix_system}.PIs.{tag}.StatePI'

                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'

                self.dop_function.new_map_str(object, 'node-path', signal)
                self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                self.dop_function.new_map_str(object, 'address', 2 * ((number - 1)))

                root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса PIs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов PIs: {traceback.format_exc()}'] = 2
            return msg  
    def pz_maps(self):
        # Зоны с тушением
        dop_pz_ptush = ['StatePZ', 'exStatePZ', 'ReadyFlags', 'TimetoNextAttack', 'AttackCounter', 'TimetoEvacuation']
        # Зоны без тушения
        #dop_pz = ['StatePZ', 'exStatePZ', 'ReadyFlags']

        msg = {}
        try:
            data = self.dop_function.connect_by_sql('pz', f'"id", "type_zone"')
            root, tree  = self.dop_function.parser_map('.PZs.')

            for value in data:
                number    = value[0]
                zone_type = value[1]

                if number == '' or number    is None: continue
                if zone_type == '' or zone_type is None: continue
                # Выбираем от типа
                #set_words = dop_pz if zone_type == 0 else dop_pz_ptush

                for key in dop_pz_ptush:
                    signal = f'Root{prefix_system}.PZs.PZ_{number}.{key}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'

                    self.dop_function.new_map_str(object, 'node-path', signal)
                    self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', 2 * ((number - 1)))

                    root.append(object)
            tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса PZs заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов PZs: {traceback.format_exc()}'] = 2
            return msg  
    
    def mklogic_DIAG_maps_param(self, variable_mod, type_mod, list_param):
        msg = {}
        try:
            root, tree  = self.dop_function.parser_map(f'.Diag.{variable_mod}.')
            data_hw = self.hardware_data(type_mod)

            for data in data_hw:
                string_name  = data['string_name']
                for i in list_param:
                    name = f'Root{prefix_system}.Diag.{variable_mod}.{string_name}.{i}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'

                    self.dop_function.new_map_str(object, 'node-path', f'{name}')
                    self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', f'{i}')
                    
                    root.append(object)
                tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса Diag.{variable_mod} заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов Diag.{variable_mod}: {traceback.format_exc()}'] = 2
            return msg  
    def mklogic_rackstates_maps(self, list_param):
        msg = {}
        try:
            root, tree  = self.dop_function.parser_map(f'.Diag.RackStates.')
            data_hw = self.dop_function.connect_count_row_sql('hardware')
            if data_hw == True:
                msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов RackStates. Количество корзин не определено'] = 2
                return msg
            for i in data_hw:
                count = i[0]

            for i in range(count):
                for j in list_param:
                    name = f'Root{prefix_system}.Diag.RackStates.rack_{i + 1}.{j}'

                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'

                    self.dop_function.new_map_str(object, 'node-path', f'{name}')
                    self.dop_function.new_map_str(object, 'table', f'Holding Registers')
                    self.dop_function.new_map_str(object, 'address', f'{i}')
                    
                    root.append(object)
                tree.write(f'{path_to_devstudio}\\OUA.xml', pretty_print=True)
            msg[f'{today} - Карта адресов: адреса Diag.RackStates заполнены'] = 1
            return msg
        except Exception:
            msg[f'{today} - Карта адресов: ошибка при заполнении карты адресов Diag.RackStates: {traceback.format_exc()}'] = 2
            return msg  

# Filling attribute DevStudio
class Filling_CodeSys():     
    def __init__(self):
        self.dop_function = General_functions()
    def write_in_file(self, list_tabl):
        msg = {}
        if len(list_tabl) == 0:             
            msg[f'{today} - Файл СУ: не выбраны атрибуты'] = 2
            return msg
        for tabl in list_tabl: 
            if tabl == 'cfg_KTPRS': 
                msg.update(self.cfg_ktprs())
                continue
            if tabl == 'cfg_VV': 
                msg.update(self.cfg_vv())
                continue
            if tabl == 'cfg_UTS': 
                msg.update(self.cfg_uts())
                continue
            if tabl == 'cfg_VSGRP': 
                msg.update(self.cfg_vsgrp())
                continue
            if tabl == 'cfg_NPS': 
                msg.update(self.cfg_nps())
                continue
            if tabl == 'cfg_RSREQ': 
                msg.update(self.cfg_rsreq())
                continue
            if tabl == 'cfg_NA': 
                msg.update(self.cfg_na())
                continue
            if tabl == 'cfg_KTPR': 
                msg.update(self.cfg_ktpr())
                continue
            if tabl == 'cfg_KTPRA': 
                msg.update(self.cfg_ktpra())
                continue
            if tabl == 'cfg_VS': 
                msg.update(self.cfg_vs())
                continue
            if tabl == 'cfg_ZD': 
                msg.update(self.cfg_zd())
                continue
            if tabl == 'cfg_DO': 
                msg.update(self.cfg_do())
                continue
            if tabl == 'cfg_DI': 
                msg.update(self.cfg_di())
                continue
            if tabl == 'cfg_AO': 
                msg.update(self.cfg_ao())
                continue
            if tabl == 'cfg_AI': 
                msg.update(self.cfg_ai())
                continue
            if tabl == 'cfg_DPS': 
                msg.update(self.cfg_dps())
                continue
            if tabl == 'cfg_DIAG': 
                msg.update(self.cfg_diag())
                continue
            if tabl == 'cfg_PIC': 
                msg.update(self.cfg_pic())
                continue
            if tabl == 'cfg_TM_TS': 
                msg.update(self.cfg_ts())
                continue
            if tabl == 'cfg_TM_TU': 
                msg.update(self.cfg_tu())
                continue
            if tabl == 'cfg_TM_TI2': 
                msg.update(self.cfg_ti2())
                continue
            if tabl == 'cfg_TM_TI4': 
                msg.update(self.cfg_ti4())
                continue
            if tabl == 'cfg_TM_TII': 
                msg.update(self.cfg_tii())
                continue
            if tabl == 'cfg_TM_TR2': 
                msg.update(self.cfg_tr2())
                continue
            if tabl == 'cfg_TM_TR4': 
                msg.update(self.cfg_tr4())
                continue
            if tabl == 'cfg_DO_sim': 
                msg.update(self.cfg_do_sim())
                continue
            if tabl == 'cfg_DI_sim': 
                msg.update(self.cfg_di_sim())
                continue
            if tabl == 'cfg_AI_sim': 
                msg.update(self.cfg_ai_sim())
                continue
        return msg
    def file_check(self, name_file):
        path_request = f'{path_su}\\{name_file}.txt'
        if not os.path.exists(path_request):
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(f'(*{name_file}*)\n')
        else:
            os.remove(path_request)
            file = codecs.open(path_request, 'w', 'utf-8')
            file.write(f'(*{name_file}*)\n')
        return file
    def ret_inp_cfg(self, inp):
        stateAI = {'Warn'  : 0,
                   'Avar'  : 1,
                   'LTMin' : 2,
                   'MTMax' : 3,
                   'AlgNdv': 4,
                   'Imit'  : 5,
                   'ExtNdv': 6,
                   'Ndv'   : 7,
                   'Init'  : 8}
        stateAIzone = {'rez_0'                  : 0,
                       'Min6'                   : 1,
                       'Min5'                   : 2,
                       'Min4'                   : 3,
                       'Min3_IsMT10Perc'        : 4,
                       'Min2_IsNdv2ndParam'     : 5,
                       'Min1_IsHighVibStat'     : 6,
                       'Norm'                   : 7,
                       'Max1_IsHighVibStatNMNWR': 8,
                       'Max2_IsHighVibNoStat'   : 9,
                       'Max3_IsAvarVibStat'     : 10,
                       'Max4_IsAvarVibStatNMNWR': 11,
                       'Max5_IsAvarVibNoStat'   : 12,
                       'Max6_IsAvar2Vib'        : 13,
                       'rez_14'                 : 14,
                       'rez_15'                 : 15}
        stateDI = {'Value'    : 0,
                   'ElInput'  : 1,
                   'O'        : 2,
                   'KZ'       : 3,
                   'NC'       : 4,
                   'Imit'     : 5,
                   'ExtNdv'   : 6,
                   'Ndv'      : 7,
                   'priority1': 8,
                   'priority2': 9,
                   'priority3': 10,
                   'rez_11'   : 11,
                   'rez_12'   : 12,
                   'Front_0_1': 13,
                   'Front_1_0': 14,
                   'CfgErr'   : 15}
        stateNPS = {'ModeNPSDst'       : 0,
                    'MNSInWork'        : 1,
                    'IsMNSOff'         : 2,
                    'IsNPSModePsl'     : 3,
                    'IsPressureReady'  : 4,
                    'NeNomFeedInterval': 5,
                    'OIPHighPressure'  : 6,
                    'KTPR_P'           : 7,
                    'KTPR_M'           : 8,
                    'CSPAlinkOK'       : 9,
                    'CSPAWorkDeny'     : 10,
                    'TSstopped'        : 11,
                    'rez_12'           : 12,
                    'stopDisp'         : 13,
                    'stopCSPA'         : 14,
                    'stopARM'          : 15}
        stateFacility = {'longGasPoint1': 0,
                         'longGasPoint2': 1,
                         'longGasPoint3': 2,
                         'longGasPoint4': 3,
                         'longGasPoint5': 4,
                         'longGasPoint6': 5,
                         'longGasPoint7': 6,
                         'longGasPoint8': 7,
                         'rez_8'        : 8,
                         'rez_9'        : 9,
                         'rez_10'       : 10,
                         'rez_11'       : 11,
                         'rez_12'       : 12,
                         'rez_13'       : 13,
                         'rez_14'       : 14,
                         'rez_15'       : 15}
        warnFacility = {'warnGasPoint1': 0,
                        'warnGasPoint2': 1,
                        'warnGasPoint3': 2,
                        'warnGasPoint4': 3,
                        'warnGasPoint5': 4,
                        'warnGasPoint6': 5,
                        'warnGasPoint7': 6,
                        'warnGasPoint8': 7,
                        'rez_8'        : 8,
                        'rez_9'        : 9,
                        'rez_10'       : 10,
                        'rez_11'       : 11,
                        'rez_12'       : 12,
                        'rez_13'       : 13,
                        'rez_14'       : 14,
                        'rez_15'       : 15}
        Facility = {'ndv2Gas'   : 0,
                    'GasLim'    : 1,
                    'GasAv'     : 2,
                    'GasKeep'   : 3,
                    'GasNdvWait': 4,
                    'GasLimWait': 5,
                    'GasNdvProt': 6,
                    'GasAvProt' : 7,
                    'ColdOn'    : 8,
                    'HotOn'     : 9,
                    'rez_10'    : 10,
                    'rez_11'    : 11,
                    'ColdOff'   : 12,
                    'HotOff'    : 13,
                    'rez_14'    : 14,
                    'rez_15'    : 15}
        vsgrpstate = {'REZ_EXIST'           : 0,
                      'REM'                 : 1,
                      'OTKL'                : 2,
                      'OTKL_BY_CMD'         : 3,
                      'VKL_AS_DOP'          : 4,
                      'PUSK_OSN'            : 5,
                      'rez_6'               : 6,
                      'rez_7'               : 7,
                      'rez_8'               : 8,
                      'rez_9'               : 9,
                      'rez_10'              : 10,
                      'rez_11'              : 11,
                      'rez_12'              : 12,
                      'rez_13'              : 13,
                      'LAST_OFF_BY_CMD_ARM ': 14,
                      'ALL_OFF_WITH_BLOCK ' : 15}
        statektpr = {'P': 0,
                     'F': 1,
                     'M': 2}
        state_na = {'MainState_1_VKL'     : 0,
                    'MainState_2_OTKL'    : 1,
                    'MainState_3_PUSK'    : 2,
                    'MainState_4_OSTANOV' : 3,
                    'SubState_1_GP'       : 4,
                    'SubState_2_GORREZ'   : 5,
                    'SubState_3_PP'       : 6,
                    'SubState_4_PO'       : 7,
                    'Mode_1_OSN'          : 8,
                    'Mode_2_TU'           : 9,
                    'Mode_3_REZ'          : 10,
                    'Mode_4_REM'          : 11,
                    'KTPRA_P'             : 12,
                    'SimAgr'              : 13,
                    'Prog_1'              : 14,
                    'Prog_2'              : 15
                    }
        state_na2 = {'HIGHVIB'      : 0,
                     'HIGHVIBNas'   : 1,
                     'QF3A'         : 2,
                     'QF1A'         : 3,
                     'BBon'         : 4,
                     'BBoff'        : 5,
                     'KTPRA_FNM'    : 6,
                     'KTPRA_M'      : 7,
                     'GMPNA_M'      : 8,
                     'BBErrOtkl_All': 9,
                     'BBErrOtkl'    : 10,
                     'BBErrOtkl1'   : 11,
                     'BBErrVkl'     : 12,
                     'GMPNA_P'      : 13,
                     'GMPNA_F'      : 14,
                     'StateAlarm_VV': 15}
        state_na3 = {'KKCAlarm1'   : 0,
                     'KKCAlarm2'   : 1,
                     'KKCAlarm3'   : 2,
                     'KKCAlarm4'   : 3,
                     'InputPath'   : 4,
                     'OutputPath'  : 5,
                     'OIPVib'      : 6,
                     'rez_7'       : 7,
                     'KTPRA_NP'    : 8,
                     'KTPR_ACHR'   : 9,
                     'KTPR_SAON'   : 10,
                     'StopWork'    : 11,
                     'StartWork'   : 12,
                     'SAR_Ramp'    : 13,
                     'needRez'     : 14,
                     'needOverhaul': 15
                     }
        state_na4 = {'StopNoCmd_1'        : 0,
                     'StopNoCmd_2'        : 1,
                     'StartNoCmd'         : 2,
                     'StateAlarm'         : 3,
                     'StateAlarm_ChRP'    : 4,
                     'StateAlarm_All'     : 5,
                     'ChRPRegError'       : 6,
                     'LogicalChRPCrash'   : 7,
                     'ZD_Unprompted_Close': 8,
                     'StopErr'            : 9,
                     'StopErr2'           : 10,
                     'StopErr_All'        : 11,
                     'StartErr'           : 12,
                     'StartErr2'          : 13,
                     'StartErr3'          : 14,
                     'StartErr_All'       : 15,
                     }
        state_na5 = {'ED_IsMT10Perc'         : 0,
                     'ED_IsNdv2ndParam'       : 1,
                     'ED_IsHighVibStat'       : 2,
                     'ED_IsHighVibNoStat'     : 3,
                     'ED_IsAvarVibStat'       : 4,
                     'ED_IsAvarVibNoStat'     : 5,
                     'ED_IsAvar2Vib'          : 6,
                     'Pump_IsMT10Perc'        : 7,
                     'Pump_IsNdv2ndParam'     : 8,
                     'Pump_IsHighVibStat'     : 9,
                     'Pump_IsHighVibStatNMNWR': 10,
                     'Pump_IsHighVibNoStat'   : 11,
                     'Pump_IsAvarVibStat'     : 12,
                     'Pump_IsAvarVibStatNMNWR': 13,
                     'Pump_IsAvarVibNoStat'   : 14,
                     'Pump_IsAvar2Vib'        : 15
                     }
        state_na6 = {'GMPNA_P_2_64': 0,
                     'rez_1': 1,
                     'rez_2': 2,
                     'rez_3': 3,
                     'rez_4': 4,
                     'rez_5': 5,
                     'rez_6': 6,
                     'rez_7': 7,
                     'rez_8': 8,
                     'rez_9': 9,
                     'rez_10': 10,
                     'rez_11': 11,
                     'RptVKL': 12,
                     'UseCT': 13,
                     'ZDin_Unprompted_Close': 14,
                     'ZDout_Unprompted_Close': 15}
        state_zd1  = {'State_1_Opening': 0,
                     'State_2_Opened' : 1,
                     'State_3_Middle' : 2,
                     'State_4_Closing': 3,
                     'State_5_Closed' : 4,
                     'Dist'           : 5,
                     'Imit'           : 6,
                     'NOT_EC'         : 7,
                     'Avar'           : 8,
                     'Diff'           : 9,
                     'WarnClose'      : 10,
                     'Blink'          : 11,
                     'KVO'            : 12,
                     'KVZ'            : 13,
                     'MPO'            : 14,
                     'MPZ'            : 15}
        state_zd2  = {'CorrCO'        : 0,
                     'CorrCZ'        : 1,
                     'VMMO'          : 2,
                     'VMMZ'          : 3,
                     'NOT_ZD_EC_KTP' : 4,
                     'Local'         : 5,
                     'Mufta'         : 6,
                     'Avar_BUR'      : 7,
                     'NeispravVU'    : 8,
                     'ErrMPO'        : 9,
                     'ErrMPZ'        : 10,
                     'EC'            : 11,
                     'RS_OK'         : 12,
                     'Close_Fail'    : 13,
                     'Open_Fail'     : 14,
                     'Stop_Fail'     : 15}
        state_zd3 = {'ECsign'          : 0,
                     'rez_1'           : 1,
                     'rez_2'           : 2,
                     'Unprompted_Open' : 3,
                     'Unprompted_Close': 4,
                     'Neisprav'        : 5,
                     'CorrCOCorrCZ'    : 6,
                     'rez_7'           : 7,
                     'Open'            : 8,
                     'Close'           : 9,
                     'Stop'            : 10,
                     'StopClose'       : 11,
                     'VMMO_save'       : 12,
                     'VMMZ_save'       : 13,
                     'Mufta_save'      : 14,
                     'Avar_BUR_save'   : 15}

        isNum = 0
        isInv = 0
        Inputvar = str(inp).split(".")
        try:
            if self.dop_function.str_find(Inputvar[0], {'NOT '}):
                isInv = 1
            Inpvr = str(Inputvar[0]).replace('NOT ', '')

            if len(Inputvar) > 2:
                if self.dop_function.str_find(Inpvr, {'stateSAR'}):
                    pInputnum = str(inp).split('.state.')[1]
                    isNum = 0
                    pInputpInputVar = (str(inp).split('.state.')[0].replace('NOT ', '')) + '.state.reg'
                    if self.dop_function.str_find(pInputpInputVar,{'stateQ'}):
                        pInputpInputVar = (str(inp).split('.state.')[0].replace('NOT ', '')) + '.reg'
                if self.dop_function.str_find(Inpvr, {'stateBUF'}):
                    pInputnum = Inputvar[2]
                    isNum = 0
                    pInputpInputVar = Inputvar[0] + '.state.reg'
                if self.dop_function.str_find(Inpvr, {'mRS'}):
                    pInputnum = Inputvar[2]
                    isNum = 0
                    #pInputpInputVar = Inputvar[1]
                    text = str(inp).replace('NOT ', '').split('.')
                    pInputpInputVar = f'{text[0]}.{text[1]}.reg'
                if self.dop_function.str_find(Inpvr, {'stateNA'}):
                    pInputnum = Inputvar[2]
                    isNum = 0
                    pInputpInputVar = Inpvr + '.' + Inputvar[1] + '.reg'
            elif len(Inputvar) > 1:
                if (Inputvar[1] in stateAI.keys()) and (self.dop_function.str_find(Inpvr, {'AI'})):
                    pInputnum = stateAI[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('AI', 'StateAI') + '.state.reg'
                    cfg = '00000000000000' + str(isNum) + str(isInv)
                    return pInputpInputVar, pInputnum, str(hex(int(cfg, 2))).replace("0x", "16#")
                if (Inputvar[1] in stateAIzone.keys()) and (self.dop_function.str_find(Inpvr, {'AI'})):
                    pInputnum = stateAIzone[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('AI', 'StateAI') + '.stateZone.reg'
                if (Inputvar[1] in vsgrpstate.keys()) and (self.dop_function.str_find(Inpvr, {'VSGRP'})):
                    pInputnum = vsgrpstate[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('VSGRP', 'stateVSGRP') + '.state.reg'
                if (Inputvar[1] in stateFacility.keys()) and (self.dop_function.str_find(Inpvr, {'Facility'})):
                    pInputnum = stateFacility[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('Facility', 'stateFacility') + '.longGas.reg'
                if (Inputvar[1] in warnFacility.keys()) and (self.dop_function.str_find(Inpvr, {'Facility'})):
                    pInputnum = warnFacility[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('Facility', 'stateFacility') + '.warnGas.reg'
                if (Inputvar[1] in Facility.keys()) and (self.dop_function.str_find(Inpvr, {'Facility'})):
                    pInputnum = Facility[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('Facility', 'stateFacility') + '.state.reg'
                if (Inputvar[1] in stateDI.keys()) and (self.dop_function.str_find(Inpvr, {'DI'})):
                    pInputnum = stateDI[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('DI', 'StateDI') + '.state.reg'
                if (Inputvar[1] in stateNPS.keys()) and (self.dop_function.str_find(Inpvr, {'NPS'})):
                    pInputnum = stateNPS[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NPS', 'stateNPS') + '.state.reg'
                if (Inputvar[1] in state_na5.keys()) and (self.dop_function.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na5[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state5.reg'
                if (Inputvar[1] in state_na.keys()) and (self.dop_function.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state1.reg'
                if (Inputvar[1] in state_na3.keys()) and (self.dop_function.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na3[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state3.reg'
                if (Inputvar[1] in state_na6.keys()) and (self.dop_function.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na6[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state6.reg'
                if (Inputvar[1] in state_na4.keys()) and (self.dop_function.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na4[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state4.reg'
                if (Inputvar[1] in state_na2.keys()) and (self.dop_function.str_find(Inpvr, {'NA'})):
                    pInputnum = state_na2[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('NA', 'stateNA') + '.state2.reg'
                if (Inputvar[1] in statektpr.keys()) and (self.dop_function.str_find(Inpvr, {'KTPR'})):
                    pInputnum = statektpr[Inputvar[1]]
                    isNum = 0
                    isInv = 0
                    pInputpInputVar = str(Inpvr).replace('KTPR', 'stateKTPR') + '.state.reg'
                if (Inputvar[1] in state_zd1.keys()) and (self.dop_function.str_find(Inpvr, {'ZD'})):
                    pInputnum = state_zd1[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('ZD', 'stateZD') + '.state1.reg'
                if (Inputvar[1] in state_zd2.keys()) and (self.dop_function.str_find(Inpvr, {'ZD'})):
                    pInputnum = state_zd2[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('ZD', 'stateZD') + '.state2.reg'
                if (Inputvar[1] in state_zd3.keys()) and (self.dop_function.str_find(Inpvr, {'ZD'})):
                    pInputnum = state_zd3[Inputvar[1]]
                    isNum = 0
                    pInputpInputVar = str(Inpvr).replace('ZD', 'stateZD') + '.state3.reg'

            if Inputvar[0][:2] == 'AI':
                if Inputvar[1] == 'Value':
                    pInputpInputVar = Inputvar
                    isNum = 1
            cfg = '00000000000000' + str(isNum) + str(isInv)
            return pInputpInputVar, pInputnum, str(hex(int(cfg, 2))).replace("0x", "16#")
        except:
            return 0, 0, 0

    def cfg_ktprs(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('ktprs', f'"id", "name", "drawdown", "prohibition_issuing_msg"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_KTPRS')

            for value in data_value:
                numbers  = value[0]
                name     = value[1]
                drawdown = value[2]
                noMsg    = value[3] 
                
                noMsg = '1' if noMsg is None else '0'
                if numbers is None: continue

                pInputvar, num, cfg = self.ret_inp_cfg(drawdown)
                cfg_txt = (f'(*{numbers} - {name}*)\n')
                if pInputvar != 0:
                    cfg_txt = cfg_txt + f'cfgKTPRS[{numbers}].pInputVar.pInputVar       REF={str(pInputvar)};\n' \
                                        f'cfgKTPRS[{numbers}].pInputVar.num               :={str(num)};\n' \
                                        f'cfgKTPRS[{numbers}].pInputVar.cfg.reg           :={str(cfg)};\n'
                cfg = '000000000000000' + str(noMsg)
                cfg_txt = cfg_txt + f"cfgKTPRS[{numbers}].cfg.reg                     :={str(hex(int(cfg, 2))).replace('0x', '16#')};\n"
                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_KTPRS заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_KTPRS: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_vv(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('vv', f'"id", "name", "VV_vkl", "VV_otkl"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_VV')

            for value in data_value:
                numbers = value[0]
                name    = value[1]
                VV_On   = value[2]
                VV_Off  = value[3]
                if numbers is None: continue

                pVVOn, pVVOnnum, pVVOncfg    = self.ret_inp_cfg(VV_On)
                pVVOff, pVVOffnum, pVVOffcfg = self.ret_inp_cfg(VV_Off)

                cfg_txt = (f'(*{numbers} - {name}*)\n')
                if pVVOn != 0:
                    cfg_txt = cfg_txt + f'cfgVV[{numbers}].pBBB.pInputVar       REF={str(pVVOn)};\n' \
                                        f'cfgVV[{numbers}].pBBB.num               :={str(pVVOnnum)};\n' \
                                        f'cfgVV[{numbers}].pBBB.cfg.reg           :={str(pVVOncfg)};\n'
                if pVVOff != 0:
                    cfg_txt = cfg_txt + f'cfgVV[{numbers}].pBBO.pInputVar       REF={str(pVVOff)};\n' \
                                        f'cfgVV[{numbers}].pBBO.num               :={str(pVVOffnum)};\n' \
                                        f'cfgVV[{numbers}].pBBO.cfg.reg           :={str(pVVOffcfg)};\n'
                write_file.write(cfg_txt)
            write_file.close()

            msg[f'{today} - Файл СУ: cfg_vv заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_vv: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_uts(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('uts', f'"id", "name", "VKL", "siren", "Does_not_require_autoshutdown", "Examination", "Kvit"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_UTS')

            for value in data_value:
                numbers      = value[0]
                name         = value[1]
                pVkl         = value[2]
                isSiren      = value[3]
                blockAutoOff = value[4]

                if numbers is None: continue
                if pVkl == '': continue

                pCheckInput, cinum, cicfg = self.ret_inp_cfg(value[5])
                pKvitInput, kvnum, kvcfg  = self.ret_inp_cfg(value[6])

                cfg_txt = (f'(*{numbers} - {name}*)\n')
                if pCheckInput != 0:
                    cfg_txt = cfg_txt + f'cfgUTS[{numbers}].pCheckInput.pInputVar       REF={str(pCheckInput)};\n' \
                                        f'cfgUTS[{numbers}].pCheckInput.num               :={str(cinum)};\n' \
                                        f'cfgUTS[{numbers}].pCheckInput.cfg.reg           :={str(cicfg)};\n'
                if pKvitInput != 0:
                    cfg_txt = cfg_txt + f'cfgUTS[{numbers}].pKvitInput.pInputVar        REF={str(pKvitInput)};\n' \
                                        f'cfgUTS[{numbers}].pKvitInput.num                :={str(kvnum)};\n' \
                                        f'cfgUTS[{numbers}].pKvitInput.cfg.reg            :={str(kvcfg)};\n'
                cfg_txt = cfg_txt + f'cfgUTS[{numbers}].pVkl                REF={str(pVkl)};\n'
                cfg_txt = cfg_txt + f'cfgUTS[{numbers}].isSiren               :={str(isSiren)};\n'
                cfg_txt = cfg_txt + f'cfgUTS[{numbers}].blockAutoOff          :={str(blockAutoOff)};\n'
                write_file.write(cfg_txt)
            write_file.close()

            msg[f'{today} - Файл СУ: cfg_uts заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_uts: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_vsgrp(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('vsgrp', f'"id", "name", "fire_or_watering", "count_auxsys_in_group", "WarnOff_flag_if_one_auxsystem_in_the_group_is_running"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_VSGRP')

            for value in data_value:
                numbers = value[0]
                name    = value[1]
                isPoz   = value[2]
                countVS = value[3]
                WarnOff = value[4]

                if name == '': continue

                isPoz   = '1' if isPoz is not None else '0'
                WarnOff = '1' if WarnOff is not None else '0'
                
                cfg = str(hex(int("000000" + str(WarnOff) + str(isPoz), 2))).replace('0x', '16#')
                cfg_txt = f'(*{name}*)\n' \
                          f'cfgVSGRP[{numbers}].countVS     :=  {countVS};\n' \
                          f'cfgVSGRP[{numbers}].cfg.reg     :=  {cfg};\n'
                
                write_file.write(cfg_txt)
            write_file.close()

            msg[f'{today} - Файл СУ: cfg_vsgrp заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_vsgrp: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_nps(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('nps', f'"id", "variable", "name", "value"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_NPS')

            for value in data_value:
                s_name = {1: 'pInputVar REF=', 2: 'num:=', 3: 'cfg.reg:='}

                numbers = value[0]
                perem   = value[1]
                name    = value[2]
                pValue  = value[3]

                if numbers is None:
                    if name == '':
                        cfg_txt = (f'(*{name}*)\n')
                        write_file.write(cfg_txt)

                if pValue == '' or pValue is None: continue

                cfg_txt = (f'(*{name}*)\n')
                if str(pValue).isdigit()                          : cfg_txt = cfg_txt + f'{perem}:={pValue};\n'
                elif self.dop_function.str_find(pValue, 'AIValue'): cfg_txt = cfg_txt + f'{perem}:={pValue};\n'
                else:
                    a = {}
                    pInput, pnum, pcfg = self.ret_inp_cfg(pValue)
                    a[1] = pInput
                    a[2] = pnum
                    a[3] = pcfg
                    for el in range(1, 4):
                        cfg_txt = cfg_txt + f'{perem}.{s_name[el]}{str(a[el])};\n'
                
                write_file.write(cfg_txt)
            write_file.close()

            msg[f'{today} - Файл СУ: cfg_nps заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_nps: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_rsreq(self):
        msg = {}
        srsreq = ['SlaveId', 'ModbusFunction', 'Address', 'Count', 'ResultOffset',
                  'SingleRequest', 'OnModifyRequest', 'RepeatOverScan', 'SkipRepeatsWhenBad', 'Enable']
        try:
            data_value = self.dop_function.connect_by_sql('rsreq', 
                                                          f'''"id", "name", "Route", "Requests", "Port", "SlaveId", "ModbusFunction", "Address", "Count", "ResultOffset",
                                                              "SingleRequest", "OnModifyRequest", "RepeatOverScan", "SkipRepeatsWhenBad", "Enable"''')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_RSREQ')

            for value in data_value:
                numbers  = value[0]
                name     = value[1]
                route    = value[2]
                requests = value[3]
                port     = value[4]

                if name is None: continue

                cfg_txt = f'(*{numbers} - {name}*)\n'
                count_column_rep = 4
                for column in srsreq:
                    count_column_rep += 1
                    if column != 'Enable': cfg_txt = cfg_txt + f'{route}.Port[{port}].Requests[{requests}].{column}:={value[count_column_rep]};\n'
                    else                 : cfg_txt = cfg_txt + f'{route}.Port[{port}].Commands[{requests}].{column}:={value[count_column_rep]};\n'          
                write_file.write(cfg_txt)
            write_file.close()

            msg[f'{today} - Файл СУ: cfg_rsreq заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_rsreq: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_na(self):
        msg = {}
        s_name = {1: 'pInputVar REF=', 2: 'num:=', 3: 'cfg.reg:='}
        s_umpna = {'pBBB1': 'ВВ Включен',
                   'pBBO1': 'ВВ Отключен',
                   'pBBB2': 'ВВ Включен дубль',
                   'pBBO2': 'ВВ Отключен дубль',
                   'pTok': 'Сила тока >  уставки холостого хода',
                   'pECx02': 'Исправность цепей включения ВВ',
                   'pECx03': 'Исправность цепей отключения ВВ',
                   'pECx03_1': 'Исправность цепей отключения ВВ дубль',
                   'pKKC[1]': 'Стоп 1',
                   'pKKC[2]': 'Стоп 2',
                   'pKKC[3]': 'Стоп 3',
                   'pKKC[4]': 'Стоп 4',
                   'pEC1': 'Сигнал «Контроль наличия напряжения в цепях оперативного тока»',
                   'pMotorCellVoltage': 'Флаг наличия напряжения в двигательной ячейке ЗРУ',
                   'pECx04': 'Тележка ВВ выкачена',
                   'pDCx01': 'Дистанционный режим управления контроллера РЗиА',
                   'pRZiALink': 'Наличие связи с контроллером РЗиА',
                   'pExcitReady': 'Состояние возбудителя ЭД',
                   'pOPCx22': 'Флаг окончания предпусковой продувки двигателя',
                   'pSafetyPressureMotor': 'Флаг наличия безопасного давления подпора воздуха в корпусе двигателя',
                   'pSafetyPressureExcit': 'Флаг наличия безопасного давления подпора воздуха в корпусе возбудителя',
                   'pBlowValveClosed': 'Флаг закрытого положения клапана продувки двигателя',
                   'pOIPOilTemperatureFreezer': 'Флаг температуры масла маслосистемы выше 10гр.С на выходе охладителя (для индивидуальной маслосистемы)',
                   'pOIPOilValueMin2': 'Флаг предельного минимального уровня масла в маслобаке (для индивидуальной маслосистемы)',
                   'pOipClosingFluidMinLevel': 'Флаг наличия минимального уровня запирающей жидкости в баке системы запирания',
                   'pOipClosingFluidPressure': 'Обобщенный флаг наличия давления запирающей жидкости к торцевому уплотнению',
                   'pGMPNA[49]': 'GMPNA_[49]',
                   'pGMPNA[50]': 'GMPNA_[50]',
                   'pGMPNA[51]': 'GMPNA_[51]',
                   'pGMPNA[52]': 'GMPNA_[52]',
                   'pGMPNA[53]': 'GMPNA_[53]',
                   'pGMPNA[54]': 'GMPNA_[54]',
                   'pGMPNA[55]': 'GMPNA_[55]',
                   'pGMPNA[56]': 'GMPNA_[56]',
                   'pGMPNA[57]': 'GMPNA_[57]',
                   'pGMPNA[58]': 'GMPNA_[58]',
                   'pGMPNA[59]': 'GMPNA_[59]',
                   'pGMPNA[60]': 'GMPNA_[60]',
                   'pGMPNA[61]': 'GMPNA_[61]',
                   'pGMPNA[62]': 'GMPNA_[62]',
                   'pGMPNA[63]': 'GMPNA_[63]',
                   'pGMPNA[64]': 'GMPNA_[64]',
                    }
        ds_umpna = {'iDelay': 'Количество сканов задержки анализа исправности цепей управления ВВ НА',
                    'nVSprMN': 'Номер агрегата вспомсистемы "пуско-резервный маслонасос" (для индивидуальной маслосистемы)',
                    'nNPS': 'Номер НПС (1 или 2), к которой относится НА',
                    'nProtACHR': 'Номер защиты АЧР в массиве станционных защит',
                    'nProtSAON': 'Номер защиты САОН в массиве станционных защит',
                    'iCounterNdv': 'Параметр для KTPRAS_1',
                    }
        out_umpna = {'pStartWork': 'Команда на включение ВВ (только для UMPNA)',
                     'pStopWork1': 'Команда на отключение ВВ (выход 1)',
                     'pStopWork2': 'Команда на отключение ВВ (выход 2)'}
        try:
            data_value = self.dop_function.connect_by_sql('umpna', 
                                                          f'''"variable", "name", "NA_Chrp", "type_NA_MNA", "pump_type_NM", 
                                                          "vv_included", "vv_disabled", "vv_double_included", "vv_double_disabled", 
                                                          "current_greater_than_noload_setting", "serviceability_of_circuits_of_inclusion_of_VV", 
                                                          "serviceability_of_circuits_of_shutdown_of_VV", "serviceability_of_circuits_of_shutdown_of_VV_double", 
                                                          "stop_1", "stop_2", "stop_3", "stop_4", "monitoring_the_presence_of_voltage_in_the_control_current", 
                                                          "voltage_presence_flag_in_the_ZRU_motor_cell", "vv_trolley_rolled_out", "remote_control_mode_of_the_RZiA_controller", 
                                                          "availability_of_communication_with_the_RZiA_controller", "the_state_of_the_causative_agent_of_ED", 
                                                          "engine_prepurge_end_flag", "flag_for_the_presence_of_safe_air_boost_pressure_in_the_en", 
                                                          "flag_for_the_presence_of_safe_air_boost_pressure_in_the_ex", "engine_purge_valve_closed_flag", 
                                                          "oil_system_oil_temperature_flag_is_above_10_at_the_cooler_ou", "flag_for_the_minimum_oil_level_in_the_oil_tank_for_an_indiv",
                                                          "flag_for_the_presence_of_the_minimum_level_of_the_barrier", "generalized_flag_for_the_presence_of_barrier_fluid_pressure", 
                                                          "gmpna_49", "gmpna_50", "gmpna_51", "gmpna_52", "gmpna_53", "gmpna_54", "gmpna_55", "gmpna_56", "gmpna_57", "gmpna_58", 
                                                          "gmpna_59", "gmpna_60", "gmpna_61", "gmpna_62", "gmpna_63", "gmpna_64", 
                                                          
                                                          "number_of_delay_scans_of_the_analysis_of_the_health_of_the", "unit_number_of_the_auxiliary_system_start_up_oil_pump", 
                                                          "NPS_number_1_or_2_which_the_AT_belongs", "achr_protection_number_in_the_array_of_station_protections", 
                                                          "saon_protection_number_in_the_array_of_station_protections", "parametr_KTPRAS_1", 

                                                          "command_to_turn_on_the_vv_only_for_UMPNA", "command_to_turn_off_the_vv_output_1", "command_to_turn_off_the_vv_output_2"''')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_NA')

            for value in data_value:
                tag       = value[0]
                name      = value[1]
                IsCMNA    = value[2] 
                MPNA_Type = value[3]
                IsNM      = value[4]
                if name is None: continue

                cfg_txt   = (f'(*{tag} - {name}*)\n')

                count_s_umpna = 4
                for perem in s_umpna:
                    count_s_umpna += 1
                    if value[count_s_umpna] is None: continue
                    a = {}
                    pInput, pnum, pcfg = self.ret_inp_cfg(value[count_s_umpna])
                    a[1] = pInput
                    a[2] = pnum
                    a[3] = pcfg
                    for el in range(1, 4):
                        cfg_txt = cfg_txt + f'cfg{tag}.{perem}.{s_name[el]}{str(a[el])};\n'
                
                count_ds_umpna = 46
                for perem in ds_umpna:
                    count_ds_umpna += 1
                    if value[count_ds_umpna] is None: continue
                    cfg_txt = cfg_txt + f'cfg{tag}.{perem}:={str(value[count_ds_umpna])};\n'
                cfg_txt = cfg_txt + f"cfg{tag}.cfg.reg:={str(hex(int('0000000000000' + str(IsNM) + str(MPNA_Type) + str(IsCMNA), 2))).replace('0x', '16#')};\n"
                
                count_out_umpna = 52
                for perem in out_umpna:
                    count_out_umpna += 1
                    if value[count_out_umpna] is None: continue
                    cfg_txt = cfg_txt + f'cfg{tag}.{perem} REF={str(value[count_out_umpna])};\n'  

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_na заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_na: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_ktpra(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('ktpra',  f'"variable", "tag", "name", "avar_parameter", "stop_type", "AVR", "close_valves", "DisableMasking"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_KTPRA')

            for value in data_value:
                KTPRA   = value[0]
                tag     = value[1]
                name    = value[2]
                a_param = value[3]

                if name is None: continue
                if tag is None : continue

                stype       = value[4] if value[4] is not None else '0'
                avr         = value[5] if value[5] is not None else '0'
                CloseValves = value[6] if value[6] is not None else '0'
                NotMasked   = value[7] if value[7] is not None else '0'

                ktpra_cfg   = str(hex(int(str(CloseValves) + str(avr) + str(NotMasked), 2))).replace('0x', '16#')

                pInput, pnum, pcfg = self.ret_inp_cfg(a_param)
                cfg_txt = (f'(*{tag} - {name}*)\n')
                if pInput != 0:
                    cfg_txt = cfg_txt + f'cfg{KTPRA}.pInput.pInputVar       REF={str(pInput)};\n' \
                                        f'cfg{KTPRA}.pInput.num               :={str(pnum)};\n' \
                                        f'cfg{KTPRA}.pInput.cfg.reg           :={str(pcfg)};\n' \
                                        f'cfg{KTPRA}.StopType                 :={str(stype)};\n' \
                                        f'cfg{KTPRA}.cfg.reg                  :={str(ktpra_cfg)};\n'
                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_ktpra заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_ktpra: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_vs(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('vs',  f'''"id", "tag", "name", "group", "number_in_group", "VKL", "OTKL", "Pressure_is_True", "Not_APV",  "MP", 
                                                          "Voltage", "Voltage_Sch", "Serviceability_of_circuits_of_inclusion", "External_alarm", "Pressure_sensor_defective"''')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_VS')

            for value in data_value:
                numbers  = value[0]
                name     = value[2]
                pVkl     = value[5]
                pOtkl    = value[6]

                grpNum   = value[3] if value[3] is not None else '0'
                numInGrp = value[4] if value[4] is not None else '0'
                PC_USE   = '1' if value[7] is not None else '0'
                noAPV    = value[8] if value[8] is not None else '0'

                if numbers is None: continue

                pMPCpInputVar, pMPCnum, pMPCcfg                         = self.ret_inp_cfg(value[9])
                pPCpInputVar, pPCnum, pPCcfg                            = self.ret_inp_cfg(value[7])
                pECpInputVar, pECnum, pECcfg                            = self.ret_inp_cfg(value[10])
                pSEC_ECpInputVar, pSEC_ECnum, pSEC_ECcfg                = self.ret_inp_cfg(value[11])
                pOPCpInputVar, pOPCnum, pOPCcfg                         = self.ret_inp_cfg(value[12])
                pDiAVARpInputVar, pDiAVARnum, pDiAVARcfg                = self.ret_inp_cfg(value[13])
                pPC_NEISPRAVpInputVar, pPC_NEISPRAVnum, pPC_NEISPRAVcfg = self.ret_inp_cfg(value[14])

                cfg_txt = (f'(*{numbers} - {name}*)\n')
                if pMPCpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pMPC.pInputVar               REF={str(pMPCpInputVar)};\n' \
                                        f'cfgVS[{numbers}].pMPC.num                       :={str(pMPCnum)};\n' \
                                        f'cfgVS[{numbers}].pMPC.cfg.reg                   :={str(pMPCcfg)};\n'
                if pPCpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pPC.pInputVar                REF={str(pPCpInputVar)};\n' \
                                        f'cfgVS[{numbers}].pPC.num                        :={str(pPCnum)};\n' \
                                        f'cfgVS[{numbers}].pPC.cfg.reg                    :={str(pPCcfg)};\n'
                if pECpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pEC.pInputVar                REF={str(pECpInputVar)};\n' \
                                        f'cfgVS[{numbers}].pEC.num                        :={str(pECnum)};\n' \
                                        f'cfgVS[{numbers}].pEC.cfg.reg                    :={str(pECcfg)};\n'
                if pSEC_ECpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pSEC_EC.pInputVar            REF={pSEC_ECpInputVar};\n' \
                                        f'cfgVS[{numbers}].pSEC_EC.num                    :={pSEC_ECnum};\n' \
                                        f'cfgVS[{numbers}].pSEC_EC.cfg.reg                :={pSEC_ECcfg};\n'
                if pOPCpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pOPC.pInputVar               REF={pOPCpInputVar};\n' \
                                        f'cfgVS[{numbers}].pOPC.num                       :={pOPCnum};\n' \
                                        f'cfgVS[{numbers}].pOPC.cfg.reg                   :={pOPCcfg};\n'
                if pDiAVARpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pDiAVAR.pInputVar            REF={pDiAVARpInputVar};\n' \
                                        f'cfgVS[{numbers}].pDiAVAR.num                    :={pDiAVARnum};\n' \
                                        f'cfgVS[{numbers}].pDiAVAR.cfg.reg                :={pDiAVARcfg};\n'
                if pPC_NEISPRAVpInputVar != 0:
                    cfg_txt = cfg_txt + f'cfgVS[{numbers}].pPC_NEISPRAV.pInputVar       REF={pPC_NEISPRAVpInputVar};\n' \
                                        f'cfgVS[{numbers}].pPC_NEISPRAV.num               :={pPC_NEISPRAVnum};\n' \
                                        f'cfgVS[{numbers}].pPC_NEISPRAV.cfg.reg           :={pPC_NEISPRAVcfg};\n'
                    
                cfgVS_unioncfgVS = str(hex(int('00000000000000' + str(noAPV) + str(PC_USE), 2))).replace('0x', '16#')

                cfg_txt = cfg_txt + f'cfgVS[{numbers}].cfgVS.reg                      :={cfgVS_unioncfgVS};\n' \
                                    f'cfgVS[{numbers}].grpNum                         :={grpNum};\n' \
                                    f'cfgVS[{numbers}].numInGrp                       :={numInGrp};\n' \
                                    f'cfgVS[{numbers}].pVkl                         REF={pVkl};\n' \
                                    f'cfgVS[{numbers}].pOtkl                        REF={pOtkl};\n'
                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_vs заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_vs: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_zd(self):
        def filling_text(notebook, number, prefix, part, val_1, val_2, val_3, bool_prefix):
            if bool_prefix is True:
                notebook = notebook + '' \
                                    f"cfgZD[{number}].{prefix}.{part}.pInputVar REF={str(val_1)};\n" \
                                    f'cfgZD[{number}].{prefix}.{part}.num:={str(val_2)};\n' \
                                    f'cfgZD[{number}].{prefix}.{part}.cfg.reg:={str(val_3)};\n'
            else:
                notebook = notebook + '' \
                                    f'cfgZD[{numbers}].{part}.pInputVar REF={str(val_1)};\n' \
                                    f'cfgZD[{numbers}].{part}.num:={str(val_2)};\n' \
                                    f'cfgZD[{numbers}].{part}.cfg.reg:={str(val_3)};\n'
            return notebook
    
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('zd',  f'''"id", "name", "exists_interface", "Open", "Close", "Stop", "Opening_stop", "Closeing_stop", "Type_BUR_ZD", 
                                                          "Is_klapan", "KVO", "KVZ", "MPO", "MPZ", "Dist", "Mufta", "Drive_failure", "No_connection", "Close_BRU", "Stop_BRU", 
                                                          "Voltage", "Voltage_in_signaling_circuits", "Voltage_CHSU", "Serviceability_opening_circuits", 
                                                          "Serviceability_closening_circuits", "VMMO", "VMMZ", "KVO_i", "KVZ_i", "MPO_i", "MPZ_i", "Dist_i", "Mufta_i", 
                                                          "Drive_failure_i", "Open_i", "Close_i", "Stop_i", "Opening_stop_i", "Closeing_stop_i"''')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_ZD')

            for value in data_value:
                numbers        = value[0]
                name           = value[1]
                rs_ok          = value[2]
                IOpOpen        = value[3]
                IOpClose       = value[4]
                IOpStop        = value[5]
                IOpStopOpen    = value[6]
                IOpStopClose   = value[7]
                typeBURtypeBUR = value[8]

                isClp          = value[9] if value[9] is not None else '0'
                freeze         = '0'

                IOpKVOpInputVar, IOpKVOnum, IOpKVOcfg_unioncfg_st                = self.ret_inp_cfg(value[10])
                IOpKVZpInputVar, IOpKVZnum, IOpKVZcfg_unioncfg_st                = self.ret_inp_cfg(value[11])
                IOpMPOpInputVar, IOpMPOnum, IOpMPOcfg_unioncfg_st                = self.ret_inp_cfg(value[12])
                IOpMPZpInputVar, IOpMPZnum, IOpMPZcfg_unioncfg_st                = self.ret_inp_cfg(value[13])
                IOpDIST_KEYpInputVar, IOpDIST_KEYnum, IOpDIST_KEYcfg_unioncfg_st = self.ret_inp_cfg(value[14])
                IOpMuftapInputVar, IOpMuftanum, IOpMuftacfg_unioncfg_st          = self.ret_inp_cfg(value[15])
                IOpAvar_BURpInputVar, IOpAvar_BURnum, IOpAvar_BURcfg_unioncfg_st = self.ret_inp_cfg(value[16])
                pNoLinkpInputVar, pNoLinknum, pNoLinkcfg_unioncfg_st             = self.ret_inp_cfg(value[17])
                pBRUClosepInputVar, pBRUClosenum, pBRUClosecfg_unioncfg_st       = self.ret_inp_cfg(value[18])
                pBRUStoppInputVar, pBRUStopnum, pBRUStopcfg_unioncfg_st          = self.ret_inp_cfg(value[19])
                pECpInputVar, pECnum, pECcfg_unioncfg_st                         = self.ret_inp_cfg(value[20])
                pECsignpInputVar, pECsignnum, pECsigncfg_unioncfg_st             = self.ret_inp_cfg(value[21])
                pZD_EC_KTPpInputVar, pZD_EC_KTPnum, pZD_EC_KTPcfg_unioncfg_st    = self.ret_inp_cfg(value[22])
                pCorrCOpInputVar, pCorrCOnum, pCorrCOcfg_unioncfg_st             = self.ret_inp_cfg(value[23])
                pCorrCZpInputVar, pCorrCZnum, pCorrCZcfg_unioncfg_st             = self.ret_inp_cfg(value[24])
                pVMMOpInputVar, pVMMOnum, pVMMOcfg_unioncfg_st                   = self.ret_inp_cfg(value[25])
                pVMMZpInputVar, pVMMZnum, pVMMZcfg_unioncfg_st                   = self.ret_inp_cfg(value[26])

                if rs_ok is True:
                    RSpKVOpInputVar, RSpKVOnum, RSpKVOcfg_unioncfg_st                = self.ret_inp_cfg(value[27])
                    RSpKVZpInputVar, RSpKVZnum, RSpKVZcfg_unioncfg_st                = self.ret_inp_cfg(value[28])
                    RSpMPOpInputVar, RSpMPOnum, RSpMPOcfg_unioncfg_st                = self.ret_inp_cfg(value[29])
                    RSpMPZpInputVar, RSpMPZnum, RSpMPZcfg_unioncfg_st                = self.ret_inp_cfg(value[30])
                    RSpDIST_KEYpInputVar, RSpDIST_KEYnum, RSpDIST_KEYcfg_unioncfg_st = self.ret_inp_cfg(value[31])
                    RSpMuftapInputVar, RSpMuftanum, RSpMuftacfg_unioncfg_st          = self.ret_inp_cfg(value[32])
                    RSpAvar_BURpInputVar, RSpAvar_BURnum, RSpAvar_BURcfg_unioncfg_st = self.ret_inp_cfg(value[33])

                    RSpOpen      = value[34]
                    RSpClose     = value[35]
                    RSpStop      = value[36]
                    RSpStopOpen  = value[37]
                    RSpStopClose = value[38]

                cfg_txt = (f'(*{numbers} - {name}*)\n')
                if rs_ok is True:
                    if RSpKVOpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'RS', 'pKVO', RSpKVOpInputVar, RSpKVOnum, RSpKVOcfg_unioncfg_st, True)
                    if RSpKVZpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'RS', 'pKVZ', RSpKVZpInputVar, RSpKVZnum, RSpKVZcfg_unioncfg_st, True)
                    if RSpMPOpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'RS', 'pMPO', RSpMPOpInputVar, RSpMPOnum, RSpMPOcfg_unioncfg_st, True)
                    if RSpMPZpInputVar != 0:cfg_txt = filling_text(cfg_txt, numbers, 'RS', 'pMPZ', RSpMPZpInputVar, RSpMPZnum, RSpMPZcfg_unioncfg_st, True)
                    if RSpDIST_KEYpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'RS', 'pDIST_KEY', RSpDIST_KEYpInputVar, RSpDIST_KEYnum, RSpDIST_KEYcfg_unioncfg_st, True)
                    if RSpMuftapInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'RS', 'pMufta', RSpMuftapInputVar, RSpMuftanum, RSpMuftacfg_unioncfg_st, True)
                    if RSpAvar_BURpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'RS', 'pAvar_BUR', RSpAvar_BURpInputVar, RSpAvar_BURnum, RSpAvar_BURcfg_unioncfg_st, True)

                    if RSpOpen != 0            : cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pOpen REF={str(RSpOpen)};\n'
                    if RSpClose is not None    : cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pClose REF={str(RSpClose)};\n'
                    if RSpStop is not None     : cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pStop REF={str(RSpStop)};\n'
                    if RSpStopOpen is not None : cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pStopOpen REF={str(RSpStopOpen)};\n'
                    if RSpStopClose is not None: cfg_txt = cfg_txt + f'cfgZD[{numbers}].RS.pStopClose REF={str(RSpStopClose)};\n'

                if IOpKVOpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'IO', 'pKVO', IOpKVOpInputVar, IOpKVOnum, IOpKVOcfg_unioncfg_st, True)
                if IOpKVZpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'IO', 'pKVZ', IOpKVZpInputVar, IOpKVZnum, IOpKVZcfg_unioncfg_st, True)
                if IOpMPOpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'IO', 'pMPO', IOpMPOpInputVar, IOpMPOnum, IOpMPOcfg_unioncfg_st, True)
                if IOpMPZpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'IO', 'pMPZ', IOpMPZpInputVar, IOpMPZnum, IOpMPZcfg_unioncfg_st, True)
                if IOpDIST_KEYpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'IO', 'pDIST_KEY', IOpDIST_KEYpInputVar, IOpDIST_KEYnum, IOpDIST_KEYcfg_unioncfg_st, True)
                if IOpMuftapInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'IO', 'pMufta', IOpMuftapInputVar, IOpMuftanum, IOpMuftacfg_unioncfg_st, True)
                if IOpAvar_BURpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, 'IO', 'pAvar_BUR', IOpAvar_BURpInputVar, IOpAvar_BURnum, IOpAvar_BURcfg_unioncfg_st, True)

                if IOpOpen != 0            : cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pOpen REF={str(IOpOpen)};\n'
                if IOpClose is not None    : cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pClose REF={str(IOpClose)};\n'
                if IOpStop is not None     : cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pStop REF={str(IOpStop)};\n'
                if IOpStopOpen is not None : cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pStopOpen REF={str(IOpStopOpen)};\n'
                if IOpStopClose is not None: cfg_txt = cfg_txt + f'cfgZD[{numbers}].IO.pStopClose REF={str(IOpStopClose)};\n'

                if pNoLinkpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pNoLink', pNoLinkpInputVar, pNoLinknum, pNoLinkcfg_unioncfg_st, False)
                if pBRUClosepInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pBRUClose', pBRUClosepInputVar, pBRUClosenum, pBRUClosecfg_unioncfg_st, False)
                if pBRUStoppInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pBRUStop', pBRUStoppInputVar, pBRUStopnum, pBRUStopcfg_unioncfg_st, False)
                if pECpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pEC', pECpInputVar, pECnum, pECcfg_unioncfg_st, False)
                if pECsignpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pECsign', pECsignpInputVar, pECsignnum, pECsigncfg_unioncfg_st, False)
                if pZD_EC_KTPpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pZD_EC_KTP', pZD_EC_KTPpInputVar, pZD_EC_KTPnum, pZD_EC_KTPcfg_unioncfg_st, False)
                if pCorrCOpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pCorrCO', pCorrCOpInputVar, pCorrCOnum, pCorrCOcfg_unioncfg_st, False)
                if pCorrCZpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pCorrCZ', pCorrCZpInputVar, pCorrCZnum, pCorrCZcfg_unioncfg_st, False)
                if pVMMOpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pVMMO', pVMMOpInputVar, pVMMOnum, pVMMOcfg_unioncfg_st, False)
                if pVMMZpInputVar != 0: cfg_txt = filling_text(cfg_txt, numbers, '', 'pVMMZ', pVMMZpInputVar, pVMMZnum, pVMMZcfg_unioncfg_st, False)

                cfg = str(hex(int('00000000000000' + str(isClp) + str(freeze), 2))).replace('0x', '16#')
                
                cfg_txt = cfg_txt + f'cfgZD[{numbers}].cfg.reg:={str(cfg)};\n'

                if typeBURtypeBUR is not None:
                    cfg_txt = cfg_txt + f'cfgZD[{numbers}].typeBUR.reg:={str(typeBURtypeBUR)};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_zd заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_zd: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_do(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('do', f'"id", "tag", "name", "pValue", "pHealth"' )
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_DO')

            for value in data_value:
                numbers = value[0]
                tag     = value[1]
                name    = value[2]
                pValue  = value[3]
                pHealth = value[4]

                if pValue is None: continue
                if tag is None: continue

                cfg_txt = f'(*{tag} - {name}*)\n' \
                          f'cfgDO[{numbers}].pValue         REF={pValue};\n' \
                          f'cfgDO[{numbers}].pHealth        REF={pHealth};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_do заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_do: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_di(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('di', f'''"id", "name", "pValue", "pHealth", "pNC_AI", "TS_ID", "priority_0", "priority_1", 
                                                                    "isModuleNC", "Msg", "isAI_Avar", "isAI_Warn", "isDI_NC", "ErrValue", "Inv"''')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_DI')

            for value in data_value:
                s = '000000000'
                #['isModuleNC', 'Msg', 'isAI_Avar', 'isAI_Warn', 'isDI_NC', 'ErrValue', 'Inv']
                for count in range(7):
                    s = s + str(value[count + 8]) if value[count + 8] is not None else s + '0'

                numbers = value[0]
                name    = value[1]
                pValue  = value[2]
                pHealth = value[3]
                
                pNC_AI    = value[4] if value[4] is not None else '0'
                TS_ID     = value[5] if value[5] is not None else '0'
                priority0 = value[6] if value[6] is not None else '0'
                priority1 = value[7] if value[7] is not None else '0'

                cfg = str(hex(int(s, 2))).replace('0x', '16#')
                
                if pValue is not None:
                    cfg_txt = f'(*{numbers} - {name}*)\n' \
                              f'cfgDI[{numbers}].pValue             REF={pValue};\n' \
                              f'cfgDI[{numbers}].pHealth            REF={pHealth};\n' \
                              f'cfgDI[{numbers}].TS_ID                :={TS_ID};\n' \
                              f'cfgDI[{numbers}].priority[0]          :={priority0};\n' \
                              f'cfgDI[{numbers}].priority[1]          :={priority1};\n' \
                              f'cfgDI[{numbers}].cfg.reg              :={cfg};\n'
                    write_file.write(cfg_txt)

                if value[4] is not None:
                    cfg_txt = f'(*{numbers} - {name}*)\n' \
                              f'cfgDI[{numbers}].pNC_AI             REF={str(pNC_AI).replace("_union.state", ".reg")};\n' \
                              f'cfgDI[{numbers}].TS_ID                :={TS_ID};\n' \
                              f'cfgDI[{numbers}].priority[0]          :={priority0};\n' \
                              f'cfgDI[{numbers}].priority[1]          :={priority1};\n' \
                              f'cfgDI[{numbers}].cfg.reg              :={cfg};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_di заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_di: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_ao(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('ao', f'"id", "name", "pValue", "pHealth"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_AO')

            for value in data_value:
                numbers = value[0]
                name    = value[1]
                pValue  = value[2]
                pHealth = value[3]

                if pValue is None: continue
                if self.dop_function.str_find(str(name).lower(), {'резерв'}): continue

                cfg_txt = f'(*{numbers} - {name}*)\n' \
                          f'cfgAO[{numbers}].pValue         REF={pValue};\n' \
                          f'cfgAO[{numbers}].pHealth        REF={pHealth};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_ao заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_ao: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_ai(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('ai', f'''"id", "name", "pValue", "pHealth", "number_ust_min_avar", "number_ust_min_pred", "number_ust_max_pred", 
                                                                    "number_ust_max_avar", "IsPumpVibration", "vibration_motor", "current_motor", "number_NA_or_aux", "fuse"''')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_AI')

            for value in data_value:
                numbers = value[0]
                name    = value[1]
                pValue  = value[2]
                pHealth = value[3]
                minAvar = value[4]
                minWarn = value[5]
                maxWarn = value[6]
                maxAvar = value[7]

                if self.dop_function.str_find(str(name).lower(), {'резерв'}): continue

                isVibroPump  = value[8] if value[8] is not None else '0'
                isVibroED    = value[9] if value[9] is not None else '0'
                isCT         = value[10] if value[10] is not None else '0'
                isPressureVS = '1' if value[11] is not None else '0'
                nNA_VS       = value[11] if value[11] is not None else '0'
                nFuse        = value[12] if value[12] is not None else '0'

                cfgAI        = str(hex(int('000000000000'+str(isPressureVS)+str(isCT)+str(isVibroED)+str(isVibroPump),2))).replace('0x','16#')
                cfgWarnAvar = '11'
                cfgWarnAvar = 2**(int(minAvar)-1) + 2**(int(minWarn)+4-1)+2**(int(maxWarn)+8-1)+2**(int(maxAvar)+12-1)
                cfgWarnAvar = '16#'+ (str(maxAvar)+str(maxWarn)+str(minWarn)+str(minAvar))

                if pValue is None: continue
                cfg_txt = f'(* {numbers} - {name} *)\n' \
                          f'cfgAI[{numbers}].pValue                 REF={pValue};\n'
                if pHealth is not None:
                    cfg_txt = cfg_txt + \
                              f'cfgAI[{numbers}].pHealth                REF={pHealth};\n' \
                              f'cfgAI[{numbers}].pHealthExt             REF={pHealth};\n'
                cfg_txt = cfg_txt + \
                          f'cfgAI[{numbers}].cfgWarnAvar.reg          :={cfgWarnAvar};\n' \
                          f'cfgAI[{numbers}].cfgAI.reg                :={str(cfgAI)};\n' \
                          f'cfgAI[{numbers}].nNA_VS                   :={nNA_VS};\n' \
                          f'cfgAI[{numbers}].nFuse                    :={nFuse};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_ai заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_ai: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_dps(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('dps', f'''"id", "control", "deblock"''')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_DPS')

            for value in data_value:
                numbers  = value[0]
                pControl = value[1]
                pDeblock = value[2]

                if pControl is None: continue
                cfg_txt = f'(*{numbers} - ДПС*)\n' \
                          f'cfgDPS[{numbers}].pControl      REF={pControl};\n'
                write_file.write(cfg_txt)
                
                if pDeblock is None: continue
                cfg_txt = f'(*{numbers} ДПС*)\n' \
                          f'cfgDPS[{numbers}].pDeblock      REF={pDeblock};\n'
                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_dps заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_dps: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_ktpr(self):
        msg = {}
        stateAI = {'Warn'  :0,
                   'Avar'  :1,
                   'LTMin' :2,
                   'MTMax' :3,
                   'AlgNdv':4,
                   'Imit'  :5,
                   'ExtNdv':6,
                   'Ndv'   :7,
                   'Init'  :8}
        stateAIzone = {'rez_0'                      :0,
                       'Min6'                       :1,
                       'Min5'                       :2,
                       'Min4'                       :3,
                       'Min3_IsMT10Perc'            :4,
                       'Min2_IsNdv2ndParam'         :5,
                       'Min1_IsHighVibStat'         :6,
                       'Norm'                       :7,
                       'Max1_IsHighVibStatNMNWRBIT;':8,
                       'Max2_IsHighVibNoStat'       :9,
                       'Max3_IsAvarVibStat'         :10,
                       'Max4_IsAvarVibStatNMNWRBIT;':11,
                       'Max5_IsAvarVibNoStat'       :12,
                       'Max6_IsAvar2Vib'            :13,
                       'rez_14'                     :14,
                       'rez_15'                     :15}
        stateDI = {'Value'    :0,
                   'ElInput'  :1,
                   'O'        :2,
                   'KZ'       :3,
                   'NC'       :4,
                   'Imit'     :5,
                   'ExtNdv'   :6,
                   'Ndv'      :7,
                   'priority1':8,
                   'priority2':9,
                   'priority3':10,
                   'rez_11'   :11,
                   'rez_12'   :12,
                   'Front_0_1':13,
                   'Front_1_0':14,
                   'CfgErr'   :15}
        stateNPS = {'ModeNPSDst'       :0,
                    'MNSInWork'        :1,
                    'IsMNSOff'         :2,
                    'IsNPSModePsl'     :3,
                    'IsPressureReady'  :4,
                    'NeNomFeedInterval':5,
                    'OIPHighPressure'  :6,
                    'KTPR_P'           :7,
                    'KTPR_M'           :8,
                    'CSPAlinkOK'       :9,
                    'CSPAWorkDeny'     :10,
                    'TSstopped'        :11,
                    'rez_12'           :12,
                    'stopDisp'         :13,
                    'stopCSPA'         :14,
                    'stopARM'          :15}
        stateFacility = {'longGasPoint1':0,
                         'longGasPoint2':1,
                         'longGasPoint3':2,
                         'longGasPoint4':3,
                         'longGasPoint5':4,
                         'longGasPoint6':5,
                         'longGasPoint7':6,
                         'longGasPoint8':7,
                         'rez_8'        :8,
                         'rez_9'        :9,
                         'rez_10'       :10,
                         'rez_11'       :11,
                         'rez_12'       :12,
                         'rez_13'       :13,
                         'rez_14'       :14,
                         'rez_15'       :15}
        warnFacility = {'warnGasPoint1':0,
                        'warnGasPoint2':1,
                        'warnGasPoint3':2,
                        'warnGasPoint4':3,
                        'warnGasPoint5':4,
                        'warnGasPoint6':5,
                        'warnGasPoint7':6,
                        'warnGasPoint8':7,
                        'rez_8'        :8,
                        'rez_9'        :9,
                        'rez_10'       :10,
                        'rez_11'       :11,
                        'rez_12'       :12,
                        'rez_13'       :13,
                        'rez_14'       :14,
                        'rez_15'       :15}
        Facility = {'ndv2Gas'   :0,
                    'GasLim'    :1,
                    'GasAv'     :2,
                    'GasKeep'   :3,
                    'GasNdvWait':4,
                    'GasLimWait':5,
                    'GasNdvProt':6,
                    'GasAvProt' :7,
                    'ColdOn'    :8,
                    'HotOn'     :9,
                    'rez_10'    :10,
                    'rez_11'    :11,
                    'ColdOff'   :12,
                    'HotOff'    :13,
                    'rez_14'    :14,
                    'rez_15'    :15}
        vsgrpstate = {'REZ_EXIST'           :0,
                      'REM'                 :1,
                      'OTKL'                :2,
                      'OTKL_BY_CMD'         :3,
                      'VKL_AS_DOP'          :4,
                      'PUSK_OSN'            :5,
                      'rez_6'               :6,
                      'rez_7'               :7,
                      'rez_8'               :8,
                      'rez_9'               :9,
                      'rez_10'              :10,
                      'rez_11'              :11,
                      'rez_12'              :12,
                      'rez_13'              :13,
                      'LAST_OFF_BY_CMD_ARM ':14,
                      'ALL_OFF_WITH_BLOCK ' :15}

        ktpr_cfg = ['Отключение ПНС с выдержкой времени до 5 с после отключения всех МНА','Автоматическая деблокировка защиты','Запрет маскирования']
        ktpr_ctrl1 = ['Закрытие задвижек на входе РП',
                      'Закрытие секущей задвижки узла подключения объекта нефтедобычи/ нефтепереработки',
                      'Закрытие задвижек на входе ФГУ',
                      'Закрытие задвижек на входе ССВД',
                      'Закрытие задвижек на выходе узла РД',
                      'Закрытие задвижек на входе узла РД',
                      'Закрытие задвижек на входе и выходе ПНА',
                      'Закрытие задвижек на входе и выходе МНА',
                      'Закрытие задвижек на входе и выходе ПНС',
                      'Закрытие задвижек на входе и выходе МНС',
                      'Закрытие задвижек между РП и ПНС',
                      'Закрытие задвижек между ПНС и МНС',
                      'Закрытие задвижек на выходе НПС',
                      'Закрытие задвижек на входе НПС']
        ktpr_ctrl2 = ['Отключение вентиляторов водоохлаждения системы оборотного водоснабжения',
                      'Отключение АВО',
                      'Отключение насосов артскважин',
                      'Отключение насосов хозяйственно-питьевого водоснабжения',
                      'Отключение насосов прокачки нефти/нефтепродукта через БИК',
                      'Отключение насосов, обеспечивающих подкачку нефти/нефтепродукта от объектов нефтедобычи/нефтепереработки',
                      'Отключение компрессоров подпора воздуха ЭД',
                      'Отключение подпорных вентиляторов электрозала',
                      'Отключение подпорных вентиляторов ЭД',
                      'Отключение беспромвальных вентиляторов электрозала',
                      'Отключение насосов откачки из емкостей ССВД',
                      'Отключение насосов откачки из емкостей сбора утечек ПНС',
                      'Отключение насосов откачки из емкостей сбора утечек МНС',
                      'Отключение насосов оборотного водоснабжения',
                      'Отключение маслонасосов после сигнала "остановлен" НА',
                      'Отключение маслонасосов']
        ktpr_ctrl3 = ['Отключение приточного вентилятора помещения СИКН',
                      'Отключение приточного вентилятора помещения БИК',
                      'Отключение приточных вентиляторов помещения компрессорной подпора воздуха ЭД и закрытие огнезадерживающих клапанов',
                      'Отключение приточного вентилятора помещения ССВД',
                      'Отключение приточного вентилятора помещения РД',
                      'Отключение приточных вентиляторов в помещении централизованной маслосистемы и закрытие огнезадерживающих клапанов',
                      'Отключение приточных вентиляторов насосного зала ПНС и закрытие огнезадерживающих клапанов',
                      'Отключение приточных вентиляторов насосного зала МНС и закрытие огнезадерживающих клапанов',
                      'Отключение крышных вентиляторов насосного зала ПНС',
                      'Отключение крышных вентиляторов насосного зала МНС',
                      'Отключение вытяжных вентиляторов в помещении ССВД',
                      'Отключение вытяжных вентиляторов в помещении РД',
                      'Отключение вытяжных вентиляторов маслоприямка в электрозале',
                      'Отключение вытяжных вентиляторов в помещении централизованной маслосистемы',
                      'Отключение вытяжных вентиляторов насосного зала ПНС',
                      'Отключение вытяжных вентиляторов насосного зала МНС']
        ktpr_ctrl4 = ['Защита по пожару',
                      'Отключение антиконденсационных электронагревателей ЭД',
                      'Отключение насосов откачки из емкостей сбора утечек всех СИКН',
                      'Отключение насосов прокачки нефти/нефтепродукта через оперативный БИК',
                      'Отключение насосов системы запирания',
                      'Отключение внешнего контура охлаждения ЧРП ПНА',
                      'Отключение внешнего контура охлаждения ЧРП МНА',
                      'Отключение воздушных охладителей системы запирания торцовых уплотнений отключенных НА',
                      'Отключение воздушных охладителей системы запирания торцовых уплотнений всех МНА',
                      'Отключение электронагревателей емкости сбора утечек СИКН',
                      'Отключение электронагревателей емкости сбора утечек ПНС',
                      'Отключение электронагревателей емкости сбора утечек МНС',
                      'Отключение электронагревателей масла',
                      'Закрытие воздушных клапанов (жалюзийных решёток) помещения компрессорной подпора воздуха ЭД',
                      'Закрытие воздушных клапанов (жалюзийных решёток) насосного зала']
        try:
            data_value = self.dop_function.connect_by_sql('ktpr', f'''"id", "tag", "name", "avar_parameter", "bitmask_protection_group_membership", "stop_type_NA", "pump_station_stop_type",
                                                          
                                                          "shutdown_PNS_a_time_delay_up_5s_after_turning", "auto_unlock_protection", "DisableMasking", 
                                                          
                                                          "closing_valves_inlet_RP", "closing_secant_valve_connection_unit__oil_production_oil", 
                                                          "closing_valves_inlet_FGU", "closing_valves_inlet_SSVD", "closing_valves_outlet_RD", "closing_valves_inlet_RD", 
                                                          "closing_valves_inlet_and_outlet_PNA", "closing_valves_inlet_and_outlet_MNA", "closing_valves_inlet_and_outlet_PNS", 
                                                          "closing_valves_inlet_and_outlet_MNS", "closing_gate_valves_between_RP_and_PNS", "closing_gate_valves_between_PNS_and_MNS", 
                                                          "closing_gate_valves_at_the_outlet_NPS", "closing_gate_valves_at_the_inlet_NPS", 
                                                          
                                                          "shutdown_of_water_cooling_fans_circulating_water_supply_system", "AVO_shutdown", "shutdown_of_art_well_pumps", 
                                                          "shutdown_domestic_and_drinking_water_pumps", "disabling_pumps_for_pumping_oil_oil_products_through_BIC", 
                                                          "shutdown_pumps_providing_oil", "shutdown_of_ED_air_compressors", "shutdown_of_retaining_fans_of_the_electrical_room", 
                                                          "shutdown_of_booster_fans_ED", "switching_off_the_electric_room_fans", "shutdown_pumps_pumping_out_from_tanks_SSVD", 
                                                          "shutdown_pumps_pumping_out_from_tanks_collection_of_leaks_PNS", "shutdown_pumps_pumping_out_from_tanks_collection_of_leaks_MNS", 
                                                          "shutdown_circulating_water_pumps", "shutdown_oil_pumps_after_signal_stopped_NA", "shutdown_oil_pumps", 
                                                          
                                                          "switching_off_the_supply_fan_of_the_SIKN_room", "switching_off_the_supply_fan_of_the_BIK_room", 
                                                          "switching_off_the_supply_fans_of_the_ED_air_compressor", "switching_off_the_supply_fan_of_the_SSVD_room", 
                                                          "switching_off_the_supply_fan_of_the_RD_room", "switch_off_the_supply_fans_in_the_centralized_oil", 
                                                          "switching_off_the_supply_fans_pumping_room_of_the_PNS", "switching_off_the_supply_fans_pumping_room_of_the_MNS", 
                                                          "shutdown_of_the_roof_fans_of_the_PNS_pump_room", "shutdown_of_the_roof_fans_of_the_MNS_pump_room", 
                                                          "shutdown_of_exhaust_fans_in_the_SSVD_room", "shutdown_of_exhaust_fans_in_the_RD_room", 
                                                          "shutdown_of_exhaust_fans_oil_pit_in_the_electrical_room", "shutdown_of_exhaust_fans_in_the_centralized_oil_system_room", 
                                                          "shutdown_of_exhaust_fans_of_the_pumping_room_PNS", "shutdown_exhaust_fans_of_the_pumping_room_of_the_MNS", 
                                                          
                                                          "fire_protection", "shutdown_of_anticondensation_electric_heaters_ED", "shutdown_of_pumping_pumps_from_leakage_collection_tanks", 
                                                          "shutdown_of_pumps_for_pumping_oil_oil_products_through", "shutdown_of_locking_system_pumps", 
                                                          "shutdown_of_the_external_cooling_circuit_ChRP_PNA", "shutdown_of_the_external_cooling_circuit_ChRP_MNA", 
                                                          "shutdown_of_air_coolers_of_the_locking_system_disc_NA", "shutdown_of_air_coolers_of_the_locking_system_MNA", 
                                                          "shutdown_of_electric_heaters_of_the_SIKN_leak_collection_tank", "shutdown_of_the_electric_heaters_of_the_leakage_collection_PNS", 
                                                          "shutdown_of_the_electric_heaters_of_the_leakage_collection_MNS", "shutdown_of_electric_oil_heaters", 
                                                          "closing_of_air_valves_louvered_grilles_of_the_compressor_room", "closing_the_air_valves_louvered_grilles_of_the_pump_room"
                                                          ''')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('сfg_KTPR')

            for value in data_value:
                numbers         = value[0]
                pInputnum       = value[0]
                tag             = value[1]
                name            = value[2]
                pInputpInputVar = value[3]

                cfg_unioncfg = '0000000000000'
                ctrl1 = '00'
                ctrl2 = ''
                ctrl3 = ''
                ctrl4 = ''
                isNum = 0
                isInv = 0
                Inputvar = str(pInputpInputVar).split(".")

                if self.dop_function.str_find(Inputvar[0], {'NOT '}):
                    isInv = 1
                b = str(Inputvar[0]).replace('NOT ', '')
                if len(Inputvar)>2:
                    if self.dop_function.str_find(Inputvar[0],'stateBUF'):
                        pInputnum = Inputvar[2]
                        isNum = 0
                        pInputpInputVar = Inputvar[0] + '.state.reg'
                if len(Inputvar)>1:
                    if Inputvar[1] in vsgrpstate.keys():
                        pInputnum = vsgrpstate[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar = str(Inputvar[0]).replace('VSGRP','stateVSGRP') + '.state.reg'
                    if Inputvar[1] in stateFacility.keys():
                        pInputnum = stateFacility[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar = str(Inputvar[0]).replace('Facility','stateFacility') + '.longGas.reg'
                    if Inputvar[1] in warnFacility.keys():
                        pInputnum = warnFacility[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar = str(Inputvar[0]).replace('Facility','stateFacility') + '.warnGas.reg'
                    if Inputvar[1] in stateAI.keys():
                        pInputnum=stateAI[Inputvar[1]]
                        isNum=0
                        pInputpInputVar=str(Inputvar[0]).replace('AI','StateAI')+'.state.reg'
                    if Inputvar[1] in Facility.keys():
                        pInputnum=Facility[Inputvar[1]]
                        isNum=0
                        pInputpInputVar=str(Inputvar[0]).replace('Facility','stateFacility')+'.state.reg'
                    if Inputvar[1] in stateAIzone.keys():
                        pInputnum=stateAIzone[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar=str(Inputvar[0]).replace('AI','StateAI')+'.stateZone.reg'
                    if Inputvar[1] in stateDI.keys():
                        pInputnum=stateDI[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar=str(Inputvar[0]).replace('DI','StateDI')+'.state.reg'
                    if Inputvar[1] in stateNPS.keys():
                        pInputnum=stateNPS[Inputvar[1]]
                        isNum = 0
                        pInputpInputVar=str(b).replace('NPS','stateNPS')+'.state.reg'

                i=0
                for count in range(7, 10):
                    i += 1
                    cfg_unioncfg = cfg_unioncfg+str(value[count]) if value[count] is not None else cfg_unioncfg+'0'
                    if i == 2: cfg_unioncfg = cfg_unioncfg + '0'

                for count in range(10, 24): ctrl1 = ctrl1+str(value[count]) if value[count] is not None else ctrl1+'0'
                for count in range(24, 40): ctrl2 = ctrl2+str(value[count]) if value[count] is not None else ctrl2+'0'
                for count in range(40, 56): ctrl3 = ctrl3+str(value[count]) if value[count] is not None else ctrl3+'0'
                for count in range(56, 71): ctrl4 = ctrl4+str(value[count]) if value[count] is not None else ctrl4+'0'

                Group        = value[4] if value[4] is not None else '0'
                NA_StopType  = value[5] if value[5] is not None else '0'
                NS_StopType  = value[6] if value[6] is not None else '0'
                
                pInputcfg_unioncfg = '00000000000000'+str(isNum)+str(isInv)
                if tag is None: continue
                cfg_txt = f'(*{numbers} - {name}*)\n' \
                          f'cfgKTPR[{numbers}].pInput.pInputVar         REF={pInputpInputVar};\n' \
                          f'cfgKTPR[{numbers}].pInput.num                 :={pInputnum};\n' \
                          f"cfgKTPR[{numbers}].pInput.cfg.reg             :={str(hex(int(pInputcfg_unioncfg,2))).replace('0x','16#')};\n" \
                          f"cfgKTPR[{numbers}].cfg.reg                    :={str(hex(int(cfg_unioncfg,2))).replace('0x','16#')};\n" \
                          f'cfgKTPR[{numbers}].Group                      :={Group};\n' \
                          f'cfgKTPR[{numbers}].NA_StopType                :={NA_StopType};\n' \
                          f'cfgKTPR[{numbers}].NS_StopType                :={NS_StopType};\n' \
                          f"cfgKTPR[{numbers}].ctrl.ctrl1.reg             :={str(hex(int(ctrl1,2))).replace('0x','16#')};\n" \
                          f"cfgKTPR[{numbers}].ctrl.ctrl2.reg             :={str(hex(int(ctrl2, 2))).replace('0x','16#')};\n" \
                          f"cfgKTPR[{numbers}].ctrl.ctrl3.reg             :={str(hex(int(ctrl3, 2))).replace('0x','16#')};\n" \
                          f"cfgKTPR[{numbers}].ctrl.ctrl4.reg             :={str(hex(int(ctrl4, 2))).replace('0x','16#')};\n"

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_ktpr заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_ktpr: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_diag(self):
        msg = {}
        # type_mk500 = {'PSU' : 'typePSU',
        #               'CPU' : 'typeCPU',
        #               'CN'  : 'typeCN',
        #               'MN'  : 'typeMN',
        #               'AI'  : 'typeAI_8ch',
        #               'AO'  : 'typeAO',
        #               'DI'  : 'typeDI',
        #               'DO'  : 'typeDO',
        #               'RS'  : 'typeRS485'}
        
        list_type = ['CPU', 'PSU', 'CN', 'MN', 'AI', 'AO', 'DI', 'RS', 'DO', 'EthEx']
        
        #list_type = ['MK-504-120', 'MK-550-024', 'MK-545-010', 'MK-546-010', 'MK-516-008A', 'MK-514-008', 'MK-521-032', 'MK-541-002', 'MK-531-032', 'MK-544-040']

        try:
            data_ss = self.dop_function.connect_by_sql('ss', f'''"id", "name", "counter_1", "number_array_stateRSreq_1", "number_NOM_1", "channel_NOM_1", 
                                                                               "counter_2", "number_array_stateRSreq_2", "number_NOM_2", "channel_NOM_2", "link_timeout"''')
            data_tmdp   = self.dop_function.connect_by_sql('tm_dp', f'''"id", "variable", "name", "link_to_link_signal", "link_to_timeout"''')

            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file_gv = self.file_check('gv_DIAG')
            write_file_cfg = self.file_check('cfg_DIAG')

            signals      = []
            signals_ss   = []
            signals_tmdp = []
            countPSUmodules     = 0
            countCPUmodules     = 0
            countCNmodules      = 0
            countMNmodules      = 0
            countAI_16CHmodules = 0
            countAI_8CHmodules  = 0
            countAOmodules      = 0
            countDImodules      = 0
            countDOmodules      = 0
            countRS485modules   = 0

            with db:
                for basket in HardWare.select().order_by(HardWare.id).dicts():
                    id_        = basket['id']
                    tag        = basket['tag']
                    uso        = basket['uso']
                    num_basket = basket['basket']
                    for key, value in basket.items():
                        if self.dop_function.str_find(value, list_type):
                            if self.dop_function.str_find(value, {'PSU'}):
                                countPSUmodules += 1
                                abs_num = countPSUmodules
                                type_modul = 'typePSU'
                            if self.dop_function.str_find(value, {'CPU'}):
                                countCPUmodules += 1
                                abs_num = countCPUmodules
                                type_modul = 'typeCPU'
                            if self.dop_function.str_find(value, {'CN'}):
                                countCNmodules += 1
                                abs_num = countCNmodules
                                type_modul = 'typeCN'
                            if self.dop_function.str_find(value, {'MN'}):
                                countMNmodules += 1
                                abs_num = countMNmodules
                                type_modul = 'typeMN'
                            if self.dop_function.str_find(value, {'AI'}):
                                countAI_8CHmodules += 1
                                abs_num = countAI_8CHmodules
                                type_modul = 'typeAI_8ch'
                            if self.dop_function.str_find(value, {'AO'}):
                                countAOmodules += 1
                                abs_num = countAOmodules
                                type_modul = 'typeAO'
                            if self.dop_function.str_find(value, {'DI'}):
                                countDImodules += 1
                                abs_num = countDImodules
                                type_modul = 'typeDI'
                            if self.dop_function.str_find(value, {'DO'}):
                                countDOmodules += 1
                                abs_num = countDOmodules
                                type_modul = 'typeDO'
                            if self.dop_function.str_find(value, {'RS'}):
                                countRS485modules += 1
                                abs_num = countRS485modules
                                type_modul = 'typeRS485'

                            number_modul = str(key).split('_')[1]
                            if int(number_modul) < 10: string_name = f'{tag}_0{number_modul}'
                            else                     : string_name = f'{tag}_{number_modul}'

                            signals.append(dict(name_uso_rus   = uso,
                                                uso_rack_modul = string_name,
                                                rack           = id_,
                                                number_modul   = f'_{string_name[-2:]}',
                                                modul_dash     = number_modul,
                                                gvd            = value,
                                                abs_num        = abs_num,
                                                type_modul     = type_modul))
            # Смежные системы
            for value in data_ss:
                number_ss, name_ss, time_out  = value[0], value[1], value[10]
                count_ss, array1_req, num1_nom, chann1_nom = value[2], value[3], value[4], value[5]
                count_SS, array2_req, num2_nom, chann2_nom = value[6], value[7], value[8], value[9]

                if number_ss is None: continue
                if name_ss   is None: continue

                signals_ss.append(dict(number_ss  = number_ss,
                                       name_ss    = name_ss,
                                       count_ss   = count_ss,
                                       array1_req = array1_req,
                                       num1_nom   = num1_nom,
                                       chann1_nom = chann1_nom,
                                       count_SS   = count_SS,
                                       array2_req = array2_req,
                                       num2_nom   = num2_nom,
                                       chann2_nom = chann2_nom,
                                       time_out   = time_out))
            # Диспетчерские пункты
            for value in data_tmdp:
                number        = value[0]
                variable      = value[1]
                name          = value[2]
                connection    = value[3]
                link_time_out = value[4]

                if number is None: continue
                if name   is None: continue

                signals_tmdp.append(dict(number       = number,
                                         variable      = variable,
                                         name          = name,
                                         link_time_out = link_time_out,
                                         connection    = connection))

            write_file_gv.write('VAR_GLOBAL CONSTANT\n' \
                            f'\tcountPSUmodules\t: UINT := {countPSUmodules};\t(* Количество модулей PSU *)\n' \
                            f'\tcountCPUmodules\t: UINT := {countCPUmodules};\t(* Количество модулей CPU *)\n' \
                            f'\tcountCNmodules\t: UINT := {countCNmodules};\t(* Количество модулей CN *)\n' \
                            f'\tcountMNmodules\t: UINT := {countMNmodules};\t(* Количество модулей MN *)\n' \
                            f'\tcountAI16modules\t: UINT := {countAI_16CHmodules};\t(* Количество 16-ти канальных модулей AI *)\n' \
                            f'\tcountAI8modules\t: UINT := {countAI_8CHmodules};\t(* Количество 8-ми канальных модулей AI *)\n' \
                            f'\tcountAOmodules\t: UINT := {countAOmodules};\t(* Количество модулей AO *)\n' \
                            f'\tcountDImodules\t: UINT := {countDImodules};\t(* Количество модулей DI *)\n' \
                            f'\tcountDOmodules\t: UINT := {countDOmodules};\t(* Количество модулей DO *)\n' \
                            f'\tcountRS485modules\t: UINT := {countRS485modules};\t(* Количество интерфейсных модулей *)\n' \
                            f'\tcontRack:=\t: UINT := {countPSUmodules};\t(* Количество всех корзин *)\n' \
                            f'\tcontSS:=\t: UINT := {countRS485modules};\t(* Количество смежных систем - ПОКА НЕ ГЕНЕРИТСЯ *)\n' \
                            f'END_VAR\n' \
                            'VAR_GLOBAL\n')

            mbusexist = {}
            for i in range(1, 33):
                count = -1
                for value in signals:
                    if value['rack'] == str(i):
                        count = count + 1
                mbusexist[i] = count

            for value in signals:
                uso_rack_modul = value['uso_rack_modul']
                type_modul = value['type_modul']
                if type_modul in ['typeAI_8ch', 'typeAO', 'typeDI', 'typeDO']:
                    if type_modul == 'typeAI_8ch':
                        t = 'AI'
                        count = '8'
                        dData = 'UINT'
                    if type_modul == 'typeDI':
                        t = 'DI'
                        count = '32'
                        dData = 'BOOL'
                    if type_modul == 'typeDO':
                        t = 'DO'
                        count = '32'
                        dData = 'BOOL'
                    if type_modul == 'typeAO':
                        t = 'AO'
                        count = '8'
                        dData = 'UINT'
                    if type_modul == 'typeAO':
                        cfg_txt = f'\t{uso_rack_modul}_Diagnostics: NftIOItfs.IOModuleDiag;\n' \
                                f'\t{uso_rack_modul}_{t}: ARRAY[1..{count}] OF {dData};\n' \
                                f'\t{uso_rack_modul}_{t}_Statuses: ARRAY[1..{count}] OF USINT;\n' \
                                '\t(*-----------------*)\n'
                    else:
                        cfg_txt = f'\t{uso_rack_modul}_Diagnostics: NftIOItfs.IOModuleDiag;\n' \
                                f'\t{uso_rack_modul}_{t}: ARRAY[1..{count}] OF {dData};\n' \
                                '\t(*-----------------*)\n'
                    write_file_gv.write(cfg_txt)
                else:

                    cfg_txt = f"\t{uso_rack_modul}:{type_modul};\n"
                    write_file_gv.write(cfg_txt)

            write_file_gv.write('\tCfgDiag:       typeCfgDiag;\t(* Конфигурация диагностики *)\n' \
                            '\tStateDiag:     typeStateDiag;\t(* Данные по диагностике, передаваемые на ВУ *)\n' \
                            '\tMemDiag:       typeMemDiag;\t(* Хранилище данных по диагностике *)\n' \
                            '\tCmdDiag:       WORD;\t(* Команда для диагностики *)\n' \
                            '\tmAI16_HEALTH:  ARRAY[1..countAI16modules] \tOF BOOL;\n' \
                            '\tmAI8_HEALTH:   ARRAY[1..countAI8modules] \tOF BOOL;\n' \
                            '\tmAO_HEALTH:    ARRAY[1..countAOmodules] \tOF BOOL;\n' \
                            '\tmDI_HEALTH:    ARRAY[1..countDImodules] \tOF BOOL;\n' \
                            '\tmDO_HEALTH:    ARRAY[1..countDOmodules] \tOF BOOL;\n' \
                            '\tprjVersion:    typePrjVersion;\n' \
                            '\tprjVersionKvit:typePrjVersion;\n' \
                            'END_VAR\n')
            write_file_gv.close()

            write_file_cfg.write('IF NOT flInit THEN\n')
            for key, value in mbusexist.items():
                if value != -1:
                    st_bin = 0
                    for b in range(0, value + 1):
                        st_bin = st_bin + 2 ** b
                    write_file_cfg.write(f"\tCfgDiag.mBUSExists[{key}] := {str(bin(st_bin)).replace('0b', '2#')};\n")

            for mod in signals:
                PortsEnbl = []
                if mod['type_modul'] == 'typePSU':
                    write_file_cfg.write(
                        f"\tCfgDiag.PSU[{mod['abs_num']}].mPSU                   REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.PSU[{mod['abs_num']}].nRack                    := {mod['rack']};\n" \
                        f"\tCfgDiag.PSU[{mod['abs_num']}].nMod                     := {mod['modul_dash']};\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeCPU':
                    PortsEnbl = str(mod['gvd']).split(';')
                    pe = PortsEnbl[1] if len(PortsEnbl) > 1 else 7
                    write_file_cfg.write(
                        f"\tCfgDiag.CPU[{mod['abs_num']}].mCPU                   REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.CPU[{mod['abs_num']}].nRack                    := {mod['rack']};\n" \
                        f"\tCfgDiag.CPU[{mod['abs_num']}].nMod                     := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.CPU[{mod['abs_num']}].PortsEnbl                := {pe};\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeCN':
                    PortsEnbl = str(mod['gvd']).split(';')
                    pe = PortsEnbl[1] if len(PortsEnbl) > 1 else 3
                    write_file_cfg.write(
                        f"\tCfgDiag.CN[{mod['abs_num']}].mCN                     REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.CN[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.CN[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.CN[{mod['abs_num']}].PortsEnbl                 := {pe};\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeMN':
                    PortsEnbl = str(mod['gvd']).split(';')
                    pe = PortsEnbl[1] if len(PortsEnbl) > 1 else 3
                    write_file_cfg.write(
                        f"\tCfgDiag.MN[{mod['abs_num']}].mMN                     REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.MN[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.MN[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.MN[{mod['abs_num']}].PortsEnbl                 := {pe};\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeAI_8ch':
                    write_file_cfg.write(
                        f"\tCfgDiag.AI8[{mod['abs_num']}].mAI8.AI                REF= {mod['uso_rack_modul']}_AI;\n" \
                        f"\tCfgDiag.AI8[{mod['abs_num']}].nRack                    := {mod['rack']};\n" \
                        f"\tCfgDiag.AI8[{mod['abs_num']}].nMod                     := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.AI8[{mod['abs_num']}].mAI8.Diagnostics       REF= {mod['uso_rack_modul']}_Diagnostics;\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeAO':
                    write_file_cfg.write(
                        f"\tCfgDiag.AO[{mod['abs_num']}].mAO.AO                  REF= {mod['uso_rack_modul']}_AO;\n" \
                        f"\tCfgDiag.AO[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.AO[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.AO[{mod['abs_num']}].mAO.Diagnostics         REF= {mod['uso_rack_modul']}_Diagnostics;\n" \
                        f"\tCfgDiag.AO[{mod['abs_num']}].mAO.AOStatuses          REF= {mod['uso_rack_modul']}_AO_Statuses;\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeDI':
                    write_file_cfg.write(
                        f"\tCfgDiag.DI[{mod['abs_num']}].mDI.DI                  REF= {mod['uso_rack_modul']}_DI;\n" \
                        f"\tCfgDiag.DI[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.DI[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.DI[{mod['abs_num']}].mDI.Diagnostics         REF= {mod['uso_rack_modul']}_Diagnostics;\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeDO':
                    write_file_cfg.write(
                        f"\tCfgDiag.DOs[{mod['abs_num']}].mDO.DOs                 REF= {mod['uso_rack_modul']}_DO;\n" \
                        f"\tCfgDiag.DOs[{mod['abs_num']}].nRack                     := {mod['rack']};\n" \
                        f"\tCfgDiag.DOs[{mod['abs_num']}].nMod                      := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.DOs[{mod['abs_num']}].mDO.Diagnostics         REF= {mod['uso_rack_modul']}_Diagnostics;\n" \
                        '\t(*---------------------*)\n')
                if mod['type_modul'] == 'typeRS485':
                    PortsEnbl = str(mod['gvd']).split(';')
                    pe = PortsEnbl[1] if len(PortsEnbl) > 1 else 3
                    write_file_cfg.write(
                        f"\tCfgDiag.RS485[{mod['abs_num']}].mRS485               REF= {mod['uso_rack_modul']};\n" \
                        f"\tCfgDiag.RS485[{mod['abs_num']}].nRack                  := {mod['rack']};\n" \
                        f"\tCfgDiag.RS485[{mod['abs_num']}].nMod                   := {mod['modul_dash']};\n" \
                        f"\tCfgDiag.RS485[{mod['abs_num']}].PortsEnbl              := {pe};\n" \
                        '(*---------------------*)\n')

            for data in signals_ss:
                number_ss  = data['number_ss']
                name_ss    = data['name_ss']
                count_ss   = data['count_ss']
                array1_req = data['array1_req']
                num1_nom   = data['num1_nom']
                chann1_nom = data['chann1_nom']
                count_SS   = data['count_SS']
                array2_req = data['array2_req']
                num2_nom   = data['num2_nom']
                chann2_nom = data['chann2_nom']
                time_out   = data['time_out']

                if time_out is None: time_out = 'tmCommon.Diag_RStimeout'
                else               : time_out = f'REF= {time_out}'

                if (not num1_nom is None) and (not chann1_nom is None):
                    write_file_cfg.write(
                        f'(*------{name_ss}--------*)\n'
                        f"\tcfgDiag.SS[{number_ss}, 1].pTimeOut	         REF= {time_out};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].RSreq                   := {array1_req}; \n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].nRS                     := {num1_nom};   \n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].chRS                    := {chann1_nom}; \n")
                elif not count_ss is None:
                    write_file_cfg.write(
                        f'(*------{name_ss}--------*)\n'
                        f"\tcfgDiag.SS[{number_ss}, 1].pTimeOut	         REF= {time_out};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].pCounter	         REF= {count_ss};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 1].RSreq                   := 0; \n")

                if (not num2_nom is None) and (not chann2_nom is None):
                    write_file_cfg.write(
                        f'(*------{name_ss}--------*)\n'
                        f"\tcfgDiag.SS[{number_ss}, 2].pTimeOut	         REF= {time_out};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].RSreq                   := {array2_req}; \n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].nRS                     := {num2_nom};   \n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].chRS                    := {chann2_nom}; \n")
                elif not count_SS is None:
                    write_file_cfg.write(
                        f'(*------{name_ss}--------*)\n'
                        f"\tcfgDiag.SS[{number_ss}, 2].pTimeOut	         REF= {time_out};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].pCounter	         REF= {count_SS};\n" \
                        f"\tcfgDiag.SS[{number_ss}, 2].RSreq                   := 0; \n")

            for data in signals_tmdp:
                number        = data['number']
                variable      = data['variable']
                name          = data['name']
                link_time_out = data['link_time_out']
                connection    = data['connection']

                if link_time_out is None: link_time_out = 'tmCommon.CSPA_t1'

                InputVar, Num, Unioncfg_st = self.ret_inp_cfg(connection)

                write_file_cfg.write(
                    f'(*------{name}--------*)\n'
                    f"\tcfgDiag.TM_DP[{number}].pLinkOk.pInputVar    REF= {InputVar};\n" \
                    f"\tcfgDiag.TM_DP[{number}].pLinkOk.num               := {Num}; \n" \
                    f"\tcfgDiag.TM_DP[{number}].pLinkOk.cfg.reg           := {Unioncfg_st};   \n" \
                    f"\tcfgDiag.TM_DP[{number}].pTimeOut                  := {link_time_out}; \n")


            write_file_cfg.write("\tflInit := TRUE;\n" \
                            "END_IF;")

            write_file_cfg.close()
            write_file_gv.close()
            msg[f'{today} - Файл СУ: cfg_diag, gv_diag заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_diag, gv_diag: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_pic(self):
        def cycle_list_pic(cfg_txt, pic_int, id_pic, event, name_pic, sort_event):
            all_symbol = 0
            count_znak = 0
            for i in list_pic: 
                if i['id_pic'] == id_pic: 
                    if i[sort_event] is None: continue
                    all_symbol += 1

            cfg_txt = cfg_txt + f"(* Pic[{id_pic}] - {name_pic} *)\n"
            cfg_txt = cfg_txt + f"ctrlPic[{id_pic}].{event} :=\n"
            pic_int = id_pic
            return cfg_txt, pic_int, count_znak, all_symbol
        msg = {}
        list_pic = []
        try:
            data_pic   = self.dop_function.connect_by_sql('pic'  , f'''"id", "name", "frame", "Pic"''')
            data_ai    = self.dop_function.connect_by_sql('ai'   , f'''"id", "name", "Pic"''')
            data_di    = self.dop_function.connect_by_sql('di'   , f'''"id", "name", "priority_0", "priority_1", "pNC_AI", "Pic"''')
            data_zd    = self.dop_function.connect_by_sql('zd'   , f'''"id", "name", "Pic"''')
            data_vs    = self.dop_function.connect_by_sql('vs'   , f'''"id", "name", "Pic"''')
            data_ss    = self.dop_function.connect_by_sql('ss'   , f'''"id", "name", "number_array_stateRSreq_1", "number_array_stateRSreq_2", "Pic"''')
            data_ktpr  = self.dop_function.connect_by_sql('ktpr' , f'''"id", "name", "Pic"''')
            data_ktpra = self.dop_function.connect_by_sql('ktpra', f'''"id", "name", "Pic"''')
            data_hw    = self.dop_function.connect_by_sql('hardware', f'''"id", "variable", "uso", "type_1", "Pic"''')
            data_tm_dp = self.dop_function.connect_by_sql('tm_dp', f'''"name", "link_to_link_signal", "Pic"''')

            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_PIC')
            cfg_txt    = ''

            for info_pic in data_pic:
                id_pic   = info_pic[0]
                name_pic = info_pic[1]
                frame    = info_pic[2]

                for value in data_ai:
                    id_ai  = value[0]
                    name   = value[1]
                    pic_ai = value[2]
                    
                    if pic_ai is None: continue

                    s_pic = str(pic_ai).split(';')
                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            list_pic.append(dict(id_pic = id_pic,
                                                 name_pic = name_pic,
                                                 value_warn = f"AIcountWarn[{id_ai}]",
                                                 value_avar = f"AIcountAvar[{id_ai}]",
                                                 name = name))
                for value in data_di:
                    id_di  = value[0]
                    name   = value[1]
                    prio_0 = value[2]
                    prio_1 = value[3]
                    pNC    = value[4]
                    pic_di = value[5]

                    if pic_di is None: continue
                    s_pic = str(pic_di).split(';')

                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            di_avar = None
                            di_warn = None
                            if prio_0 == 3 or prio_1 == 3:
                                di_avar = f"BYTE_TO_UDINT(DIcountAvar[{id_di}])"
                                if pNC is not None:
                                    di_warn = f"BYTE_TO_UDINT(DIcountWarn[{id_di}])"
                            else:
                                di_warn = f"BYTE_TO_UDINT(DIcountWarn[{id_di}])"
                            
                            list_pic.append(dict(id_pic     = id_pic,
                                                 name_pic   = name_pic,
                                                 value_warn = di_warn,
                                                 value_avar = di_avar,
                                                 name       = name))
                for value in data_zd:
                    id_zd  = value[0]
                    name   = value[1]
                    pic_zd = value[2]

                    if pic_zd is None: continue
                    s_pic = str(pic_zd).split(';')

                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            list_pic.append(dict(id_pic = id_pic,
                                                 name_pic = name_pic,
                                                 value_warn = f"BOOL_TO_UDINT(stateZD[{id_zd}].state2.bits.NeispravVU)",
                                                 value_avar = f"BOOL_TO_UDINT(stateZD[{id_zd}].state1.bits.Avar)+\nBOOL_TO_UDINT(stateZD[{id_zd}].state1.bits.NOT_EC)",
                                                 name = name))
                for value in data_vs:
                    id_vs  = value[0]
                    name   = value[1]
                    pic_vs = value[2]

                    if pic_vs is None: continue
                    s_pic = str(pic_vs).split(';')

                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            list_pic.append(dict(id_pic = id_pic,
                                                 name_pic = name_pic,
                                                 value_warn = None,
                                                 value_avar = f"BOOL_TO_UDINT(stateVS[{id_vs}].state1.bits.NEISPRAV)+\nBOOL_TO_UDINT(stateVS[{id_vs}].state1.bits.MPC_CEPI_VKL)+\nBOOL_TO_UDINT(NOT stateVS[{id_vs}].state1.bits.EC)",
                                                 name = name))
                for value in data_ktpra:
                    id_ktpra  = value[0]
                    name      = value[1]
                    pic_ktpra = value[2]

                    if pic_ktpra is None: continue
                    s_pic = str(pic_ktpra).split(';')

                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            list_pic.append(dict(id_pic = id_pic,
                                                 name_pic = name_pic,
                                                 value_warn = None,
                                                 value_avar = f"BOOL_TO_UDINT(state{id_ktpra}.state.bits.F AND (NOT state{id_ktpra}.state.bits.M))",
                                                 name = name))
                for value in data_ktpr:
                    id_ktpr  = value[0]
                    name     = value[1]
                    pic_ktpr = value[2]

                    if pic_ktpr is None: continue
                    s_pic = str(pic_ktpr).split(';')

                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            list_pic.append(dict(id_pic = id_pic,
                                                 name_pic = name_pic,
                                                 value_warn = None,
                                                 value_avar = f"BOOL_TO_UDINT((state{id_ktpr}.state.bits.F) AND (NOT state{id_ktpr}.state.bits.M))",
                                                 name = name))       
                for value in data_pic:
                    id_pic_d = value[0]
                    name     = value[1]
                    pic_pic  = value[2]

                    if pic_pic is None: continue

                    if id_pic == id_pic_d:
                        s_pic = str(pic_pic).split(';') 
                        for pic_num in s_pic:
                            if str(pic_num) == str(id_pic):
                                list_pic.append(dict(id_pic = id_pic,
                                                     name_pic = name_pic,
                                                     value_warn = f"ctrlPic[{id_pic_d}].countWarn",
                                                     value_avar = f"ctrlPic[{id_pic_d}].countAvar",
                                                     name = name)) 
                for value in data_ss:
                    id_ss   = value[0]
                    name    = value[1]
                    array_1 = value[2]
                    array_2 = value[3]
                    pic_ss  = value[4]

                    if pic_ss is None: continue
                    s_pic = str(pic_ss).split(';') 
                    
                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            ss_warn = None
                            if not array_1 is None and not array_2 is None:
                                ss_warn = f"BOOL_TO_UDINT(NOT stateDIAG.SS[{id_ss}].bits.link1Ok)+\nBOOL_TO_UDINT(NOT stateDIAG.SS[{id_ss}].bits.link2Ok)"
                                
                            list_pic.append(dict(id_pic = id_pic,
                                                 name_pic = name_pic,
                                                 value_warn = ss_warn,
                                                 value_avar = f"BOOL_TO_UDINT(NOT stateDIAG.SS[{id_ss}].bits.linkOk)",
                                                 name = name))
                for value in data_tm_dp:
                    name    = value[0]
                    link    = value[1]
                    pic_tmdp= value[2]

                    if pic_tmdp is None: continue
                    s_pic = str(pic_tmdp).split(';') 
                    
                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            list_pic.append(dict(id_pic = id_pic,
                                                 name_pic = name_pic,
                                                 value_warn = None,
                                                 value_avar = f"TM_DP_linkOk.{str(link).split('.state.')[1]}",
                                                 name = name))
                for value in data_hw:
                    var_hw = value[1]
                    uso    = value[2]
                    pic_hw = value[4]

                    if pic_hw is None: continue
                    s_pic = str(pic_hw).split(';') 
                    
                    for pic_num in s_pic:
                        if str(pic_num) == str(id_pic):
                            list_pic.append(dict(id_pic = id_pic,
                                                 name_pic = name_pic,
                                                 value_warn = None,
                                                 value_avar = f"{var_hw}",
                                                 name = uso))
                # Сеть УСО
                if self.dop_function.str_find(str(frame).lower(), {'net_uso'}) or self.dop_function.str_find(str(name_pic).lower(), {'сеть усо'}):
                    count_mn = 0
                    count_cn = 0
                    for value in data_hw:
                        var_hw = value[1]
                        uso    = value[2]
                        type_1 = value[3]

                        if type_1 == 'MK-546-010':
                            count_mn += 1
                            value_avar = f"BOOL_TO_UDINT(stateDIAG.diagMN[{count_mn}].ports_State.bits.eP1NotLink)+\nBOOL_TO_UDINT(stateDIAG.diagMN[{count_mn}].ports_State.bits.eP2NotLink)"
                        if type_1 == 'MK-545-010':
                            count_cn += 1
                            value_avar = f"BOOL_TO_UDINT(stateDIAG.diagCN[{count_cn}].ports_State.bits.eP1NotLink)+\nBOOL_TO_UDINT(stateDIAG.diagCN[{count_cn}].ports_State.bits.eP2NotLink)"
                        
                        list_pic.append(dict(id_pic = id_pic,
                                             name_pic = name_pic,
                                             value_warn = None,
                                             value_avar = value_avar,
                                             name = uso))

            cfg_txt = cfg_txt + '(* Желтые рамки *)\n'
            pic_int = 0
            for data in list_pic:
                id_pic     = data['id_pic']
                name_pic   = data['name_pic']
                value_warn = data['value_warn']
                name       = data['name']
                if value_warn is None: continue
                if pic_int != id_pic:
                    cfg_txt, pic_int, count_znak, all_symbol = cycle_list_pic(cfg_txt, pic_int, id_pic, 'countWarn', name_pic, 'value_warn')
                count_znak += 1
                znak = ';' if count_znak == all_symbol else '+'
                cfg_txt = cfg_txt + f"  {value_warn} {znak}           (* {name} *)\n"

            cfg_txt = cfg_txt + '(* Красные рамки *)\n'
            pic_int = 0
            for data in list_pic:
                id_pic     = data['id_pic']
                name_pic   = data['name_pic']
                value_avar = data['value_avar']
                name       = data['name']
                if value_avar is None: continue
                if pic_int != id_pic:
                    cfg_txt, pic_int, count_znak, all_symbol = cycle_list_pic(cfg_txt, pic_int, id_pic, 'countAvar', name_pic, 'value_avar')
                count_znak += 1
                znak = ';' if count_znak == all_symbol else '+'
                cfg_txt = cfg_txt + f"  {value_avar} {znak}           (* {name} *)\n"

            write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_pic заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_pic: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_ts(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('tm_ts', f'"id", "name", "addr_object", "link_value"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_TM_TS')

            for data in data_value:
                number    = data[0]
                name      = data[1]
                adress    = data[2]
                reference = data[3]
                
                pInput, pnum, pcfg = self.ret_inp_cfg(reference)

                if pInput != 0:
                    cfg_txt = f'(* {adress} - {name} *)\n' \
                            f'cfgTM_TS[{number}].pInputVar          REF={str(pInput)};\n' \
                            f'cfgTM_TS[{number}].num                  :={str(pnum)};\n' \
                            f'cfgTM_TS[{number}].cfg.reg              :={str(pcfg)};\n'

                    write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_tm_ts заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_tm_ts: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_tu(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('tm_tu', f'"id", "name", "variable_change", "change_bit", "addr_object"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_TM_TU')

            for data in data_value:
                number  = data[0]
                name    = data[1]
                mut_var = data[2]
                bits    = data[3]
                adress  = data[4]

                if mut_var is None: continue
                if bits    is None: continue

                cfg_txt = f'(* {adress} - {name} *)\n' \
                        f'cfgTM_TU[{number}].pVal       REF={str(mut_var)}.reg;\n' \
                        f'cfgTM_TU[{number}].iBit         :={str(bits)};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_tm_tu заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_tm_tu: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_ti2(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('tm_ti2', f'"id", "name", "variable_value", "variable_status", "addr_object"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('Cfg_TM_TI2')

            for data in data_value:
                number  = data[0]
                name    = data[1]
                value   = data[2]
                status  = data[3]
                adress  = data[4]

                if name is None: continue
                if status is None:
                    cfg_txt = f'(* {adress} - {name} *)\n' \
                            f'cfgTM_TI2[{number}].pVal              REF={str(value)};\n'
                else:
                    cfg_txt = f'(* {adress} - {name} *)\n' \
                            f'cfgTM_TI2[{number}].pVal              REF={str(value)};\n' \
                            f'cfgTM_TI2[{number}].pState            REF={str(status)};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_tm_ti2 заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_tm_ti2: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_ti4(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('tm_ti4', f'"id", "name", "variable_value", "variable_status", "variable_Aiparam", "addr_object"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('Cfg_TM_TI4')

            for data in data_value:
                number   = data[0]
                name     = data[1]
                value    = data[2]
                status   = data[3]
                ai_param = data[4]
                adress   = data[5]

                if value is None: continue

                if (not status is None) and (not ai_param is None):
                    cfg_txt = f'(* {adress} - {name} *)\n' \
                            f'cfgTM_TI4[{number}].pVal              REF={str(value)};\n' \
                            f'cfgTM_TI4[{number}].pState            REF={str(status)}.reg;\n' \
                            f'cfgTM_TI4[{number}].pAIparam          REF={str(ai_param)};\n'
                if (status is None) and (not ai_param is None):
                    cfg_txt = f'(* {adress} - {name} *)\n' \
                            f'cfgTM_TI4[{number}].pVal              REF={str(value)};\n' \
                            f'cfgTM_TI4[{number}].pAIparam          REF={str(ai_param)};\n'
                if (not status is None) and (ai_param is None):
                    cfg_txt = f'(* {adress} - {name} *)\n' \
                            f'cfgTM_TI4[{number}].pVal              REF={str(value)};\n' \
                            f'cfgTM_TI4[{number}].pState            REF={str(status)}.reg;\n'
                if (status is None) and (ai_param is None):
                    cfg_txt = f'(* {adress} - {name} *)\n' \
                            f'cfgTM_TI4[{number}].pVal              REF={str(value)};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_tm_ti4 заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_tm_ti4: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_tii(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('tm_tii', f'"id", "name", "variable_value", "variable_status", "addr_object"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_TM_TII')

            for data in data_value:
                number  = data[0]
                name    = data[1]
                value   = data[2]
                status  = data[3]
                adress  = data[4]

                if value is None: continue

                if status is None:
                    cfg_txt = f'(* {name} - {adress} *)\n' \
                            f'cfgTM_TII[{number}].pVal              REF={str(value)};\n'
                else:
                    cfg_txt = f'(* {name} - {adress} *)\n' \
                            f'cfgTM_TII[{number}].pVal              REF={str(value)};\n' \
                            f'cfgTM_TII[{number}].pState            REF={str(status)};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_tm_tii заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_tm_tii: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_tr2(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('tm_tr2', f'"id", "name", "variable_change", "descriptionTR4", "addr_object"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('Cfg_TM_TR2')

            for data in data_value:
                number = data[0]
                name   = data[1]
                value  = data[2]
                sign   = data[3]
                adress = data[4]

                if value is None: continue

                cfg_txt = f'(* {adress} - {name} - {sign} *)\n' \
                        f'cfgTM_TR2[{number}].pVal              REF={str(value)};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_tm_tr2 заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_tm_tr2: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_tr4(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('tm_tr4', f'"id", "name", "variable_change", "descriptionTR4", "addr_object"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_TM_TR4')

            for data in data_value:
                number = data[0]
                name   = data[1]
                value  = data[2]
                sign   = data[3]
                adress = data[4]

                if value is None: continue

                cfg_txt = f'(* {adress} - {name} - {sign} *)\n' \
                        f'cfgTM_TR4[{number}].pVal              REF={str(value)};\n'

                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_tm_tr4 заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_tm_tr4: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_do_sim(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('do', f'"id", "tag", "name", "pValue", "pHealth"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('Cfg_DO_sim')

            for value in data_value:
                numbers = value[0]
                tag     = value[1]
                name    = value[2]
                pValue  = value[3]
                pHealth = value[4]

                if self.dop_function.str_find(pHealth,'mDO'):
                    pValue = str(pValue)[str(pValue).index('['):]
                    pHealth = str(pHealth)[str(pHealth).index('['):]
                    cfg_txt = f'(*{tag} {name}*)\n' \
                              f'cfgDO[{numbers}].pValue         REF=simDO_bool{pHealth}{pValue};\n'

                    write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_do_sim заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_do_sim: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_di_sim(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('di', f'"id", "tag", "name", "pValue", "pHealth"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('Cfg_DI_sim')

            for value in data_value:
                numbers = value[0]
                tag     = value[1]
                name    = value[2]
                pValue  = value[3]
                pHealth = value[4]

                if self.dop_function.str_find(pHealth,'mDO'):
                    pValue = str(pValue)[str(pValue).index('['):]
                    pHealth = str(pHealth)[str(pHealth).index('['):]
                    cfg_txt = f'(*{tag} {name}*)\n' \
                              f'cfgDI[{numbers}].pValue         REF=simDI_bool{pHealth}{pValue};\n'
                    
                    write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_di_sim заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_di_sim: {traceback.format_exc()}'] = 2
            return msg  
    def cfg_ai_sim(self):
        msg = {}
        try:
            data_value = self.dop_function.connect_by_sql('di', f'"id", "tag", "name", "pValue", "pHealth"')
            # Проверяем файл на наличие в папке, если есть удаляем и создаем новый
            write_file = self.file_check('cfg_AI_sim')

            for value in data_value:
                numbers = value[0]
                tag     = value[1]
                name    = value[2]
                pValue  = value[3]
                pHealth = value[4]

                if pHealth is None: continue

                ch = str(pValue).split('[')
                modul = str(pHealth).split('[')
                num = f'[{str(modul[1]).replace("]","")}][{str(ch[1]).replace("]","")}]'
                cfg_txt = f'(*{tag} {name}*)\n' \
                          f'cfgAI[{numbers}].pValue         REF=simAI{num};\n' \
                    
                write_file.write(cfg_txt)
            write_file.close()
            msg[f'{today} - Файл СУ: cfg_ai_sim заполнен'] = 1
            return msg
        except Exception:
            msg[f'{today} - Файл СУ: ошибка при заполнении cfg_ai_sim: {traceback.format_exc()}'] = 2
            return msg  

# Work with filling in the table 
class Filling_HardWare():
    def __init__(self):
        self.cursor = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals по количеству корзин и модулю
    def getting_modul(self, kk_is_True):
        msg = {}
        list_type = {'CPU': 'MK-504-120', 'PSU': 'MK-550-024', 'CN': 'MK-545-010', 'MN' : 'MK-546-010', 'AI'    : 'MK-516-008A',
                     'AO' : 'MK-514-008', 'DI' : 'MK-521-032', 'RS': 'MK-541-002', 'DO' : 'MK-531-032', 'EthEx' : 'MK-544-040'}
        with db:
            try:
                if self.dop_function.empty_table('signals'): 
                    msg[f'{today} - Таблица: signals пустая! Заполни таблицу!'] = 2
                    return msg

                self.cursor.execute(f'''SELECT DISTINCT uso 
                                        FROM signals
                                        ORDER BY uso''')
                list_uso = self.cursor.fetchall()

                temp_flag    = False
                test_s       = []
                count_basket = 0
                count_AI, count_AO, count_EthEx = 0, 0, 0
                count_DI, count_DO, count_RS = 0, 0, 0 
                for uso in list_uso:
                    self.cursor.execute(f"""SELECT DISTINCT basket 
                                            FROM signals
                                            WHERE uso='{uso[0]}'
                                            ORDER BY basket""")
                    list_basket = self.cursor.fetchall()

                    # ЦК в количестве 2 - ONE!
                    if temp_flag is False:
                        for i in range(2):
                            uso_kk = uso[0]
                            test_s.append(dict(uso = uso[0], variable = f'countsErrDiag[{i + 1}]', tag = '',
                                                basket  = i + 1, powerLink_ID ='', Pic = '',
                                                type_0  = f'MK-550-024', variable_0 = f'PSU',   type_1 = f'MK-546-010', variable_1 = f'MN;3',
                                                type_2  = f'MK-504-120', variable_2 = f'CPU;7', type_3 = f'',           variable_3 = f'',
                                                type_4  = f'',           variable_4 = f'',      type_5 = f'',           variable_5 = f'',
                                                type_6  = f'',           variable_6 = f'',      type_7 = f'',           variable_7 = f'',
                                                type_8  = f'',           variable_8 = f'',      type_9 = f'',           variable_9 = f'',
                                                type_10 = f'',           variable_10= f'',      type_11= f'',           variable_11= f'',
                                                type_12 = f'',           variable_12= f'',      type_13= f'',           variable_13= f'',
                                                type_14 = f'',           variable_14= f'',      type_15= f'',           variable_15= f'',
                                                type_16 = f'',           variable_16= f'',      type_17= f'',           variable_17= f'',
                                                type_18 = f'',           variable_18= f'',      type_19= f'',           variable_19= f'',
                                                type_20 = f'',           variable_20= f'',      type_21= f'',           variable_21= f'',
                                                type_22 = f'',           variable_22= f'',      type_23= f'',           variable_23= f'',
                                                type_24 = f'',           variable_24= f'',      type_25= f'',           variable_25= f'',
                                                type_26 = f'',           variable_26= f'',      type_27= f'',           variable_27= f'',
                                                type_28 = f'',           variable_28= f'',      type_29= f'',           variable_29= f'',
                                                type_30 = f'',           variable_30= f'',      type_31= f'',           variable_31= f'',
                                                type_32 = f'',           variable_32= f''))
                        temp_flag = True
                    for basket in list_basket:
                        count_basket     += 1
                        list_hw           = {}
                        list_hw['uso']    = uso[0]    
                        list_hw['basket'] = basket[0] 

                        # Если в проекте есть КК
                        if kk_is_True and count_basket == 3:
                            for i in range(4, 6, 1):
                                test_s.append(dict(uso        = uso_kk,
                                                   variable   = f'countsErrDiag[{i + 1}]',
                                                   tag        = '',
                                                   basket     = i + 1,
                                                   type_0     = 'MK-550-024',
                                                   variable_0 = f'PSU',
                                                   type_1     = f'MK-544-040',
                                                   variable_1 = f'EthEx;3',
                                                   type_2     = f'MK-504-120',
                                                   variable_2 = f'CPU;7'))

                        self.cursor.execute(f"""SELECT DISTINCT module, type_signal 
                                                FROM signals
                                                WHERE uso='{uso[0]}' AND basket={basket[0]}
                                                ORDER BY module""")
                        req_modul = self.cursor.fetchall()
                        for i in req_modul:
                            if i[1] is None or i[1] == '' or i[1] == ' ': 
                                type_kod = 'Неопределен!'
                                type_mod = 'Неопределен!'
                                msg[f'{today} - Таблица: Hardware. {uso[0]}.A{basket[0]}.{i[0]} тип не определен!'] = 2
                            else:
                                for key, value in list_type.items():
                                    if str(i[1]).find(key) != -1: 
                                        if key == 'AI': 
                                            count_AI += 1
                                            type_mod = f'{key}[{count_AI}]'
                                        elif key == 'AO': 
                                            count_AO += 1
                                            type_mod = f'{key}[{count_AO}]'
                                        elif key == 'DI': 
                                            count_DI += 1
                                            type_mod = f'{key}[{count_DI}]'
                                        elif key == 'DO': 
                                            count_DO += 1
                                            type_mod = f'{key}[{count_DO}]'
                                        elif key == 'RS': 
                                            count_RS += 1
                                            type_mod = f'RS[{count_RS}];3'
                                        elif key == 'EthEx': 
                                            count_EthEx += 1
                                            type_mod = f'{key}[{count_EthEx}]'
                                        else:
                                            type_mod = key
                                        type_kod = value

                            if   kk_is_True and (count_basket == 1 or count_basket == 2): list_hw[f'id'] = count_basket + 2
                            else: list_hw[f'id'] = count_basket + 4
                            
                            list_hw[f'variable']        = f'countsErrDiag[]'
                            list_hw[f'tag']             = ''
                            list_hw[f'powerLink_ID']    = count_basket
                            list_hw[f'type_0']          = 'MK-550-024'
                            list_hw[f'variable_0']      = 'PSU'
                            list_hw[f'type_1']          = 'MK-545-010'
                            list_hw[f'variable_1']      = 'CN;3'
                            list_hw[f'type_{i[0]}']     = type_kod
                            list_hw[f'variable_{i[0]}'] = type_mod
                        test_s.append(list_hw)

                # Checking for the existence of a database
                HardWare.insert_many(test_s).execute()
                msg[f'{today} - Таблица: hardware заполнена'] = 1
            except Exception:
                msg[f'{today} - Таблица: hardware, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: hardware, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу HardWare
    def column_check(self):
        list_default = ['variable', 'tag', 'uso', 'basket', 'powerLink_ID', 'Pic', 
                        'type_0',  'variable_0',  'type_1',  'variable_1',  'type_2',  'variable_2', 
                        'type_3',  'variable_3',  'type_4',  'variable_4',  'type_5',  'variable_5', 
                        'type_6',  'variable_6',  'type_7',  'variable_7',  'type_8',  'variable_8',
                        'type_9',  'variable_9',  'type_10', 'variable_10', 'type_11', 'variable_11', 
                        'type_12', 'variable_12', 'type_13', 'variable_13', 'type_14', 'variable_14', 
                        'type_15', 'variable_15', 'type_16', 'variable_16', 'type_17', 'variable_17',
                        'type_18', 'variable_18', 'type_19', 'variable_19', 'type_20', 'variable_20', 
                        'type_21', 'variable_21', 'type_22', 'variable_22', 'type_23', 'variable_23', 
                        'type_24', 'variable_24', 'type_25', 'variable_25', 'type_26', 'variable_26',
                        'type_27', 'variable_27', 'type_28', 'variable_28', 'type_29', 'variable_29', 
                        'type_30', 'variable_30', 'type_31', 'variable_31', 'type_32', 'variable_32']
        
        self.dop_func = General_functions()
        msg = self.dop_func.column_check(HardWare, 'hardware', list_default)
        return msg
class Filling_USO():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        temp = False
        count_USO        = 0
        list_diag_signal = []
        with db:
            try:
                if self.dop_function.empty_table('ai') or self.dop_function.empty_table('di'): 
                    msg[f'{today} - Таблицы: ai или di пустые! Заполни таблицу!'] = 2
                    return msg
                try:
                    self.cursor.execute(f'''SELECT * FROM ai''')
                    self.cursor.execute(f'''SELECT * FROM di''')
                except:
                    msg[f'{today} - Таблицы: ai или di не найдены!'] = 2
                    return msg
                self.cursor.execute(f'''SELECT DISTINCT uso 
                                        FROM signals 
                                        ORDER BY uso''')
                list_uso = self.cursor.fetchall()
                for uso in list_uso:
                    count_DI  = 0
                    count_USO += 1
                    list_diag = {}

                    list_diag['id']       = f'{count_USO}'
                    list_diag['variable'] = f'USO[{count_USO}]'
                    list_diag['name']     = f'{uso[0]}'

                    self.cursor.execute(f"""SELECT variable, "name"
                                            FROM ai
                                            WHERE "name" LIKE '%{uso[0]}%'""")
                    current_ai = self.cursor.fetchall()
                    try:
                        if len(current_ai) == 0: raise
                        for ai in current_ai:
                            list_diag['temperature']  = f'{ai[0]}'
                            break
                    except:
                        list_diag['temperature']  = ''
                        msg[f'{today} - Таблица: uso. Температура в шкафу {uso[0]} не найдена!'] = 2

                    self.cursor.execute(f"""SELECT variable, name
                                            FROM di
                                            WHERE name LIKE '%{uso[0]}%' AND 
                                                 (name LIKE '%двер%' OR name LIKE '%Двер%')""")
                    current_door = self.cursor.fetchall()
                    try:
                        if len(current_door) == 0: raise
                        for door in current_door:
                            list_diag['door']  = f'{door[0]}.Value'
                            break
                    except:
                        list_diag['temperature']  = ''
                        msg[f'{today} - Таблица: uso. Сигнал открытой двери шкафа {uso[0]} не найден!'] = 2

                    self.cursor.execute(f"""SELECT variable, name
                                            FROM di
                                            WHERE name LIKE '%{uso[0]}%' AND 
                                                 (name NOT LIKE '%двер%') AND (name NOT LIKE '%Двер%') 
                                            ORDER BY name""")
                    current_di = self.cursor.fetchall()
                    try:
                        for di in current_di:
                            count_DI += 1
                            list_diag[f'signal_{count_DI}']  = f'{di[0]}.Value'
                    except:
                        list_diag[f'signal_{count_DI}']  = ''

                    # При первом заполнение необходимо использовать все колонки
                    if temp is False:
                        for i in range(count_DI + 1, 33):
                            list_diag[f'signal_{i}']  = ''
                        temp = True

                    list_diag_signal.append(list_diag)
                # Checking for the existence of a database
                USO.insert_many(list_diag_signal).execute()
                msg[f'{today} - Таблица: uso заполнена'] = 1
            except Exception:
                msg[f'{today} - Таблица: uso, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: uso, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу USO
    def column_check(self):
        list_default = ['variable', 'name', 'temperature', 'door',
                        'signal_1', 'signal_2', 'signal_3', 'signal_4', 'signal_5', 'signal_6', 'signal_7', 'signal_8', 
                        'signal_9', 'signal_10', 'signal_11', 'signal_12', 'signal_13', 'signal_14', 'signal_15', 'signal_16',
                        'signal_17', 'signal_18', 'signal_19', 'signal_20', 'signal_21', 'signal_22', 'signal_23', 'signal_24',
                        'signal_25', 'signal_26', 'signal_27', 'signal_28', 'signal_29', 'signal_30', 'signal_31', 'signal_32'
                        ]
        msg = self.dop_function.column_check(USO, 'uso', list_default)
        return msg 
class Filling_AI():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        list_AI = []
        count_AI = 0
        dop_analog = {'Аварийное отключение'  : ['', 'мА', 'Сигналы с контролем цепи', 'Сигнализаторы', 'Сигналы с контролем цепи', [4, 20], 0],
                      'Аварийный максимальный': ['', 'мА', 'Сигналы с контролем цепи', 'Сигнализаторы', 'Сигналы с контролем цепи', [4, 20], 0],
                      'Аварийный минимальный' : ['', 'мА', 'Сигналы с контролем цепи', 'Сигнализаторы', 'Сигналы с контролем цепи', [4, 20], 0],
                      'объем'                 : ['V', 'м3', '', '', '', [None, None], 0], 
                      'объём'                 : ['V', 'м3', '', '', '', [None, None], 0],
                      'перепад'               : ['dP', 'МПа', 'Аналоги (макс1 = макс.уставка)', 'Перепад давления', '', [0, 1], 2],
                      'давлени'               : ['P', 'МПа', 'Аналоги (макс1 = повышенная)', 'Давления', '', [0, 6], 2],
                      'загазованность'        : ['Газ', '%', 'Загазованность', 'Загазованность', '', [0, 100], 0],
                      'вертик'                : ['Xверт', 'мм/с', 'Вибрации', '', '', [0, 30], 0],
                      'горизонт'              : ['Xгор', 'мм/с', 'Вибрации', '', '', [0, 30], 0],
                      'осевая'                : ['Xос', 'мм/с', 'Вибрации', '', '', [0, 30], 0],
                      'попереч'               : ['Xпоп', 'мм/с', 'Вибрации', '', '', [0, 30], 0],
                      'осевое'                : ['Xoc', 'мм/с', 'Вибрации', 'Осевые смещения', '', [0, 30], 0],
                      'сила'                  : ['I', 'A', 'Аналоги (макс1 = повышенная)', 'Общестанционные', '', [0, 1000], 0],
                      'температура'           : ['T', '°C', 'Аналоги (макс1 = повышенная)', 'Температуры', '', [-50, 100], 0],
                      'уровень'               : ['L', 'мм', 'Аналоги (макс1 = макс.уставка)', 'Уровни', '', [200, 1000], 0],
                      'утечк'                 : ['L', 'мм', 'Сигналы с контролем цепи', 'Сигнализаторы', 'Сигналы с контролем цепи', [4, 20], 0],
                      'расход'                : ['Q', 'м3/ч', 'Аналоги (макс1 = макс.уставка)', '', '', [0, 1000], 0],
                      'положени'              : ['Q', '%', '', '', '', [0, 100], 0],
                      'затоплен'              : ['L', 'мА', 'Сигналы с контролем цепи', 'Сигнализаторы', 'Сигналы с контролем цепи', [4, 20], 0],
                      'частот'                : ['F', 'Гц', '', 'Уровни', '', [0, 100], 0],
                      'процен'                : ['Q', '%', 'Аналоги (макс1 = макс.уставка)', '', '', [0, 100], 0],
                      'заслон'                : ['Q', '%', 'Аналоги (макс1 = макс.уставка)', '', '', [0, 100], 0],
                     }
        with db:
            try:
                if self.dop_function.empty_table('signals') or self.dop_function.empty_table('hardware'): 
                    msg[f'{today} - Таблицы: signals или hardware пустые! Заполни таблицу!'] = 2
                    return msg
                
                for row_sql in Signals.select().order_by(Signals.id).dicts():
                    id_s        = row_sql['id']    
                    uso_s       = row_sql['uso']    
                    tag         = row_sql['tag']
                    description = row_sql['description']
                    type_signal = row_sql['type_signal']
                    scheme      = row_sql['schema']
                    basket_s    = row_sql['basket']
                    module_s    = row_sql['module']
                    channel_s   = row_sql['channel']

                    if tag == 'None' or tag is None: tag = ''
                    tag_eng = self.dop_function.translate(tag)

                    if self.dop_function.str_find(type_signal, {'AI'}) or self.dop_function.str_find(scheme, {'AI'}):
                        count_AI += 1
                        # Выбор между полным заполнением или обновлением
                        if self.dop_function.empty_table('ai'):
                            msg[f'{today} - Таблица: ai пуста, идет заполнение'] = 1
                        else:
                            msg[f'{today} - Таблица: ai не пуста, идет обновление'] = 1

                        coincidence = AI.select().where(AI.uso     == uso_s,
                                                        AI.basket  == basket_s,
                                                        AI.module  == module_s,
                                                        AI.channel == channel_s)
                        if bool(coincidence):
                            exist_tag  = AI.select().where(AI.tag  == tag)
                            exist_name = AI.select().where(AI.name == description)

                            if not bool(exist_tag):
                                self.cursor.execute(f"""SELECT id, "tag" 
                                                        FROM ai
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}""")
                                for id_, tag_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: ai, у сигнала обновлен tag: id = {id_}, ({tag_}) {tag}'] = 2
                                self.cursor.execute(f'''UPDATE ai
                                                        SET Tag='{tag}' 
                                                        WHERE uso='{uso_s}' AND 
                                                            basket={basket_s} AND 
                                                            module={module_s} AND 
                                                            channel={channel_s}''')
        
                            if not bool(exist_name):
                                self.cursor.execute(f'''SELECT id, "name" 
                                                        FROM ai
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                                for id_, name_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: ai, у сигнала обновлено name: id = {id_}, ({name_}) {description}'] = 2
                                self.cursor.execute(f'''UPDATE ai
                                                        SET Name='{description}' 
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                            continue
                        # Сквозной номер модуля
                        try:
                            for through_module_number in HardWare.select().dicts():
                                uso_h    = through_module_number['uso']
                                basket_h = through_module_number['basket']

                                isdigit_num = ''
                                if uso_s == uso_h and basket_s == basket_h:
                                    type_mod = through_module_number[f'variable_{module_s}']
                                    isdigit_num  = re.findall('\d+', str(type_mod))
                                    
                                    try   : isdigit_num = isdigit_num[0]
                                    except: 
                                        msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                                    break
                        except Exception:
                            msg[f'{today} - Таблица: ai, ошибка при заполнении. Заполнение продолжится: {traceback.format_exc()}'] = 2
                            msg[f'{today} - Таблица: signals, ошибка в строке. Строка пропускается: {row_sql}'] = 2
                            continue

                        sign             = ''
                        unit             = ''
                        rule             = ''
                        group_analog     = ''
                        group_ust_analog = ''
                        eng_min          = None
                        eng_max          = None
                        value_precision  = 0

                        for key, short in dop_analog.items():
                            if self.dop_function.str_find(str(description).lower(), {key}):
                                sign = short[0]
                                unit = short[1]
                                rule = short[2]
                                group_analog = short[3]
                                group_ust_analog = short[4]
                                eng_min = short[5][0]
                                eng_max = short[5][1]
                                value_precision = short[6]
                                break

                        flag_MPa_kgccm2 = 1 if self.dop_function.str_find(str(description).lower(), {'давлен'}) else 0

                        IsPumpVibration = 1 if self.dop_function.str_find(str(description).lower(), {'вибрац'}) and self.dop_function.str_find(str(description).lower(), {'насос'}) else None
                        vibration_motor = 1 if self.dop_function.str_find(str(description).lower(), {'вибрац'}) and self.dop_function.str_find(str(description).lower(), {'эд'}) else None
                        current_motor = 1 if self.dop_function.str_find(str(description).lower(), {'сила тока'}) else None

                        current_motor = 1 if self.dop_function.str_find(str(description).lower(), {'сила тока'}) else None
                        
                        if self.dop_function.str_find(str(description).lower(), {'вибрац'}) and self.dop_function.str_find(str(description).lower(), {'насос'}):
                            SigMask = '0111_1111_0111_0001'
                            MsgMask = '0111_1111_0111_0001'
                            CtrlMask = '0000_1111_0000_0000'
                        elif self.dop_function.str_find(str(description).lower(), {'вибрац'}) and \
                            (self.dop_function.str_find(str(description).lower(), {'эд'} or self.dop_function.str_find(str(description).lower(), {'двигат'}))):
                            SigMask = '0111_0110_0111_0001'
                            MsgMask = '0111_0110_0111_0001'
                            CtrlMask = '0000_1101_0000_0000'
                        elif self.dop_function.str_find(str(description).lower(), {'аварийное откл'}) or \
                             self.dop_function.str_find(str(description).lower(), {'аварийн'}) or \
                             self.dop_function.str_find(str(description).lower(), {'затоплен'}) or \
                             self.dop_function.str_find(str(description).lower(), {'утечк'}) or \
                             self.dop_function.str_find(str(description).lower(), {'пожар'}):
                            SigMask = '0100_0100_0000_0001'
                            MsgMask = '0100_0100_0000_0001'
                            CtrlMask = '0000_1111_0000_1111'
                        elif self.dop_function.str_find(str(description).lower(), {'температура нефт'}) or \
                             self.dop_function.str_find(str(description).lower(), {'уровень неф'}) or \
                             self.dop_function.str_find(str(description).lower(), {'температура охл'}) or \
                             self.dop_function.str_find(str(description).lower(), {'давление неф'}) or \
                             self.dop_function.str_find(str(description).lower(), {'пожар'}):
                            SigMask = '0100_0000_0000_0001'
                            MsgMask = '0100_0000_0000_0001'
                            CtrlMask = '0000_0000_0000_0000'
                        elif self.dop_function.str_find(str(description).lower(), {'давление на вых'}) or \
                             self.dop_function.str_find(str(description).lower(), {'давление мас'}) or \
                             self.dop_function.str_find(str(description).lower(), {'перепад давл'}):
                            SigMask = '0100_0000_0001_0001'
                            MsgMask = '0100_0000_0001_0001'
                            CtrlMask = '0000_0000_0000_0000'
                        elif self.dop_function.str_find(str(description).lower(), {'сила тока'}):
                            SigMask = '0100_0000_0010_0001'
                            MsgMask = '0100_0000_0010_0001'
                            CtrlMask = '0000_0000_0000_0000'
                        elif self.dop_function.str_find(str(description).lower(), {'Температура возд'}):
                            SigMask = '0100_0010_0010_0001'
                            MsgMask = '0100_0010_0010_0001'
                            CtrlMask = '0000_0000_0000_0000'
                        elif self.dop_function.str_find(str(description).lower(), {'температура горяч'}) or \
                             self.dop_function.str_find(str(description).lower(), {'температура задн'}) or \
                             self.dop_function.str_find(str(description).lower(), {'температура корп'}) or \
                             self.dop_function.str_find(str(description).lower(), {'температура упор'}) or \
                             self.dop_function.str_find(str(description).lower(), {'температура холо'}):
                            SigMask = '0100_0101_0000_0001'
                            MsgMask = '0100_0101_0000_0001'
                            CtrlMask = '0000_1111_0000_1111'
                        elif self.dop_function.str_find(str(description).lower(), {'загазован'}) or \
                             self.dop_function.str_find(str(description).lower(), {'Температура желез'}) or \
                             self.dop_function.str_find(str(description).lower(), {'температура обмо'}):
                            SigMask = '0100_0110_0000_0001'
                            MsgMask = '0100_0111_1000_0001'
                            CtrlMask = '0000_1111_0000_1111'
                        else:
                            SigMask = '0000_0000_0000_0000'
                            MsgMask = '0000_0000_0000_0000'
                            CtrlMask = '0000_0000_0000_0000'

                        if isdigit_num == '':
                            msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2

                        msg[f'{today} - Таблица: ai, добавлен новый сигнал: {row_sql}'] = 1
                        list_AI.append(dict(id = count_AI,
                                            variable = f'AI[{count_AI}]',
                                            tag = tag,
                                            name = description,
                                            pValue = f'mAI8[{isdigit_num}, {module_s}]',
                                            pHealth = f'mAI8_HEALTH[{isdigit_num}]',
                                            AnalogGroupId = group_analog,
                                            SetpointGroupId = group_ust_analog,
                                            Egu = unit,
                                            sign_VU = sign,
                                            IsOilPressure = flag_MPa_kgccm2,
                                            number_NA_or_aux = None,
                                            IsPumpVibration = IsPumpVibration,
                                            vibration_motor = vibration_motor,
                                            current_motor = current_motor,
                                            aux_outlet_pressure = None,
                                            number_ust_min_avar = None,
                                            number_ust_min_pred = None,
                                            number_ust_max_pred = None,
                                            number_ust_max_avar = None,
                                            LoLimField = 4000,
                                            HiLimField = 20000,
                                            LoLimEng = eng_min,
                                            HiLimEng = eng_max,
                                            LoLim = 3900,
                                            HiLim = 20100,
                                            Histeresis = 0,
                                            TimeFilter = 0,
                                            Min6 = None, Min5 = None, Min4 = None, Min3 = None, Min2 = None, Min1 = None,
                                            Max1 = None, Max2 = None, Max3 = None, Max4 = None, Max5 = None, Max6 = None,
                                            Precision = value_precision,
                                            SigMask = SigMask, 
                                            MsgMask = MsgMask, 
                                            CtrlMask = CtrlMask,
                                            Pic = '', TrendingGroup = None, DeltaT = 0, PhysicEgu = 'мкА', 
                                            RuleName = rule, fuse = '', uso = uso_s, basket = basket_s, module = module_s, channel = channel_s, tag_eng = tag_eng,
                                            AlphaHMI = '', AlphaHMI_PIC1 = '', AlphaHMI_PIC1_Number_kont = '', AlphaHMI_PIC2 = '', 
                                            AlphaHMI_PIC2_Number_kont = '', AlphaHMI_PIC3 = '', AlphaHMI_PIC3_Number_kont = '', 
                                            AlphaHMI_PIC4 = '', AlphaHMI_PIC4_Number_kont = ''))

                # Checking for the existence of a database
                AI.insert_many(list_AI).execute()
            except Exception:
                msg[f'{today} - Таблица: ai, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: ai, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу AI
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'pValue', 'pHealth', 'AnalogGroupId',
                        'SetpointGroupId',  'Egu',  'sign_VU',  'IsOilPressure',  'number_NA_or_aux',  
                        'IsPumpVibration',  'vibration_motor',  'current_motor',  'aux_outlet_pressure', 
                        'number_ust_min_avar',  'number_ust_min_pred',  'number_ust_max_pred',  'number_ust_max_avar', 
                        'LoLimField',  'HiLimField',  'LoLimEng', 'HiLimEng', 'LoLim', 'HiLim', 
                        'Histeresis', 'TimeFilter', 
                        'Min6', 'Min5', 'Min4', 'Min3', 'Min2', 'Min1', 'Max1', 'Max2', 'Max3', 'Max4', 'Max5', 'Max6', 
                        'SigMask', 'MsgMask', 'CtrlMask', 'Precision', 'Pic', 'TrendingGroup', 'DeltaT', 
                        'PhysicEgu', 'RuleName', 'fuse', 'uso', 'basket', 'module', 'channel', 'tag_eng', 'AlphaHMI', 'AlphaHMI_PIC1', 
                        'AlphaHMI_PIC1_Number_kont', 'AlphaHMI_PIC2', 'AlphaHMI_PIC2_Number_kont',
                        'AlphaHMI_PIC3', 'AlphaHMI_PIC3_Number_kont', 'AlphaHMI_PIC4', 'AlphaHMI_PIC4_Number_kont']
    
        msg = self.dop_function.column_check(AI, 'ai', list_default)
        return msg 
class Filling_AO():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        list_AO = []
        count_AO = 0
        
        with db:
            try:
                if self.dop_function.empty_table('signals') or self.dop_function.empty_table('hardware'): 
                    msg[f'{today} - Таблицы: signals или hardware пустые! Заполни таблицу!'] = 2
                    return msg
                
                for row_sql in Signals.select().order_by(Signals.id).dicts():
                    id_s        = row_sql['id']  
                    uso_s       = row_sql['uso']    
                    tag         = row_sql['tag']
                    description = row_sql['description']
                    type_signal = row_sql['type_signal']
                    scheme      = row_sql['schema']
                    basket_s    = row_sql['basket']
                    module_s    = row_sql['module']
                    channel_s   = row_sql['channel']

                    if tag == 'None' or tag is None: tag = ''
                    tag_eng = self.dop_function.translate(tag)

                    if self.dop_function.str_find(type_signal, {'AO'}) or self.dop_function.str_find(scheme, {'AO'}):
                        count_AO += 1
                        # Выбор между полным заполнением или обновлением
                        if self.dop_function.empty_table('ao'):
                            msg[f'{today} - Таблица: ao пуста, идет заполнение'] = 1
                        else:
                            msg[f'{today} - Таблица: ao не пуста, идет обновление'] = 1

                        coincidence = AO.select().where(AO.uso     == uso_s,
                                                        AO.basket  == basket_s,
                                                        AO.module  == module_s,
                                                        AO.channel == channel_s)
                        if bool(coincidence):
                            exist_tag  = AO.select().where(AO.tag == tag)
                            exist_name = AO.select().where(AO.name == description)

                            if not bool(exist_tag):
                                self.cursor.execute(f'''SELECT id, tag 
                                                        FROM ao
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                                for id_, tag_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: ao, у сигнала обновлен tag: id = {id_}, ({tag_}) {tag}'] = 2
                                self.cursor.execute(f'''UPDATE ao
                                                        SET tag='{tag}' 
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
        
                            if not bool(exist_name):
                                self.cursor.execute(f'''SELECT id, name 
                                                        FROM ao
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                                for id_, name_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: ao, у сигнала обновлено name: id = {id_}, ({name_}) {description}'] = 2
                                self.cursor.execute(f'''UPDATE ao
                                                        SET name='{description}' 
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                            continue

                        # Сквозной номер модуля
                        try:
                            for through_module_number in HardWare.select().dicts():
                                tag_h    = through_module_number['tag']
                                uso_h    = through_module_number['uso']
                                basket_h = through_module_number['basket']

                                isdigit_num = ''
                                if uso_s == uso_h and basket_s == basket_h:
                                    type_mod = through_module_number[f'variable_{module_s}']
                                    isdigit_num  = re.findall('\d+', str(type_mod))

                                    try   : isdigit_num = isdigit_num[0]
                                    except: 
                                        msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                                    break

                            if module_s < 10: prefix = f'0{module_s}' 
                            else            : prefix = f'{module_s}'
                        except Exception:
                            msg[f'{today} - Таблица: ao, ошибка при заполнении. Заполнение продолжится: {traceback.format_exc()}'] = 2
                            msg[f'{today} - Таблица: signals, ошибка в строке. Строка пропусается: {row_sql}'] = 2
                            continue
                        
                        if isdigit_num == '':
                            msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                        
                        msg[f'{today} - Таблица: ao, добавлен новый сигнал: {row_sql}'] = 1
                        list_AO.append(dict(id = count_AO,
                                            variable = f'AO[{count_AO}]',
                                            tag = tag,
                                            name = description,
                                            pValue = f'{tag_h}_{prefix}_AO[{channel_s}]',
                                            pHealth = f'mAO_HEALTH[{isdigit_num}]',
                                            uso = uso_s, 
                                            basket = basket_s, 
                                            module = module_s, 
                                            channel = channel_s,
                                            tag_eng = tag_eng,
                                            ))

                # Checking for the existence of a database
                AO.insert_many(list_AO).execute()
            except Exception:
                msg[f'{today} - Таблица: ao, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: ao, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу AO
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'pValue', 'pHealth', 'uso', 'basket', 'module', 'channel', 'tag_eng']
        msg = self.dop_function.column_check(AO, 'ao', list_default)
        return msg 
class Filling_DI():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        list_DI = []
        count_DI = 0
        with db:
            try:
                if self.dop_function.empty_table('signals') or self.dop_function.empty_table('hardware'): 
                    msg[f'{today} - Таблицы: signals или hardware пустые! Заполни таблицу!'] = 2
                    return msg
                
                for row_sql in Signals.select().order_by(Signals.id).dicts():
                    id_s       = row_sql['id'] 
                    uso_s       = row_sql['uso']    
                    tag         = row_sql['tag']
                    description = row_sql['description']
                    type_signal = row_sql['type_signal']
                    scheme      = row_sql['schema']
                    basket_s    = row_sql['basket']
                    module_s    = row_sql['module']
                    channel_s   = row_sql['channel']

                    if tag == 'None' or tag is None: tag = ''
                    tag_eng = self.dop_function.translate(tag)

                    if self.dop_function.str_find(type_signal, {'DI'}) or self.dop_function.str_find(scheme, {'DI'}):
                        count_DI += 1
                        # Выбор между полным заполнением или обновлением
                        if self.dop_function.empty_table('di'):
                            msg[f'{today} - Таблица: di пуста, идет заполнение'] = 1
                        else:
                            msg[f'{today} - Таблица: di не пуста, идет обновление'] = 1

                        coincidence = DI.select().where(DI.uso     == uso_s,
                                                        DI.basket  == basket_s,
                                                        DI.module  == module_s,
                                                        DI.channel == channel_s)
                        if bool(coincidence):
                            exist_tag  = DI.select().where(DI.tag  == tag)
                            exist_name = DI.select().where(DI.name == description)

                            if not bool(exist_tag):
                                self.cursor.execute(f'''SELECT id, tag 
                                                        FROM di
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                                for id_, tag_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: di, у сигнала обновлен tag: id = {id_}, ({tag_}) {tag}'] = 2
                                self.cursor.execute(f'''UPDATE di
                                                        SET tag='{tag}' 
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
        
                            if not bool(exist_name):
                                self.cursor.execute(f'''SELECT id, name 
                                                        FROM di
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                                for id_, name_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: di, у сигнала обновлено name: id = {id_}, ({name_}) {description}'] = 2
                                self.cursor.execute(f'''UPDATE di
                                                        SET name='{description}' 
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                            continue
                        # Сквозной номер модуля
                        try:
                            for through_module_number in HardWare.select().dicts():
                                tag_h    = through_module_number['tag']
                                uso_h    = through_module_number['uso']
                                basket_h = through_module_number['basket']

                                isdigit_num = ''
                                if uso_s == uso_h and basket_s == basket_h:
                                    type_mod = through_module_number[f'variable_{module_s}']
                                    isdigit_num  = re.findall('\d+', str(type_mod))

                                    try   : 
                                        isdigit_num = isdigit_num[0]
                                        if tag_h == '':
                                            msg[f'{today} - В таблице hardware не заполнен tag: {id_s}, {description}, "pValue" некорректно заполнено'] = 2
                                    except: 
                                        msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                                    break

                            if module_s < 10: prefix = f'0{module_s}' 
                            else            : prefix = f'{module_s}'
                        except Exception:
                            msg[f'{today} - Таблица: di, ошибка при заполнении. Заполнение продолжится: {traceback.format_exc()}'] = 2
                            msg[f'{today} - Таблица: signals, ошибка в строке. Строка пропусается: {row_sql}'] = 2
                            continue

                        if self.dop_function.str_find(str(tag).lower(), {'csc'}) : group_diskrets = 'Диагностика'
                        elif self.dop_function.str_find(str(tag).lower(), {'ec'}): group_diskrets = 'Электроснабжение'
                        else: group_diskrets = 'Общие'

                        if isdigit_num == '':
                            msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                        
                        msg[f'{today} - Таблица: di, добавлен новый сигнал: {row_sql}'] = 1

                        list_DI.append(dict(id = count_DI,
                                            variable = f'DI[{count_DI}]',
                                            tag = tag,
                                            name = description,
                                            pValue = f'{tag_h}_{prefix}_DI[{channel_s}]',
                                            pHealth = f'mDI_HEALTH[{str(isdigit_num)}]',
                                            Inv = 0,
                                            ErrValue = 0,
                                            priority_0 = 1,
                                            priority_1 = 1,
                                            Msg = 1,
                                            tabl_msg = 'TblDiscretes',
                                            group_diskrets = group_diskrets,
                                            short_title = description,
                                            uso = uso_s, basket = basket_s, module = module_s, channel = channel_s, tag_eng = tag_eng,))

                # Checking for the existence of a database
                DI.insert_many(list_DI).execute()
            except Exception:
                msg[f'{today} - Таблица: di, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: di, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу DI
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'pValue', 'pHealth', 'Inv',
                        'ErrValue', 'priority_0', 'priority_1', 'Msg', 'isDI_NC',  
                        'isAI_Warn', 'isAI_Avar', 'pNC_AI',  'TS_ID', 
                        'isModuleNC',  'Pic',  'tabl_msg',  'group_diskrets', 
                        'msg_priority_0',  'msg_priority_1', 'short_title', 'uso', 'basket', 'module', 'channel', 'tag_eng',
                        'AlphaHMI', 'AlphaHMI_PIC1', 'AlphaHMI_PIC1_Number_kont', 'AlphaHMI_PIC2',
                        'AlphaHMI_PIC2_Number_kont','AlphaHMI_PIC3', 'AlphaHMI_PIC3_Number_kont', 
                        'AlphaHMI_PIC4', 'AlphaHMI_PIC4_Number_kont']
        msg = self.dop_function.column_check(DI, 'di', list_default)
        return msg 
class Filling_DO():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        list_DO = []
        count_DO = 0
        with db:
            try:
                if self.dop_function.empty_table('signals') or self.dop_function.empty_table('hardware'): 
                    msg[f'{today} - Таблицы: signals или hardware пустые! Заполни таблицу!'] = 2
                    return msg
                
                for row_sql in Signals.select().order_by(Signals.id).dicts():
                    id_s       = row_sql['id'] 
                    uso_s       = row_sql['uso']    
                    tag         = row_sql['tag']
                    description = row_sql['description']
                    type_signal = row_sql['type_signal']
                    scheme      = row_sql['schema']
                    basket_s    = row_sql['basket']
                    module_s    = row_sql['module']
                    channel_s   = row_sql['channel']

                    if tag == 'None' or tag is None: tag = ''
                    tag_eng = self.dop_function.translate(tag)

                    if self.dop_function.str_find(type_signal, {'DO'}) or self.dop_function.str_find(scheme, {'DO'}):
                        count_DO += 1
                        # Выбор между полным заполнением или обновлением
                        if self.dop_function.empty_table("do"):
                            msg[f'{today} - Таблица: do пуста, идет заполнение'] = 1
                        else:
                            msg[f'{today} - Таблица: do не пуста, идет обновление'] = 1

                        coincidence = DO.select().where(DO.uso    == uso_s,
                                                        DO.basket == basket_s,
                                                        DO.module == module_s,
                                                        DO.channel== channel_s)
                        if bool(coincidence):
                            exist_tag  = DO.select().where(DO.tag == tag)
                            exist_name = DO.select().where(DO.name == description)

                            if not bool(exist_tag):
                                self.cursor.execute(f'''SELECT id, tag 
                                                        FROM "do"
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                                for id_, tag_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: do, у сигнала обновлен tag: id = {id_}, ({tag_}) {tag}'] = 2
                                self.cursor.execute(f'''UPDATE "do"
                                                        SET tag='{tag}' 
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
        
                            if not bool(exist_name):
                                self.cursor.execute(f'''SELECT id, name 
                                                        FROM "do"
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                                for id_, name_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: do, у сигнала обновлено name: id = {id_}, ({name_}) {description}'] = 2
                                self.cursor.execute(f'''UPDATE "do"
                                                        SET name='{description}' 
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                            continue
                        # Сквозной номер модуля
                        try:
                            for through_module_number in HardWare.select().dicts():
                                tag_h    = through_module_number['tag']
                                uso_h    = through_module_number['uso']
                                basket_h = through_module_number['basket']

                                isdigit_num = ''
                                if uso_s == uso_h and basket_s == basket_h:
                                    type_mod = through_module_number[f'variable_{module_s}']
                                    isdigit_num  = re.findall('\d+', str(type_mod))

                                    try   : isdigit_num = isdigit_num[0]
                                    except: 
                                        msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                                    break

                            if module_s < 10: prefix = f'0{module_s}' 
                            else            : prefix = f'{module_s}'
                        except Exception:
                            msg[f'{today} - Таблица: do, ошибка при заполнении. Заполнение продолжится: {traceback.format_exc()}'] = 2
                            msg[f'{today} - Таблица: signals, ошибка в строке. Строка пропусается: {row_sql}'] = 2
                            continue

                        if isdigit_num == '':
                            msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                        
                        msg[f'{today} - Таблица: do, добавлен новый сигнал: {row_sql}'] = 1

                        list_DO.append(dict(id = count_DO,
                                            variable = f'DO[{count_DO}]',
                                            tag = tag,
                                            name = description,
                                            pValue = f'{tag_h}_{prefix}_DO[{channel_s}]',
                                            pHealth = f'mDO_HEALTH[{str(isdigit_num)}]',
                                            short_title = description, tabl_msg = '',
                                            uso = uso_s, basket = basket_s, module = module_s, channel = channel_s, tag_eng = tag_eng,
                                            AlphaHMI = '', AlphaHMI_PIC1 = '', AlphaHMI_PIC1_Number_kont = '', AlphaHMI_PIC2 = '', 
                                            AlphaHMI_PIC2_Number_kont = '', AlphaHMI_PIC3 = '', AlphaHMI_PIC3_Number_kont = '', 
                                            AlphaHMI_PIC4 = '', AlphaHMI_PIC4_Number_kont = ''))

                # Checking for the existence of a database
                DO.insert_many(list_DO).execute()
            except Exception:
                msg[f'{today} - Таблица: do, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: do, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу DO
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'pValue', 'pHealth', 'short_title', 'tabl_msg', 'uso', 'basket', 'module', 'channel', 'tag_eng',
                        'AlphaHMI', 'AlphaHMI_PIC1', 'AlphaHMI_PIC1_Number_kont', 'AlphaHMI_PIC2',
                        'AlphaHMI_PIC2_Number_kont','AlphaHMI_PIC3', 'AlphaHMI_PIC3_Number_kont', 
                        'AlphaHMI_PIC4', 'AlphaHMI_PIC4_Number_kont']
        msg = self.dop_function.column_check(DO, 'do', list_default)
        return msg 
class Filling_RS():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        list_RS = []
        count_RS = 0
        with db:
            try:
                if self.dop_function.empty_table('signals') or self.dop_function.empty_table('hardware'): 
                    msg[f'{today} - Таблицы: signals или hardware пустые! Заполни таблицу!'] = 2
                    return msg
                
                for row_sql in Signals.select().order_by(Signals.id).dicts():
                    id_s       = row_sql['id'] 
                    uso_s       = row_sql['uso']    
                    tag         = row_sql['tag']
                    description = row_sql['description']
                    type_signal = row_sql['type_signal']
                    scheme      = row_sql['schema']
                    basket_s    = row_sql['basket']
                    module_s    = row_sql['module']
                    channel_s   = row_sql['channel']

                    tag_translate = self.dop_function.translate(str(tag))
                    if tag_translate == 'None': tag_translate = ''

                    if self.dop_function.str_find(type_signal, {'RS'}) or self.dop_function.str_find(scheme, {'RS'}):
                        count_RS += 1
                        # Выбор между полным заполнением или обновлением
                        if self.dop_function.empty_table("rs"):
                            msg[f'{today} - Таблица: rs пуста, идет заполнение'] = 1
                        else:
                            msg[f'{today} - Таблица: rs не пуста, идет обновление'] = 1

                        coincidence = RS.select().where(RS.uso    == uso_s,
                                                        RS.basket == basket_s,
                                                        RS.module == module_s,
                                                        RS.channel== channel_s)
                        if bool(coincidence):
                            exist_name = RS.select().where(RS.name == description)
        
                            if not bool(exist_name):
                                self.cursor.execute(f'''SELECT id, name 
                                                        FROM "rs"
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                                for id_, name_ in self.cursor.fetchall():
                                    msg[f'{today} - Таблица: rs, у сигнала обновлено name: id = {id_}, ({name_}) {description}'] = 2
                                self.cursor.execute(f'''UPDATE "rs"
                                                        SET name='{description}' 
                                                        WHERE uso='{uso_s}' AND 
                                                              basket={basket_s} AND 
                                                              module={module_s} AND 
                                                              channel={channel_s}''')
                            continue

                        # Сквозной номер модуля
                        try:
                            for through_module_number in HardWare.select().dicts():
                                tag_h    = through_module_number['tag']
                                uso_h    = through_module_number['uso']
                                basket_h = through_module_number['basket']

                                isdigit_num = ''
                                if uso_s == uso_h and basket_s == basket_h:
                                    type_mod = through_module_number[f'variable_{module_s}']
                                    isdigit_num  = re.findall('\d+', str(type_mod))

                                    try   : isdigit_num = isdigit_num[0]
                                    except: 
                                        msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                                    break

                            if module_s < 10: prefix = f'0{module_s}' 
                            else            : prefix = f'{module_s}'
                        except Exception:
                            msg[f'{today} - Таблица: do, ошибка при заполнении. Заполнение продолжится: {traceback.format_exc()}'] = 2
                            msg[f'{today} - Таблица: signals, ошибка в строке. Строка пропусается: {row_sql}'] = 2
                            continue

                        if isdigit_num == '':
                            msg[f'{today} - В таблице hardware не найден модуль сигнала: {id_s}, {tag}, {description}, {uso_s}_A{basket_s}_{module_s}_{channel_s}, "pValue" не заполнен'] = 2
                        
                        msg[f'{today} - Таблица: rs, добавлен новый сигнал: {row_sql}'] = 1

                        list_RS.append(dict(id = count_RS,
                                            variable = f'RS[{count_RS}]',
                                            tag = tag_translate,
                                            name = description,
                                            pValue = f'{tag_h}_{prefix}.COM_CH[{channel_s}]',
                                            pHealth = f'mDO_HEALTH[{str(isdigit_num)}]',
                                            Pic = '',
                                            uso = uso_s, basket = basket_s, module = module_s, channel = channel_s))

                # Checking for the existence of a database
                RS.insert_many(list_RS).execute()
            except Exception:
                msg[f'{today} - Таблица: rs, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: rs, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу RS
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'array_number_modul', 'pValue', 'pHealth', 'Pic', 'uso', 'basket', 'module', 'channel']
        msg = self.dop_function.column_check(RS, 'rs', list_default)
        return msg 
class Filling_KTPRP():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_KTPRP = []
        with db:
            for i in range(1, 31):
                list_KTPRP.append(dict(id = i,
                                       variable = f'KTPRP[{i}]',
                                       tag = '',
                                       name = 'Резерв',
                                       Number_PZ = '',
                                       Type = '',
                                       Pic = ''))

            # Checking for the existence of a database
            KTPRP.insert_many(list_KTPRP).execute()

        msg[f'{today} - Таблица: ktprp подготовлена'] = 1
        return(msg)
    # Заполняем таблицу KTPRP
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'Number_PZ', 'Type', 'Pic', 
                         'number_list_VU', 'number_protect_VU']
        msg = self.dop_function.column_check(KTPRP, 'ktprp', list_default)
        return msg 
class Filling_KTPR():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_KTPR = []
        with db:
            for i in range(1, 97):
                list_KTPR.append(dict(id = i,
                                      variable = f'KTPR[{i}]',
                                      tag = '',
                                      name = 'Резерв',
                                      avar_parameter = '',
                                    #   prohibition_masking = '',
                                    #   auto_unlock_protection = '',
                                    #   shutdown_PNS_a_time_delay_up_5s_after_turning = '',
                                    #   bitmask_protection_group_membership = '',
                                    #   stop_type_NA = '',
                                    #   pump_station_stop_type = '',
                                    #   closing_gate_valves_at_the_inlet_NPS = '',
                                    #   closing_gate_valves_at_the_outlet_NPS = '',
                                    #   closing_gate_valves_between_PNS_and_MNS = '',
                                    #   closing_gate_valves_between_RP_and_PNS = '',
                                    #   closing_valves_inlet_and_outlet_MNS = '',
                                    #   closing_valves_inlet_and_outlet_PNS = '',
                                    #   closing_valves_inlet_and_outlet_MNA = '',
                                    #   closing_valves_inlet_and_outlet_PNA = '',
                                    #   closing_valves_inlet_RD = '',
                                    #   closing_valves_outlet_RD = '',
                                    #   closing_valves_inlet_SSVD = '',
                                    #   closing_valves_inlet_FGU = '',
                                    #   closing_secant_valve_connection_unit__oil_production_oil = '',
                                    #   closing_valves_inlet_RP = '',
                                    #   reserve_protect_14 = '',
                                    #   reserve_protect_15 = '',
                                    #   shutdown_oil_pumps_after_signal_stopped_NA = '',
                                    #   shutdown_circulating_water_pumps = '',
                                    #   shutdown_pumps_pumping_out_from_tanks_collection_of_leaks_MNS = '',
                                    #   shutdown_pumps_pumping_out_from_tanks_collection_of_leaks_PNS = '',
                                    #   shutdown_pumps_pumping_out_from_tanks_SSVD = '',
                                    #   switching_off_the_electric_room_fans = '',
                                    #   shutdown_of_booster_fans_ED = '',
                                    #   shutdown_of_retaining_fans_of_the_electrical_room = '',
                                    #   shutdown_of_ED_air_compressors = '',
                                    #   shutdown_pumps_providing_oil = '',
                                    #   disabling_pumps_for_pumping_oil_oil_products_through_BIC = '',
                                    #   shutdown_domestic_and_drinking_water_pumps = '',
                                    #   shutdown_of_art_well_pumps = '',
                                    #   AVO_shutdown = '',
                                    #   shutdown_of_water_cooling_fans_circulating_water_supply_system = '',
                                    #   shutdown_exhaust_fans_of_the_pumping_room_of_the_MNS = '',
                                    #   shutdown_of_exhaust_fans_of_the_pumping_room_PNS = '',
                                    #   shutdown_of_exhaust_fans_in_the_centralized_oil_system_room = '',
                                    #   shutdown_of_exhaust_fans_oil_pit_in_the_electrical_room = '',
                                    #   shutdown_of_exhaust_fans_in_the_RD_room = '',
                                    #   shutdown_of_exhaust_fans_in_the_SSVD_room = '',
                                    #   shutdown_of_the_roof_fans_of_the_MNS_pump_room = '',
                                    #   shutdown_of_the_roof_fans_of_the_PNS_pump_room = '',
                                    #   switching_off_the_supply_fans_pumping_room_of_the_MNS = '',
                                    #   switching_off_the_supply_fans_pumping_room_of_the_PNS = '',
                                    #   switch_off_the_supply_fans_in_the_centralized_oil = '',
                                    #   switching_off_the_supply_fan_of_the_RD_room = '',
                                    #   switching_off_the_supply_fan_of_the_SSVD_room = '',
                                    #   switching_off_the_supply_fans_of_the_ED_air_compressor = '',
                                    #   switching_off_the_supply_fan_of_the_BIK_room = '',
                                    #   switching_off_the_supply_fan_of_the_SIKN_room = '',  
                                    #   closing_the_air_valves_louvered_grilles_of_the_pump_room = '',
                                    #   closing_of_air_valves_louvered_grilles_of_the_compressor_room = '',
                                    #   shutdown_of_electric_oil_heaters = '',
                                    #   shutdown_of_the_electric_heaters_of_the_leakage_collection_MNS = '',
                                    #   shutdown_of_the_electric_heaters_of_the_leakage_collection_PNS = '',
                                    #   shutdown_of_electric_heaters_of_the_SIKN_leak_collection_tank = '',
                                    #   shutdown_of_air_coolers_of_the_locking_system_MNA = '',
                                    #   shutdown_of_air_coolers_of_the_locking_system_disc_NA = '',
                                    #   shutdown_of_the_external_cooling_circuit_ChRP_MNA = '',
                                    #   shutdown_of_the_external_cooling_circuit_ChRP_PNA = '',
                                    #   shutdown_of_locking_system_pumps = '',
                                    #   shutdown_of_pumps_for_pumping_oil_oil_products_through = '',
                                    #   shutdown_of_pumping_pumps_from_leakage_collection_tanks = '',
                                    #   shutdown_of_anticondensation_electric_heaters_ED = '',
                                    #   fire_protection = '',
                                    #   reserve_aux_15 = '',
                                    #   value_ust = '',
                                    #   Pic = '',
                                      group_ust = 'Временные уставки общестанционных защит',
                                      rule_map_ust = 'Временные уставки',
                                    #   number_list_VU = '',
                                    #   number_protect_VU = ''
                                    ))

            # Checking for the existence of a database
            KTPR.insert_many(list_KTPR).execute()

        msg[f'{today} - Таблица: ktpr подготовлена'] = 1
        return(msg)
    # Заполняем таблицу KTPR
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 
                        'avar_parameter', 'DisableMasking', 'auto_unlock_protection', 'shutdown_PNS_a_time_delay_up_5s_after_turning',
                        'bitmask_protection_group_membership', 'stop_type_NA', 'pump_station_stop_type',
                        'closing_gate_valves_at_the_inlet_NPS', 'closing_gate_valves_at_the_outlet_NPS', 'closing_gate_valves_between_PNS_and_MNS',
                        'closing_gate_valves_between_RP_and_PNS', 'closing_valves_inlet_and_outlet_MNS', 'closing_valves_inlet_and_outlet_PNS',
                        'closing_valves_inlet_and_outlet_MNA', 'closing_valves_inlet_and_outlet_PNA', 'closing_valves_inlet_RD',
                        'closing_valves_outlet_RD', 'closing_valves_inlet_SSVD', 'closing_valves_inlet_FGU',
                        'closing_secant_valve_connection_unit__oil_production_oil', 'closing_valves_inlet_RP', 'reserve_protect_14', 'reserve_protect_15',
                        'shutdown_oil_pumps', 'shutdown_oil_pumps_after_signal_stopped_NA', 'shutdown_circulating_water_pumps',
                        'shutdown_pumps_pumping_out_from_tanks_collection_of_leaks_MNS', 'shutdown_pumps_pumping_out_from_tanks_collection_of_leaks_PNS',
                        'shutdown_pumps_pumping_out_from_tanks_SSVD', 'switching_off_the_electric_room_fans', 'shutdown_of_booster_fans_ED', 
                        'shutdown_of_retaining_fans_of_the_electrical_room', 'shutdown_of_ED_air_compressors', 
                        'shutdown_pumps_providing_oil', 
                        'disabling_pumps_for_pumping_oil_oil_products_through_BIC', 'shutdown_domestic_and_drinking_water_pumps', 'shutdown_of_art_well_pumps',
                        'AVO_shutdown', 'shutdown_of_water_cooling_fans_circulating_water_supply_system',
                        'shutdown_exhaust_fans_of_the_pumping_room_of_the_MNS', 'shutdown_of_exhaust_fans_of_the_pumping_room_PNS',
                        'shutdown_of_exhaust_fans_in_the_centralized_oil_system_room', 'shutdown_of_exhaust_fans_oil_pit_in_the_electrical_room', 
                        'shutdown_of_exhaust_fans_in_the_RD_room', 'shutdown_of_exhaust_fans_in_the_SSVD_room',
                        'shutdown_of_the_roof_fans_of_the_MNS_pump_room', 'shutdown_of_the_roof_fans_of_the_PNS_pump_room',
                        'switching_off_the_supply_fans_pumping_room_of_the_MNS', 'switching_off_the_supply_fans_pumping_room_of_the_PNS',
                        'switch_off_the_supply_fans_in_the_centralized_oil', 'switching_off_the_supply_fan_of_the_RD_room',
                        'switching_off_the_supply_fan_of_the_SSVD_room', 'switching_off_the_supply_fans_of_the_ED_air_compressor',
                        'switching_off_the_supply_fan_of_the_BIK_room', 'switching_off_the_supply_fan_of_the_SIKN_room',
                        'closing_the_air_valves_louvered_grilles_of_the_pump_room', 'closing_of_air_valves_louvered_grilles_of_the_compressor_room',
                        'shutdown_of_electric_oil_heaters', 'shutdown_of_the_electric_heaters_of_the_leakage_collection_MNS',
                        'shutdown_of_the_electric_heaters_of_the_leakage_collection_PNS', 'shutdown_of_electric_heaters_of_the_SIKN_leak_collection_tank',
                        'shutdown_of_air_coolers_of_the_locking_system_MNA', 'shutdown_of_air_coolers_of_the_locking_system_disc_NA',
                        'shutdown_of_the_external_cooling_circuit_ChRP_MNA', 'shutdown_of_the_external_cooling_circuit_ChRP_PNA', 'shutdown_of_locking_system_pumps',
                        'shutdown_of_pumps_for_pumping_oil_oil_products_through',
                        'shutdown_of_pumping_pumps_from_leakage_collection_tanks', 'shutdown_of_anticondensation_electric_heaters_ED', 'fire_protection', 'reserve_aux_15', 
                        'value_ust', 'Pic', 'group_ust', 'rule_map_ust', 'number_list_VU', 'number_protect_VU']
        msg = self.dop_function.column_check(KTPR, 'ktpr', list_default)
        return msg 
class Filling_KTPRA():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_ktpra = []
        count_defence = 0
        with db:
            try:
                for i in range(1, 5):
                    for k in range(1, 97):
                        count_defence += 1
                        list_ktpra.append(dict(id = count_defence,
                                               id_num = k,
                                                variable = f'KTPRA[{i}][{k}]',
                                                tag  = '',
                                                name = f'Резерв',
                                                # NA = '',
                                                # avar_parameter = '',
                                                # stop_type = '',
                                                # AVR = '',
                                                # close_valves = '',
                                                DisableMasking = False,
                                                #value_ust = '',
                                                Pic = '',
                                                group_ust = f'Tm - Агрегатные защиты МНА{i}',
                                                rule_map_ust = 'Временные уставки',
                                                #number_list_VU = ,
                                                #number_protect_VU = '',
                                                number_pump_VU = i
                                                ))
                # Checking for the existence of a database
                KTPRA.insert_many(list_ktpra).execute()
            except Exception:
                msg[f'{today} - Таблица: ktpra, ошибка при заполнении: {traceback.format_exc()}'] = 2
                return msg
        msg[f'{today} - Таблица: ktpra подготовлена'] = 1
        return(msg)
    # Заполняем таблицу KTPRA
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 
                        'NA', 'avar_parameter', 'stop_type', 'AVR', 'close_valves',
                        'DisableMasking', 'value_ust', 'Pic', 
                        'group_ust', 'rule_map_ust', 'number_list_VU', 'number_protect_VU', 'number_pump_VU']
        msg = self.dop_function.column_check(KTPRA, 'ktpra', list_default)
        return msg 
class Filling_KTPRS():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        list_KTPRS = []
        with db:
            for i in range(1, 21):
                list_KTPRS.append(dict(id = i,
                                       variable = f'KTPRS[{i}]',
                                       tag  = '',
                                       name = 'Резерв',
                                       drawdown = '',
                                       reference_to_value = '',
                                       Pic = ''))

            # Checking for the existence of a database
            KTPRS.insert_many(list_KTPRS).execute()

        msg[f'{today} - Таблица: ktprs подготовлена'] = 1
        return(msg)
    # Заполняем таблицу KTPRS
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'drawdown', 'reference_to_value', 'priority_msg_0', 
                        'priority_msg_1', 'prohibition_issuing_msg', 'Pic']
        msg = self.dop_function.column_check(KTPRS, 'ktprs', list_default)
        return msg 
class Filling_GMPNA():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        list_GMPNA = []
        count_defence = 0
        try:
            with db:
                for i in range(1, 5):
                    for k in range(1, 65):
                        count_defence += 1
                        list_GMPNA.append(dict(id = count_defence,
                                               id_num = k,
                                                variable = f'GMPNA[{i}][{k}]',
                                                tag  = '',
                                                name = 'Резерв',
                                                name_for_Chrp_in_local_mode = '',
                                                NA = '',
                                                #used_time_ust = '',
                                                #value_ust = '',
                                                group_ust = f'Tm - Агрегатные готовности МНА{i}',
                                                rule_map_ust = 'Временные уставки',
                                                # number_list_VU = '',
                                                # number_protect_VU = '',
                                                number_pump_VU = i))

                # Checking for the existence of a database
                GMPNA.insert_many(list_GMPNA).execute()
        except Exception:
            msg[f'{today} - Таблица: gmpna, ошибка при заполнении: {traceback.format_exc()}'] = 2
            return msg
        msg[f'{today} - Таблица: gmpna подготовлена'] = 1
        return(msg)
    # Заполняем таблицу GMPNA
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'name_for_Chrp_in_local_mode', 'NA', 'used_time_ust', 'value_ust', 'group_ust', 
                        'rule_map_ust', 'number_list_VU', 'number_protect_VU', 'number_pump_VU']
        msg = self.dop_function.column_check(GMPNA, 'gmpna', list_default)
        return msg 
class Filling_UMPNA():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы AI и DI 
    def getting_modul(self, count_NA):
        msg = {}
        with db:
            try:
                try:
                    if self.dop_function.empty_table('di') or self.dop_function.empty_table('ai'): 
                        msg[f'{today} - Таблицы: ai или di пустые! Заполни таблицы!'] = 2
                        return msg
                except:
                    msg[f'{today} - Таблицы: ai или di отсутсвует!'] = 2
                    return msg

                self.cursor.execute(f'''SELECT Count (*) FROM "umpna"''')
                row_count = self.cursor.fetchall()[0][0]

                for i in range(1, count_NA + 1):

                    if row_count < i:
                        list_UMPNA = []
                        msg[f'{today} - Таблица: umpna, отсутствует NA[{i}] идет заполнение'] = 3

                        vv_included = self.dop_function.search_signal(DI, "di", f"MBC{i}01-1")
                        vv_double_included = self.dop_function.search_signal(DI, "di", f'MBC{i}01-2')
                        vv_disabled = self.dop_function.search_signal(DI, "di", f'MBC{i}02-1')
                        vv_double_disabled = self.dop_function.search_signal(DI, "di", f'MBC{i}02-2')
                        current_greater_than_noload_setting = self.dop_function.search_signal(AI, 'ai', f'CT{i}01')
                        serviceability_of_circuits_of_inclusion_of_VV = self.dop_function.search_signal(DI, "di", f'ECB{i}01')
                        serviceability_of_circuits_of_shutdown_of_VV = self.dop_function.search_signal(DI, "di", f'ECO{i}01-1')
                        serviceability_of_circuits_of_shutdown_of_VV_double = self.dop_function.search_signal(DI, "di", f'ECO{i}01-2')
                        stop_1 = self.dop_function.search_signal(DI, "di", f'KKC{i}01')
                        stop_2 = self.dop_function.search_signal(DI, "di", f'KKC{i}02')
                        monitoring_the_presence_of_voltage_in_the_control_current_circuits = self.dop_function.search_signal(DI, "di", f'EC{i}08')
                        vv_trolley_rolled_out = self.dop_function.search_signal(DI, "di", f'EC{i}04')
                        command_to_turn_on_the_vv_only_for_UMPNA = self.dop_function.search_signal(DO, 'do', f'ABB{i}01')
                        command_to_turn_off_the_vv_output_1 = self.dop_function.search_signal(DO, 'do', f'ABO{i}01-1')
                        command_to_turn_off_the_vv_output_2 = self.dop_function.search_signal(DO, 'do', f'ABO{i}01-2')

                        list_UMPNA.append(dict(
                            id = i,
                            variable = f'NA[{i}]',
                            tag = '',
                            name ='',
                            vv_included = vv_included,
                            vv_double_included = vv_double_included,
                            vv_disabled = vv_disabled,
                            vv_double_disabled = vv_double_disabled,
                            current_greater_than_noload_setting = current_greater_than_noload_setting,
                            serviceability_of_circuits_of_inclusion_of_VV = serviceability_of_circuits_of_inclusion_of_VV,
                            serviceability_of_circuits_of_shutdown_of_VV = serviceability_of_circuits_of_shutdown_of_VV,
                            serviceability_of_circuits_of_shutdown_of_VV_double = serviceability_of_circuits_of_shutdown_of_VV_double,
                            stop_1 = f'NOT {stop_1}',
                            stop_2 = f'NOT {stop_2}',
                            stop_3 ='',
                            stop_4 ='',
                            monitoring_the_presence_of_voltage_in_the_control_current = monitoring_the_presence_of_voltage_in_the_control_current_circuits,
                            voltage_presence_flag_in_the_ZRU_motor_cell ='',
                            vv_trolley_rolled_out = vv_trolley_rolled_out,
                            remote_control_mode_of_the_RZiA_controller ='',
                            availability_of_communication_with_the_RZiA_controller ='',
                            the_state_of_the_causative_agent_of_ED ='',
                            engine_prepurge_end_flag ='',
                            flag_for_the_presence_of_safe_air_boost_pressure_in_the_en ='',
                            flag_for_the_presence_of_safe_air_boost_pressure_in_the_ex ='',
                            engine_purge_valve_closed_flag ='',
                            oil_system_oil_temperature_flag_is_above_10_at_the_cooler_ou ='',
                            flag_for_the_minimum_oil_level_in_the_oil_tank_for_an_indiv ='',
                            flag_for_the_presence_of_the_minimum_level_of_the_barrier ='',
                            generalized_flag_for_the_presence_of_barrier_fluid_pressure ='',
                            command_to_turn_on_the_vv_only_for_UMPNA = command_to_turn_on_the_vv_only_for_UMPNA,
                            command_to_turn_off_the_vv_output_1 = command_to_turn_off_the_vv_output_1,
                            command_to_turn_off_the_vv_output_2 = command_to_turn_off_the_vv_output_2,
                            NA_Chrp ='',
                            type_NA_MNA ='',
                            pump_type_NM ='',
                            parametr_KTPRAS_1 ='',
                            number_of_delay_scans_of_the_analysis_of_the_health_of_the ='',
                            unit_number_of_the_auxiliary_system_start_up_oil_pump ='',
                            NPS_number_1_or_2_which_the_AT_belongs ='',
                            achr_protection_number_in_the_array_of_station_protections ='',
                            saon_protection_number_in_the_array_of_station_protections ='',
                            gmpna_49 ='',
                            gmpna_50 ='',
                            gmpna_51 ='',
                            gmpna_52 ='',
                            gmpna_53 ='',
                            gmpna_54 ='',
                            gmpna_55 ='',
                            gmpna_56 ='',
                            gmpna_57 ='',
                            gmpna_58 ='',
                            gmpna_59 ='',
                            gmpna_60 ='',
                            gmpna_61 ='',
                            gmpna_62 ='',
                            gmpna_63 ='',
                            gmpna_64 ='',
                            Pic ='',
                            tabl_msg = 'TblPumpsUMPNA',
                            replacement_uso_signal_vv_1 ='',
                            replacement_uso_signal_vv_2 =''))
                            
                        # Checking for the existence of a database
                        UMPNA.insert_many(list_UMPNA).execute()
                        msg[f'{today} - Таблица: umpna, NA[{i}] заполнен'] = 1

                    else:

                        msg[f'{today} - Таблица: umpna, NA[{i}] идет обновление'] = 3

                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"MBC{i}01-1"), i, UMPNA.vv_included, 'vv_included'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"MBC{i}01-2"), i, UMPNA.vv_double_included, 'vv_double_included'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"MBC{i}02-1"), i, UMPNA.vv_disabled, 'vv_disabled'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"MBC{i}02-2"), i, UMPNA.vv_double_disabled, 'vv_double_disabled'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(AI, "ai", f"CT{i}01"), i, UMPNA.current_greater_than_noload_setting, 'current_greater_than_noload_setting'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"ECB{i}01"), i, UMPNA.serviceability_of_circuits_of_inclusion_of_VV, 'serviceability_of_circuits_of_inclusion_of_VV'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"ECO{i}01-1"), i, UMPNA.serviceability_of_circuits_of_shutdown_of_VV, 'serviceability_of_circuits_of_shutdown_of_VV'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"ECO{i}01-2"), i, UMPNA.serviceability_of_circuits_of_shutdown_of_VV_double, 'serviceability_of_circuits_of_shutdown_of_VV_double'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            f'NOT {self.dop_function.search_signal(DI, "di", f"KKC{i}01")}', i, UMPNA.stop_1, 'stop_1'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            f'NOT {self.dop_function.search_signal(DI, "di", f"KKC{i}02")}', i, UMPNA.stop_2, 'stop_2'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"EC{i}08"), i, UMPNA.monitoring_the_presence_of_voltage_in_the_control_current, 'monitoring_the_presence_of_voltage_in_the_control_current'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DI, "di", f"EC{i}04"), i, UMPNA.vv_trolley_rolled_out, 'vv_trolley_rolled_out'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DO, 'do', f"ABB{i}01"), i, UMPNA.command_to_turn_on_the_vv_only_for_UMPNA, 'command_to_turn_on_the_vv_only_for_UMPNA'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DO, 'do', f"ABO{i}01-1"), i, UMPNA.command_to_turn_off_the_vv_output_1, 'command_to_turn_off_the_vv_output_1'))
                        msg.update(self.dop_function.update_signal(UMPNA, 'umpna', 
                            self.dop_function.search_signal(DO, 'do', f"ABO{i}01-2"), i, UMPNA.command_to_turn_off_the_vv_output_2, 'command_to_turn_off_the_vv_output_2'))
                        
                        msg[f'{today} - Таблица: umpna, сигналы NA[{i}] обновлены'] = 1
                
                self.cursor.execute(f'''SELECT name FROM "umpna"''')
                for i in self.cursor.fetchall():
                    if i[0] is None or i[0] == '' or i[0] == ' ':
                        msg[f'{today} - Таблица: umpna, необходимо заполнить название НА!'] = 3
            except Exception:
                msg[f'{today} - Таблица: umpna, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: umpna, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу UMPNA
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'vv_included', 'vv_double_included', 'vv_disabled', 
                        'vv_double_disabled', 'current_greater_than_noload_setting', 'serviceability_of_circuits_of_inclusion_of_VV',
                        'serviceability_of_circuits_of_shutdown_of_VV', 'serviceability_of_circuits_of_shutdown_of_VV_double',
                        'stop_1', 'stop_2', 'stop_3', 'stop_4',
                        'monitoring_the_presence_of_voltage_in_the_control_current', 'voltage_presence_flag_in_the_ZRU_motor_cell',
                        'vv_trolley_rolled_out', 'remote_control_mode_of_the_RZiA_controller', 
                        'availability_of_communication_with_the_RZiA_controller','the_state_of_the_causative_agent_of_ED',
                        'engine_prepurge_end_flag', 'flag_for_the_presence_of_safe_air_boost_pressure_in_the_en',
                        'flag_for_the_presence_of_safe_air_boost_pressure_in_the_ex', 'engine_purge_valve_closed_flag',
                        'oil_system_oil_temperature_flag_is_above_10_at_the_cooler_ou', 
                        'flag_for_the_minimum_oil_level_in_the_oil_tank_for_an_indiv', 
                        'flag_for_the_presence_of_the_minimum_level_of_the_barrier',
                        'generalized_flag_for_the_presence_of_barrier_fluid_pressure', 'command_to_turn_on_the_vv_only_for_UMPNA',
                        'command_to_turn_off_the_vv_output_1', 'command_to_turn_off_the_vv_output_2', 'NA_Chrp', 'type_NA_MNA',
                        'pump_type_NM','parametr_KTPRAS_1', 'number_of_delay_scans_of_the_analysis_of_the_health_of_the',
                        'unit_number_of_the_auxiliary_system_start_up_oil_pump', 'NPS_number_1_or_2_which_the_AT_belongs',
                        'achr_protection_number_in_the_array_of_station_protections','saon_protection_number_in_the_array_of_station_protections', 
                        'gmpna_49', 'gmpna_50', 'gmpna_51', 'gmpna_52','gmpna_53', 'gmpna_54', 'gmpna_55', 'gmpna_56',
                        'gmpna_57','gmpna_58', 'gmpna_59', 'gmpna_60', 'gmpna_61', 'gmpna_62','gmpna_63', 'gmpna_64', 'Pic', 'tabl_msg',
                        'replacement_uso_signal_vv_1', 'replacement_uso_signal_vv_2']
        msg = self.dop_function.column_check(UMPNA, 'umpna', list_default)
        return msg 
class Filling_tmNA_UMPNA():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        count_NA = 0
        count_row = 0
        list_tmna_umpna = []
        time_ust = [('Время на проверку корректности состояния цепей контроля ВВ  (с отключенным контролем по току)' , 'T1noCT', '3'), 
                    ('Время на проверку корректности состояния цепей контроля ВВ (с включенным контролем по току)', 'T1CT', '6'),
                    ('Время удержания команды "Включить ВВ", после получения включенного состояния ВВ', 'T2', '1'),
                    ('Максимальное время удержания команды "Отключить ВВ"', 'T3', '1'),
                    ('Время до выдачи команды Включить ВВ, необходимое для отработки САР во время рамповой функции', 'T4', '3'),
                    ('Время на открытие выкидной задвижки перед выдачей команды на включение НА по программе П2', 'T5', '5'),
                    ('Контрольное время выполнения процесса отключения ВВ НА', 'T6', '4'),
                    ('Время на подъем силы тока ЭД после включения ВВ НА', 'T7', '3'),
                    ('Время полного хода выкидной задвижки (используется в алгоритмах программы пуска П2)', 'T8', '600'),
                    ('Время пускового режима работы ЭД при пуске', 'T9', '30'),
                    ('Время перед выдачей команды повторого отключения ВВ при невыполнении программы остановки', 'T10', '3'),
                    ('Контрольное время выполнения процесса включения ВВ НА', 'T11', '4'),
                    ('Время снижения силы тока ЭД после отключения ВВ НА', 'T12', '3'),
                    ('Время фильтрации сигналов цепей включения/отключения', 'T13', '3'),
                    ('Время пускового режима работы насоса при пуске', 'T14', '300'),
                    ('Длительность выдачи команды в САР для рамповой функции', 'T15', '5'),
                    ('Время на выполнение АВР при получении сигнала "Электрозащита" после получения состояния ВВ отключен', 'T16', '3'),
                    ('Время, через которое будет выдана команда "Стоп" при отсутствии электрозащиты', 'T17', '3'),
                    ('Резерв', 'T18', '0'), 
                    ('Размер колеса насосного агрегата', 'WheelSize', '1')] 
        with db:
            try:
                if self.dop_function.empty_table('umpna'): 
                    msg[f'{today} - Таблицы: umpna пустая! Заполни таблицу!'] = 2
                    return msg
                self.cursor.execute(f'''SELECT name FROM umpna''')
                for i in self.cursor.fetchall():
                    count_NA += 1
                    if i[0] is None or i[0] == '' or i[0] == ' ':
                        msg[f'{today} - Таблица: umpna, необходимо заполнить название НА!'] = 3
                    else:
                        for ust in time_ust:
                            count_row += 1
                            list_tmna_umpna.append(dict(id = count_row, 
                                                        variable = f'tmNA_UMPNA[{count_NA}].{ust[1]}',
                                                        tag  = f'HNA{count_NA}_{ust[1]}',
                                                        name = f'{i[0]}. {ust[0]}',
                                                        unit = 'с',
                                                        used = '1',
                                                        value_ust = f'{ust[2]}',
                                                        minimum = '0',
                                                        maximum = '65535',
                                                        group_ust = 'Временные уставки МНА',
                                                        rule_map_ust = 'Временные уставки'))
                        msg[f'{today} - Таблица: umpna_tm, заполнен НА_{count_NA}'] = 1
                            
                # Checking for the existence of a database
                tmNA_UMPNA.insert_many(list_tmna_umpna).execute()
            except Exception:
                msg[f'{today} - Таблица: umpna_tm, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: umpna_tm, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу tmNA_UMPNA
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 
                        'unit', 'used', 'value_ust', 'value_real_ust', 'minimum', 'maximum', 'group_ust', 'rule_map_ust']
        msg = self.dop_function.column_check(tmNA_UMPNA, 'umpna_tm', list_default)
        return msg 
class Filling_tmNA_UMPNA_narab():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы Signals 
    def getting_modul(self):
        msg = {}
        count_NA = 0
        count_row = 0
        list_tmna_umpna = []
        time_ust = [('Время наработки с момента запуска до перехода на резерв' , 'operatingTimeSinceSwitchingOnSet', '36000', 'ч.', 'Наработки. До перехода на резерв'), 
                    ('Время наработки до капитального ремонта', 'operatingTimeBeforeOverhaulSet', '432000', 'ч.', 'Наработки. До капитального ремонта'),
                    ('Время наработки до технического обслуживания (уставка предупредительная)', 'operatingTimeTOSetWarn', '36000', 'ч.', 'Наработки. До ТО (предупредительная)'),
                    ('Время наработки до технического обслуживания (уставка аварийная)', 'operatingTimeTOSet', '432000', 'ч.', 'Наработки. До ТО (аварийная)'),
                    ('Время наработки до среднего технического обслуживания', 'operatingTimeMidTOSet', '36000', 'ч.', 'Наработки. До среднего ТО'),
                    ('Время наработки ЭД до планового текущего ремонта', 'operatingTimeEDSet', '432000', 'ч.', 'Наработки. До планового текущего ремонта'),
                    ('Количество пусков ЭД', 'numOfStartSet', '10', '', 'Наработки. Количество пусков')] 
        with db:
            try:
                if self.dop_function.empty_table('umpna'): 
                    msg[f'{today} - Таблицы: umpna пустая! Заполни таблицу!'] = 2
                    return msg
                self.cursor.execute(f'''SELECT name FROM umpna''')
                for i in self.cursor.fetchall():
                    count_NA += 1
                    if i[0] is None or i[0] == '' or i[0] == ' ':
                        msg[f'{today} - Таблица: umpna, необходимо заполнить название НА!'] = 3
                    else:
                        for ust in time_ust:
                            count_row += 1
                            list_tmna_umpna.append(dict(id = count_row, 
                                                        variable = f'statisticNA[{count_NA}].{ust[1]}',
                                                        tag  = f'SNA{count_NA}_{ust[1]}',
                                                        name = f'{i[0]}. {ust[0]}',
                                                        unit = ust[3],
                                                        used = '1',
                                                        value_ust = f'{ust[2]}',
                                                        minimum = '0',
                                                        maximum = '65535',
                                                        group_ust = ust[4],
                                                        rule_map_ust = 'Наработки'))
                        msg[f'{today} - Таблица: umpna_tm, заполнен НА_{count_NA}'] = 1
                            
                # Checking for the existence of a database
                tmNA_UMPNA_narab.insert_many(list_tmna_umpna).execute()
            except Exception:
                msg[f'{today} - Таблица: umpna_narab_tm, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: umpna_narab_tm, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу tmNA_UMPNA
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 
                        'unit', 'used', 'value_ust', 'minimum', 'maximum', 'group_ust', 'rule_map_ust']
        msg = self.dop_function.column_check(tmNA_UMPNA_narab, 'umpna_narab_tm', list_default)
        return msg 
class Filling_ZD():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы AI и DI 
    def getting_modul(self):
        msg = {}
        array_di_tag_zd = ('OKC', 'CKC', 'ODC', 'CDC', 'MC', 'OPC', 'DCK', 'MCO', 'MCC', 'KKCC', 'KKCS', 'EC', 'OFC', 'CFC')
        array_do_tag_zd = ('DOB', 'DKB', 'DCB', 'DCOB', 'DCCB')
        with db:
            try:
                try:
                    if self.dop_function.empty_table('di') or self.dop_function.empty_table('do'): 
                        msg[f'{today} - Таблицы: di или do пустая! Заполни таблицу!'] = 2
                        return msg
                except:
                    msg[f'{today} - Таблицы: di или do пустая! Заполни таблицу!'] = 2
                    return msg
                
                # Новый список задвижек из таблицы DI
                self.cursor.execute(f"""SELECT name 
                                        FROM di
                                        WHERE name LIKE '%задвижк%' OR name LIKE '%Задвижк%' OR 
                                              name LIKE '%клап%' OR name LIKE '%Клап%' OR
                                              name LIKE '%клоп%' OR name LIKE '%КЛОП%' OR
                                              name LIKE '%кран шар%' OR name LIKE '%Кран шар%'
                                        ORDER BY id""")
                name_zd_new = self.cursor.fetchall()
                list_zd_name_split = []
                for i in name_zd_new: 
                    zd_default = str(i[0]).split(' - ')[0]
                    zd_default = str(zd_default).split('. ')[0]
                    zd_default = zd_default.replace('. Клапан закрыт', '').replace('. клапан закрыт', '').replace('.Клапан закрыт', '').replace('.клапан закрыт', '')
                    zd_default = zd_default.replace('. Клапан открыт', '').replace('. клапан открыт', '').replace('.Клапан открыт', '').replace('.клапан открыт', '')
                    zd_default = zd_default.replace('открыт', '').replace('Открыт', '')
                    zd_default = zd_default.replace('закрыт', '').replace('Закрыт', '')
                    list_zd_name_split.append(zd_default)
                unique_name = set(list_zd_name_split)

                # Существующий список задвижек из таблицы ZD
                self.cursor.execute(f'''SELECT name FROM zd''')
                name_zd_old = self.cursor.fetchall()
                tabl_zd_name = []
                for i in name_zd_old:
                    tabl_zd_name.append(i[0])

                # Количество строк в таблице
                self.cursor.execute(f'''SELECT COUNT(*) FROM zd''')
                count_row = self.cursor.fetchall()[0][0]
                        
                for name in sorted(unique_name):
                    list_zd = []

                    kvo, kvz, mpo, mpz, mufta, error, dist, vmmo, vmmz = '', '', '', '', '', '', '', '', ''
                    close_bru, stop_bru, voltage, isp_opening_chain, isp_closing_chain   = '', '', '', '', ''
                    open_zd, close_zd, stop_zd, open_stop, close_stop = '', '', '', '', ''

                    for tag in array_di_tag_zd:
                        self.cursor.execute(f"""SELECT id, tag_eng, name 
                                                FROM di
                                                WHERE (name LIKE '%{name} %' or name LIKE '%{name}%') AND tag_eng LIKE '%{tag}%'""")
                        
                        try: 
                            number_id = self.cursor.fetchall()[0][0]
                        except: continue

                        if tag == 'OKC':   kvo = f'DI[{number_id}].Value'
                        if tag == 'CKC':   kvz = f'DI[{number_id}].Value'
                        if tag == 'ODC':   mpo = f'DI[{number_id}].Value'
                        if tag == 'CDC':   mpz = f'DI[{number_id}].Value'
                        if tag == 'MC' : mufta = f'DI[{number_id}].Value'
                        if tag == 'OPC': error = f'DI[{number_id}].Value'
                        if tag == 'DCK':  dist = f'DI[{number_id}].Value'
                        if tag == 'MCO':  vmmo = f'DI[{number_id}].Value'
                        if tag == 'MCC':  vmmz = f'DI[{number_id}].Value'
                        if tag == 'KKCC': close_bru = f'DI[{number_id}].Value'
                        if tag == 'KKCS': stop_bru  = f'DI[{number_id}].Value'
                        if tag == 'EC' :  voltage = f'DI[{number_id}].Value'
                        if tag == 'OFC':  isp_opening_chain = f'DI[{number_id}].Value'
                        if tag == 'CFC':  isp_closing_chain = f'DI[{number_id}].Value'

                    for tag in array_do_tag_zd:    
                        self.cursor.execute(f"""SELECT id, tag_eng, name 
                                                FROM "do"
                                                WHERE name LIKE '%{name} %' AND tag_eng LIKE '%{tag}%'""")
                        
                        try   : number_id = self.cursor.fetchall()[0][0]
                        except: continue
                        
                        if tag == 'DOB' : open_zd    = f'ctrlDO[{number_id}]'
                        if tag == 'DKB' : close_zd   = f'ctrlDO[{number_id}]'
                        if tag == 'DCB' : stop_zd    = f'ctrlDO[{number_id}]'
                        if tag == 'DCOB': open_stop  = f'ctrlDO[{number_id}]'
                        if tag == 'DCCB': close_stop = f'ctrlDO[{number_id}]'
                    
                    if kvo == '' and kvz == '': continue

                    if self.dop_function.str_find(str(name).lower, {'клапа'}) or self.dop_function.str_find(str(name).lower, {'клоп'}): klapan = 1
                    else: klapan = 0

                    if name in tabl_zd_name:
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.KVO, 'KVO', kvo))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.KVZ, 'KVZ', kvz))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.MPO, 'MPO', mpo))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.MPZ, 'MPZ', mpz))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Mufta, 'Mufta', mufta))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Drive_failure, 'Drive_failure', error))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Dist, 'Dist', dist))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.VMMO, 'VMMO', vmmo))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.VMMZ, 'VMMZ', vmmz))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Close_BRU, 'Close_BRU', close_bru))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Stop_BRU, 'Stop_BRU', stop_bru))

                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Voltage, 'Voltage', voltage))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Serviceability_opening_circuits, 'Serviceability_opening_circuits', isp_opening_chain))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Serviceability_closening_circuits, 'Serviceability_closening_circuits', isp_closing_chain))

                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Open, 'Open', open_zd))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Close, 'Close', close_zd))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Stop, 'Stop', stop_zd))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Opening_stop, 'Opening_stop', open_stop))
                        msg.update(self.dop_function.update_signal_dop(ZD, "zd", name, ZD.Closeing_stop, 'Closeing_stop', close_stop))
                    else:
                        count_row += 1
        
                        msg[f'{today} - Таблица: zd, добавлена новая задвижка: ZD[{count_row}], {name}'] = 1
                        list_zd.append(dict(id = count_row,
                                            variable = f'ZD[{count_row}]',
                                            tag = '',
                                            name = name,
                                            short_name = '',
                                            exists_interface = 0,
                                            KVO = kvo,KVZ = kvz,MPO = mpo,MPZ = mpz,Dist = dist,Mufta = mufta,Drive_failure = error,
                                            Open = open_zd,Close = close_zd,Stop = stop_zd,Opening_stop = open_stop,Closeing_stop = close_stop,
                                            KVO_i = '',KVZ_i = '',MPO_i = '',MPZ_i = '',Dist_i = '',Mufta_i = '',Drive_failure_i = '',
                                            Open_i = '',Close_i = '',Stop_i = '',Opening_stop_i = '',Closeing_stop_i = '',
                                            No_connection = '',
                                            Close_BRU = close_bru,Stop_BRU = stop_bru,
                                            Voltage = voltage,Voltage_CHSU = '',Voltage_in_signaling_circuits = '',
                                            Serviceability_opening_circuits = isp_opening_chain,Serviceability_closening_circuits = isp_closing_chain,
                                            VMMO = vmmo,VMMZ = vmmz,
                                            Freeze_on_suspicious_change = '',
                                            Is_klapan = klapan,
                                            Opening_percent = '',
                                            Pic = '',Type_BUR_ZD = '', tabl_msg='TblValves',
                                            AlphaHMI = '',AlphaHMI_PIC1 = '',AlphaHMI_PIC1_Number_kont = '',
                                            AlphaHMI_PIC2 = '',AlphaHMI_PIC2_Number_kont = '',AlphaHMI_PIC3 = '',
                                            AlphaHMI_PIC3_Number_kont = '',AlphaHMI_PIC4 = '',AlphaHMI_PIC4_Number_kont = ''))

                        # Checking for the existence of a database
                        ZD.insert_many(list_zd).execute()
                if len(msg) == 0: msg[f'{today} - Таблица: zd, обновление завершено, изменений не обнаружено!'] = 1
                
                # Существование ZD в таблице ZD
                for zd in tabl_zd_name:
                    if zd not in unique_name:
                        msg[f'{today} - Таблица: zd, {zd} не существует в таблице DI'] = 3
            except Exception:
                msg[f'{today} - Таблица: zd, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: zd, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу ZD
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'short_name', 'exists_interface', 'KVO', 'KVZ', 'MPO', 'MPZ', 'Dist',
                        'Mufta', 'Drive_failure', 'Open', 'Close', 'Stop', 'Opening_stop', 'Closeing_stop', 'KVO_i', 'KVZ_i',
                        'MPO_i', 'MPZ_i', 'Dist_i', 'Mufta_i', 'Drive_failure_i', 'Open_i', 'Close_i', 'Stop_i', 'Opening_stop_i',
                        'Closeing_stop_i', 'No_connection', 'Close_BRU', 'Stop_BRU', 'Voltage', 'Voltage_CHSU', 
                        'Voltage_in_signaling_circuits', 'Serviceability_opening_circuits', 'Serviceability_closening_circuits', 'VMMO', 'VMMZ', 
                        'Freeze_on_suspicious_change', 'Is_klapan', 'Opening_percent', 'Pic', 'Type_BUR_ZD', 'tabl_msg',
                        'AlphaHMI', 'AlphaHMI_PIC1', 'AlphaHMI_PIC1_Number_kont', 'AlphaHMI_PIC2',
                        'AlphaHMI_PIC2_Number_kont','AlphaHMI_PIC3', 'AlphaHMI_PIC3_Number_kont', 
                        'AlphaHMI_PIC4', 'AlphaHMI_PIC4_Number_kont']
        msg = self.dop_function.column_check(ZD, 'zd', list_default)
        return msg 
class Filling_ZD_tm():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы ZD 
    def getting_modul(self):
        msg = {}
        count_ZD = 0
        count_row = 0
        list_zd_tm = []
        time_ust = [('Время переходных процессов' , 'T1', '500', 'мс'), 
                    ('Ожидание прихода сигнала от МП после команды на открытие/закрытие', 'T2', '2', 'c'),
                    ('Время схода с КВЗ/КВО', 'T3', '10', 'c'),
                    ('Время хода вала', 'T4', '60', 'c'),
                    ('Время на отключение МП в крайних положениях', 'T5', '2', 'c'),
                    ('Время на возврат напряжения при стопе по месту', 'T6', '3', 'c'),
                    ('Время выполнения команды на открытие при имитации задвижки', 'T7', '20', 'c'),
                    ('Время на подачу команды СТОП', 'T8', '2', 'c'),
                    ('Время на ожидание возврата концевиков после команды стоп', 'T9', '1', 'c'),
                    ('Время на выполнение команд с БРУ', 'T10', '3', 'c'),
                    ('Время на проверку несправности цепей включения', 'T11', '2', 'c'),
                    ('Время на проверку несправности цепей отключения', 'T12', '2', 'c'),
                    ('Время рассогласования между сигналами по физическому и интерфейсному каналу', 'T13', '2', 'c'),
                    ('Время на задержку при подозрительных переходах', 'T14', '2', 'c'),
                    ('Резерв', 'T15', '0', 'c')] 
        with db:
            try:
                if self.dop_function.empty_table('zd'): 
                    msg[f'{today} - Таблицы: zd пустая! Заполни таблицу!'] = 2
                    return msg
                
                self.cursor.execute(f'''SELECT name FROM zd''')
                for i in self.cursor.fetchall():
                    count_ZD += 1
                    for ust in time_ust:
                        count_row += 1
                        used = '0' if ust[0] == 'Резерв' else '1' 
                        list_zd_tm.append(dict(id = count_row, 
                                                variable = f'tmZD[{count_ZD}].{ust[1]}',
                                                tag  = f'HZD{count_ZD}_{ust[1]}',
                                                name = f'{i[0]}. {ust[0]}',
                                                unit = ust[3],
                                                used = used,
                                                value_ust = f'{ust[2]}',
                                                minimum = '0',
                                                maximum = '65535',
                                                group_ust = 'Временные уставки задвижек',
                                                rule_map_ust = 'Временные уставки'))
                            
                # Checking for the existence of a database
                ZD_tm.insert_many(list_zd_tm).execute()
            except Exception:
                msg[f'{today} - Таблица: zd_tm, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: zd_tm, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу zd_tm
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'unit', 'used', 
                        'value_ust', 'minimum', 'maximum', 'group_ust', 'rule_map_ust']
        msg = self.dop_function.column_check(ZD_tm, 'zd_tm', list_default)
        return msg 
class Filling_VS():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы AI и DI 
    def getting_modul(self):
        msg = {}
        array_di_tag_vs = ('MPC', 'МРС',  'ЕC', 'EC')
        array_do_tag_vs = ('ABB', 'АВВ', 'АВО', 'ABO')
        array_tag_opc_vs = ('авар', 'Авар', 'исправн', 'Исправн')

        with db:
            try:
                try:
                    if self.dop_function.empty_table('di') or self.dop_function.empty_table('do'): 
                        msg[f'{today} - Таблицы: di или do пустая! Заполни таблицу!'] = 2
                        return msg
                except:
                    msg[f'{today} - Таблицы: di или do пустая! Заполни таблицу!'] = 2
                    return msg
                
                # Новый список вспомсистем из таблицы DI
                self.cursor.execute(f"""SELECT tag_eng, name 
                                        FROM di
                                        WHERE tag_eng LIKE '%MPC%' or tag_eng LIKE '%МРС%'""")
                vs_name = self.cursor.fetchall()
                list_vs_name_split = []
                for i in vs_name: 
                    if   self.dop_function.str_find(i[1], {'- сигнал от МП'}):
                        list_vs_name_split.append(str(i[1]).split('- сигнал от МП')[0].strip())
                    elif self.dop_function.str_find(i, {'-сигнал от МП'}):
                        list_vs_name_split.append(str(i[1]).split('-сигнал от МП')[0].strip())
                    elif self.dop_function.str_find(i, {'- включен'}):
                        list_vs_name_split.append(str(i[1]).split('- включен')[0].strip())
                    elif self.dop_function.str_find(i, {'-включен'}):
                        list_vs_name_split.append(str(i[1]).split('-включен')[0].strip())
                    elif self.dop_function.str_find(i, {'.Включен'}):
                        list_vs_name_split.append(str(i[1]).split('.Включен')[0].strip())
                    elif self.dop_function.str_find(i, {'. Включен'}):
                        list_vs_name_split.append(str(i[1]).split('. Включен')[0].strip())

                unique_name = set(list_vs_name_split)

                # Существующий список вспомсистем из таблицы VS
                self.cursor.execute(f'''SELECT name FROM vs''')
                exists_vs = self.cursor.fetchall()
                tabl_vs_name = []
                for i in exists_vs:
                    tabl_vs_name.append(i[0])

                # Количество строк в таблице
                self.cursor.execute(f'''SELECT COUNT(*) FROM vs''')
                count_row = self.cursor.fetchall()[0][0]
                        
                for name in sorted(unique_name):
                    list_vs = []
                    mp, voltage, isp_opening_chain, open_vs, close_vs, error = '', '', '', '', '', ''

                    # Принадлежность OPC тега
                    for tag in array_tag_opc_vs:  
                        self.cursor.execute(f"""SELECT id, tag_eng, name 
                                                FROM di
                                                WHERE name LIKE '%{name}%' AND name LIKE '%{tag}%' AND tag_eng LIKE '%OPC%'""")
                        
                        try   : number_id = self.cursor.fetchall()[0][0]
                        except: continue

                        if tag == 'авар': 
                            error = f'DI[{number_id}].Value'
                        elif tag == 'Авар' : 
                            error = f'DI[{number_id}].Value'
                        elif tag == 'исправн' : 
                            isp_opening_chain = f'DI[{number_id}].Value'
                        elif tag == 'Исправн' : 
                            isp_opening_chain = f'DI[{number_id}].Value'

                    for tag in array_di_tag_vs:
                        self.cursor.execute(f"""SELECT id, tag_eng, name
                                                FROM di
                                                WHERE name LIKE '%{name}%' AND tag_eng LIKE '%{tag}%'""")
                        
                        try   : number_id = self.cursor.fetchall()[0][0]
                        except: continue

                        if tag == 'MPC' or tag == 'МРС': mp      = f'DI[{number_id}].Value'
                        if tag == 'EC' or tag == 'ЕC'  : voltage = f'DI[{number_id}].Value'
                        
                    for tag in array_do_tag_vs:    
                        self.cursor.execute(f"""SELECT id, tag_eng, name 
                                                FROM "do"
                                                WHERE name LIKE '%{name}%' AND tag_eng LIKE '%{tag}%'""")
                        
                        try   : number_id = self.cursor.fetchall()[0][0]
                        except: continue
                        
                        if tag == 'ABB' or tag == 'АВВ': open_vs  = f'ctrlDO[{number_id}]'
                        if tag == 'ABO' or tag == 'АВО': close_vs = f'ctrlDO[{number_id}]'

                    # Давление на выходе
                    new_name = str(name).strip()
                    new_name = str(new_name).replace('ой', 'ом')
                    new_name = str(new_name).replace('сос', 'соса')
                    new_name = str(new_name).replace('ой', 'ого')
                    new_name = str(new_name).replace('ый', 'ом')
                    new_name = str(new_name).replace('ор', 'оре')
                    new_name = str(new_name).replace('ель', 'еля')
                    new_name = str(new_name).replace('Нас', 'нас')
                    new_name = str(new_name).replace('Масл', 'масл')
                    new_name = str(new_name).replace('Венти', 'венти')
                    new_name = str(new_name).replace('Погр', 'погр')
                    new_name = str(new_name).replace('Подп', 'подп')
                    new_name = str(new_name).replace('Прит', 'прит')
                    new_name = str(new_name).replace('Вытяж', 'вытяж')

                    self.cursor.execute(f"""SELECT id, "name" 
                                            FROM ai
                                            WHERE "name" LIKE '%{new_name}%' and ("name" LIKE '%Давл%' or  "name" LIKE '%давл%')""")
                    try: 
                        number_id = self.cursor.fetchall()[0][0]
                        pressure_norm = f'AI[{number_id}].Norm'
                        pressure_ndv  = f'AI[{number_id}].Ndv'
                    except:
                        pressure_norm = f''
                        pressure_ndv  = f''

                    if name in tabl_vs_name:
                        msg.update(self.dop_function.update_signal_dop(VS, "vs", name, VS.MP, 'MP', mp))
                        msg.update(self.dop_function.update_signal_dop(VS, "vs", name, VS.Voltage, 'Voltage', voltage))
                        msg.update(self.dop_function.update_signal_dop(VS, "vs", name, VS.Serviceability_of_circuits_of_inclusion, 'Serviceability_of_circuits_of_inclusion', isp_opening_chain))
                        msg.update(self.dop_function.update_signal_dop(VS, "vs", name, VS.External_alarm, 'External_alarm', error))

                        msg.update(self.dop_function.update_signal_dop(VS, "vs", name, VS.VKL, 'VKL', open_vs))
                        msg.update(self.dop_function.update_signal_dop(VS, "vs", name, VS.OTKL, 'OTKL', close_vs))

                        msg.update(self.dop_function.update_signal_dop(VS, "vs", name, VS.Pressure_is_True, 'Pressure_is_True', pressure_norm))
                        msg.update(self.dop_function.update_signal_dop(VS, "vs", name, VS.Pressure_sensor_defective, 'Pressure_sensor_defective', pressure_ndv))

                    else:
                        count_row += 1
                        
                        msg[f'{today} - Таблица: vs, добавлена новая вспомсистема: VS[{count_row}], {name}'] = 1
                        list_vs.append(dict(id = count_row, 
                                            variable = f'VS[{count_row}]',
                                            tag = '',
                                            name = name,
                                            short_name = '',
                                            MP = mp,
                                            Pressure_is_True = pressure_norm,
                                            Voltage = voltage,
                                            Voltage_Sch = '',
                                            Serviceability_of_circuits_of_inclusion = isp_opening_chain,
                                            External_alarm = error,
                                            Pressure_sensor_defective = pressure_ndv,
                                            VKL = open_vs,
                                            OTKL = close_vs,
                                            Not_APV = '0',
                                            Pic = '',
                                            tabl_msg = 'TblAuxSyses',
                                            Is_klapana_interface_auxsystem = '0',
                                            
                                            AlphaHMI = '',AlphaHMI_PIC1 = '',AlphaHMI_PIC1_Number_kont = '',
                                            AlphaHMI_PIC2 = '',AlphaHMI_PIC2_Number_kont = '',AlphaHMI_PIC3 = '',
                                            AlphaHMI_PIC3_Number_kont = '',AlphaHMI_PIC4 = '',AlphaHMI_PIC4_Number_kont = ''))

                        # Checking for the existence of a database
                        VS.insert_many(list_vs).execute()
                if len(msg) == 0: msg[f'{today} - Таблица: vs, обновление завершено, изменений не обнаружено!'] = 1
                
                # Существование вспомсистемы в таблице VS
                for vs in tabl_vs_name:
                    if vs not in unique_name:
                        msg[f'{today} - Таблица: vs, {vs} не существует в таблице DI'] = 3
            except Exception:
                msg[f'{today} - Таблица: vs, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: vs, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу VS
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'short_name', 'group', 'number_in_group', 'MP', 'Pressure_is_True', 'Voltage', 'Voltage_Sch', 
                        'Serviceability_of_circuits_of_inclusion', 'External_alarm', 'Pressure_sensor_defective', 'VKL', 'OTKL', 'Not_APV',
                        'Pic', 'tabl_msg', 'Is_klapana_interface_auxsystem',
                        'AlphaHMI', 'AlphaHMI_PIC1', 'AlphaHMI_PIC1_Number_kont', 'AlphaHMI_PIC2',
                        'AlphaHMI_PIC2_Number_kont','AlphaHMI_PIC3', 'AlphaHMI_PIC3_Number_kont', 
                        'AlphaHMI_PIC4', 'AlphaHMI_PIC4_Number_kont']
        msg = self.dop_function.column_check(VS, 'vs', list_default)
        return msg 
class Filling_VS_tm():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы VS
    def getting_modul(self):
        msg = {}
        count_VS = 0
        count_row = 0
        list_vs_tm = []

        time_ust = [('Выдержка времени на ожидание срабатывания / исчезновения МП после включения / отключения' , 'T1', '2', 'с'), 
                    ('Выдержка времени на ожидание набора давления после появления сигнала МП в процессе пуска агрегата вспомсистемы', 'T2', '10', 'c'),
                    ('Выдержка времени на ожидание спада давления после снятия сигнала МП в процессе остановки агрегата вспомсистемы', 'T3', '5', 'c'),
                    ('Выдержка времени на возврат напряжения при стопе по месту', 'T4', '3', 'c'),
                    ('Выдержка времени на контроль давления во время работы', 'T5', '5', 'c'),
                    ('Выдержка времени для перевода неработающего агрегата вспомсистемы в режим ремонтный при исчезновении напряжения в схеме управления', 'T6', '40', 'c'),
                    ('Выдержка времени на запаздывание сигналов исчезновения МП и сигнала наличия напряжения от СШ (при кратковременных исчезновениях напряжения на секции шин)', 'T7', '0', 'c'),
                    ('Выдержка времени на перевод пожарного насоса в ремонтный режим при неисправности цепей включения', 'T8', '40', 'c')] 
        with db:
            try:
                if self.dop_function.empty_table('vs'): 
                    msg[f'{today} - Таблицы: vs пустая! Заполни таблицу!'] = 2
                    return msg
                
                self.cursor.execute(f'''SELECT name FROM vs''')
                for i in self.cursor.fetchall():
                    count_VS += 1
                    for ust in time_ust:
                        count_row += 1
                        used = '0' if ust[0] == 'Резерв' else '1' 
                        list_vs_tm.append(dict(id = count_row, 
                                                variable = f'tmVS[{count_VS}].{ust[1]}',
                                                tag  = f'HVS{count_VS}_{ust[1]}',
                                                name = f'{i[0]}. {ust[0]}',
                                                unit = ust[3],
                                                used = used,
                                                value_ust = f'{ust[2]}',
                                                minimum = '0',
                                                maximum = '65535',
                                                group_ust = 'Временные уставки вспомсистем',
                                                rule_map_ust = 'Временные уставки'))
                            
                # Checking for the existence of a database
                VS_tm.insert_many(list_vs_tm).execute()
            except Exception:
                msg[f'{today} - Таблица: vs_tm, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: vs_tm, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу vs_tm
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'unit', 'used', 
                        'value_ust', 'minimum', 'maximum', 'group_ust', 'rule_map_ust']
        msg = self.dop_function.column_check(VS_tm, 'vs_tm', list_default)
        return msg 
class Filling_VSGRP():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Заполняем таблицу VSGRP
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'fire_or_watering', 'count_auxsys_in_group', 'Number_of_auxsystem_in_group',
                        'WarnOff_flag_if_one_auxsystem_in_the_group_is_running']
        msg = self.dop_function.column_check(VSGRP, 'vsgrp', list_default)
        msg[f'{today} - Таблица: vs_grp подготовлена'] = 1
        return msg 
class Filling_VSGRP_tm():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы VS
    def getting_modul(self):
        msg = {}
        count_VSGRP = 0
        count_row = 0
        list_vsgrp_tm = []

        time_ust = [('Выдержка времени на выполнение АПВ' , 'T1', '1', 'с')] 
        with db:
            if self.dop_function.empty_table('vs'): 
                msg[f'{today} - Таблицы: VSGRP пустая! Заполни таблицу!'] = 2
                return msg
            
            self.cursor.execute(f'''SELECT name FROM vsgrp''')
            for i in self.cursor.fetchall():
                count_VSGRP += 1
                for ust in time_ust:
                    count_row += 1
                    used = '0' if ust[0] == 'Резерв' else '1' 
                    list_vsgrp_tm.append(dict(id = count_row, 
                                              variable = '',
                                              tag  = f'HVSGRP{count_VSGRP}_{ust[1]}',
                                              name = f'{i[0]}. {ust[0]}',
                                              unit = ust[3],
                                              used = used,
                                              value_ust = f'{ust[2]}',
                                              minimum = '0',
                                              maximum = '65535',
                                              group_ust = 'Временные уставки групп вспомсистем',
                                              rule_map_ust = 'Временные уставки'))
                        
            # Checking for the existence of a database
            VSGRP_tm.insert_many(list_vsgrp_tm).execute()
        msg[f'{today} - Таблица: vsgrp_tm, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу vsgrp_tm
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'unit', 'used', 
                        'value_ust', 'minimum', 'maximum', 'group_ust', 'rule_map_ust']
        msg = self.dop_function.column_check(VSGRP_tm, 'vsgrp_tm', list_default)
        return msg 
class Filling_UTS():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы AI и DI 
    def getting_modul(self, bool_uts_upts):
        if bool_uts_upts: 
            tabl_used  = 'upts'
            model_used = UPTS
            variable   = 'UPTS' 
        else:
            tabl_used  = 'uts'
            model_used = UTS
            variable   = 'UTS' 

        msg = {}
        list_uts = []
        with db:
            try:
                try:
                    if self.dop_function.empty_table('do'): 
                        msg[f'{today} - Таблица: do пустая! Заполни таблицу!'] = 2
                        return msg
                except:
                    msg[f'{today} - Таблица: do пустая! Заполни таблицу!'] = 2
                    return msg
                
                # Новый список из таблицы DI
                self.cursor.execute(f"""SELECT id, tag, name, uso, basket, module, channel
                                        FROM "do"
                                        WHERE (name LIKE '%Табл%' AND tag LIKE '%BB%')  OR
                                              (name LIKE '%Сирен%' AND tag LIKE '%BB%') OR
                                              (name LIKE '%Звон%' AND tag LIKE '%BB%')  OR
                                              (name LIKE '%табл%' AND tag LIKE '%BB%')  OR
                                              (name LIKE '%сирен%' AND tag LIKE '%BB%') OR
                                              (name LIKE '%звон%' AND tag LIKE '%BB%') OR
                                              (name LIKE '%ОТВ%' OR name LIKE '%отв%') OR
                                              (name LIKE '%сигнализац%')
                                        ORDER BY tag""")
                list_uts_do = self.cursor.fetchall()
                # Количество строк в таблице
                self.cursor.execute(f"""SELECT COUNT(*) FROM {tabl_used}""")
                count_row = self.cursor.fetchall()[0][0]

                for uts_do in list_uts_do:
                    name_uts = str(uts_do[2]).replace('Включение звонка Авария', 'Звонок Авария')
                    name_uts = str(name_uts).replace('Включение звонка "Авария"', 'Звонок "Авария"')
                    name_uts = str(name_uts).replace('Включение звонка авария', 'Звонок авария')
                    name_uts = str(name_uts).replace('Включение звонка "авария"', 'Звонок "авария"')
                    name_uts = str(name_uts).replace(' - включить', '')
                    name_uts = str(name_uts).replace('-включить', '')
                    name_uts = str(name_uts).replace('Включение сирены', 'Сирены')
                    name_uts = str(name_uts).replace('Включение табло', 'Табло')
                    name_uts = str(name_uts).replace('Включение/отключение сирены', 'Сирены')
                    name_uts = str(name_uts).replace('Включение/отключение табло', 'Табло')
                    name_uts = str(name_uts).replace('Включение звуковой сигнализации', 'Звуковая сигнализация')

                    coincidence = model_used.select().where(model_used.uso     == uts_do[3],
                                                            model_used.basket  == uts_do[4],
                                                            model_used.module  == uts_do[5],
                                                            model_used.channel == uts_do[6])
                    if bool(coincidence):
                        exist_vkl  = model_used.select().where(model_used.VKL  == f'ctrlDO[{uts_do[0]}]')
                        exist_tag  = model_used.select().where(model_used.tag  == uts_do[1])
                        exist_name = model_used.select().where(model_used.name == name_uts)

                        if not bool(exist_vkl):
                            self.cursor.execute(f"""SELECT id, tag 
                                                    FROM "do"
                                                    WHERE uso='{uts_do[3]}' AND 
                                                          basket={uts_do[4]} AND 
                                                          module={uts_do[5]} AND 
                                                          channel={uts_do[6]}""")
                            for id_, vkl_ in self.cursor.fetchall():
                                msg[f'{today} - Таблица: {tabl_used}, у сигнала в таблице do обновлена команда включить: id = {id_}, ({vkl_}) ctrlDO[{uts_do[0]}]'] = 3
                            self.cursor.execute(f"""UPDATE {tabl_used}
                                                    SET "VKL"='ctrlDO[{uts_do[0]}]'
                                                    WHERE uso='{uts_do[3]}' AND 
                                                          basket={uts_do[4]} AND 
                                                          module={uts_do[5]} AND 
                                                          channel={uts_do[6]}""")

                        if not bool(exist_tag):
                            self.cursor.execute(f"""SELECT id, tag 
                                                    FROM "do"
                                                    WHERE uso='{uts_do[3]}' AND 
                                                          basket={uts_do[4]} AND 
                                                          module={uts_do[5]} AND 
                                                          channel={uts_do[6]}""")
                            for id_, tag_ in self.cursor.fetchall():
                                msg[f'{today} - Таблица: {tabl_used}, у сигнала в таблице do обновлен tag: id = {id_}, ({tag_}) {uts_do[1]}'] = 3
                            self.cursor.execute(f"""UPDATE {tabl_used}
                                                    SET "tag"='{uts_do[1]}' 
                                                    WHERE uso='{uts_do[3]}' AND 
                                                          basket={uts_do[4]} AND 
                                                          module={uts_do[5]} AND 
                                                          channel={uts_do[6]}""")

                        if not bool(exist_name):
                            self.cursor.execute(f"""SELECT id, name 
                                                    FROM "do"
                                                    WHERE uso='{uts_do[3]}' AND 
                                                          basket={uts_do[4]} AND 
                                                          module={uts_do[5]} AND 
                                                          channel={uts_do[6]}""")
                            for id_, name_ in self.cursor.fetchall():
                                msg[f'{today} - Таблица: {tabl_used}, у сигнала в таблице do обновлено name: id = {id_}, ({uts_do[1]}), {name_uts}'] = 3
                            self.cursor.execute(f"""UPDATE {tabl_used}
                                                    SET "name"='{name_uts}' 
                                                    WHERE uso='{uts_do[3]}' AND 
                                                          basket={uts_do[4]} AND 
                                                          module={uts_do[5]} AND 
                                                          channel={uts_do[6]}""")
                        continue
                    count_row += 1
                    msg[f'{today} - Таблица: {tabl_used}, добавлен новый сигнал: id = {uts_do[0]}, ({uts_do[1]}), {name_uts}'] = 1
                    siren = 1 if self.dop_function.str_find(str(name_uts).lower(), {'сирен'}) else 0
                    list_uts.append(dict(id = count_row,
                                         variable = f'{variable}[{count_row}]',
                                         tag = f'{uts_do[1]}',
                                         name = f'{name_uts}',
                                         short_name = '',
                                         location = '',
                                         VKL = f'ctrlDO[{uts_do[0]}]',
                                         siren = siren, 
                                         Does_not_require_autoshutdown = '0', 
                                         Serviceability_of_circuits_of_inclusion = '',
                                         Examination = '', 
                                         Kvit = '',
                                         Pic = '',
                                         number_list_VU = '',
                                         order_number_for_VU = '', 
                                         uso = f'{uts_do[3]}', 
                                         basket =  uts_do[4], 
                                         module =  uts_do[5], 
                                         channel = uts_do[6]))

                # Checking for the existence of a database
                model_used.insert_many(list_uts).execute()
            except Exception:
                msg[f'{today} - Таблица: {tabl_used}, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: {tabl_used}, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу UTS
    def column_check(self, bool_uts_upts):
        list_default = ['variable', 'tag', 'name', 'short_name', 'location', 'VKL', 'Serviceability_of_circuits_of_inclusion', 'siren', 'Does_not_require_autoshutdown', 'Examination',
                        'Kvit', 'Pic', 'number_list_VU', 'order_number_for_VU', 'uso', 'basket', 'module', 'channel']
        if bool_uts_upts:
            msg = self.dop_function.column_check(UPTS, 'upts', list_default)
        else:
            msg = self.dop_function.column_check(UTS, 'uts', list_default)
        return msg 
class Filling_UTS_tm():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы UTS_tm
    def getting_modul(self):
        msg = {}
        count_UTS = 0
        count_row = 0
        list_uts_tm = []

        time_ust = [('Время непрерывной работы' , 'T1', '1', 'с'), 
                    ('Время паузы работы', 'T2', '1', 'c')] 
        with db:
            try:
                if self.dop_function.empty_table('uts'): 
                    msg[f'{today} - Таблицы: uts пустая! Заполни таблицу!'] = 2
                    return msg
                
                self.cursor.execute(f'''SELECT name FROM uts''')
                for i in self.cursor.fetchall():
                    count_UTS += 1
                    for ust in time_ust:
                        count_row += 1
                        used = '0' if ust[0] == 'Резерв' else '1' 
                        list_uts_tm.append(dict(id = count_row, 
                                                variable = '',
                                                tag  = f'HUTS[{count_UTS}]_{ust[1]}',
                                                name = f'{i[0]}. {ust[0]}',
                                                unit = ust[3],
                                                used = used,
                                                value_ust = f'{ust[2]}',
                                                minimum = '0',
                                                maximum = '65535',
                                                group_ust = 'Временные уставки сирен и табло',
                                                rule_map_ust = 'Временные уставки'))
                            
                # Checking for the existence of a database
                UTS_tm.insert_many(list_uts_tm).execute()
            except Exception:
                msg[f'{today} - Таблица: uts_tm, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: uts_tm, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу uts_tm
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'unit', 'used', 
                        'value_ust', 'minimum', 'maximum', 'group_ust', 'rule_map_ust']
        msg = self.dop_function.column_check(UTS_tm, 'uts_tm', list_default)
        return msg 
class Filling_VV():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы DI 
    def getting_modul(self):
        msg = {}
        list_vv = []
        with db:
            try:
                try:
                    if self.dop_function.empty_table('di'): 
                        msg[f'{today} - Таблица: di пустая! Заполни таблицу!'] = 2
                        return msg
                except:
                    msg[f'{today} - Таблица: di пустая! Заполни таблицу!'] = 2
                    return msg
                
                # Cписок ВВ из таблицы DI
                self.cursor.execute(f'''SELECT id, tag_eng, name
                                        FROM di
                                        WHERE (name LIKE '%ввода%' AND (tag_eng LIKE '%MBC%' OR tag_eng LIKE '%EC%')) OR
                                              (name LIKE '%СВВ%' AND (tag_eng LIKE '%MBC%' OR tag_eng LIKE '%EC%')) OR
                                              (name LIKE '%ССВ%' AND (tag_eng LIKE '%MBC%' OR tag_eng LIKE '%EC%'))''')
                list_vv_di = self.cursor.fetchall()

                # Существующий список из таблицы VV
                self.cursor.execute(f'''SELECT name FROM vv''')
                name_vv_old = self.cursor.fetchall()
                tabl_vv_name = []
                for i in name_vv_old:
                    tabl_vv_name.append(i[0])

                # Количество строк в таблице
                self.cursor.execute(f'''SELECT COUNT(*) FROM vv''')
                count_row = self.cursor.fetchall()[0][0]
                
                # Короткое имя
                list_name_vv = []
                for vv_di in list_vv_di:
                    name_vv = vv_di[2]

                    if self.dop_function.str_find(name_vv, {'включ'}) : name = str(name_vv).replace('включен', '')
                    if self.dop_function.str_find(name_vv, {'отключ'}): name = str(name_vv).replace('отключен', '')
                    
                    try   : list_name_vv.append(str(name).split('.')[1].strip())
                    except: list_name_vv.append(str(name))
                set_name_vv = set(list_name_vv)

                for set_name in sorted(set_name_vv):
                    vkl_vv  = ''
                    otkl_vv = ''
                    self.cursor.execute(f"""SELECT id, name
                                            FROM di
                                            WHERE name LIKE '%{set_name}%'""")
                    list_vv_signals = self.cursor.fetchall()
                    for signal in list_vv_signals:
                        id_vv   = signal[0]
                        name_vv = signal[1]
                    
                        if self.dop_function.str_find(name_vv, {'включ'}) : vkl_vv  = f'DI[{id_vv}].Value'
                        if self.dop_function.str_find(name_vv, {'отключ'}): otkl_vv = f'DI[{id_vv}].Value'

                    if set_name in tabl_vv_name:
                        msg.update(self.dop_function.update_signal_dop(VV, "vv", set_name, VV.VV_vkl, 'VV_vkl', vkl_vv))
                        msg.update(self.dop_function.update_signal_dop(VV, "vv", set_name, VV.otkl_vv, 'otkl_vv', otkl_vv))
                    else:
                        msg[f'{today} - Таблица: vv, добавлен новый сигнал: id = {id_vv}, {name_vv}'] = 3
                        count_row += 1
                        list_vv.append(dict(id = count_row, 
                                            variable = f'VV[{count_row}]',
                                            name = set_name,
                                            VV_vkl  = vkl_vv,
                                            VV_otkl = otkl_vv,
                                            Pic = ''))

                # Checking for the existence of a database
                VV.insert_many(list_vv).execute()
                msg[f'{today} - Таблица: vv заполнена'] = 1
            except Exception:
                 msg[f'{today} - Таблица: vv, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: vv, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу VV
    def column_check(self):
        list_default = ['variable', 'name', 'VV_vkl', 'VV_otkl', 'Pic']
        msg = self.dop_function.column_check(VV, 'vv', list_default)
        return msg 
class Filling_PI():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы AI и DI 
    def getting_modul(self):
        msg = {}
        list_pi = []
        with db:
            try:
                try:
                    if self.dop_function.empty_table('ai'): 
                        msg[f'{today} - Таблица: ai пустая! Заполни таблицу!'] = 2
                        return msg
                except:
                    msg[f'{today} - Таблица: ai пустая! Заполни таблицу!'] = 2
                    return msg
                
                # Новый список из таблицы DI
                self.cursor.execute(f"""SELECT id, "tag", "name"
                                        FROM ai
                                        WHERE ("name" LIKE '%адрес%' AND "name" LIKE '%пусков%') OR
                                              ("name" LIKE '%пожар%' AND "name" LIKE '%дымов%')  OR
                                              ("name" LIKE '%теплов%') 
                                        ORDER BY "tag" """)
                list_pi_ai = self.cursor.fetchall()

                # Существующий список из таблицы PI
                self.cursor.execute(f'''SELECT name FROM pi''')
                name_pi_old = self.cursor.fetchall()
                tabl_pi_name = []
                for i in name_pi_old:
                    tabl_pi_name.append(i[0])

                # Количество строк в таблице
                self.cursor.execute(f'''SELECT COUNT(*) FROM pi''')
                count_row = self.cursor.fetchall()[0][0]

                for new_list_pi in list_pi_ai:
                    number_ai = new_list_pi[0]
                    tag_ai    = new_list_pi[1]
                    name_ai   = new_list_pi[2]

                    # Type PI
                    if self.dop_function.str_find(name_ai, {'адресн'}) : type_pi  = '4'
                    elif self.dop_function.str_find(name_ai, {'дымов'}): type_pi  = '3'
                    elif self.dop_function.str_find(name_ai, {'теплов'}): type_pi  = '5'
                    else: type_pi = ''
                    # Attention
                    if self.dop_function.str_find(name_ai, {'шле'}) or self.dop_function.str_find(name_ai, {'шлейф'}): 
                        attention  = f'stateAI[{number_ai}].state.reg'
                    else: 
                        attention = ''
                    # Reset
                    try:
                        self.cursor.execute(f"""SELECT id, tag
                                                FROM "do"
                                                WHERE tag LIKE '%{tag_ai}%'""")
                        list_pi_do = self.cursor.fetchall()
                        ctrl_DO = f'ctrlDO[{list_pi_do[0][0]}]'
                    except Exception:
                        ctrl_DO = ''

                    fire_0  = f'stateAI[{number_ai}].state.reg'
                    fault_1 = f'stateAI[{number_ai}].state.reg'
                    fault_2 = f'stateAI[{number_ai}].state.reg'

                    if name_ai in tabl_pi_name:
                        msg.update(self.dop_function.update_signal_dop(PI, "pi", name_ai, PI.tag, 'tag', tag_ai))
                        msg.update(self.dop_function.update_signal_dop(PI, "pi", name_ai, PI.name, 'name', name_ai))
                        msg.update(self.dop_function.update_signal_dop(PI, "pi", name_ai, PI.Type_PI, 'Type_PI', type_pi))
                        msg.update(self.dop_function.update_signal_dop(PI, "pi", name_ai, PI.Fire_0, 'Fire_0', fire_0))
                        msg.update(self.dop_function.update_signal_dop(PI, "pi", name_ai, PI.Attention_1, 'Attention_1', attention))
                        msg.update(self.dop_function.update_signal_dop(PI, "pi", name_ai, PI.Fault_1_glass_pollution_broken_2, 'Fault_1_glass_pollution_broken_2', fault_1))
                        msg.update(self.dop_function.update_signal_dop(PI, "pi", name_ai, PI.Fault_2_fault_KZ_3, 'Fault_2_fault_KZ_3', fault_2))
                        msg.update(self.dop_function.update_signal_dop(PI, "pi", name_ai, PI.Reset_Link, 'Reset_Link', ctrl_DO))
                    else:
                        msg[f'{today} - Таблица: pi, добавлен новый сигнал: id = {number_ai}, {name_ai}'] = 3
                        count_row += 1
                        list_pi.append(dict(id = count_row, 
                                            variable = f'PI[{count_row}]',
                                            tag = tag_ai,
                                            name = name_ai,
                                            Type_PI = type_pi,
                                            Fire_0 = f'stateAI[{number_ai}].state.reg',
                                            Attention_1 = attention,
                                            Fault_1_glass_pollution_broken_2 = f'stateAI[{number_ai}].state.reg',
                                            Fault_2_fault_KZ_3 = f'stateAI[{number_ai}].state.reg',
                                            Yes_connection_4 = '',
                                            Frequency_generator_failure_5 = '',
                                            Parameter_loading_error_6 = '',
                                            Communication_error_module_IPP_7 = '',
                                            Supply_voltage_fault_8 = '',
                                            Optics_contamination_9 = '',
                                            IK_channel_failure_10 = '',
                                            UF_channel_failure_11 = '',
                                            Loading_12 = '',
                                            Test_13 = '',
                                            Reserve_14 = '',
                                            Reset_Link = ctrl_DO,
                                            Reset_Request = '0',
                                            Through_loop_number_for_interface = '0',
                                            location = '',
                                            Pic = '',
                                            Normal = ''))

                # Checking for the existence of a database
                PI.insert_many(list_pi).execute()
                if len(msg) == 0: msg[f'{today} - Таблица: pi, обновление завершено, изменений не обнаружено!'] = 1
            except Exception:
                msg[f'{today} - Таблица: pi, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: pi, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу VS
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'Type_PI', 'Fire_0', 'Attention_1', 'Fault_1_glass_pollution_broken_2', 
                        'Fault_2_fault_KZ_3', 'Yes_connection_4', 'Frequency_generator_failure_5', 
                        'Parameter_loading_error_6', 'Communication_error_module_IPP_7', 'Supply_voltage_fault_8', 'Optics_contamination_9',
                        'IK_channel_failure_10', 'UF_channel_failure_11', 'Loading_12', 'Test_13', 'Reserve_14',
                        'Reset_Link', 'Reset_Request', 'Through_loop_number_for_interface', 'location', 'Pic','Normal']
        msg = self.dop_function.column_check(PI, 'pi', list_default)
        return msg 
class Filling_PZ_tm():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Получаем данные с таблицы PZ_tm
    def getting_modul(self):
        msg = {}
        count_PZ = 0
        count_row = 0
        list_pz_tm = []

        time_ust = [('Задержка атаки' , 'T1', '10', 'с'), 
                    ('Задержка на возникновение запроса об остановке тушения', 'T2', '30', 'c'),
                    ('Длительность атаки', 'T3', '60', 'c'),
                    ('Контроль процесса пуска тушения', 'T4', '20', 'c'),
                    ('Контроль процесса останова тушения', 'T5', '20', 'c'),
                    ('Инерционность системы', 'T6', '50', 'c'),
                    ('Задержка включения насосов с момента окончания задержки атаки', 'T7', '0', 'c'),
                    ('Выдержка времения на включение следующего насоса при включении нескольких насосов', 'T8', '10', 'c'),
                    ('Задержка открытия задвижек с момента окончания задержки атаки', 'T9', '10', 'c'),] 
        with db:
            try:                
                try:
                    if self.dop_function.empty_table('pz'): 
                        msg[f'{today} - Таблица: pz пустая или не существует! Заполни таблицу!'] = 2
                        return msg
                except:
                    msg[f'{today} - Таблица: pz пустая или не существует!'] = 2
                    return msg
                
                self.cursor.execute(f'''SELECT name FROM pz''')
                for i in self.cursor.fetchall():
                    count_PZ += 1
                    count = 0
                    for ust in time_ust:
                        count_row += 1
                        count += 1
                        used = '0' if ust[0] == 'Резерв' else '1' 
                        list_pz_tm.append(dict(id = count_row, 
                                                variable = f'HPZ[{count}].{count_PZ}',
                                                tag  = f'HUTS[{count_PZ}]_{ust[1]}',
                                                name = f'{i[0]}. {ust[0]}',
                                                unit = ust[3],
                                                used = used,
                                                value_ust = f'{ust[2]}',
                                                minimum = '0',
                                                maximum = '65535',
                                                group_ust = 'Временные уставки пожарных зон',
                                                rule_map_ust = 'Временные уставки'))
                            
                # Checking for the existence of a database
                PZ_tm.insert_many(list_pz_tm).execute()
            except Exception:
                msg[f'{today} - Таблица: pz_tm, ошибка при заполнении: {traceback.format_exc()}'] = 2
            msg[f'{today} - Таблица: pz_tm, выполнение кода завершено!'] = 1
        return(msg)
    # Заполняем таблицу pz_tm
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'unit', 'used', 
                        'value_ust', 'minimum', 'maximum', 'group_ust', 'rule_map_ust']
        msg = self.dop_function.column_check(PZ_tm, 'pz_tm', list_default)
        return msg 
class Filling_DPS():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'control', 'deblock', 
                        'actuation', 'actuation_transmitter', 'malfunction', 'voltage']
        msg = self.dop_function.column_check(DPS, 'dps', list_default)
        return msg 
class Filling_TM_DP():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    # Заполняем таблицу pz_tm
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'link_to_link_signal', 'link_to_timeout', 'Pic']
        msg = self.dop_function.column_check(TM_DP, 'tm_dp', list_default)
        return msg 
class Filling_TM_TS():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_tmts = []
        with db:
            try:
                for i in range(1, 2544):
                    list_tmts.append(dict(id = i,
                                          variable = f'TM_TS[{i}]',
                                          tag = '',
                                          name  = '',
                                          function_ASDU = 'M_SP_TB_1 (30)',
                                          addr_object = 4095 + i,
                                          link_value = ''))
                    
                # Checking for the existence of a database
                TM_TS.insert_many(list_tmts).execute()
            except Exception:
                msg[f'{today} - Таблица: tm_ts, ошибка при заполнении: {traceback.format_exc()}'] = 2
                return msg
        msg[f'{today} - Таблица: tm_ts подготовлена'] = 1
        return(msg)
    # Заполняем таблицу TM_TS
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'function_ASDU', 'addr_object', 'link_value']
        msg = self.dop_function.column_check(TM_TS, 'tm_ts', list_default)
        return msg 
class Filling_TM_TI4():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_tmti4 = []
        with db:
            try:
                for i in range(1, 108):
                    list_tmti4.append(dict(id = i,
                                          variable = f'TM_TI4[{i}]',
                                          tag = '',
                                          name  = '',
                                          function_ASDU = 'M_ME_TF_1 (36)',
                                          addr_object = 16433 + i))
                    
                # Checking for the existence of a database
                TM_TI4.insert_many(list_tmti4).execute()
            except Exception:
                msg[f'{today} - Таблица: tm_ti4, ошибка при заполнении: {traceback.format_exc()}'] = 2
                return msg
        msg[f'{today} - Таблица: tm_ti4 подготовлена'] = 1
        return(msg)
    # Заполняем таблицу TM_TI4
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'function_ASDU', 'addr_object', 'variable_value', 'variable_status', 'variable_Aiparam']
        msg = self.dop_function.column_check(TM_TI4, 'tm_ti4', list_default)
        return msg 
class Filling_TM_TI2():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_tmti2 = []
        with db:
            try:
                for i in range(1, 50):
                    list_tmti2.append(dict(id = i,
                                          variable = f'TM_TI2[{i}]',
                                          tag = '',
                                          name  = '',
                                          function_ASDU = 'M_ME_TE_1 (35)',
                                          addr_object = 16383 + i))
                    
                # Checking for the existence of a database
                TM_TI2.insert_many(list_tmti2).execute()
            except Exception:
                msg[f'{today} - Таблица: tm_ti2, ошибка при заполнении: {traceback.format_exc()}'] = 2
                return msg
        msg[f'{today} - Таблица: tm_ti2 подготовлена'] = 1
        return(msg)
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'function_ASDU', 'addr_object', 'variable_value', 'variable_status']
        msg = self.dop_function.column_check(TM_TI2, 'tm_ti2', list_default)
        return msg 
class Filling_TM_TII():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_tmtii = []
        with db:
            try:
                for i in range(1, 54):
                    list_tmtii.append(dict(id = i,
                                          variable = f'TM_TII[{i}]',
                                          tag = '',
                                          name  = '',
                                          function_ASDU = 'M_BO_TB_1 (33)',
                                          addr_object = 40959 + i))
                    
                # Checking for the existence of a database
                TM_TII.insert_many(list_tmtii).execute()
            except Exception:
                msg[f'{today} - Таблица: tm_tii, ошибка при заполнении: {traceback.format_exc()}'] = 2
                return msg
        msg[f'{today} - Таблица: tm_tii подготовлена'] = 1
        return(msg)
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'function_ASDU', 'addr_object', 'variable_value', 'variable_status']
        msg = self.dop_function.column_check(TM_TII, 'tm_tii', list_default)
        return msg 
class Filling_TM_TU():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_tmtu = []
        with db:
            try:
                for i in range(1, 240):
                    list_tmtu.append(dict(id = i,
                                          variable = f'TM_TU[{i}]',
                                          tag = '',
                                          name  = '',
                                          function_ASDU = 'C_SC_NA_1 (45)',
                                          addr_object = 24797 + i))
                    
                # Checking for the existence of a database
                TM_TU.insert_many(list_tmtu).execute()
            except Exception:
                msg[f'{today} - Таблица: tm_tu, ошибка при заполнении: {traceback.format_exc()}'] = 2
                return msg
        msg[f'{today} - Таблица: tm_tu подготовлена'] = 1
        return(msg)
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'function_ASDU', 'addr_object', 'variable_change', 'change_bit', 'descriptionTU']
        msg = self.dop_function.column_check(TM_TU, 'tm_tu', list_default)
        return msg 
class Filling_TM_TR4():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_tmtr4 = []
        with db:
            try:
                for i in range(1, 10):
                    list_tmtr4.append(dict(id = i,
                                          variable = f'TM_TR4[{i}]',
                                          tag = '',
                                          name  = '',
                                          function_ASDU = 'C_SE_NC_1 (50)',
                                          addr_object = 32767 + i))
                    
                # Checking for the existence of a database
                TM_TR4.insert_many(list_tmtr4).execute()
            except Exception:
                msg[f'{today} - Таблица: tm_tr4, ошибка при заполнении: {traceback.format_exc()}'] = 2
                return msg
        msg[f'{today} - Таблица: tm_tr4 подготовлена'] = 1
        return(msg)
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'function_ASDU', 'addr_object', 'variable_change', 'descriptionTR4']
        msg = self.dop_function.column_check(TM_TR4, 'tm_tr4', list_default)
        return msg 
class Filling_TM_TR2():
    def __init__(self):
        self.cursor   = db.cursor()
        self.dop_function = General_functions()
    def getting_modul(self):
        msg = {}
        list_tmtr2 = []
        with db:
            try:
                for i in range(1, 10):
                    list_tmtr2.append(dict(id = i,
                                          variable = f'TM_TR2[{i}]',
                                          tag = '',
                                          name  = '',
                                          function_ASDU = 'C_SE_NB_1 (49)',
                                          addr_object = 32787 + i))
                    
                # Checking for the existence of a database
                TM_TR2.insert_many(list_tmtr2).execute()
            except Exception:
                msg[f'{today} - Таблица: tm_tr2, ошибка при заполнении: {traceback.format_exc()}'] = 2
                return msg
        msg[f'{today} - Таблица: tm_tr2 подготовлена'] = 1
        return(msg)
    def column_check(self):
        list_default = ['variable', 'tag', 'name', 'function_ASDU', 'addr_object', 'variable_change', 'descriptionTR4']
        msg = self.dop_function.column_check(TM_TR2, 'tm_tr2', list_default)
        return msg 
